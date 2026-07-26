#!/usr/bin/env python3
"""
LitCall 全库交叉审计 — 对比 Zotero / Obsidian / Excel / processed_log 四库 DOI 一致性。

用法:
    python audit_libraries.py           # 打印摘要报告
    python audit_libraries.py --json    # JSON 输出
    python audit_libraries.py --full    # 打印所有缺失条目详情
"""

import asyncio
import io
import json
import sys
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).parent.absolute()
sys.path.insert(0, str(SCRIPT_DIR))

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(
        sys.stdout.buffer, encoding="utf-8", errors="replace"
    )

from litcall.core.paths import PROCESSED_LOG, EXCEL_PATH, OBSIDIAN_DIR
from litcall.stores.zotero import ZoteroStore
from litcall.stores.obsidian import ObsidianStore
from litcall.stores.base import norm_doi

OUTPUT_JSON = "--json" in sys.argv
OUTPUT_FULL = "--full" in sys.argv


async def cross_library_audit() -> dict:
    """全库交叉审计。"""
    from litcall.core.config import config

    report = {
        "timestamp": datetime.now().isoformat(),
        "summary": {},
        "issues": [],
    }

    # ── 1. processed_log ──
    log_dois: set = set()
    try:
        if PROCESSED_LOG.exists():
            records = json.loads(PROCESSED_LOG.read_text(encoding="utf-8"))
            for rec in records:
                d = norm_doi(rec.get("doi", ""))
                if d:
                    log_dois.add(d)
    except Exception as e:
        print(f"审计: 读取 processed_log 失败: {e}")
    report["summary"]["processed_log"] = len(log_dois)

    # ── 2. Zotero (litcall 集合) ──
    try:
        zs = ZoteroStore()
        items = await zs.list_collection_items()
        zotero_dois = {norm_doi(i.get("doi", "")) for i in items if i.get("doi")}
        report["summary"]["zotero"] = len(zotero_dois)
    except Exception as e:
        print(f"审计: Zotero 查询失败: {e}")
        zotero_dois = set()
        report["summary"]["zotero"] = 0

    # ── 3. Excel ──
    excel_dois: set = set()
    try:
        if EXCEL_PATH.exists():
            import openpyxl
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active
            for row_idx in range(2, ws.max_row + 1):
                doi_val = norm_doi(str(ws.cell(row_idx, 10).value or ""))
                if doi_val:
                    excel_dois.add(doi_val)
            wb.close()
    except Exception as e:
        print(f"审计: Excel 读取失败: {e}")
    report["summary"]["excel"] = len(excel_dois)

    # ── 4. Obsidian ──
    try:
        os_store = ObsidianStore()
        obsidian_dois = os_store.list_dois()
        report["summary"]["obsidian"] = len(obsidian_dois)
    except Exception as e:
        print(f"审计: Obsidian 查询失败: {e}")
        obsidian_dois = set()
        report["summary"]["obsidian"] = 0

    # ── 5. 交叉比对 ──
    all_dois = log_dois | zotero_dois | excel_dois | obsidian_dois
    report["summary"]["union"] = len(all_dois)

    for doi in sorted(all_dois):
        missing = []
        if doi not in log_dois:
            missing.append("processed_log")
        if doi not in zotero_dois:
            missing.append("zotero")
        if doi not in excel_dois:
            missing.append("excel")
        if doi not in obsidian_dois:
            missing.append("obsidian")

        if missing:
            severity = "critical" if len(missing) >= 3 else (
                "warning" if len(missing) >= 2 else "info"
            )
            report["issues"].append({
                "doi": doi,
                "missing_stores": missing,
                "severity": severity,
            })

    report["summary"]["issues"] = len(report["issues"])
    report["summary"]["clean"] = len(all_dois) - len(report["issues"])
    return report


def main():
    result = asyncio.run(cross_library_audit())

    if OUTPUT_JSON:
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    s = result["summary"]
    print("=" * 70)
    print("  LitCall 全库交叉审计")
    print("=" * 70)
    print(f"  processed_log : {s.get('processed_log', 0):>5} 条")
    print(f"  Zotero        : {s.get('zotero', 0):>5} 条")
    print(f"  Excel         : {s.get('excel', 0):>5} 条")
    print(f"  Obsidian      : {s.get('obsidian', 0):>5} 条")
    print(f"  ─────────────────────────")
    print(f"  并集          : {s.get('union', 0):>5} DOI")
    print(f"  完全一致      : {s.get('clean', 0):>5} DOI")
    print(f"  存在问题      : {s.get('issues', 0):>5} DOI")
    print("=" * 70)

    if result["issues"]:
        critical = [i for i in result["issues"] if i["severity"] == "critical"]
        warnings = [i for i in result["issues"] if i["severity"] == "warning"]
        infos = [i for i in result["issues"] if i["severity"] == "info"]

        if critical:
            print(f"\n  🔴 严重 ({len(critical)}):")
            for i in critical:
                print(f"     {i['doi']} → 缺失: {', '.join(i['missing_stores'])}")

        if warnings and OUTPUT_FULL:
            print(f"\n  🟡 警告 ({len(warnings)}):")
            for i in warnings:
                print(f"     {i['doi']} → 缺失: {', '.join(i['missing_stores'])}")

        if infos and OUTPUT_FULL:
            print(f"\n  🔵 信息 ({len(infos)}):")
            for i in infos:
                print(f"     {i['doi']} → 缺失: {', '.join(i['missing_stores'])}")

    if s.get("issues", 0) == 0:
        print("\n  ✅ 四库完全一致！")
    print()


if __name__ == "__main__":
    main()
