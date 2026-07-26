"""ExcelStore — Excel 文献汇总表读写。

文件: litcall文献汇总.xlsx
功能: 读写、文件锁、自愈（从 Obsidian 恢复）

铁律 #1: Excel 损坏 → 从 Obsidian 重建（不用 processed_log）。
        从 processed_log 恢复会产生 DOI-only 骨架行（Bug #22）。
铁律 #8: 已存在 DOI → 填充空字段，不跳过。
"""

import logging
import os
import shutil
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from litcall.core.paths import EXCEL_PATH, EXCEL_LOCK_FILE, OBSIDIAN_DIR
from litcall.stores.base import AbstractStore, PaperData, VerifyResult, norm_doi
from litcall.stores.obsidian import ObsidianStore

logger = logging.getLogger(__name__)

# ── Excel 表头 ──
EXCEL_HEADERS = [
    "序号", "标题", "作者", "第一作者", "通讯作者", "年份", "期刊",
    "影响因子", "分区", "doi", "阅读方式", "阅读日期", "关键词",
    "研究背景与动机", "研究问题", "变量汇总", "研究方法",
    "方法论详解", "研究结果", "讨论与结论", "创新点",
    "局限与展望", "图表分析", "PDF路径", "入库时间",
]

# 字段 → 列索引
_FIELD_TO_COL = {h: i for i, h in enumerate(EXCEL_HEADERS)}


class ExcelStore(AbstractStore):
    """Excel 文献汇总表存储后端。"""

    name = "excel"

    def __init__(self, path: Optional[Path] = None):
        self._path = path or EXCEL_PATH
        self._lock_path = EXCEL_LOCK_FILE
        self._lock_path.parent.mkdir(parents=True, exist_ok=True)
        self._ensure_valid()

    # ── 文件锁 ──

    def _acquire_lock(self, timeout: float = 30.0) -> bool:
        """获取 Excel 文件锁（os.O_CREAT | os.O_EXCL）。"""
        deadline = time.time() + timeout
        while time.time() < deadline:
            try:
                fd = os.open(str(self._lock_path),
                             os.O_CREAT | os.O_EXCL | os.O_RDWR)
                os.close(fd)
                return True
            except OSError:
                time.sleep(0.5)
        return False

    def _release_lock(self):
        """释放 Excel 文件锁。"""
        try:
            self._lock_path.unlink(missing_ok=True)
        except Exception:
            pass

    # ── 有效性检查 + 自愈 ──

    def _validate(self) -> bool:
        """检测 Excel 文件是否可正常打开。"""
        try:
            import zipfile
            with zipfile.ZipFile(str(self._path), 'r') as zf:
                for name in zf.namelist():
                    zf.read(name)
            return True
        except Exception:
            return False

    def _rebuild_from_obsidian(self) -> bool:
        """从 Obsidian frontmatter 重建 Excel（铁律: 不从 processed_log 恢复）。"""
        if not OBSIDIAN_DIR.exists():
            return False
        try:
            md_files = list(OBSIDIAN_DIR.glob("**/*.md"))
            if not md_files:
                return False

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(EXCEL_HEADERS)

            seq = 1
            recovered = 0
            for md in md_files:
                try:
                    text = md.read_text(encoding="utf-8")
                    fm = ObsidianStore.parse_frontmatter(text)
                    doi = norm_doi(fm.get("doi", ""))
                    if not doi:
                        continue

                    row = ["" for _ in range(len(EXCEL_HEADERS))]
                    row[0] = str(seq)
                    # FM key → Excel column
                    mapping = {
                        "title": "标题", "author": "作者",
                        "first_author": "第一作者",
                        "corresponding_author": "通讯作者",
                        "year": "年份", "journal": "期刊",
                        "impact_factor": "影响因子", "quartile": "分区",
                        "doi": "doi", "keywords": "关键词",
                        "reading_method": "阅读方式",
                        "reading_date": "阅读日期",
                        "background": "研究背景与动机",
                        "research_question": "研究问题",
                        "variables": "变量汇总",
                        "method": "研究方法",
                        "method_details": "方法论详解",
                        "results": "研究结果",
                        "discussion": "讨论与结论",
                        "innovation": "创新点",
                        "limitations": "局限与展望",
                        "figure_analysis": "图表分析",
                        "pdf_path": "PDF路径",
                    }
                    for fm_k, excel_k in mapping.items():
                        if excel_k in _FIELD_TO_COL and fm_k in fm:
                            row[_FIELD_TO_COL[excel_k]] = fm[fm_k]
                    row[_FIELD_TO_COL["入库时间"]] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    ws.append(row)
                    seq += 1
                    recovered += 1
                except Exception:
                    continue

            if recovered == 0:
                wb.close()
                return False

            self._atomic_save(wb)
            wb.close()
            logger.info(f"Excel 从 Obsidian 重建成功: {recovered} 行")
            return True
        except Exception as e:
            logger.error(f"Excel 从 Obsidian 重建失败: {e}")
            return False

    def _ensure_valid(self):
        """确保 Excel 文件可用（损坏时自愈）。"""
        if not self._path.exists():
            # 创建空 Excel
            wb = openpyxl.Workbook()
            wb.active.append(EXCEL_HEADERS)
            self._atomic_save(wb)
            wb.close()
            return

        if not self._validate():
            logger.warning("Excel 文件损坏，尝试从 Obsidian 恢复...")
            damaged = Path(str(self._path) + f".damaged.{datetime.now():%Y%m%d_%H%M%S}")
            shutil.copy(self._path, damaged)
            logger.info(f"损坏文件已备份: {damaged}")

            if not self._rebuild_from_obsidian():
                # 最终回退：空白 Excel
                wb = openpyxl.Workbook()
                wb.active.append(EXCEL_HEADERS)
                self._atomic_save(wb)
                wb.close()
                logger.warning("无法重建，创建空白 Excel")

    def _atomic_save(self, wb: openpyxl.Workbook):
        """原子保存：先写临时文件再 replace。"""
        tmp = Path(str(self._path) + ".tmp")
        wb.save(str(tmp))
        tmp.replace(self._path)

    # ── AbstractStore 接口 ──

    async def write(self, paper: PaperData) -> bool:
        """写入或更新 Excel 行。

        铁律 #8: 已存在 DOI → 填充空字段，不跳过。
        """
        if not self._acquire_lock():
            logger.error("无法获取 Excel 文件锁")
            return False
        try:
            self._ensure_valid()
            wb = openpyxl.load_workbook(str(self._path))
            ws = wb.active

            ndoi = norm_doi(paper.doi)

            # 查找已有行
            existing_row = None
            doi_col = _FIELD_TO_COL.get("doi", 9)
            for row in ws.iter_rows(min_row=2, values_only=False):
                cell_doi = norm_doi(str(row[doi_col].value or ""))
                if cell_doi == ndoi:
                    existing_row = row
                    break

            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if existing_row:
                # 填充空字段
                mapping = {
                    "标题": paper.title, "作者": paper.authors,
                    "第一作者": paper.first_author,
                    "通讯作者": paper.corresponding_author,
                    "年份": paper.year, "期刊": paper.journal,
                    "影响因子": paper.impact_factor, "分区": paper.quartile,
                    "doi": paper.doi, "关键词": paper.keywords,
                    "阅读方式": paper.reading_method,
                    "阅读日期": paper.reading_date or now[:10],
                    "研究背景与动机": paper.background,
                    "研究问题": paper.research_question,
                    "变量汇总": paper.variables,
                    "研究方法": paper.method,
                    "方法论详解": paper.method_details,
                    "研究结果": paper.results,
                    "讨论与结论": paper.discussion,
                    "创新点": paper.innovation,
                    "局限与展望": paper.limitations,
                    "图表分析": paper.figure_analysis,
                    "PDF路径": Path(paper.pdf_path).name if paper.pdf_path else "",
                }
                for field, value in mapping.items():
                    if field in _FIELD_TO_COL and value:
                        cell = existing_row[_FIELD_TO_COL[field]]
                        if not cell.value:
                            cell.value = value
                existing_row[_FIELD_TO_COL["入库时间"]].value = now
                logger.info(f"Excel: 填充已有行 {ndoi}")
            else:
                # 新行
                row_data = ["" for _ in range(len(EXCEL_HEADERS))]
                row_data[0] = str(ws.max_row)  # 序号
                row_data[_FIELD_TO_COL["标题"]] = paper.title
                row_data[_FIELD_TO_COL["作者"]] = paper.authors
                row_data[_FIELD_TO_COL["第一作者"]] = paper.first_author
                row_data[_FIELD_TO_COL["通讯作者"]] = paper.corresponding_author
                row_data[_FIELD_TO_COL["年份"]] = paper.year
                row_data[_FIELD_TO_COL["期刊"]] = paper.journal
                row_data[_FIELD_TO_COL["影响因子"]] = paper.impact_factor
                row_data[_FIELD_TO_COL["分区"]] = paper.quartile
                row_data[_FIELD_TO_COL["doi"]] = paper.doi
                row_data[_FIELD_TO_COL["关键词"]] = paper.keywords
                row_data[_FIELD_TO_COL["阅读方式"]] = paper.reading_method
                row_data[_FIELD_TO_COL["阅读日期"]] = paper.reading_date or now[:10]
                row_data[_FIELD_TO_COL["研究背景与动机"]] = paper.background
                row_data[_FIELD_TO_COL["研究问题"]] = paper.research_question
                row_data[_FIELD_TO_COL["变量汇总"]] = paper.variables
                row_data[_FIELD_TO_COL["研究方法"]] = paper.method
                row_data[_FIELD_TO_COL["方法论详解"]] = paper.method_details
                row_data[_FIELD_TO_COL["研究结果"]] = paper.results
                row_data[_FIELD_TO_COL["讨论与结论"]] = paper.discussion
                row_data[_FIELD_TO_COL["创新点"]] = paper.innovation
                row_data[_FIELD_TO_COL["局限与展望"]] = paper.limitations
                row_data[_FIELD_TO_COL["图表分析"]] = paper.figure_analysis
                row_data[_FIELD_TO_COL["PDF路径"]] = Path(paper.pdf_path).name if paper.pdf_path else ""
                row_data[_FIELD_TO_COL["入库时间"]] = now
                ws.append(row_data)
                logger.info(f"Excel: 新增行 {ndoi}")

            self._atomic_save(wb)
            wb.close()
            return True
        except Exception as e:
            logger.error(f"ExcelStore.write 失败: {e}")
            return False
        finally:
            self._release_lock()

    async def verify(self, doi: str) -> VerifyResult:
        """验证 DOI 存在于 Excel 中。"""
        try:
            ndoi = norm_doi(doi)
            self._ensure_valid()
            wb = openpyxl.load_workbook(str(self._path), read_only=True)
            ws = wb.active
            doi_col = _FIELD_TO_COL.get("doi", 9)
            title_col = _FIELD_TO_COL.get("标题", 1)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if norm_doi(str(row[doi_col] or "")) == ndoi:
                    wb.close()
                    return VerifyResult(
                        store_name=self.name, ok=True,
                        found_doi=ndoi,
                        found_title=str(row[title_col] or ""),
                    )
            wb.close()
            return VerifyResult(
                store_name=self.name, ok=False,
                detail=f"DOI {ndoi} 不在 Excel 中",
            )
        except Exception as e:
            return VerifyResult(
                store_name=self.name, ok=False,
                detail=f"验证异常: {e}",
            )

    async def delete(self, doi: str) -> bool:
        """删除 Excel 行（用于回滚）。"""
        if not self._acquire_lock():
            return False
        try:
            ndoi = norm_doi(doi)
            self._ensure_valid()
            wb = openpyxl.load_workbook(str(self._path))
            ws = wb.active
            doi_col = _FIELD_TO_COL.get("doi", 9)
            for row in ws.iter_rows(min_row=2):
                if norm_doi(str(row[doi_col].value or "")) == ndoi:
                    ws.delete_rows(row[0].row)
                    self._atomic_save(wb)
                    wb.close()
                    return True
            wb.close()
            return False
        except Exception as e:
            logger.error(f"ExcelStore.delete 失败: {e}")
            return False
        finally:
            self._release_lock()

    def count(self) -> int:
        try:
            self._ensure_valid()
            wb = openpyxl.load_workbook(str(self._path), read_only=True)
            count = wb.active.max_row - 1  # 减去表头
            wb.close()
            return max(0, count)
        except Exception:
            return 0

    def list_dois(self) -> Set[str]:
        dois = set()
        try:
            self._ensure_valid()
            wb = openpyxl.load_workbook(str(self._path), read_only=True)
            ws = wb.active
            doi_col = _FIELD_TO_COL.get("doi", 9)
            for row in ws.iter_rows(min_row=2, values_only=True):
                d = norm_doi(str(row[doi_col] or ""))
                if d:
                    dois.add(d)
            wb.close()
        except Exception:
            pass
        return dois

    # ── 扩展方法 ──

    def rebuild(self) -> bool:
        """强制从 Obsidian 重建 Excel。"""
        return self._rebuild_from_obsidian()
