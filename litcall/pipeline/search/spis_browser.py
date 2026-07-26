"""
spis_browser.py — SPIS Playwright 浏览器自动化
打开 SPIS、执行搜索、翻页、详情页补全、VPN 检测。

所有 Playwright 浏览器操作集中于此模块。
页面内容解析在 spis_parser.py 中。
"""

import asyncio
import datetime
import json
import logging
import random
import re
import shutil
import subprocess as _sp
import sys
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

from litcall.core.config import config, load_config, save_config
from litcall.core.paths import BASE_DIR, PDF_DIR, CONFIG_PATH, PROCESSED_LOG, EXCEL_PATH
from litcall.pipeline.search.spis_parser import (
    scrape_page,
    _extract_from_detail,
    _check_detail_download_url,
    _auto_submit_literature_help,
    _parse_article_year,
)

logger = logging.getLogger(__name__)

# ============================================================================
# 模块级常量
# ============================================================================

MAX_PAGES_PER_KEYWORD = 10
MIN_SLEEP = 3.0
MAX_SLEEP = 8.0
ENRICH_FROM_DETAIL = True  # 是否从详情页采集完整元数据（较慢但完整）

BROWSER_USER_DATA = BASE_DIR / "browser_data"

# ============================================================================
# 键盘控制（Windows only）
# ============================================================================

_IS_WINDOWS = sys.platform == "win32"
_keyboard_state = "running"  # running | paused | terminated

if _IS_WINDOWS:
    import msvcrt


def _poll_keyboard():
    """非阻塞轮询键盘，检测 P / G / E 按键。仅 Windows。"""
    global _keyboard_state
    if not _IS_WINDOWS:
        return
    try:
        while msvcrt.kbhit():
            key = msvcrt.getch()
            if key in (b'p', b'P'):
                if _keyboard_state != "paused":
                    _keyboard_state = "paused"
                    print("\n⏸  [P] 检索已暂停！按 G 继续，按 E 终止")
            elif key in (b'g', b'G'):
                if _keyboard_state == "paused":
                    _keyboard_state = "running"
                    print("▶  [G] 检索已继续")
            elif key in (b'e', b'E'):
                _keyboard_state = "terminated"
                print("\n⏹  [E] 检索终止信号已接收，正在退出当前环节...")
    except Exception:
        pass


async def _kb_wait_if_paused():
    """暂停时阻塞等待，轮询 G 或 E。"""
    global _keyboard_state
    while _keyboard_state == "paused":
        await asyncio.sleep(0.3)
        _poll_keyboard()
    if _keyboard_state == "terminated":
        raise KeyboardInterrupt("用户按 E 终止")


def _kb_is_terminated() -> bool:
    return _keyboard_state == "terminated"


# ============================================================================
# Agent 信号文件
# ============================================================================

class AgentSignalError(Exception):
    """Agent 收到终止信号时抛出。调用方应捕获并优雅退出。"""
    pass


async def _check_signal_files(run_logger=None) -> str:
    """检查 运行日志/.pause 和 .terminate 信号文件。

    由 Streamlit 写入信号，Worker 子进程轮询检测：
      - .terminate → 标记 run_logger status=terminated → 抛 AgentSignalError
      - .pause     → 标记 run_logger _paused=true → 阻塞轮询直到 .pause 消失或 .terminate 出现
      - 无信号     → 返回 "ok"

    Args:
        run_logger: AgentRunLogger 实例（可选），用于实时更新运行状态 JSON
    Returns:
        "ok" | "terminated" | "resumed"
    Raises:
        AgentSignalError: 收到终止信号
    """
    from litcall.core.paths import BASE_DIR as _BASE_DIR
    pause_file = _BASE_DIR / "运行日志" / ".pause"
    terminate_file = _BASE_DIR / "运行日志" / ".terminate"

    if terminate_file.exists():
        if run_logger:
            run_logger._data["status"] = "terminated"
            run_logger._save()
        terminate_file.unlink(missing_ok=True)
        pause_file.unlink(missing_ok=True)
        logger.info("⏹ Agent 已终止 (收到 .terminate 信号)")
        raise AgentSignalError("用户终止")

    if pause_file.exists():
        if run_logger:
            run_logger._data["_paused"] = True
            run_logger._save()
        logger.info("⏸ Agent 已暂停 (检测到 .pause 信号文件)")
        while pause_file.exists():
            await asyncio.sleep(1)
            if terminate_file.exists():
                if run_logger:
                    run_logger._data["status"] = "terminated"
                    run_logger._save()
                terminate_file.unlink(missing_ok=True)
                pause_file.unlink(missing_ok=True)
                logger.info("⏹ Agent 在暂停期间被终止")
                raise AgentSignalError("用户在暂停期间终止")
        if run_logger:
            run_logger._data["_paused"] = False
            run_logger._save()
        logger.info("▶ Agent 已继续 (.pause 信号文件已移除)")
        return "resumed"

    return "ok"


# ============================================================================
# 关键词游标管理
# ============================================================================

def _get_category_index(scope: str, flat_index: int) -> tuple:
    """将 scope 标签 + 扁平索引转换为 (category_index, keyword_index_within_category)。"""
    if scope == "宽":
        return (0, flat_index)
    elif scope == "窄":
        return (1, flat_index - len(config.get("keywords", {}).get("broad", [])))
    elif scope == "中":
        broad_len = len(config.get("keywords", {}).get("broad", []))
        narrow_len = len(config.get("keywords", {}).get("narrow", []))
        return (2, flat_index - broad_len - narrow_len)
    return (0, flat_index)


def _read_keyword_cursor() -> dict:
    """从 config.json 读取关键词游标。"""
    return config.get("_keyword_cursor", {"category_index": 0, "keyword_index": 0})


def _write_keyword_cursor(category_index: int, keyword_index: int):
    """原子写入关键词游标到 config.json（临时文件 + rename 防并发写坏）。"""
    cfg = load_config()
    cfg["_keyword_cursor"] = {
        "category_index": category_index,
        "keyword_index": keyword_index,
        "last_updated": datetime.datetime.now().isoformat(),
    }
    save_config(cfg)


def _get_next_keyword(cursor: dict) -> tuple:
    """从游标位置获取下一个关键词。

    Returns:
        (keyword_str, scope_tag, category_index, keyword_index)
        全部穷尽时返回 (None, None, -1, -1)
    """
    categories = [
        ("宽", config.get("keywords", {}).get("broad", [])),
        ("窄", config.get("keywords", {}).get("narrow", [])),
        ("中", config.get("keywords", {}).get("chinese", [])),
    ]

    ci = cursor.get("category_index", 0)
    ki = cursor.get("keyword_index", 0)

    while ci < len(categories):
        scope, kw_list = categories[ci]
        if ki < len(kw_list):
            return (kw_list[ki], scope, ci, ki)
        ci += 1
        ki = 0

    return (None, None, -1, -1)


# ============================================================================
# 关键词打印
# ============================================================================

def print_keywords() -> List[tuple]:
    """打印关键词矩阵并返回扁平列表 [(keyword, scope_tag), ...]."""
    keywords = config.get("keywords", {})
    broad = keywords.get("broad", [])
    narrow = keywords.get("narrow", [])
    chinese = keywords.get("chinese", [])
    all_kw = [(kw, "宽") for kw in broad] + [(kw, "窄") for kw in narrow] + [(kw, "中") for kw in chinese]
    print("\n===== 关键词矩阵 =====")
    for idx, (kw, scope) in enumerate(all_kw, 1):
        print(f"[{idx}] ({scope}) {kw}")
    print("========================\n")
    return all_kw


# ============================================================================
# 去重数据加载
# ============================================================================

def load_processed_log() -> Tuple[Set[str], Set[str]]:
    """读取 processed_log.json，返回 (doi_set, title_set)。"""
    from litcall.pipeline.search.dedup import normalize_title

    dois: Set[str] = set()
    titles: Set[str] = set()
    if not PROCESSED_LOG.exists():
        return dois, titles
    try:
        with open(PROCESSED_LOG, "r", encoding="utf-8") as f:
            records = json.load(f)
        for rec in records:
            doi = rec.get("doi", "").strip().lower()
            if doi:
                dois.add(doi)
            title = normalize_title(rec.get("title", ""))
            if title:
                titles.add(title)
    except Exception as e:
        logger.warning(f"读取处理日志失败: {e}")
    return dois, titles


def _load_excel_dois() -> set[str]:
    """读取 Excel 中已有的所有 DOI，用于去重过滤。"""
    import openpyxl
    dois: set[str] = set()
    DOI_COL = 9  # NOTE_FIELDS 中 "doi" 的 0-based 索引
    if not EXCEL_PATH.exists():
        return dois
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        for row_idx in range(2, ws.max_row + 1):
            doi_val = str(ws.cell(row_idx, DOI_COL + 1).value or "").strip().lower()
            if doi_val:
                dois.add(doi_val)
        wb.close()
    except Exception as e:
        logger.warning(f"读取 Excel DOI 失败: {e}")
    return dois


def _load_excel_titles() -> set[str]:
    """读取 Excel 中已有的所有标题（normalize 后），用于去重过滤。"""
    import openpyxl
    from litcall.pipeline.search.dedup import normalize_title

    titles: set[str] = set()
    TITLE_COL = 1  # NOTE_FIELDS 中 "标题" 的 0-based 索引
    if not EXCEL_PATH.exists():
        return titles
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        for row_idx in range(2, ws.max_row + 1):
            title_val = str(ws.cell(row_idx, TITLE_COL + 1).value or "").strip()
            if title_val:
                titles.add(normalize_title(title_val))
        wb.close()
    except Exception as e:
        logger.warning(f"读取 Excel 标题失败: {e}")
    return titles


# ============================================================================
# Zotero 去重数据获取
# ============================================================================

_ZOTERO_COLLECTION_KEY_CACHE: Optional[str] = None


async def _get_zotero_collection_key() -> Optional[str]:
    """获取 Zotero 中 litcall 集合的 key（带缓存）。"""
    global _ZOTERO_COLLECTION_KEY_CACHE
    if _ZOTERO_COLLECTION_KEY_CACHE:
        return _ZOTERO_COLLECTION_KEY_CACHE

    zotero = config.get("zotero", {})
    user_id = zotero.get("user_id")
    api_key = zotero.get("api_key")
    collection_name = zotero.get("collection_name", "litcall")
    if not user_id or not api_key:
        return None

    base = f"https://api.zotero.org/users/{user_id}"
    headers = {"Zotero-API-Key": api_key}
    coll_url = f"{base}/collections"
    resp = await _async_request("GET", coll_url, headers=headers)
    if resp.status_code == 403:
        logger.error("Zotero API 403 Forbidden — 请检查:")
        logger.error("  1. API Key 是否正确: https://www.zotero.org/settings/keys")
        logger.error("  2. 创建 Key 时是否勾选了「Allow library access」和「Allow write access」")
        logger.error("  3. user_id 是否为纯数字（在 Zotero 设置页面可找到）")
        logger.error(f"  当前配置: user_id={user_id}, api_key={api_key[:6]}...")
        return None
    if resp.status_code != 200:
        logger.error(f"获取 Zotero 集合列表失败: {resp.status_code}")
        return None

    collections = resp.json()
    matching = [c for c in collections if c["data"]["name"] == collection_name]

    if len(matching) == 1:
        _ZOTERO_COLLECTION_KEY_CACHE = matching[0]["key"]
        return _ZOTERO_COLLECTION_KEY_CACHE

    if len(matching) > 1:
        logger.warning(f"发现 {len(matching)} 个同名 collection '{collection_name}'，选择条目最多的")
        best = None
        best_items = -1
        for c in matching:
            num_items = c.get("meta", {}).get("numItems", 0)
            logger.info(f"  {c['key']}: {num_items} items")
            if num_items > best_items:
                best_items = num_items
                best = c
        if best:
            _ZOTERO_COLLECTION_KEY_CACHE = best["key"]
            logger.info(f"  选择 {best['key']} ({best_items} items)")
            return _ZOTERO_COLLECTION_KEY_CACHE

    logger.info(f"集合 '{collection_name}' 不存在，正在创建...")
    create_resp = await _async_request("POST", coll_url, headers=headers,
                                       json=[{"name": collection_name}])
    if create_resp.status_code == 200:
        resp2 = await _async_request("GET", coll_url, headers=headers)
        if resp2.status_code == 200:
            for c in resp2.json():
                if c["data"]["name"] == collection_name:
                    _ZOTERO_COLLECTION_KEY_CACHE = c["key"]
                    return _ZOTERO_COLLECTION_KEY_CACHE
    logger.error("创建 Zotero 集合失败")
    return None


async def fetch_zotero_existing_dois() -> Tuple[Set[str], Set[str]]:
    """从 Zotero 获取 litcall 集合中已有文献的 DOI 和标题集合（用于去重）。"""
    from litcall.pipeline.search.dedup import normalize_title

    zotero = config.get("zotero", {})
    user_id = zotero.get("user_id")
    api_key = zotero.get("api_key")
    if not user_id or not api_key:
        logger.error("Zotero 凭据未配置")
        return set(), set()

    coll_key = await _get_zotero_collection_key()
    if not coll_key:
        return set(), set()

    base = f"https://api.zotero.org/users/{user_id}"
    headers = {"Zotero-API-Key": api_key}
    items = []
    start = 0
    while True:
        url = f"{base}/collections/{coll_key}/items?limit=100&start={start}"
        resp = await _async_request("GET", url, headers=headers)
        if resp.status_code != 200:
            break
        batch = resp.json()
        if not batch:
            break
        items.extend(batch)
        start += 100

    dois: Set[str] = set()
    titles: Set[str] = set()
    for item in items:
        data = item.get("data", {})
        doi = data.get("DOI", "").strip().lower()
        if doi:
            dois.add(doi)
        title = normalize_title(data.get("title", ""))
        if title:
            titles.add(title)
    return dois, titles


# ============================================================================
# HTTP 工具
# ============================================================================

def _safe_http_request(method: str, url: str, **kwargs):
    """同步 HTTP 请求（带 SSL 回退）。"""
    import requests
    kwargs.setdefault("timeout", 30)
    try:
        return requests.request(method, url, **kwargs)
    except requests.exceptions.SSLError:
        kwargs["verify"] = False
        return requests.request(method, url, **kwargs)
    except requests.exceptions.ConnectionError as e:
        raise ConnectionError(f"无法连接到 {url}: {e}") from e
    except requests.exceptions.Timeout as e:
        raise TimeoutError(f"请求超时 {url}: {e}") from e


async def _async_request(method: str, url: str, **kwargs):
    """异步 HTTP 请求（在线程池中执行同步请求）。"""
    return await asyncio.to_thread(_safe_http_request, method, url, **kwargs)


# ============================================================================
# Chrome 僵尸进程清理
# ============================================================================

def _kill_stale_chrome():
    """清理上次崩溃可能留下的僵尸 Chrome/Chromium 进程 + 锁文件，避免 Playwright 无法创建新标签页。"""
    if not _IS_WINDOWS:
        return
    try:
        killed = 0
        # 用 cmd.exe /c 执行，避免 Git Bash 的路径翻译干扰 /F /IM 等参数
        for exe in ('chrome.exe', 'chromium.exe'):
            try:
                result = _sp.run(
                    ['cmd.exe', '/c', f'taskkill /F /IM {exe}'],
                    capture_output=True, text=True, timeout=10,
                )
                if result.returncode == 0:
                    for line in result.stdout.splitlines():
                        if 'SUCCESS' in line:
                            killed += 1
            except Exception:
                pass
        if killed > 0:
            logger.info(f"清理 {killed} 个残留 Chrome/Chromium 进程")

        # 清理 browser_data 目录中的 Chrome 锁文件（SingletonLock/SingletonSocket）
        browser_data = BASE_DIR / "browser_data"
        if browser_data.exists():
            for lock_name in ('SingletonLock', 'SingletonSocket', 'lockfile'):
                lock_path = browser_data / lock_name
                if lock_path.exists():
                    try:
                        lock_path.unlink()
                        logger.info(f"清理 Chrome 锁文件: {lock_path}")
                    except Exception:
                        pass
            # Default 子目录中也可能有锁
            default_dir = browser_data / "Default"
            if default_dir.exists():
                for lock_name in ('SingletonLock', 'SingletonSocket'):
                    lock_path = default_dir / lock_name
                    if lock_path.exists():
                        try:
                            lock_path.unlink()
                        except Exception:
                            pass
    except Exception:
        pass  # 清理失败不影响主流程


# ═══════════════════════════════════════════════════════════════════
# 浏览器控制：打开 SPIS & 搜索
# ═══════════════════════════════════════════════════════════════════

async def open_spis_autonomous(
    keyword: str,
    year_start: int = 2025,
    year_end: int = 2026,
    headless: bool = False,
) -> tuple:
    """【Agent 模式】自主打开 SPIS 并执行搜索 — 无需人工干预。

    关键特性：
    - 使用 persistent context 保存登录 cookies（首次需手动登录）
    - 自动在搜索框输入关键词并点击搜索
    - 自动设置年份筛选（2025-2026）
    - 登录态过期时自动提示

    Returns: (playwright, browser_context, page)
    """
    _kill_stale_chrome()
    BROWSER_USER_DATA.mkdir(parents=True, exist_ok=True)

    p = await async_playwright().start()
    try:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(BROWSER_USER_DATA),
            headless=headless,
            viewport={"width": 1280, "height": 800},
            locale="zh-CN",
        )
    except Exception as e:
        logger.error(f"浏览器启动失败: {e}")
        await p.stop()
        raise

    page = await context.new_page()

    try:
        # ── 导航到 SPIS ──
        try:
            await page.goto("https://spis.hnlat.com", timeout=120000, wait_until="domcontentloaded")
        except Exception as e:
            logger.error(f"无法打开 SPIS: {e}")
            await context.close()
            await p.stop()
            raise SystemExit(1)

        await asyncio.sleep(random.uniform(2.0, 3.0))

        # ── 检测是否需要登录（VPN 断开时 SPIS 显示登录页）──
        login_needed = await page.evaluate("""() => {
            const body = document.body.innerText || '';
            // VPN未连接时 SPIS 显示登录页，特征文字：
            //   "账号登录" "手机号登录" "微信扫码" "第三方" "当前IP"
            const loginMarkers = ['账号登录', '手机号登录', '微信扫码', '当前IP'];
            const found = loginMarkers.filter(m => body.includes(m));
            if (found.length >= 2) return true;  // 至少匹配2个 → 确认是登录页
            if (window.location.href.includes('login') || window.location.href.includes('auth')) return true;
            return false;
        }""")

        if login_needed:
            logger.warning("SPIS 显示登录页（VPN 可能已断开）！请手动连接 VPN。")
            logger.info("   连接 VPN 后自动继续... 等待最多 5 分钟。")
            # 等待 VPN 恢复（轮询检测页面变化）
            for _ in range(60):
                await asyncio.sleep(5)
                still_login = await page.evaluate("""() => {
                    const body = document.body.innerText || '';
                    const loginMarkers = ['账号登录', '手机号登录', '微信扫码', '当前IP'];
                    const found = loginMarkers.filter(m => body.includes(m));
                    return found.length >= 2 || window.location.href.includes('login');
                }""")
                if not still_login:
                    logger.info("VPN 已恢复，继续...")
                    await asyncio.sleep(2)
                    break
            else:
                logger.error("VPN 恢复超时（5分钟），跳过该关键词。")
                await context.close()
                await p.stop()
                return {
                    "collected": [], "by_keyword": {}, "with_links": [],
                    "without_links": [], "help_submitted": 0,
                }

        # ── 等待搜索页面加载 ──
        await asyncio.sleep(random.uniform(1.5, 2.5))

        # ── 输入关键词 ──
        search_input = None
        for selector in [
            ".spis-search-box .search-c .input .ant-input",
            ".spis-search-box .ant-input",
            ".search-c input.ant-input",
            "input.ant-input[type='text']",
        ]:
            try:
                search_input = await page.wait_for_selector(selector, timeout=5000)
                if search_input:
                    logger.info(f"  搜索框定位成功: {selector}")
                    break
            except Exception:
                continue

        if not search_input:
            logger.error("找不到 SPIS 搜索输入框，请确认页面已加载且 DOM 未变更。")
            await context.close()
            await p.stop()
            raise RuntimeError("搜索输入框未找到")

        # 清空 + 输入关键词（模拟真人打字节奏）
        await search_input.click()
        await asyncio.sleep(random.uniform(0.3, 0.6))
        await search_input.fill("")
        await asyncio.sleep(random.uniform(0.2, 0.4))
        await search_input.type(keyword, delay=random.randint(60, 100))
        logger.info(f"  已输入关键词: {keyword}")
        await asyncio.sleep(random.uniform(0.8, 1.2))

        # ── 尝试设置年份筛选 ──
        year_filter_set = await _set_year_filter_autonomous(page, year_start, year_end)

        # ── 点击搜索按钮 ──
        search_clicked = False
        for selector in [
            ".spis-search-box .search-button .ant-btn",
            ".spis-search-box .search-button",
            ".search-button .ant-btn",
            "button.search-button",
        ]:
            try:
                btn = await page.query_selector(selector)
                if btn:
                    await btn.click()
                    logger.info(f"  已点击搜索按钮: {selector}")
                    search_clicked = True
                    break
            except Exception:
                continue

        if not search_clicked:
            # 回退：按 Enter 触发搜索
            logger.info("  搜索按钮未找到，尝试按 Enter 触发搜索...")
            await search_input.press("Enter")

        # ── 等待搜索结果加载 ──
        try:
            await page.wait_for_selector("article.article", timeout=20000)
            logger.info("  搜索结果已加载")
        except Exception:
            logger.warning("  搜索结果加载超时，页面可能无结果或结构变更")

        await asyncio.sleep(random.uniform(1.0, 2.0))

        # 统计结果数
        try:
            article_count = await page.evaluate(
                "() => document.querySelectorAll('article.article').length"
            )
            logger.info(f"  当前页显示 {article_count} 篇文章")
        except Exception:
            pass

        return p, context, page

    except Exception as e:
        logger.error(f"自主搜索异常: {e}")
        try:
            await context.close()
        except Exception:
            pass
        try:
            await p.stop()
        except Exception:
            pass
        raise


async def _set_year_filter_autonomous(page, year_start: int, year_end: int) -> bool:
    """尝试在 SPIS 高级搜索中设置年份筛选。成功返回 True。"""
    try:
        # ── 尝试打开高级搜索弹窗 ──
        advanced_btn = None
        for sel in [
            ".spis-search-box .button .ant-btn",
            ".spis-search-box button",
            "button:has-text('高级')",
            "button:has-text('专业')",
        ]:
            try:
                btns = await page.query_selector_all(sel)
                for b in btns:
                    text = (await b.inner_text()).strip()
                    if "高级" in text or "专业" in text:
                        advanced_btn = b
                        break
                if advanced_btn:
                    break
            except Exception:
                continue

        if advanced_btn:
            await advanced_btn.click()
            await asyncio.sleep(random.uniform(1.0, 1.5))
            logger.info("  已打开高级搜索")

            # ── 找年份选择器 ──
            year_picker = None
            for sel in [
                ".search-modal-box .year-row .ant-picker",
                ".search-filter .year-row .ant-picker",
                ".ant-picker",
            ]:
                try:
                    pickers = await page.query_selector_all(sel)
                    for picker in pickers:
                        if await picker.is_visible():
                            year_picker = picker
                            break
                    if year_picker:
                        break
                except Exception:
                    continue

            if year_picker:
                await year_picker.click()
                await asyncio.sleep(random.uniform(0.5, 1.0))
                # 尝试设置年份范围
                try:
                    # 清空并输入起始年
                    await page.keyboard.press("Control+a")
                    await asyncio.sleep(0.2)
                    year_text = f"{year_start}-{year_end}"
                    await page.keyboard.type(year_text, delay=60)
                    await asyncio.sleep(0.5)
                    await page.keyboard.press("Enter")
                    logger.info(f"  已设置年份筛选: {year_text}")
                    # 关闭高级搜索弹窗
                    await asyncio.sleep(0.5)
                    await page.keyboard.press("Escape")
                    await asyncio.sleep(0.5)
                    return True
                except Exception as e:
                    logger.warning(f"  年份选择器交互异常: {e}")

    except Exception as e:
        logger.warning(f"  年份筛选设置失败: {e}（将使用后置过滤）")

    return False


async def open_spis_and_wait(keyword: str, index: int, total: int):
    """【人工模式】打开 SPIS 并等待用户手动完成搜索后按回车。"""
    _kill_stale_chrome()
    p = await async_playwright().start()
    browser = await p.chromium.launch(headless=False)
    context = await browser.new_context()
    page = await context.new_page()
    try:
        await page.goto("https://spis.hnlat.com", timeout=120000)
    except Exception as e:
        print(f"\n[ERROR] 无法打开 SPIS 网站: {e}")
        print("可能原因：")
        print("  1. 未连接校园 VPN → 请先连接 VPN 后重试")
        print("  2. SPIS 服务器暂时不可用 → 稍后重试")
        print("  3. 网络不通 → 在浏览器中手动访问 https://spis.hnlat.com 确认")
        await browser.close()
        await p.stop()
        raise SystemExit(1)
    print("\n" + "─" * 50)
    print(f">>> 人工环节：关键词组合 {index}/{total} <<<")
    print(f"请搜索: {keyword}")
    print("1. 如尚未登录，请手动登录 SPIS")
    print("2. 将关键词粘贴到搜索框，执行搜索")
    print("3. 确认已看到搜索结果列表后，按回车继续")
    print("─" * 50)
    input()
    return p, browser, context, page


# ═══════════════════════════════════════════════════════════════════
# 翻页
# ═══════════════════════════════════════════════════════════════════

async def _click_next_page(page, current_pg: int) -> bool:
    """翻到下一页。成功返回 True，失败（已到最后一页）返回 False。"""
    # 取当前页码
    current_num = current_pg + 1
    try:
        active_el = await page.query_selector("button.pagination-number.active")
        if active_el:
            txt = (await active_el.inner_text()).strip()
            if txt.isdigit():
                current_num = int(txt)
    except Exception:
        pass

    next_num = current_num + 1

    # ── 方式1: '>' 箭头 ──
    all_arrows = await page.query_selector_all("button.pagination-arrow")
    for arrow in all_arrows:
        try:
            txt = (await arrow.inner_text()).strip()
            if txt == ">":
                cls = (await arrow.get_attribute("class") or "").lower()
                if "first" not in cls and "disabled" not in cls:
                    await arrow.scroll_into_view_if_needed()
                    await asyncio.sleep(0.3)
                    await arrow.click()
                    # 轮询等激活
                    for _ in range(20):
                        await asyncio.sleep(0.5)
                        new_active = await page.query_selector("button.pagination-number.active")
                        if new_active:
                            new_txt = (await new_active.inner_text()).strip()
                            if new_txt.isdigit() and int(new_txt) == next_num:
                                logger.info(f"  -> 翻页: {current_num} -> {new_txt}")
                                await asyncio.sleep(1)
                                return True
                    logger.warning(f"  箭头点击后未检测到页码变化")
        except Exception:
            continue

    # ── 方式2: 页码按钮 ──
    all_nums = await page.query_selector_all("button.pagination-number")
    for btn in all_nums:
        try:
            txt = (await btn.inner_text()).strip()
            cls = (await btn.get_attribute("class") or "").lower()
            if txt == str(next_num) and "active" not in cls:
                await btn.scroll_into_view_if_needed()
                await asyncio.sleep(0.3)
                await btn.click()
                for _ in range(20):
                    await asyncio.sleep(0.5)
                    new_active = await page.query_selector("button.pagination-number.active")
                    if new_active:
                        new_txt = (await new_active.inner_text()).strip()
                        if new_txt.isdigit() and int(new_txt) == next_num:
                            logger.info(f"  -> 翻页: {current_num} -> {new_txt}")
                            await asyncio.sleep(1)
                            return True
        except Exception:
            continue

    logger.info(f"  翻页失败（第{current_num}页可能是最后一页）")
    return False


# ═══════════════════════════════════════════════════════════════════
# VPN 检测
# ═══════════════════════════════════════════════════════════════════

async def _check_vpn_connected(page) -> bool:
    """检测 SPIS 是否已通过 VPN 连通（学校 IP 自动识别）。

    返回 True = VPN 连通，False = SPIS 显示登录页或不可达。
    关键：先验证页面是否真的加载了 SPIS，避免把空白错误页误判为已连通。
    """
    try:
        # ── 第0步：检查页面 URL 是否在 SPIS 域名 ──
        url = (await page.evaluate("() => window.location.href")).strip()
        if not url or "spis" not in url.lower():
            return False  # 页面根本没加载到 SPIS（空白页/错误页/超时）

        # ── 第1步：正向检测 SPIS 特征（学术检索平台的特征文字）──
        is_spis = await page.evaluate("""() => {
            const body = document.body.innerText || '';
            const spisMarkers = ['SPIS', '学术', '高级检索', 'scholar'];
            return spisMarkers.some(m => body.includes(m));
        }""")
        if not is_spis:
            return False  # 没有 SPIS 特征 → 不是 SPIS 页面

        # ── 第2步：排除登录页 ──
        is_login_page = await page.evaluate("""() => {
            const body = document.body.innerText || '';
            const loginMarkers = ['账号登录', '手机号登录', '微信扫码', '当前IP'];
            const found = loginMarkers.filter(m => body.includes(m));
            return found.length >= 2;
        }""")
        if is_login_page:
            return False  # 确认是登录页 → VPN 未连

        # ── 第3步：有 SPIS 特征且不是登录页 → VPN 已连通 ──
        return True
    except Exception:
        return False


async def _wait_for_vpn(
    headless: bool = False,
    timeout: int = 300,
    progress_callback=None,
) -> tuple:
    """等待 VPN 连接。每 15 秒检测一次 SPIS 是否可访问。

    Args:
        headless: 是否无头模式
        timeout: 最长等待秒数（默认 300 = 5 分钟）
        progress_callback: Streamlit 进度回调

    Returns:
        (connected: bool, browser_page, browser_context, playwright_instance)
        - connected=True  → 复用返回的 page/context/playwright（VPN 已通）
        - connected=False → 全部为 None（超时，调用方应关闭资源）

    调用方负责关闭返回的 context/playwright。
    """
    from datetime import datetime
    from playwright.async_api import async_playwright as _async_playwright

    logger.info("检测 VPN 连接状态（SPIS 学校 IP 认证）...")
    if progress_callback:
        progress_callback("vpn_check", "检测 VPN 连接状态...", {"status": "checking"})

    # ── 清理残留 Chrome 进程 ──
    _kill_stale_chrome()

    playwright = None
    browser = None
    context = None
    page = None

    try:
        playwright = await _async_playwright().start()

        # VPN 检测用普通 launch（可靠，不需要 cookie 持久化）
        browser = await playwright.chromium.launch(
            headless=headless,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
        )
        context = await browser.new_context()
        page = await context.new_page()
    except Exception as e:
        logger.warning(f"Chrome 启动失败: {type(e).__name__}: {e}")
        logger.warning("  跳过 VPN 检测，直接进入 Phase 2（深度阅读已有 PDF）")
        # 清理已分配的资源
        if context is not None:
            try:
                await context.close()
            except Exception:
                pass
        if browser is not None:
            try:
                await browser.close()
            except Exception:
                pass
        if playwright is not None:
            try:
                await playwright.stop()
            except Exception:
                pass
        return False, None, None, None

    # 首次打开 SPIS
    try:
        await page.goto(
            "https://spis.hnlat.com",
            timeout=60000,
            wait_until="domcontentloaded",
        )
    except Exception as e:
        logger.warning(f"SPIS 无法访问: {e}")

    start = datetime.now()
    deadline = start.timestamp() + timeout

    while datetime.now().timestamp() < deadline:
        # ── 信号检查：允许用户在 VPN 等待期间暂停/终止 ──
        try:
            await _check_signal_files(run_logger=None)
        except AgentSignalError:
            logger.info("VPN 等待期间收到终止信号")
            try:
                await page.close()
                await context.close()
            except Exception:
                pass
            raise

        connected = await _check_vpn_connected(page)
        elapsed = (datetime.now() - start).total_seconds()

        if connected:
            logger.info(f"VPN 已连接，SPIS 可访问（耗时 {elapsed:.0f} 秒）")
            if progress_callback:
                progress_callback("vpn_ok", f"VPN 已连接（{elapsed:.0f} 秒）", {"status": "connected"})
            return True, page, context, playwright

        remaining = timeout - elapsed
        logger.info(f"VPN 未连接 — SPIS 显示登录页，{remaining:.0f} 秒后超时跳过检索")
        if progress_callback:
            progress_callback(
                "vpn_waiting",
                f"等待 VPN 连接... {remaining:.0f} 秒后超时",
                {"status": "waiting", "remaining": int(remaining)},
            )

        # 每 15 秒刷新页面重试
        await asyncio.sleep(15)
        try:
            await page.reload(timeout=30000, wait_until="domcontentloaded")
        except Exception:
            try:
                await page.goto(
                    "https://spis.hnlat.com",
                    timeout=30000,
                    wait_until="domcontentloaded",
                )
            except Exception:
                pass  # 网络不通，继续等

    # 超时
    logger.warning("VPN 连接超时（5分钟），跳过 Phase 1 检索，继续 Phase 2+3")
    if progress_callback:
        progress_callback("vpn_timeout", "VPN 超时 — 跳过检索", {"status": "timeout"})

    try:
        await page.close()
        await context.close()
    except Exception:
        pass

    return False, None, None, None


# ═══════════════════════════════════════════════════════════════════
# 详情页补全
# ═══════════════════════════════════════════════════════════════════

async def _enrich_articles_autonomous(page, articles: List[Dict[str, str]], run_logger=None) -> List[Dict[str, str]]:
    """【Agent 模式】详情页补全 + 下载链接检测 + 文献求助，无需键盘交互。

    run_logger: AgentRunLogger 实例，用于信号文件检查（暂停/终止）。
    """
    from litcall.pipeline.search.journal_filter import (
        journal_in_whitelist,
        _find_whitelist_journal_in_text,
        _resolve_truncated_journal,
    )

    if not articles:
        return articles

    # 残缺标题优先
    garbled = [a for a in articles if a.get("is_garbled")]
    normal = [a for a in articles if not a.get("is_garbled")]
    articles[:] = garbled + normal
    if garbled:
        logger.info(f"  详情页补全 ({len(articles)} 篇, {len(garbled)} 篇残缺优先)...")
    else:
        logger.info(f"  详情页补全 ({len(articles)} 篇)...")

    enriched_list = []
    list_url = page.url
    help_email = config.get("help_email", "18922596828@163.com")

    for i, art in enumerate(articles):
        # ── 信号检查：每篇文章处理前检查暂停/终止 ──
        if run_logger:
            try:
                await _check_signal_files(run_logger)
            except AgentSignalError:
                logger.info(f"详情补全收到终止信号，已处理 {len(enriched_list)}/{len(articles)} 篇")
                raise

        logger.info(f"  详情 [{i+1}/{len(articles)}]: {art['title'][:60]}...")
        enriched = dict(art)

        # 恢复列表页 DOM
        try:
            await page.wait_for_selector("article.article", timeout=10000)
        except Exception:
            enriched_list.append(enriched)
            continue

        article_els = await page.query_selector_all("article.article")

        # 标题指纹匹配
        art_el = None
        target_title = art["title"]
        for el in article_els:
            try:
                el_text = await el.inner_text()
                if target_title[:40] in el_text.strip()[:80].replace("\n", " "):
                    art_el = el
                    break
            except Exception:
                continue
        if not art_el:
            if i < len(article_els):
                art_el = article_els[i]
            else:
                enriched_list.append(enriched)
                continue

        current_url = page.url

        # 找可点击元素
        click_el = None
        for sel in ["div.d-t.jump", "div.allow-ai", "div.jump", "div.d-t"]:
            try:
                el = await art_el.query_selector(sel)
                if el:
                    click_el = el
                    break
            except Exception:
                continue
        if not click_el:
            click_el = art_el

        # ── 点击 + 轮询 ──
        new_page = None
        url_changed = False
        try:
            async with page.context.expect_page(timeout=5000) as new_page_info:
                await click_el.click(force=True)
                for _ in range(15):
                    await asyncio.sleep(0.5)
                    try:
                        if page.url != current_url:
                            url_changed = True
                            break
                    except Exception:
                        pass
                if not url_changed:
                    try:
                        new_page = await new_page_info.value
                    except Exception:
                        pass
        except Exception:
            pass

        if not url_changed and not new_page:
            for _ in range(20):
                await asyncio.sleep(0.5)
                try:
                    if page.url != current_url:
                        url_changed = True
                        break
                except Exception:
                    pass

        # ── 提取数据 ──
        detail_tab = new_page if new_page else page
        if new_page:
            try:
                await new_page.wait_for_selector("h1, h2, .title, [class*='detail'], strong", timeout=15000)
                await asyncio.sleep(1.0)
                await _extract_from_detail(new_page, enriched)
            except Exception as e:
                logger.warning(f"  新标签提取异常: {e}")
        elif url_changed:
            try:
                await page.wait_for_selector("h1, h2, .title, [class*='detail'], strong", timeout=15000)
            except Exception:
                pass
            await asyncio.sleep(1.0)
            await _extract_from_detail(page, enriched)
        else:
            modal_sel = ".modal, .dialog, .drawer, [role='dialog']"
            try:
                modal_el = await page.query_selector(modal_sel)
            except Exception:
                modal_el = None
            if modal_el:
                await asyncio.sleep(0.5)
                await _extract_from_detail(page, enriched)
                try:
                    await page.keyboard.press("Escape")
                except Exception:
                    pass

        # ── 期刊名白名单补全（移植自助手模式 _enrich_articles_from_detail）──
        journal = enriched.get("journal", "")
        if not journal or not journal_in_whitelist(journal):
            # 从 detail_tab 全文搜索白名单期刊
            try:
                full_text = await detail_tab.evaluate("() => document.body.innerText")
                found = _find_whitelist_journal_in_text(full_text[:500])
                if found:
                    logger.info(f"  期刊补全: '{journal}' -> '{found}'")
                    enriched["journal"] = found
                    journal = found
            except Exception:
                pass
        if journal and not journal_in_whitelist(journal):
            resolved = _resolve_truncated_journal(journal)
            if resolved:
                logger.info(f"  截断修复: '{journal}' -> '{resolved}'")
                enriched["journal"] = resolved

        # ── 下载链接检测 -> 分流 ──
        download_url = await _check_detail_download_url(detail_tab)
        has_download = download_url is not None

        if has_download:
            enriched["download_url"] = download_url
            logger.info(f"  下载链接: {download_url[:80]}...")
            if new_page:
                try:
                    await new_page.close()
                except Exception:
                    pass
            elif url_changed:
                try:
                    await page.go_back(timeout=10000)
                except Exception:
                    await page.goto(list_url, timeout=15000)
                await asyncio.sleep(random.uniform(1.0, 2.0))
        else:
            # 无下载链接 → 自动文献求助
            logger.info(f"  无下载链接 -> 自动文献求助")
            help_ok = await _auto_submit_literature_help(detail_tab, enriched, help_email)
            enriched["help_submitted"] = help_ok
            if new_page:
                try:
                    await asyncio.sleep(random.uniform(1.0, 2.0))
                    await new_page.close()
                except Exception:
                    pass
            elif url_changed:
                try:
                    await page.go_back(timeout=10000)
                except Exception:
                    await page.goto(list_url, timeout=15000)
                await asyncio.sleep(random.uniform(1.0, 2.0))

        enriched_list.append(enriched)

    # 统计
    improved_titles = sum(1 for a, e in zip(articles, enriched_list)
                          if e.get("title", "") != a.get("title", ""))
    improved_dois = sum(1 for a, e in zip(articles, enriched_list)
                        if e.get("doi") and not a.get("doi"))
    with_links = sum(1 for a in enriched_list if a.get("download_url"))
    help_count = sum(1 for a in enriched_list if a.get("help_submitted"))
    logger.info(f"  详情补全完成: 标题+{improved_titles} DOI+{improved_dois} "
                f"下载链接{with_links} 文献求助{help_count}")

    return enriched_list


async def _enrich_articles_from_detail(page, articles: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """逐篇点击文章标题打开详情页，采集完整元数据。

    V14 改进：
    - 有下载链接的详情页不关闭，留待后续手动操作。
    - 无下载链接的文献自动通过「文献求助」提交至预设邮箱。
    - 模拟真人操作节奏，避免触发反爬。
    """
    if not articles:
        return articles

    # 优先处理残缺标题（is_garbled）的文章 — 原地重排
    garbled = [a for a in articles if a.get("is_garbled")]
    normal = [a for a in articles if not a.get("is_garbled")]
    articles[:] = garbled + normal
    if garbled:
        logger.info(f"开始详情页补全 ({len(articles)} 篇, 其中 {len(garbled)} 篇标题残缺优先)...")
    else:
        logger.info(f"开始详情页补全 ({len(articles)} 篇)...")
    logger.info("  键盘: P=暂停 G=继续 E=终止")
    logger.info("  功能: 自动检测下载链接，无下载则自动文献求助 -> 18922596828@163.com")
    enriched_list = []
    list_url = page.url
    help_email = config.get("help_email", "18922596828@163.com")
    download_pages = []  # 保存有下载链接的详情页，不关闭
    help_submitted = 0

    for i, art in enumerate(articles):
        # ── 键盘控制 ──
        _poll_keyboard()
        await _kb_wait_if_paused()
        if _kb_is_terminated():
            logger.info("用户终止详情补全，保留已处理结果")
            break

        logger.info(f"详情 [{i+1}/{len(articles)}]: {art['title'][:60]}...")
        enriched = dict(art)

        # 每次循环重新查询 DOM
        try:
            await page.wait_for_selector("article.article", timeout=10000)
        except Exception:
            logger.warning("  列表页 DOM 未恢复，跳过")
            enriched_list.append(enriched)
            continue

        article_els = await page.query_selector_all("article.article")

        # 标题文本指纹匹配
        art_el = None
        target_title = art["title"]
        for el in article_els:
            try:
                el_text = await el.inner_text()
                if target_title[:40] in el_text.strip()[:80].replace("\n", " "):
                    art_el = el
                    break
            except Exception:
                continue
        if not art_el:
            if i < len(article_els):
                art_el = article_els[i]
            else:
                logger.warning(f"  找不到文章元素 #{i+1}")
                enriched_list.append(enriched)
                continue

        current_url = page.url

        # 找到可点击元素
        click_el = None
        for sel in ["div.d-t.jump", "div.allow-ai", "div.jump", "div.d-t"]:
            try:
                el = await art_el.query_selector(sel)
                if el:
                    click_el = el
                    break
            except Exception:
                continue
        if not click_el:
            click_el = art_el

        # ── 点击 + 轮询等导航 ──
        new_page = None
        url_changed = False

        try:
            async with page.context.expect_page(timeout=5000) as new_page_info:
                await click_el.click(force=True)
                # 同时轮询 URL 变化（SPA 路由跳转）
                for _ in range(15):
                    await asyncio.sleep(0.5)
                    try:
                        if page.url != current_url:
                            url_changed = True
                            break
                    except Exception:
                        pass
                    _poll_keyboard()
                    await _kb_wait_if_paused()
                if not url_changed:
                    try:
                        new_page = await new_page_info.value
                    except Exception:
                        pass
        except Exception:
            pass

        # ── 如果还没变化，继续等 ──
        if not url_changed and not new_page:
            for _ in range(20):
                await asyncio.sleep(0.5)
                try:
                    if page.url != current_url:
                        url_changed = True
                        break
                except Exception:
                    pass
                _poll_keyboard()

        logger.info(f"  URL变化={url_changed} 新标签={new_page is not None} "
                    f"({current_url[:60]} -> {page.url[:60] if url_changed else '(未变)'})")

        # ── 确定当前操作的 tab（新标签页 or 当前页）──
        detail_tab = new_page if new_page else page

        # ── 提取数据 ──
        if new_page:
            try:
                await new_page.wait_for_selector(
                    "h1, h2, .title, [class*='detail'], strong",
                    timeout=15000
                )
                await asyncio.sleep(1.0)
                await _extract_from_detail(new_page, enriched)
            except Exception as e:
                logger.warning(f"  新标签页提取异常: {e}")
        elif url_changed:
            try:
                await page.wait_for_selector(
                    "h1, h2, .title, [class*='detail'], strong",
                    timeout=15000
                )
            except Exception:
                pass
            await asyncio.sleep(1.0)
            await _extract_from_detail(page, enriched)
        else:
            # 检查弹窗
            modal_sel = ".modal, .dialog, .drawer, [role='dialog']"
            try:
                modal_el = await page.query_selector(modal_sel)
            except Exception:
                modal_el = None
            if modal_el:
                await asyncio.sleep(0.5)
                await _extract_from_detail(page, enriched)
                try:
                    await page.keyboard.press("Escape")
                except Exception:
                    pass
            else:
                logger.info(f"  所有策略均未触发导航，保留列表页数据")

        # ── V14: 检测下载链接 -> 决定关闭 or 文献求助 ──
        download_url = await _check_detail_download_url(detail_tab)
        has_download = download_url is not None

        if has_download:
            logger.info(f"  下载链接: {download_url[:80]}...")
            # 有新标签页的就留着不关
            if new_page:
                download_pages.append(new_page)
                logger.info(f"  保留详情页（有下载链接）")
            elif url_changed:
                # SPA 跳转的情况，先回到列表页
                logger.info(f"  有下载链接，返回列表页继续")
                try:
                    await page.go_back(timeout=10000)
                except Exception:
                    await page.goto(list_url, timeout=15000)
                await asyncio.sleep(random.uniform(1.0, 2.0))
        else:
            # 无下载链接 → 自动文献求助
            logger.info(f"  无下载链接 -> 自动文献求助")
            if new_page:
                # 在新标签页中操作文献求助
                help_ok = await _auto_submit_literature_help(new_page, enriched, help_email)
                if help_ok:
                    help_submitted += 1
                # 文献求助提交后关闭标签页
                try:
                    await asyncio.sleep(random.uniform(1.0, 2.0))
                    await new_page.close()
                except Exception:
                    pass
            elif url_changed:
                # 当前页操作文献求助
                help_ok = await _auto_submit_literature_help(page, enriched, help_email)
                if help_ok:
                    help_submitted += 1
                # 返回列表页
                await asyncio.sleep(random.uniform(1.0, 2.0))
                try:
                    await page.go_back(timeout=10000)
                except Exception:
                    await page.goto(list_url, timeout=15000)
                await asyncio.sleep(random.uniform(1.0, 2.0))
            else:
                try:
                    if new_page:
                        await new_page.close()
                except Exception:
                    pass

        enriched_list.append(enriched)

    # ── 清理：关闭所有保留的下载详情页 ──
    if download_pages:
        logger.info(f"关闭 {len(download_pages)} 个保留的下载详情页...")
        for dp in download_pages:
            try:
                await dp.close()
            except Exception:
                pass

    # 统计（分别统计标题补全、DOI补全、期刊补全）
    improved_titles = sum(1 for a, e in zip(articles, enriched_list)
                          if e.get("title", "") != a.get("title", ""))
    improved_dois = sum(1 for a, e in zip(articles, enriched_list)
                        if e.get("doi") and not a.get("doi"))
    improved_journals = sum(1 for a, e in zip(articles, enriched_list)
                            if e.get("journal") != a.get("journal"))
    # 统计严重截断标题的修复率
    garbled_count = sum(1 for a in articles if a.get("is_garbled"))
    garbled_fixed = sum(1 for a, e in zip(articles, enriched_list)
                        if a.get("is_garbled") and len(e.get("title", "")) > len(a.get("title", "")))
    if garbled_count:
        logger.info(f"详情补全完成: 标题补全 {improved_titles}, DOI补全 {improved_dois}, "
                    f"期刊补全 {improved_journals}, 残缺修复 {garbled_fixed}/{garbled_count}, "
                    f"文献求助 {help_submitted}")
    else:
        logger.info(f"详情补全完成: 标题补全 {improved_titles}, DOI补全 {improved_dois}, "
                    f"期刊补全 {improved_journals}, 文献求助 {help_submitted}")
    return enriched_list


# ═══════════════════════════════════════════════════════════════════
# 关键词级检索（人工模式 & 自主模式）
# ═══════════════════════════════════════════════════════════════════

async def scrape_keyword_until_full(
    page,
    keyword: str,
    global_doi_set: Set[str],
    global_title_set: Set[str],
    target_count: int,
    max_pages: int = MAX_PAGES_PER_KEYWORD,
) -> List[Dict[str, str]]:
    """【人工模式】翻页收集论文，直到目标篇数或连续3页无新。"""
    from litcall.pipeline.search.dedup import normalize_title, is_title_duplicate
    from litcall.pipeline.search.journal_filter import journal_in_whitelist

    new_collected = []
    consecutive_no_new = 0
    for pg in range(max_pages):
        _poll_keyboard()
        await _kb_wait_if_paused()
        if _kb_is_terminated():
            break
        await asyncio.sleep(random.uniform(MIN_SLEEP, MAX_SLEEP))
        page_articles = await scrape_page(page)
        if ENRICH_FROM_DETAIL and page_articles:
            page_articles = await _enrich_articles_from_detail(page, page_articles)
            # V15: 详情页补全后严格过滤 — 期刊必须在白名单内，确保高价值文献
            filtered_before = len(page_articles)
            filtered_out = []
            kept = []
            for a in page_articles:
                j = a.get("journal", "")
                if journal_in_whitelist(j):
                    kept.append(a)
                else:
                    filtered_out.append(a)
                    logger.info(f"[V15过滤] {a['title'][:60]} | 期刊不在白名单: {j}")
            if filtered_out:
                logger.info(f"V15严格过滤: {len(filtered_out)}/{filtered_before} 篇非白名单期刊已移除，"
                           f"保留 {len(kept)} 篇高价值文献")
            page_articles = kept
        page_added = 0
        for art in page_articles:
            doi = art["doi"].strip().lower()
            norm_title = normalize_title(art["title"])
            if doi and doi in global_doi_set:
                logger.info(f"[去重DOI] {art['title'][:50]} | DOI 已存在，跳过")
                continue
            title_dup = False
            dup_title = ""
            dup_char_sim = 0.0
            dup_word_jac = 0.0
            for gt in global_title_set:
                is_dup, char_sim, word_jac = is_title_duplicate(norm_title, gt)
                if is_dup:
                    title_dup = True
                    dup_title = gt
                    dup_char_sim = char_sim
                    dup_word_jac = word_jac
                    break
            if title_dup:
                logger.info(f"[去重标题] {art['title'][:60]} | "
                           f"字符相似{dup_char_sim*100:.0f}% 词Jaccard{dup_word_jac*100:.0f}% -> 与: {dup_title[:60]}")
                continue
            new_collected.append(art)
            if doi:
                global_doi_set.add(doi)
            global_title_set.add(norm_title)
            page_added += 1
            logger.info(f"新文献 (累计{len(new_collected)}): {art['title'][:120]} | {art.get('journal','')} | {art.get('year','')}")

        if page_added == 0:
            consecutive_no_new += 1
            if consecutive_no_new >= 3:
                logger.info(f"关键词 [{keyword}] 连续 {consecutive_no_new} 页无新文献，停止翻页。")
                break
        else:
            consecutive_no_new = 0

        # ── 达到目标后提示用户：继续翻同一关键词，还是停止 ──
        if len(new_collected) >= target_count:
            global _keyboard_state
            print(f"\n  本关键词已收集 {len(new_collected)} 篇（目标 {target_count} 篇），"
                  f"当前第 {pg+1}/{max_pages} 页。")
            print(f"  浏览器保持打开，可手动下载PDF。")
            print(f"  [G] 继续翻页检索（同一关键词）  [E] 停止本关键词")
            _keyboard_state = "paused"
            while _keyboard_state == "paused":
                await asyncio.sleep(0.3)
                _poll_keyboard()
            if _keyboard_state == "terminated":
                break
            # G 被按下：继续当前关键词后续页面，不再每页提示
            target_count = 9999

        # ── 翻页（V9：精确选择器，基于 SPIS 真实 HTML） ──
        # SPIS 分页结构:
        #   <button class="pagination-arrow first disabled"><</button>
        #   <button class="pagination-number active">N</button>
        #   <button class="pagination-number">N+1</button> ...
        #   <button class="pagination-arrow">></button>

        # 取当前页码
        current_num = pg + 1  # 回退默认值
        try:
            active_el = await page.query_selector("button.pagination-number.active")
            if active_el:
                txt = (await active_el.inner_text()).strip()
                if txt.isdigit():
                    current_num = int(txt)
        except Exception as e:
            logger.debug(f"取当前页码失败: {e}")

        next_num = current_num + 1
        clicked = False

        # ── 方式1：找文本为 '>' 的 pagination-arrow 按钮 ──
        #     不用 :not(.first) 避免 CSS 兼容问题，直接按文本匹配
        all_arrows = await page.query_selector_all("button.pagination-arrow")
        next_arrow = None
        for arrow in all_arrows:
            try:
                txt = (await arrow.inner_text()).strip()
                if txt == ">":
                    cls = (await arrow.get_attribute("class") or "").lower()
                    if "first" not in cls and "disabled" not in cls:
                        next_arrow = arrow
                        break
            except Exception:
                continue

        if next_arrow:
            try:
                await next_arrow.scroll_into_view_if_needed()
                await asyncio.sleep(0.3)
                await next_arrow.click()
                logger.debug(f"已点击 > 箭头，等待翻页到第{next_num}页...")
                # 轮询等 active 页码变成 next_num（最多 10 秒）
                for _ in range(20):
                    await asyncio.sleep(0.5)
                    new_active = await page.query_selector("button.pagination-number.active")
                    if new_active:
                        new_txt = (await new_active.inner_text()).strip()
                        if new_txt.isdigit() and int(new_txt) == next_num:
                            clicked = True
                            logger.info(f"翻页成功 -> 箭头 (第{current_num}页->第{new_txt}页)")
                            break
                if not clicked:
                    logger.warning(f"箭头点击后未检测到页码变化 (期望第{next_num}页)")
            except Exception as e:
                logger.warning(f"箭头翻页异常: {e}")

        # ── 方式2：点击具体页码按钮 ──
        if not clicked:
            target_btn = None
            all_nums = await page.query_selector_all("button.pagination-number")
            for btn in all_nums:
                try:
                    txt = (await btn.inner_text()).strip()
                    if txt == str(next_num) and "active" not in ((await btn.get_attribute("class") or "").lower()):
                        target_btn = btn
                        break
                except Exception:
                    continue

            if target_btn:
                try:
                    await target_btn.scroll_into_view_if_needed()
                    await asyncio.sleep(0.3)
                    await target_btn.click()
                    logger.debug(f"已点击页码按钮 {next_num}，等待激活...")
                    for _ in range(20):
                        await asyncio.sleep(0.5)
                        new_active = await page.query_selector("button.pagination-number.active")
                        if new_active:
                            new_txt = (await new_active.inner_text()).strip()
                            if new_txt.isdigit() and int(new_txt) == next_num:
                                clicked = True
                                logger.info(f"翻页成功 -> 页码按钮 (第{current_num}页->第{new_txt}页)")
                                break
                    if not clicked:
                        logger.warning(f"页码按钮点击后未检测到激活 (期望第{next_num}页)")
                except Exception as e:
                    logger.warning(f"页码翻页异常: {e}")
            else:
                logger.info(f"第{next_num}页按钮不在当前分页窗口内（仅显示前几页），且箭头不可用")

        if not clicked:
            logger.info(f"翻页失败 (第{current_num}页->第{next_num}页)，停止该关键词。")
            break

        # 等文章加载
        try:
            await page.wait_for_selector("article.article", timeout=15000)
        except PlaywrightTimeout:
            pass
        await asyncio.sleep(1)
    return new_collected


async def _scrape_keyword_exhaustive(
    page,
    keyword: str,
    global_doi_set: Set[str],
    global_title_set: Set[str],
    year_start: int = 2025,
    year_end: int = 2026,
    target_per_kw: int = 5,
    max_pages: int = MAX_PAGES_PER_KEYWORD,
    global_paper_limit: int = None,
    global_collected_count: list = None,
    journal_filter: list = None,
    run_logger = None,
) -> List[Dict[str, str]]:
    """【Agent 模式】翻页收集论文，达到全局上限 or 连续3页无新 or 翻完 max_pages 即停止。

    global_paper_limit: 本次运行总论文上限（跨关键词），None=不限制
    global_collected_count: 可变列表 [count] 用于跨关键词累加计数
    journal_filter: 可选，仅保留这些期刊的论文（覆盖白名单）
    """
    from litcall.pipeline.search.dedup import normalize_title, is_title_duplicate
    from litcall.pipeline.search.journal_filter import journal_in_whitelist

    new_collected = []
    consecutive_no_new = 0

    for pg in range(max_pages):
        # ── 信号检查 ──
        if run_logger:
            await _check_signal_files(run_logger)

        await asyncio.sleep(random.uniform(MIN_SLEEP, MAX_SLEEP))
        page_articles = await scrape_page(page)

        # ── 后置年份过滤（兜底：SPIS 年份筛选未生效时）──
        if page_articles:
            before_filter = len(page_articles)
            page_articles = [
                a for a in page_articles
                if _parse_article_year(a.get("year", ""), year_start, year_end)
            ]
            if len(page_articles) < before_filter:
                logger.info(f"  年份过滤 {year_start}-{year_end}: {before_filter} -> {len(page_articles)} 篇")

        # ── 详情页补全（含下载链接检测 + 文献求助）──
        if ENRICH_FROM_DETAIL and page_articles:
            page_articles = await _enrich_articles_autonomous(page, page_articles, run_logger)

            # ── 期刊过滤 ──
            kept = []
            for a in page_articles:
                j = a.get("journal", "")
                # 如果用户指定了 journal_filter，用它；否则用白名单
                if journal_filter:
                    matched = any(
                        jf.lower() in j.lower() or j.lower() in jf.lower()
                        for jf in journal_filter
                    )
                    if matched:
                        kept.append(a)
                    else:
                        logger.info(f"[期刊过滤] {a['title'][:60]} | 不在指定期刊: {j}")
                elif journal_in_whitelist(j):
                    kept.append(a)
                else:
                    logger.info(f"[V15过滤] {a['title'][:60]} | 期刊不在白名单: {j}")
            if len(kept) < len(page_articles):
                filter_name = "期刊过滤" if journal_filter else "V15过滤"
                logger.info(f"  {filter_name}: {len(page_articles) - len(kept)} 篇移除，保留 {len(kept)} 篇")
            page_articles = kept

        # ── 去重 ──
        page_added = 0
        for art in page_articles:
            # ── 全局上限检查（去重循环内检查，避免超收）──
            if global_paper_limit is not None and global_collected_count is not None:
                if global_collected_count[0] >= global_paper_limit:
                    logger.info(f"[Agent] 已收满全局 {global_paper_limit} 篇，停止。")
                    return new_collected

            doi = art["doi"].strip().lower()
            norm_title = normalize_title(art["title"])
            if doi and doi in global_doi_set:
                logger.info(f"[去重DOI] {art['title'][:50]}")
                continue
            title_dup = False
            for gt in global_title_set:
                is_dup, _, _ = is_title_duplicate(norm_title, gt)
                if is_dup:
                    title_dup = True
                    break
            if title_dup:
                logger.info(f"[去重标题] {art['title'][:60]}")
                continue
            new_collected.append(art)
            if doi:
                global_doi_set.add(doi)
            global_title_set.add(norm_title)
            page_added += 1
            if global_collected_count is not None:
                global_collected_count[0] += 1
            logger.info(f"[{len(new_collected)}/{global_collected_count[0] if global_collected_count else '?'}] {art['title'][:120]} | {art.get('journal','')} | {art.get('year','')}")

            # ── 全局达标检查 ──
            if global_paper_limit is not None and global_collected_count is not None:
                if global_collected_count[0] >= global_paper_limit:
                    logger.info(f"[Agent] 已收满全局 {global_paper_limit} 篇，完成。")
                    return new_collected

        # ── 翻页终止判断（连续3页无新即停止）──
        if page_added == 0:
            consecutive_no_new += 1
            if consecutive_no_new >= 3:
                logger.info(f"[Agent] 关键词 [{keyword}] 连续 {consecutive_no_new} 页无新文献，穷尽完成。")
                break
        else:
            consecutive_no_new = 0

        # ── 翻页 ──
        if not await _click_next_page(page, pg):
            logger.info(f"[Agent] 关键词 [{keyword}] 翻页结束（第{pg+1}页是最后一页）")
            break

    logger.info(f"[Agent] 关键词 [{keyword}] 完成: {len(new_collected)} 篇新文献")
    return new_collected


async def _scrape_new_articles_autonomous(
    year_start: int = 2025,
    year_end: int = 2026,
    headless: bool = False,
    keyword_cursor: dict = None,
    global_paper_limit: int = 5,
    max_pages_per_kw: int = 10,
    journal_filter: list = None,
    keyword_override: str = None,
    run_logger = None,
) -> dict:
    """【Agent 自主模式】从游标位置开始，逐关键词检索，直到收满 global_paper_limit 篇。

    keyword_cursor: 从 config.json 读的游标，指定从哪个关键词开始
    keyword_override: 若提供，忽略游标，仅用此一个词检索（游标不更新）
    journal_filter: 用户指定的期刊白名单（覆盖默认白名单）
    global_paper_limit: 本次运行总共收多少篇
    """
    # ── 关键词列表：override 优先，否则从游标开始 ──
    if keyword_override:
        all_keywords = [(keyword_override, "自定义")]
        cursor_ci, cursor_ki = 0, 0
        is_override = True
    else:
        all_keywords = print_keywords()
        if not all_keywords:
            return {"collected": [], "by_keyword": {}, "with_links": [], "without_links": [], "help_submitted": 0}
        if keyword_cursor is None:
            keyword_cursor = _read_keyword_cursor()
        cursor_ci = keyword_cursor.get("category_index", 0)
        cursor_ki = keyword_cursor.get("keyword_index", 0)
        is_override = False

    # ── 构建去重集合 ──
    try:
        zotero_dois, zotero_titles = await fetch_zotero_existing_dois()
    except Exception as e:
        logger.warning(f"Zotero 连接失败（{e}），仅使用本地日志去重。")
        zotero_dois, zotero_titles = set(), set()
    log_dois, log_titles = load_processed_log()
    excel_dois = _load_excel_dois()
    excel_titles = _load_excel_titles()
    global_doi_set = zotero_dois | log_dois | excel_dois
    global_title_set = zotero_titles | log_titles | excel_titles
    logger.info(f"[Agent] 去重集合: DOI={len(global_doi_set)} 标题={len(global_title_set)}")
    logger.info(f"[Agent] 全局上限: {global_paper_limit} 篇 | 每词最大翻页: {max_pages_per_kw}")

    all_collected = []
    by_keyword = {}
    papers_with_links = []
    papers_without_links = []
    total_help_submitted = 0
    global_count = [0]  # 可变计数器，跨关键词传递
    keywords_exhausted = False

    # ── 从游标位置开始迭代 ──
    skipped_before_cursor = 0
    for idx, (keyword, scope) in enumerate(all_keywords, 1):
        # 非 override 模式下，跳过游标之前的关键词
        if not is_override:
            kw_cat, kw_idx_in_cat = _get_category_index(scope, idx - 1)
            if kw_cat < cursor_ci or (kw_cat == cursor_ci and kw_idx_in_cat < cursor_ki):
                skipped_before_cursor += 1
                continue

        # ── 全局上限检查 ──
        if global_count[0] >= global_paper_limit:
            logger.info(f"[Agent] 全局已收满 {global_paper_limit} 篇，停止。")
            break

        logger.info(f"\n{'='*60}")
        logger.info(f"[Agent] 关键词 [{idx}/{len(all_keywords)}]: {keyword}")
        logger.info(f"   游标: cat={cursor_ci} ki={cursor_ki} | 已收: {global_count[0]}/{global_paper_limit}")
        logger.info(f"{'='*60}")

        if run_logger:
            run_logger.log_keyword_progress(keyword, cursor_ci, cursor_ki, global_count[0])

        p, context, page = await open_spis_autonomous(
            keyword, year_start=year_start, year_end=year_end, headless=headless
        )
        kw_collected = []
        try:
            kw_collected = await _scrape_keyword_exhaustive(
                page, keyword, global_doi_set, global_title_set,
                year_start=year_start, year_end=year_end,
                max_pages=max_pages_per_kw,
                global_paper_limit=global_paper_limit,
                global_collected_count=global_count,
                journal_filter=journal_filter,
                run_logger=run_logger,
            )
            all_collected.extend(kw_collected)
            by_keyword[keyword] = kw_collected
            logger.info(f"[Agent] 关键词 [{keyword}] 贡献 {len(kw_collected)} 篇，累计 {len(all_collected)}/{global_paper_limit}")

            if kw_collected:
                print(f"\n{'='*70}")
                print(f"  [{idx}/{len(all_keywords)}] 「{keyword}」— {len(kw_collected)}篇:")
                print(f"{'='*70}")
                for j, art in enumerate(kw_collected, 1):
                    doi_str = f"  DOI: {art['doi']}" if art.get('doi') else ""
                    link_info = "有下载" if art.get('download_url') else "已求助"
                    print(f"  [{j}] {link_info} {art['title'][:100]}")
                    print(f"      期刊: {art.get('journal', '?')} | {art.get('year', '?')}{doi_str}")
                print(f"{'='*70}\n")

        finally:
            await context.close()
            await p.stop()

        # ── 分流 ──
        for art in kw_collected:
            if art.get("download_url"):
                papers_with_links.append(art)
            else:
                papers_without_links.append(art)
        total_help_submitted += sum(1 for a in kw_collected if a.get("help_submitted"))

        # ── 更新游标（override 模式不更新）──
        if not is_override:
            kw_cat, kw_idx_in_cat = _get_category_index(scope, idx - 1)
            _write_keyword_cursor(kw_cat, kw_idx_in_cat + 1)
            cursor_ci = kw_cat
            cursor_ki = kw_idx_in_cat + 1  # 下一个

        # ── 信号检查 ──
        if run_logger:
            await _check_signal_files(run_logger)

    # ── 检查是否全部穷尽 ──
    if not is_override:
        next_kw, _, _, _ = _get_next_keyword(_read_keyword_cursor())
        if next_kw is None:
            keywords_exhausted = True
            logger.info("[Agent] 所有关键词已遍历完毕！")
            if run_logger:
                run_logger.log_keywords_exhausted()

    # ── 保存待处理清单（有下载链接的论文）──
    if papers_with_links:
        pending_file = PDF_DIR / "pending_manual.json"
        # 合并已有清单
        existing_pending = []
        if pending_file.exists():
            try:
                with open(pending_file, "r", encoding="utf-8") as f:
                    existing_pending = json.load(f)
            except Exception:
                pass
        # 去重后保存
        existing_titles = {a.get("title", "")[:80] for a in existing_pending}
        new_to_add = [a for a in papers_with_links if a.get("title", "")[:80] not in existing_titles]
        combined = existing_pending + new_to_add
        tmp_file = Path(str(pending_file) + ".tmp")
        with open(tmp_file, "w", encoding="utf-8") as f:
            json.dump(combined, f, ensure_ascii=False, indent=2)
        tmp_file.replace(pending_file)
        logger.info(f"[Agent] pending_manual.json: 原有 {len(existing_pending)} + 新增 {len(new_to_add)} = {len(combined)} 篇")

    result = {
        "collected": all_collected,
        "by_keyword": by_keyword,
        "with_links": papers_with_links,
        "without_links": papers_without_links,
        "help_submitted": total_help_submitted,
        "keywords_exhausted": keywords_exhausted,
    }
    return result
