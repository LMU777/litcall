#!/usr/bin/env python3
"""
literature_agent.py — LitCall 学术文本深度阅读与知识管理智能体
Academic Text Harvesting & Empirical Note-taking Agent
- 渐进式关键词检索，收集满 20 篇新文献后停止。
- Zotero + processed_log.json 双重去重。
- 公开 OA 自动下载；付费文献记录待办清单。
- 处理后询问是否清理已完成的 PDF，避免堆积。
- 期刊白名单已扩展至管理、心理类顶刊。
- 防幻觉流水线: 自检 + 二次验证 + 变量交叉校验 + 人工复核队列。
"""

import asyncio
import datetime
import json
import logging
import os
import hashlib
import io
import re
import shutil
import sys
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
import random
import unicodedata
import urllib.parse

import aiohttp
import requests
import openpyxl
import fitz  # PyMuPDF
import numpy as np
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

# ============================================================================
# 键盘控制：P 暂停 / G 继续 / E 终止
# ============================================================================
_IS_WINDOWS = sys.platform == "win32"
_keyboard_state = "running"  # running | paused | terminated

if _IS_WINDOWS:
    import msvcrt


def _poll_keyboard():
    """非阻塞轮询键盘，检测 P / G / E 按键。仅 Windows。"""
    global _keyboard_state
    if not _IS_WINDOWS:
        return
    try:
        while msvcrt.kbhit():
            key = msvcrt.getch()
            if key in (b'p', b'P'):
                if _keyboard_state != "paused":
                    _keyboard_state = "paused"
                    print("\n⏸  [P] 检索已暂停！按 G 继续，按 E 终止")
            elif key in (b'g', b'G'):
                if _keyboard_state == "paused":
                    _keyboard_state = "running"
                    print("▶  [G] 检索已继续")
            elif key in (b'e', b'E'):
                _keyboard_state = "terminated"
                print("\n⏹  [E] 检索终止信号已接收，正在退出当前环节...")
    except Exception:
        pass


async def _kb_wait_if_paused():
    """暂停时阻塞等待，轮询 G 或 E。"""
    global _keyboard_state
    while _keyboard_state == "paused":
        await asyncio.sleep(0.3)
        _poll_keyboard()
    if _keyboard_state == "terminated":
        raise KeyboardInterrupt("用户按 E 终止")


def _kb_is_terminated() -> bool:
    return _keyboard_state == "terminated"

# ============================================================================
# 路径常量
# ============================================================================
SCRIPT_DIR = Path(__file__).parent.absolute()
BASE_DIR = SCRIPT_DIR  # 自动从脚本位置推断，不再硬编码
PDF_DIR = BASE_DIR / "新论文待处理"
NOTES_DIR = PDF_DIR / "notes"  # 深度阅读笔记 JSON 暂存目录
EXCEL_PATH = BASE_DIR / "agent文献汇总.xlsx"
LOG_FILE = BASE_DIR / "运行日志" / "Claude_Semi_auto_paper_log.txt"
RUNS_DIR = BASE_DIR / "运行日志" / "runs"
OBSIDIAN_DIR = BASE_DIR / "agent抓取"
CONFIG_PATH = SCRIPT_DIR / "config.json"
PROCESSED_LOG = SCRIPT_DIR / "processed_log.json"

# ============================================================================
# Agent 信号文件 & 关键词游标
# ============================================================================

class AgentSignalError(Exception):
    """Agent 收到终止信号时抛出。调用方应捕获并优雅退出。"""
    pass


async def _check_signal_files(run_logger=None) -> str:
    """检查 运行日志/.pause 和 .terminate 信号文件。

    由 Streamlit 写入信号，Worker 子进程轮询检测：
      - .terminate → 标记 run_logger status=terminated → 抛 AgentSignalError
      - .pause     → 标记 run_logger _paused=true → 阻塞轮询直到 .pause 消失或 .terminate 出现
      - 无信号     → 返回 "ok"

    Args:
        run_logger: AgentRunLogger 实例（可选），用于实时更新运行状态 JSON
    Returns:
        "ok" | "terminated" | "resumed"
    Raises:
        AgentSignalError: 收到终止信号
    """
    pause_file = BASE_DIR / "运行日志" / ".pause"
    terminate_file = BASE_DIR / "运行日志" / ".terminate"

    if terminate_file.exists():
        if run_logger:
            run_logger._data["status"] = "terminated"
            run_logger._save()
        terminate_file.unlink(missing_ok=True)
        pause_file.unlink(missing_ok=True)
        logger.info("⏹ Agent 已终止 (收到 .terminate 信号)")
        raise AgentSignalError("用户终止")

    if pause_file.exists():
        if run_logger:
            run_logger._data["_paused"] = True
            run_logger._save()
        logger.info("⏸ Agent 已暂停 (检测到 .pause 信号文件)")
        while pause_file.exists():
            await asyncio.sleep(1)
            if terminate_file.exists():
                if run_logger:
                    run_logger._data["status"] = "terminated"
                    run_logger._save()
                terminate_file.unlink(missing_ok=True)
                pause_file.unlink(missing_ok=True)
                logger.info("⏹ Agent 在暂停期间被终止")
                raise AgentSignalError("用户在暂停期间终止")
        if run_logger:
            run_logger._data["_paused"] = False
            run_logger._save()
        logger.info("▶ Agent 已继续 (.pause 信号文件已移除)")
        return "resumed"

    return "ok"


# ── 关键词游标 ──

def _get_category_index(scope: str, flat_index: int) -> tuple:
    """将 scope 标签 + 扁平索引转换为 (category_index, keyword_index_within_category)。"""
    if scope == "宽":
        return (0, flat_index)
    elif scope == "窄":
        return (1, flat_index - len(config.get("keywords", {}).get("broad", [])))
    elif scope == "中":
        broad_len = len(config.get("keywords", {}).get("broad", []))
        narrow_len = len(config.get("keywords", {}).get("narrow", []))
        return (2, flat_index - broad_len - narrow_len)
    return (0, flat_index)


def _read_keyword_cursor() -> dict:
    """从 config.json 读取关键词游标。"""
    return config.get("_keyword_cursor", {"category_index": 0, "keyword_index": 0})


def _write_keyword_cursor(category_index: int, keyword_index: int):
    """原子写入关键词游标到 config.json（临时文件 + rename 防并发写坏）。"""
    from datetime import datetime
    config["_keyword_cursor"] = {
        "category_index": category_index,
        "keyword_index": keyword_index,
        "last_updated": datetime.now().isoformat(),
    }
    tmp = CONFIG_PATH.with_suffix(".json.tmp")
    try:
        tmp.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(CONFIG_PATH)
    except Exception as e:
        logger.warning(f"写入关键词游标失败: {e}")


def _get_next_keyword(cursor: dict) -> tuple:
    """从游标位置获取下一个关键词。

    Returns:
        (keyword_str, scope_tag, category_index, keyword_index)
        全部穷尽时返回 (None, None, -1, -1)
    """
    categories = [
        ("宽", config.get("keywords", {}).get("broad", [])),
        ("窄", config.get("keywords", {}).get("narrow", [])),
        ("中", config.get("keywords", {}).get("chinese", [])),
    ]

    ci = cursor.get("category_index", 0)
    ki = cursor.get("keyword_index", 0)

    while ci < len(categories):
        scope, kw_list = categories[ci]
        if ki < len(kw_list):
            return (kw_list[ki], scope, ci, ki)
        ci += 1
        ki = 0

    return (None, None, -1, -1)


def _advance_keyword_cursor(cursor: dict):
    """推进游标到下一个关键词（写回 config.json）。"""
    ci = cursor.get("category_index", 0)
    ki = cursor.get("keyword_index", 0)

    categories = [
        config.get("keywords", {}).get("broad", []),
        config.get("keywords", {}).get("narrow", []),
        config.get("keywords", {}).get("chinese", []),
    ]

    ki += 1
    while ci < len(categories) and ki >= len(categories[ci]):
        ci += 1
        ki = 0

    _write_keyword_cursor(ci, ki)
    return {"category_index": ci, "keyword_index": ki}

# ============================================================================\
# 扩展版期刊白名单（硬编码，60+ 种）
# 可通过 config.json 中 "journal_whitelist_extra" 字段追加自定义期刊
# ============================================================================
WHITELIST = [
    # ── 营销类核心 ──
    "Journal of Marketing",
    "Journal of Marketing Research",
    "Marketing Science",
    "Journal of Consumer Research",
    "Journal of the Academy of Marketing Science",
    "Journal of Consumer Psychology",
    # 营销类扩展（高影响力）
    "Journal of Business Research",
    "Journal of Retailing",
    "Industrial Marketing Management",
    "Journal of Service Research",
    "Journal of Advertising",
    "Psychology & Marketing",
    "European Journal of Marketing",
    "International Journal of Research in Marketing",
    "Journal of Interactive Marketing",
    "Journal of Research in Interactive Marketing",
    "Journal of Public Policy & Marketing",
    "Journal of International Marketing",
    "International Journal of Market Research",
    "Marketing Letters",
    "Quantitative Marketing and Economics",
    "Journal of Advertising Research",
    "International Journal of Advertising",
    "Journal of Brand Management",
    "Journal of Product & Brand Management",
    "Journal of Services Marketing",
    "Journal of Marketing Management",
    "Journal of Macromarketing",
    "Consumption Markets & Culture",
    "Journal of the Association for Consumer Research",
    "Journal of Consumer Behaviour",
    "Journal of Consumer Affairs",
    "Journal of Retailing and Consumer Services",
    # ── 管理类 (UTD24/FT50 全覆盖) ──
    "Academy of Management Journal",
    "Academy of Management Review",
    "Academy of Management Perspectives",
    "Academy of Management Annals",
    "Administrative Science Quarterly",
    "Organization Science",
    "Strategic Management Journal",
    "Management Science",
    "Journal of Management",
    "Journal of Management Studies",
    "Journal of International Business Studies",
    "Journal of Operations Management",
    "Production and Operations Management",
    "Information Systems Research",
    "MIS Quarterly",
    # 管理类扩展
    "Harvard Business Review",
    "MIT Sloan Management Review",
    "California Management Review",
    "Journal of World Business",
    "Global Strategy Journal",
    "Long Range Planning",
    "Research Policy",
    "Technovation",
    "Journal of Product Innovation Management",
    "Decision Sciences",
    "Manufacturing & Service Operations Management",
    "Omega",
    "European Journal of Operational Research",
    "Management and Organization Review",
    "Asia Pacific Journal of Management",
    "Journal of Supply Chain Management",
    # ── 工商管理/创业/伦理 ──
    "Journal of Business Ethics",
    "Business Ethics Quarterly",
    "Business & Society",
    "Entrepreneurship Theory and Practice",
    "Journal of Business Venturing",
    "Human Relations",
    "Journal of Organizational Behavior",
    "Leadership Quarterly",
    "Organizational Research Methods",
    # ── 应用心理学 / 消费者行为 ──
    "Journal of Applied Psychology",
    "Personnel Psychology",
    "Organizational Behavior and Human Decision Processes",
    "Journal of Personality and Social Psychology",
    "Psychological Science",
    "Journal of Experimental Psychology: General",
    "Journal of Experimental Psychology: Applied",
    "Journal of Experimental Social Psychology",
    "Personality and Social Psychology Bulletin",
    "Psychological Review",
    "Annual Review of Psychology",
    "Cognition",
    "Emotion",
    "Judgment and Decision Making",
    "Journal of Behavioral Decision Making",
    "Social Psychological and Personality Science",
    "Journal of Economic Psychology",
    "Appetite",
    "Food Quality and Preference",
    # ── 信息系统 / 技术 ──
    "Journal of the Association for Information Systems",
    "European Journal of Information Systems",
    "Information & Management",
    "Decision Support Systems",
    "Journal of Information Technology",
    "Journal of Management Information Systems",
    "Information Systems Journal",
    # ── 人机交互 / 计算机交叉（AI+营销高频发表阵地）──
    "Computers in Human Behavior",
    "Business Horizons",
    # ── 知识管理 / 信息管理 ──
    "Journal of Knowledge Management",
    "International Journal of Information Management",
    "Journal of Information Science",
    "Information Processing & Management",
    "Knowledge Management Research & Practice",
    # ── 旅游/酒店（AI+营销/服务高频发表阵地）──
    "Tourism Management",
    "Annals of Tourism Research",
    "Journal of Travel Research",
    "International Journal of Hospitality Management",
    "International Journal of Contemporary Hospitality Management",
    "Journal of Hospitality Marketing & Management",
    # ── 服务质量/运营（Emerald 等）──
    "International Journal of Quality and Service Sciences",
    "Journal of Service Theory and Practice",
    "Managing Service Quality",
    "International Journal of Quality & Reliability Management",
    "The TQM Journal",
    "Benchmarking",
    # ── 营销扩展（补充）──
    "Journal of Strategic Marketing",
    "Journal of Marketing Communications",
    "Marketing Theory",
    "Journal of Fashion Marketing and Management",
    "International Marketing Review",
    "Journal of Business & Industrial Marketing",
    "Journal of Consumer Marketing",
    "Journal of Personal Selling & Sales Management",
    "International Review of Retail Distribution and Consumer Research",
    # ── 管理扩展 ──
    "British Journal of Management",
    "International Journal of Management Reviews",
    "Scandinavian Journal of Management",
    "Technological Forecasting and Social Change",
    "Journal of Cleaner Production",
    "Business Strategy and the Environment",
    "Corporate Social Responsibility and Environmental Management",
    # ── 创新/科技管理 ──
    "Journal of Innovation & Knowledge",
    "IEEE Transactions on Engineering Management",
    "R&D Management",
    "Journal of Engineering and Technology Management",
    # ── 信息系统扩展 ──
    "Journal of Strategic Information Systems",
    "Information and Organization",
    "Journal of Database Management",
    "Information Systems Frontiers",
    # ── 社科/综合（⚠️ 全学科巨型期刊，可能收录非营销/AI论文，建议人工复核）──
    "PLOS ONE",
    "Scientific Reports",
    "Heliyon",
    "SAGE Open",
]

TARGET_NEW_COUNT = 10
MAX_PAGES_PER_KEYWORD = 10
MIN_SLEEP = 3.0
MAX_SLEEP = 8.0
ENRICH_FROM_DETAIL = True  # 是否从详情页采集完整元数据（较慢但完整）

# ============================================================================
# 配置加载
# ============================================================================
if not CONFIG_PATH.exists():
    print(f"配置文件 {CONFIG_PATH} 不存在，请从 config.example.json 复制并填写。")
    sys.exit(1)

with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    config = json.load(f)

# ============================================================================
# 影响因子配置加载 (journal_if.json)
# ============================================================================
JOURNAL_IF_PATH = SCRIPT_DIR / "journal_if.json"
_journal_if_map: Dict[str, float] = {}

def _load_journal_if_map() -> Dict[str, float]:
    """加载 journal_if.json，构建 期刊名→IF 的扁平映射表"""
    global _journal_if_map
    if _journal_if_map:
        return _journal_if_map
    if not JOURNAL_IF_PATH.exists():
        logger.warning(f"影响因子配置文件不存在: {JOURNAL_IF_PATH}，将无法自动填充 IF")
        return {}
    try:
        with open(JOURNAL_IF_PATH, "r", encoding="utf-8") as f:
            if_data = json.load(f)
        for category, journals in if_data.items():
            if category.startswith("_"):
                continue
            if isinstance(journals, dict):
                for jn, if_val in journals.items():
                    if isinstance(if_val, (int, float)) and if_val > 0:
                        _journal_if_map[jn.lower().strip()] = float(if_val)
        logger.info(f"已加载 {len(_journal_if_map)} 条期刊影响因子")
    except Exception as e:
        logger.warning(f"加载影响因子配置失败: {e}")
    return _journal_if_map

# ============================================================================
# 日志配置
# ============================================================================
LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
_handlers = [logging.FileHandler(LOG_FILE, encoding="utf-8")]
try:
    _handlers.append(logging.StreamHandler(io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")))
except (ValueError, OSError, AttributeError):
    # 子进程/重定向环境下 stdout.buffer 可能不可用，退化为 stderr
    try:
        _handlers.append(logging.StreamHandler(sys.stderr))
    except Exception:
        pass  # 连 stderr 都不可用就只用文件日志

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    handlers=_handlers,
)
logger = logging.getLogger(__name__)

# 模块加载时初始化（必须在 logger 之后）
_journal_if_map = _load_journal_if_map()

# ============================================================================
# 笔记生成 API 模式诊断
# ============================================================================
_deepseek_key = config.get("deepseek_api_key", "")
_deepseek_model = config.get("deepseek_model", "deepseek-chat")
if not _deepseek_key:
    logger.error("!!! DeepSeek API Key 未配置！请在 config.json 中设置 deepseek_api_key。")
    logger.error("!!! 笔记生成将无法工作（无回退模式）。")
else:
    logger.info(f"笔记生成：DeepSeek V4 Pro ({_deepseek_model}) — 无回退模式")

NOTE_FIELDS = [
    "序号",           # 自动递增（上一行序号+1）
    "标题",
    "作者",
    "第一作者",
    "通讯作者",
    "年份",
    "期刊",
    "影响因子",
    "分区",
    "doi",
    "阅读方式",       # 用户手动填写（精读/略读/挑读）
    "阅读日期",       # 用户手动填写
    "关键词",
    "研究背景与动机",
    "研究问题",
    "变量汇总",
    "研究方法",
    "方法论详解",       # 建模逻辑、计量方法、识别策略等技术细节深度讲解
    "研究结果",
    "讨论与结论",
    "创新点",
    "局限与展望",
    "图表分析",        # Gemini Vision + DeepSeek 图表智能分析
]

_zotero_collection_key_cache: Optional[str] = None
_non_interactive: bool = False  # --yes 标志，跳过所有交互式确认，Phase 2 使用倒计时


# ============================================================================
# 工具函数
# ============================================================================
def clean_filename(text: str, max_len: int = 80) -> str:
    if not text:
        return "untitled"
    cleaned = re.sub(r"[^\w\s一-鿿-]", "_", text)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned[:max_len] if len(cleaned) > max_len else cleaned


def normalize_title(title: str) -> str:
    if not title:
        return ""
    title = re.sub(r"[^\w\s]", "", title).lower().strip()
    return re.sub(r"\s+", " ", title)


def _word_set(title: str) -> set:
    """返回标题的小写词集合（过滤 <3 字符的噪声词如 'a', 'in', 'of' 等保留）。"""
    return set(w for w in title.lower().split() if len(w) >= 1)


def word_jaccard(a: str, b: str) -> float:
    """词级 Jaccard 相似度。两个标题共享的独特词比例。"""
    wa = _word_set(normalize_title(a))
    wb = _word_set(normalize_title(b))
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / len(wa | wb)


def title_similarity(a: str, b: str) -> float:
    """字符级相似度（SequenceMatcher）。保留用于向后兼容。"""
    import difflib
    return difflib.SequenceMatcher(None, normalize_title(a), normalize_title(b)).ratio()


def is_title_duplicate(a: str, b: str) -> Tuple[bool, float, float]:
    """综合判断两个标题是否为同一篇文献。

    使用双重防线：
    1. 字符级相似度 (SequenceMatcher) — 检测截断/OCR 变体
    2. 词级 Jaccard — 防止共享长短语的不同论文被误杀

    返回 (is_dup, char_sim, word_jac)。
    判定规则：字符级 > 0.80 且 词级 > 0.50 才视为重复。
    """
    char_sim = title_similarity(a, b)
    word_jac = word_jaccard(a, b)
    # 两个条件必须同时满足：
    # - 字符级高：标题文本确实相似
    # - 词级高：排除仅共享通用前缀的不同论文
    is_dup = char_sim > 0.80 and word_jac > 0.50
    return is_dup, char_sim, word_jac


def extract_doi_from_pdf(pdf_path: Path) -> Optional[str]:
    """从 PDF 前 3 页提取 DOI"""
    try:
        doc = fitz.open(pdf_path)
        for i in range(min(3, len(doc))):
            text = doc[i].get_text()
            m = re.search(r"(10\.\d{4,}/[^\s]+)", text)
            if m:
                doc.close()
                return m.group(1).strip().rstrip(".")
        doc.close()
    except Exception:
        pass
    return None


# ============================================================================
# 合并白名单：硬编码 + config.json 扩展
# ============================================================================
def _get_effective_whitelist() -> List[str]:
    """合并硬编码白名单与 config.json 中的 journal_whitelist_extra"""
    extra = config.get("journal_whitelist_extra", [])
    if extra:
        logger.info(f"加载自定义期刊 {len(extra)} 种: {extra}")
    seen = set(w.lower() for w in WHITELIST)
    merged = list(WHITELIST)
    for jn in extra:
        if jn.lower() not in seen:
            merged.append(jn)
            seen.add(jn.lower())
    return merged


# 构建开头词索引：{"journal of": ["Journal of Marketing", "Journal of Consumer Research", ...]}
def _build_prefix_index(whitelist: List[str]) -> Dict[str, List[str]]:
    idx: Dict[str, List[str]] = {}
    for w in whitelist:
        words = w.lower().split()
        for n in range(2, min(len(words) + 1, 6)):
            prefix = " ".join(words[:n])
            idx.setdefault(prefix, []).append(w)
    return idx


_EFFECTIVE_WHITELIST = _get_effective_whitelist()
_PREFIX_INDEX = _build_prefix_index(_EFFECTIVE_WHITELIST)


def _find_whitelist_journal_in_text(text: str) -> Optional[str]:
    """在任意文本中搜索白名单期刊名，返回匹配到的完整期刊名（最长匹配优先）。

    用于解决 SPIS 截断/异形格式导致的期刊名提取失败。
    直接搜索文本中是否包含白名单期刊名，不依赖任何格式假设。
    返回白名单中的原始大小写期刊名，或 None。"""
    if not text:
        return None
    text_lower = text.lower()
    best_match: Optional[str] = None
    best_len = 0
    for w in _EFFECTIVE_WHITELIST:
        wl = w.lower()
        if wl in text_lower:
            if len(wl) > best_len:
                best_match = w
                best_len = len(wl)
    return best_match


def _resolve_truncated_journal(partial: str) -> Optional[str]:
    """尝试从截断的期刊名片段补全为完整白名单期刊名。

    V14 改进：后缀匹配必须唯一 — 只有当 partial 唯一匹配到一个白名单期刊时才返回。
    避免 "International Journal of" → 误匹配 "International Journal of Contemporary Hospitality Management"。
    """
    if not partial or len(partial) < 5:
        return None
    partial_lower = partial.lower().strip()

    # 策略1：白名单期刊名以 partial 结尾（partial 是后半截）
    # V14: 必须唯一匹配
    suffix_matches = []
    for w in _EFFECTIVE_WHITELIST:
        wl = w.lower()
        if wl.endswith(partial_lower) and len(wl) > len(partial_lower):
            suffix_matches.append(w)
    if len(suffix_matches) == 1:
        logger.debug(f"截断期刊补全(后缀匹配,唯一): '{partial}' → '{suffix_matches[0]}'")
        return suffix_matches[0]
    elif len(suffix_matches) > 1:
        logger.debug(f"截断期刊补全(后缀匹配,歧义): '{partial}' 匹配 {len(suffix_matches)} 个，不猜测")

    # 策略2：取 partial 最后 2-3 个实词搜索
    words = partial_lower.split()
    for n in [3, 2]:
        if len(words) >= n:
            suffix = " ".join(words[-n:])
            if len(suffix) >= 8:
                word_matches = []
                for w in _EFFECTIVE_WHITELIST:
                    wl = w.lower()
                    if suffix in wl and len(wl) > len(partial_lower):
                        word_matches.append(w)
                if len(word_matches) == 1:
                    logger.debug(f"截断期刊补全(尾词搜索,唯一): '{partial}' → '{word_matches[0]}'")
                    return word_matches[0]
                elif len(word_matches) > 1:
                    logger.debug(f"截断期刊补全(尾词搜索,歧义): '{partial}' 匹配 {len(word_matches)} 个")

    return None


def journal_in_whitelist(journal: str) -> bool:
    """增强匹配：首词前缀 → 前缀索引 → 逐级回退 → 子串双向包含 → 尾缀匹配

    V7 改进：新增策略0.5，专门处理 SPIS 截断导致的单/双词期刊名前缀
    （如 "California" → "California Management Review"）。"""
    if not journal:
        return False
    # 去除 SPIS 截断标记（前后 "…" / "..." 及空白）
    j = journal.lower().strip().strip("….").strip().rstrip(",")

    # 策略0: 精确匹配（快速路径）
    for w in _EFFECTIVE_WHITELIST:
        if w.lower() == j:
            return True

    # ── 策略0.5（V14改进）：首词前缀匹配 ──
    # 专门处理 SPIS 单/双词截断场景。
    # V14 严格条件：
    #   1. j 是 wl 的前缀，且 j≥5字符
    #   2. 长度比≥35%（排除太短的通用前缀如 "International Journal of"）
    #   3. 排除通用开头词：只匹配有独特性的前缀
    #   4. 该前缀在所有白名单期刊中只匹配唯一一本（避免歧义）
    # "california"(10) / "california management review"(28) = 35.7% → 通过 ✓
    # "psychology"(10) / "psychology & marketing"(22) = 45.5% → 通过 ✓
    # "international journal of"(24) → 多个匹配，歧义过大 → 拒绝 ✓
    # "the"(3) → len<5 拒绝 ✓
    if len(j) >= 5:
        # 统计有多少白名单期刊以 j 开头
        matches = []
        for w in _EFFECTIVE_WHITELIST:
            wl = w.lower()
            if wl.startswith(j) and len(j) / len(wl) >= 0.35:
                matches.append(w)
        # 只有当 j 唯一匹配一个白名单期刊时才通过
        # 这避免了 "International Journal of" 匹配到错误的期刊
        if len(matches) == 1:
            w = matches[0]
            wl = w.lower()
            logger.debug(f"首词前缀匹配(V14): '{journal}' → '{w}' ({len(j)/len(wl)*100:.0f}%)")
            return True
        elif len(matches) > 1:
            logger.debug(f"首词前缀歧义(V14): '{journal}' 匹配 {len(matches)} 个期刊，拒绝猜测。"
                        f"候选项: {matches[:5]}")

    # 策略1: 前缀索引精确匹配（处理截断，如 "Journal of Consumer" → "Journal of Consumer Research"）
    candidates = _PREFIX_INDEX.get(j)
    if candidates:
        for c in candidates:
            c_lower = c.lower()
            if c_lower.startswith(j):
                logger.debug(f"前缀匹配: '{journal}' → '{c}'")
                return True
            if j.startswith(c_lower):
                logger.debug(f"前缀匹配: '{journal}' → '{c}'")
                return True

    # 策略2: 检查 j 本身的逐级前缀（被截断时 j 可能不在索引中但 j[:n] 在）
    j_words = j.split()
    for n in range(len(j_words) - 1, 1, -1):
        partial = " ".join(j_words[:n])
        if partial in _PREFIX_INDEX:
            for c in _PREFIX_INDEX[partial]:
                c_lower = c.lower()
                if c_lower.startswith(partial) and j in c_lower:
                    logger.debug(f"部分前缀匹配: '{journal}' → '{c}'")
                    return True

    # 策略3: 子串双向包含（兜底），V7 门槛降至 30%
    for w in _EFFECTIVE_WHITELIST:
        wl = w.lower()
        if wl in j:
            return True
        if j in wl:
            if len(j) / len(wl) >= 0.30:
                return True
        prefix = " ".join(wl.split()[:4])
        if len(prefix) >= 15 and prefix in j:
            return True

    # 策略4（V7新增）：尾缀匹配 — 白名单刊名以 j 结尾
    # 例: "Management Review" → "California Management Review"
    for w in _EFFECTIVE_WHITELIST:
        wl = w.lower()
        if wl.endswith(j) and len(j) >= 10 and len(j) / len(wl) >= 0.30:
            logger.debug(f"尾缀匹配(V7): '{journal}' → '{w}'")
            return True

    return False


def safe_http_request(method: str, url: str, **kwargs) -> requests.Response:
    kwargs.setdefault("timeout", 30)
    try:
        return requests.request(method, url, **kwargs)
    except requests.exceptions.SSLError:
        kwargs["verify"] = False
        return requests.request(method, url, **kwargs)
    except requests.exceptions.ConnectionError as e:
        raise ConnectionError(f"无法连接到 {url}: {e}") from e
    except requests.exceptions.Timeout as e:
        raise TimeoutError(f"请求超时 {url}: {e}") from e


async def async_request(method: str, url: str, **kwargs) -> requests.Response:
    return await asyncio.to_thread(safe_http_request, method, url, **kwargs)


# ============================================================================
# 处理日志（processed_log.json）
# ============================================================================
def load_processed_log() -> Tuple[Set[str], Set[str]]:
    dois = set()
    titles = set()
    if not PROCESSED_LOG.exists():
        return dois, titles
    try:
        with open(PROCESSED_LOG, "r", encoding="utf-8") as f:
            records = json.load(f)
        for rec in records:
            doi = rec.get("doi", "").strip().lower()
            if doi:
                dois.add(doi)
            title = normalize_title(rec.get("title", ""))
            if title:
                titles.add(title)
    except Exception as e:
        logger.warning(f"读取处理日志失败: {e}")
    return dois, titles


def save_processed_log(articles: List[Dict]):
    existing = []
    if PROCESSED_LOG.exists():
        try:
            with open(PROCESSED_LOG, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            pass
    seen = set()
    for rec in existing:
        key = (rec.get("doi", ""), rec.get("title", ""))
        seen.add(key)
    for art in articles:
        key = (art.get("doi", ""), art.get("title", ""))
        if key not in seen:
            existing.append({"doi": art.get("doi", ""), "title": art.get("title", "")})
            seen.add(key)
    with open(PROCESSED_LOG, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)


# ============================================================================
# Zotero 集合 key 缓存
# ============================================================================
async def get_zotero_collection_key() -> Optional[str]:
    global _zotero_collection_key_cache
    if _zotero_collection_key_cache:
        return _zotero_collection_key_cache

    zotero = config.get("zotero", {})
    user_id = zotero.get("user_id")
    api_key = zotero.get("api_key")
    collection_name = zotero.get("collection_name", "agent抓取")
    if not user_id or not api_key:
        return None

    base = f"https://api.zotero.org/users/{user_id}"
    headers = {"Zotero-API-Key": api_key}
    coll_url = f"{base}/collections"
    resp = await async_request("GET", coll_url, headers=headers)
    if resp.status_code == 403:
        logger.error("Zotero API 403 Forbidden — 请检查:")
        logger.error("  1. API Key 是否正确: https://www.zotero.org/settings/keys")
        logger.error("  2. 创建 Key 时是否勾选了「Allow library access」和「Allow write access」")
        logger.error("  3. user_id 是否为纯数字（在 Zotero 设置页面可找到）")
        logger.error(f"  当前配置: user_id={user_id}, api_key={api_key[:6]}...")
        return None
    if resp.status_code != 200:
        logger.error(f"获取 Zotero 集合列表失败: {resp.status_code}")
        return None

    collections = resp.json()
    for c in collections:
        if c["data"]["name"] == collection_name:
            _zotero_collection_key_cache = c["key"]
            return _zotero_collection_key_cache

    logger.info(f"集合 '{collection_name}' 不存在，正在创建...")
    create_resp = await async_request("POST", coll_url, headers=headers,
                                      json=[{"name": collection_name}])
    if create_resp.status_code == 200:
        resp2 = await async_request("GET", coll_url, headers=headers)
        if resp2.status_code == 200:
            for c in resp2.json():
                if c["data"]["name"] == collection_name:
                    _zotero_collection_key_cache = c["key"]
                    return _zotero_collection_key_cache
    logger.error("创建 Zotero 集合失败")
    return None


# ============================================================================
# Zotero 去重数据获取
# ============================================================================
async def fetch_zotero_existing_dois() -> Tuple[Set[str], Set[str]]:
    zotero = config.get("zotero", {})
    user_id = zotero.get("user_id")
    api_key = zotero.get("api_key")
    if not user_id or not api_key:
        logger.error("Zotero 凭据未配置")
        return set(), set()

    coll_key = await get_zotero_collection_key()
    if not coll_key:
        return set(), set()

    base = f"https://api.zotero.org/users/{user_id}"
    headers = {"Zotero-API-Key": api_key}
    items = []
    start = 0
    while True:
        url = f"{base}/collections/{coll_key}/items?limit=100&start={start}"
        resp = await async_request("GET", url, headers=headers)
        if resp.status_code != 200:
            break
        batch = resp.json()
        if not batch:
            break
        items.extend(batch)
        start += 100

    dois = set()
    titles = set()
    for item in items:
        data = item.get("data", {})
        doi = data.get("DOI", "").strip().lower()
        if doi:
            dois.add(doi)
        title = normalize_title(data.get("title", ""))
        if title:
            titles.add(title)
    return dois, titles


# ============================================================================
# Step 1: 打印关键词
# ============================================================================
def print_keywords():
    keywords = config.get("keywords", {})
    broad = keywords.get("broad", [])
    narrow = keywords.get("narrow", [])
    chinese = keywords.get("chinese", [])
    all_kw = [(kw, "宽") for kw in broad] + [(kw, "窄") for kw in narrow] + [(kw, "中") for kw in chinese]
    print("\n===== 关键词矩阵 =====")
    for idx, (kw, scope) in enumerate(all_kw, 1):
        print(f"[{idx}] ({scope}) {kw}")
    print("========================\n")
    return all_kw


# ============================================================================
# Step 2: SPIS 浏览器控制 — 自主搜索 & 人工搜索
# ============================================================================

# 持久化浏览器用户数据目录（保存 SPIS 登录态）
BROWSER_USER_DATA = BASE_DIR / "browser_data"
BROWSER_USER_DATA.mkdir(parents=True, exist_ok=True)


async def open_spis_autonomous(
    keyword: str,
    year_start: int = 2025,
    year_end: int = 2026,
    headless: bool = False,
) -> tuple:
    """【Agent 模式】自主打开 SPIS 并执行搜索 — 无需人工干预。

    关键特性：
    - 使用 persistent context 保存登录 cookies（首次需手动登录）
    - 自动在搜索框输入关键词并点击搜索
    - 自动设置年份筛选（2025-2026）
    - 登录态过期时自动提示

    Returns: (playwright, browser_context, page)
    """
    p = await async_playwright().start()
    try:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(BROWSER_USER_DATA),
            headless=headless,
            viewport={"width": 1280, "height": 800},
            locale="zh-CN",
        )
    except Exception as e:
        logger.error(f"浏览器启动失败: {e}")
        await p.stop()
        raise

    page = await context.new_page()

    try:
        # ── 导航到 SPIS ──
        try:
            await page.goto("https://spis.hnlat.com", timeout=120000, wait_until="domcontentloaded")
        except Exception as e:
            logger.error(f"无法打开 SPIS: {e}")
            await context.close()
            await p.stop()
            raise SystemExit(1)

        await asyncio.sleep(random.uniform(2.0, 3.0))

        # ── 检测是否需要登录（VPN 断开时 SPIS 显示登录页）──
        login_needed = await page.evaluate("""() => {
            const body = document.body.innerText || '';
            // VPN未连接时 SPIS 显示登录页，特征文字：
            //   "账号登录" "手机号登录" "微信扫码" "第三方" "当前IP"
            const loginMarkers = ['账号登录', '手机号登录', '微信扫码', '当前IP'];
            const found = loginMarkers.filter(m => body.includes(m));
            if (found.length >= 2) return true;  // 至少匹配2个 → 确认是登录页
            if (window.location.href.includes('login') || window.location.href.includes('auth')) return true;
            return false;
        }""")

        if login_needed:
            logger.warning("⚠ SPIS 显示登录页（VPN 可能已断开）！请手动连接 VPN。")
            logger.info("   连接 VPN 后自动继续... 等待最多 5 分钟。")
            # 等待 VPN 恢复（轮询检测页面变化）
            for _ in range(60):
                await asyncio.sleep(5)
                still_login = await page.evaluate("""() => {
                    const body = document.body.innerText || '';
                    const loginMarkers = ['账号登录', '手机号登录', '微信扫码', '当前IP'];
                    const found = loginMarkers.filter(m => body.includes(m));
                    return found.length >= 2 || window.location.href.includes('login');
                }""")
                if not still_login:
                    logger.info("✓ VPN 已恢复，继续...")
                    await asyncio.sleep(2)
                    break
            else:
                logger.error("VPN 恢复超时（5分钟），跳过该关键词。")
                await context.close()
                await p.stop()
                return {
                    "collected": [], "by_keyword": {}, "with_links": [],
                    "without_links": [], "help_submitted": 0,
                }

        # ── 等待搜索页面加载 ──
        await asyncio.sleep(random.uniform(1.5, 2.5))

        # ── 输入关键词 ──
        search_input = None
        for selector in [
            ".spis-search-box .search-c .input .ant-input",
            ".spis-search-box .ant-input",
            ".search-c input.ant-input",
            "input.ant-input[type='text']",
        ]:
            try:
                search_input = await page.wait_for_selector(selector, timeout=5000)
                if search_input:
                    logger.info(f"  搜索框定位成功: {selector}")
                    break
            except Exception:
                continue

        if not search_input:
            logger.error("找不到 SPIS 搜索输入框，请确认页面已加载且 DOM 未变更。")
            await context.close()
            await p.stop()
            raise RuntimeError("搜索输入框未找到")

        # 清空 + 输入关键词（模拟真人打字节奏）
        await search_input.click()
        await asyncio.sleep(random.uniform(0.3, 0.6))
        await search_input.fill("")
        await asyncio.sleep(random.uniform(0.2, 0.4))
        await search_input.type(keyword, delay=random.randint(60, 100))
        logger.info(f"  ✓ 已输入关键词: {keyword}")
        await asyncio.sleep(random.uniform(0.8, 1.2))

        # ── 尝试设置年份筛选 ──
        year_filter_set = await _set_year_filter_autonomous(page, year_start, year_end)

        # ── 点击搜索按钮 ──
        search_clicked = False
        for selector in [
            ".spis-search-box .search-button .ant-btn",
            ".spis-search-box .search-button",
            ".search-button .ant-btn",
            "button.search-button",
        ]:
            try:
                btn = await page.query_selector(selector)
                if btn:
                    await btn.click()
                    logger.info(f"  ✓ 已点击搜索按钮: {selector}")
                    search_clicked = True
                    break
            except Exception:
                continue

        if not search_clicked:
            # 回退：按 Enter 触发搜索
            logger.info("  搜索按钮未找到，尝试按 Enter 触发搜索...")
            await search_input.press("Enter")

        # ── 等待搜索结果加载 ──
        try:
            await page.wait_for_selector("article.article", timeout=20000)
            logger.info("  ✓ 搜索结果已加载")
        except Exception:
            logger.warning("  ⚠ 搜索结果加载超时，页面可能无结果或结构变更")

        await asyncio.sleep(random.uniform(1.0, 2.0))

        # 统计结果数
        try:
            article_count = await page.evaluate(
                "() => document.querySelectorAll('article.article').length"
            )
            logger.info(f"  📊 当前页显示 {article_count} 篇文章")
        except Exception:
            pass

        return p, context, page

    except Exception as e:
        logger.error(f"自主搜索异常: {e}")
        try:
            await context.close()
        except Exception:
            pass
        try:
            await p.stop()
        except Exception:
            pass
        raise


async def _set_year_filter_autonomous(page, year_start: int, year_end: int) -> bool:
    """尝试在 SPIS 高级搜索中设置年份筛选。成功返回 True。"""
    try:
        # ── 尝试打开高级搜索弹窗 ──
        advanced_btn = None
        for sel in [
            ".spis-search-box .button .ant-btn",
            ".spis-search-box button",
            "button:has-text('高级')",
            "button:has-text('专业')",
        ]:
            try:
                btns = await page.query_selector_all(sel)
                for b in btns:
                    text = (await b.inner_text()).strip()
                    if "高级" in text or "专业" in text:
                        advanced_btn = b
                        break
                if advanced_btn:
                    break
            except Exception:
                continue

        if advanced_btn:
            await advanced_btn.click()
            await asyncio.sleep(random.uniform(1.0, 1.5))
            logger.info("  ✓ 已打开高级搜索")

            # ── 找年份选择器 ──
            year_picker = None
            for sel in [
                ".search-modal-box .year-row .ant-picker",
                ".search-filter .year-row .ant-picker",
                ".ant-picker",
            ]:
                try:
                    pickers = await page.query_selector_all(sel)
                    for picker in pickers:
                        if await picker.is_visible():
                            year_picker = picker
                            break
                    if year_picker:
                        break
                except Exception:
                    continue

            if year_picker:
                await year_picker.click()
                await asyncio.sleep(random.uniform(0.5, 1.0))
                # 尝试设置年份范围
                try:
                    # 清空并输入起始年
                    await page.keyboard.press("Control+a")
                    await asyncio.sleep(0.2)
                    year_text = f"{year_start}-{year_end}"
                    await page.keyboard.type(year_text, delay=60)
                    await asyncio.sleep(0.5)
                    await page.keyboard.press("Enter")
                    logger.info(f"  ✓ 已设置年份筛选: {year_text}")
                    # 关闭高级搜索弹窗
                    await asyncio.sleep(0.5)
                    await page.keyboard.press("Escape")
                    await asyncio.sleep(0.5)
                    return True
                except Exception as e:
                    logger.warning(f"  年份选择器交互异常: {e}")

    except Exception as e:
        logger.warning(f"  年份筛选设置失败: {e}（将使用后置过滤）")

    return False


async def open_spis_and_wait(keyword: str, index: int, total: int):
    p = await async_playwright().start()
    browser = await p.chromium.launch(headless=False)
    context = await browser.new_context()
    page = await context.new_page()
    try:
        await page.goto("https://spis.hnlat.com", timeout=120000)
    except Exception as e:
        print(f"\n[ERROR] 无法打开 SPIS 网站: {e}")
        print("可能原因：")
        print("  1. 未连接校园 VPN → 请先连接 VPN 后重试")
        print("  2. SPIS 服务器暂时不可用 → 稍后重试")
        print("  3. 网络不通 → 在浏览器中手动访问 https://spis.hnlat.com 确认")
        await browser.close()
        await p.stop()
        raise SystemExit(1)
    print("\n" + "─" * 50)
    print(f">>> 人工环节：关键词组合 {index}/{total} <<<")
    print(f"请搜索: {keyword}")
    print("1. 如尚未登录，请手动登录 SPIS")
    print("2. 将关键词粘贴到搜索框，执行搜索")
    print("3. 确认已看到搜索结果列表后，按回车继续")
    print("─" * 50)
    input()
    return p, browser, context, page


# ============================================================================
# Step 3: 抓取页面
# ============================================================================
def _looks_like_author(text: str) -> bool:
    """检测文本是否看起来像作者名而非期刊名。

    用于策略1/2的候选过滤，避免将"AK Pradeep, A Appel"或"P Gentsch -"误判为期刊。"""
    if not text:
        return True
    # 规则1: 3+逗号分隔 → 多作者列表（如 "Smith J, Wang L, Zhang K"）
    if text.count(",") >= 2 and not any(
        w.lower() in text.lower()
        for w in ["journal", "review", "research", "marketing", "management",
                   "science", "psychology", "quarterly", "studies", "annals",
                   "letters", "report", "bulletin", "ethics"]
    ):
        return True
    # 规则2: 以 " -" 结尾 → 作者名后跟破折号被截断（如 "P Gentsch -"）
    if text.rstrip().endswith(" -"):
        return True
    # 规则3: 单作者模式 "Initial Surname" 或 "Initials Surname"
    # 匹配像 "A Mari", "P Gentsch", "MA Upadhyay" 等单人作者
    stripped = text.rstrip().rstrip("-").strip()
    if re.match(r"^[A-Z][a-z]{0,2}\s+[A-Z][a-z]+\s*$", stripped):
        return True
    # 规则4: 逗号分隔的 "Initial Surname, Initial Surname" 模式
    parts = [p.strip() for p in stripped.split(",")]
    if len(parts) >= 2 and all(
        re.match(r"^[A-Z][a-z]{0,2}\s+[A-Z][a-z]+$", p) for p in parts
    ):
        return True
    return False


def parse_meta(raw_text: str) -> Tuple[str, str]:
    """多策略解析 SPIS 元数据行，提取期刊名和年份。

    核心改进（V6）：
    - 策略0（最高优先级）：直接在文本中搜索白名单期刊名，彻底绕开 SPIS 格式变异。
    - 策略1-3（回退）：正则提取，用于不在白名单中但仍可能有价值的期刊。
    - 通过白名单匹配来规避"提取到作者名而非期刊名"的经典误识别。"""
    if not raw_text:
        return "", ""
    text = unicodedata.normalize('NFKC', raw_text).replace("–", " - ").replace("—", " - ")

    # 提取年份（第一个 202x-203x）
    year = ""
    for m in re.finditer(r"\b(20\d\d)\b", text):
        year = m.group(1)
        break

    # ── 策略0（新增，最高优先级）：直接在文本中搜索白名单期刊名 ──
    # SPIS 常截断期刊名，且正则容易把作者名误判为期刊名。
    # 直接在原始文本中搜索白名单，如果找到则立即返回，无需依赖格式假设。
    journal = _find_whitelist_journal_in_text(text)
    if journal:
        return journal, year

    # ── 策略1: " - Journal Name, 2024" 或 " - Journal Name 2024" ──
    journal = ""
    m = re.search(r"\s[-–—]\s(.+?),?\s*20\d\d", text)
    if m:
        candidate = m.group(1).strip().rstrip("…").rstrip(".").rstrip(",").strip()
        # 二次校验：如果提取结果看起来不像期刊名（太短/像作者名），尝试白名单匹配
        if len(candidate) >= 3:
            resolved = _find_whitelist_journal_in_text(candidate)
            if resolved:
                return resolved, year
            # 防护：排除明显是作者名的候选
            if _looks_like_author(candidate):
                pass  # 看起来像作者名，不采用
            else:
                journal = candidate

    # ── 策略2: "Journal Name, 2024" 格式（无前导破折号）──
    if not journal or len(journal) < 5:
        m = re.search(r"^(.+?),?\s*20\d\d", text.strip())
        if m:
            candidate = m.group(1).strip().rstrip("…").rstrip(".").rstrip(",").strip()
            if len(candidate) > len(journal):
                # 防护：排除明显是作者名的候选（与策略1保持一致）
                if _looks_like_author(candidate):
                    pass  # 作者名，不采用
                else:
                    resolved = _find_whitelist_journal_in_text(candidate)
                    if resolved:
                        return resolved, year
                    journal = candidate

    # ── 策略3: 从括号中提取 ──
    if not journal or len(journal) < 5:
        m = re.search(r"[\(（]([^)）]{5,80})[\)）]", text)
        if m:
            candidate = m.group(1).strip().rstrip("…").rstrip(".").rstrip(",").strip()
            if len(candidate) > len(journal):
                resolved = _find_whitelist_journal_in_text(candidate)
                if resolved:
                    return resolved, year
                journal = candidate

    # 清理：移除首尾的省略号、逗号、句号、空白
    journal = re.sub(r"^[…\.\s]+", "", journal)
    journal = re.sub(r"[,\.…\s]+$", "", journal)

    return journal, year


async def scrape_page(page) -> List[Dict[str, str]]:
    articles = []
    article_els = await page.query_selector_all("article.article")
    for art in article_els:
        title = ""
        for sel in ["div.d-t.jump", "div.allow-ai", "div.jump", "div[title]"]:
            el = await art.query_selector(sel)
            if el:
                title = (await el.get_attribute("title") or "").strip()
                if not title:
                    title = (await el.inner_text()).strip()
                if title:
                    break
        if not title:
            continue

        # ── 标题清洗 ──
        # 1. 去掉 SPIS 序号前缀（"1、", "24、", "29、" 等）
        title = re.sub(r'^\d+[、.．]\s*', '', title)
        # 2. 去掉末尾的省略号和多余空白（但保留用于检测截断）
        title_raw = title  # 保存原始版本用于截断检测
        title = title.rstrip("…。.，, \t")
        # 3. 尝试从 <a> 标签 title 属性获取更完整的标题（SPIS 显示可能截断）
        a_els = await art.query_selector_all("a")
        for a_el in a_els:
            a_title = (await a_el.get_attribute("title") or "").strip()
            a_text = (await a_el.inner_text()).strip()
            # 如果有完整标题且比当前标题更长（去序号前缀后），采用之
            for candidate in [a_title, a_text]:
                clean_candidate = re.sub(r'^\d+[、.．]\s*', '', candidate).rstrip("…。.，, \t")
                if clean_candidate and len(clean_candidate) > len(title):
                    # 候选标题应该包含原标题的核心词（避免匹配到无关文本）
                    if len(title) < 20 or title[:20] in clean_candidate:
                        title = clean_candidate
                        break

        if not title:
            continue

        # ── 增强标题清洗（V11）──
        # 4. 去除尾部作者混入模式
        #    "Title: X. Wang et al" → "Title"
        #    "Title - Author Name, Author Name" → "Title"
        title = re.sub(r'\s*[:：]\s*[A-Z][.\s]*[A-Za-z]+(?:\s+et\s+al\.?)?\s*$', '', title)
        title = re.sub(r'\s*[-–—]\s*[A-Z][a-z]+\s+[A-Z][a-z]+.*$', '', title)
        # 5. 去除括号内的作者年份引用 "(Smith 2024)" 或 "(Wang et al. 2023)"
        title = re.sub(r'\s*[\(（]\s*[A-Z][a-z]+(?:\s+et\s+al\.?)?\s*,?\s*\d{4}\s*[\)）]\s*$', '', title)
        # 6. 如果标题以省略号开头（SPIS严重截断），标记为残缺
        is_garbled = title_raw.startswith("…") or title_raw.startswith("...") or title.startswith("…")

        # 主路径：div.d-a 文本
        meta_el = await art.query_selector("div.d-a")
        meta_text = await meta_el.inner_text() if meta_el else ""
        journal, year = parse_meta(meta_text)

        # 备选路径：当解析不到期刊名时，尝试从 <a> 标签 href/title 提取
        if not journal or len(journal) < 4:
            a_els = await art.query_selector_all("a")
            for a_el in a_els:
                href = (await a_el.get_attribute("href") or "").strip()
                a_title = (await a_el.get_attribute("title") or "").strip()
                a_text = (await a_el.inner_text()).strip()
                for candidate in [a_title, a_text]:
                    if candidate and len(candidate) > 4 and not candidate.startswith("http"):
                        j2, y2 = parse_meta(candidate)
                        if j2 and len(j2) >= len(journal):
                            journal = j2
                        if y2 and not year:
                            year = y2

        # 备选路径2：从 article 全文提取
        if not journal or len(journal) < 4:
            full_text = await art.inner_text()
            # 在全文前 200 字符中寻找可能含期刊名的行
            lines = full_text.split("\n")
            for line in lines[:5]:
                j3, y3 = parse_meta(line)
                if j3 and len(j3) > len(journal):
                    journal = j3
                if y3 and not year:
                    year = y3

        doi = ""
        doi_el = await art.query_selector("a[href*='doi.org']")
        if doi_el:
            doi = (await doi_el.get_attribute("href") or "").strip()
        if not doi:
            full_text = await art.inner_text()
            m = re.search(r"(10\.\d{4,}/[^\s]+)", full_text)
            if m:
                doi = m.group(1).strip()

        # ── 期刊名白名单匹配与补全 ──
        # 如果正则提取没有得到期刊名，或提取结果不在白名单中，
        # 则搜索 article 全文来查找白名单期刊名。
        # 这能解决 SPIS 截断/格式异化导致的"作者名被误判为期刊名"问题。
        if not journal or not journal_in_whitelist(journal):
            # 从 article 全文（前 500 字符足够包含期刊元数据）搜索白名单期刊
            full_text = await art.inner_text()
            found = _find_whitelist_journal_in_text(full_text[:500])
            if found:
                logger.debug(f"全文搜索补全期刊: '{journal}' → '{found}'")
                journal = found
                # 如果还没年份，尝试从全文提取
                if not year:
                    _, year = parse_meta(full_text[:200])

        # ── 期刊白名单过滤（V12：放宽，截断刊名不再直接过滤）──
        # 策略：先尝试补全截断刊名 → 白名单匹配 → 模糊通过 → 最后才过滤
        if journal and not journal_in_whitelist(journal):
            # 尝试补全截断的期刊名
            resolved = _resolve_truncated_journal(journal)
            if resolved:
                logger.debug(f"列表页期刊补全: '{journal}' → '{resolved}'")
                journal = resolved

        if journal and not journal_in_whitelist(journal):
            # 刊名不在白名单 → 判断是否看起来像正经学术期刊（非会议/非垃圾）
            j_lower = journal.lower().strip()
            j_words = j_lower.split()
            is_likely_journal = (
                len(j_words) >= 3  # ≥3 词：通常是完整期刊名
                and not any(w in j_lower for w in [
                    "conference", "proceedings", "symposium", "workshop",
                    "preprint", "archive", "arxiv", "ssrn",
                    "procedia", "ieee", "lecture notes",
                ])
                and not j_lower.startswith("advances in")  # 很多垃圾刊以这个开头
            )
            if is_likely_journal:
                # 看起来像正经期刊，放行 → 详情页补全时会获取完整刊名
                logger.info(f"⚠ [待确认] {title[:50]} | 期刊待详情页确认: {journal}")
            else:
                logger.info(f"✗ [过滤] {title[:50]} | 期刊不在白名单: {journal}")
                continue
        if not journal:
            # 诊断日志：打印原始元数据文本，定位 parse_meta 失败原因
            full_diag = await art.inner_text()
            logger.warning(f"✗ [空期刊] {title[:50]} | 无法识别期刊名，跳过")
            logger.warning(f"  >>> div.d-a 原始文本: {repr(meta_text[:300])}")
            logger.warning(f"  >>> article 全文(前300字符): {repr(full_diag[:300])}")
            continue
        if year:
            min_year = config.get("min_year", 2023)
            if int(year) < min_year:
                logger.info(f"✗ [过滤] {title[:50]} | 年份过旧: {year}")
                continue
        else:
            if journal:  # 有期刊但无年份，同样诊断
                full_diag = await art.inner_text()
                logger.warning(f"⚠ [空年份] {title[:50]} | 无法识别年份")
                logger.warning(f"  >>> div.d-a 原始文本: {repr(meta_text[:300])}")
                logger.warning(f"  >>> article 全文(前300字符): {repr(full_diag[:300])}")

        log_prefix = "⚠ [残缺]" if is_garbled else "✓ [通过]"
        logger.info(f"{log_prefix} {title[:80]} | {journal} | {year}")
        articles.append({
            "title": title,
            "journal": journal,
            "year": year,
            "doi": doi,
            "is_garbled": is_garbled,  # 标记严重截断，供后续优先补全
        })
    return articles


# ============================================================================
# 详情页补全：打开每篇文章的详情页，获取完整标题、期刊、DOI
# ============================================================================
async def _extract_from_detail(page, result: Dict[str, str]):
    """从 SPIS 详情页提取完整元数据，原地写入 result dict。

    基于 SPIS 真实 DOM 结构（V12）：
      - 标题: a.article-title 的 title 属性（完整标题，非视觉截断）
      - 期刊: div.summary-label"来源：" → 相邻 a.jump-link 的 title
      - DOI:  div.summary-label"DOI：" → 相邻 a.jump-link 的 title/href
      - 年份: 来源行内的 <span>
    """
    old_title = result.get("title", "")

    # 等待 React 渲染
    try:
        await page.wait_for_function(
            """() => {
                const root = document.getElementById('root');
                return root && root.innerText.trim().length > 50;
            }""",
            timeout=15000
        )
    except Exception:
        pass
    await asyncio.sleep(1.0)

    try:
        full_text = await page.inner_text("body")
    except Exception:
        return

    # ── 1. 提取完整标题（V12：优先用 title 属性而非 inner_text）──
    title = ""
    title_el = await page.query_selector("a.article-title")
    if title_el:
        # title 属性包含 SPIS 完整标题（inner_text 可能被视觉截断）
        attr_title = (await title_el.get_attribute("title") or "").strip()
        inner_title = (await title_el.inner_text()).strip()
        # 取两者中较长者（title 属性通常更完整）
        if attr_title and len(attr_title) > 15:
            title = attr_title
            logger.info(f"  [诊断] a.article-title[title] → {title[:120]}")
        elif inner_title and len(inner_title) > 15:
            title = inner_title
            logger.info(f"  [诊断] a.article-title[innerText] → {title[:120]}")

    # 回退：其他选择器
    if not title or len(title) < 15:
        for sel in ["h1", ".paper-title", "[class*='detailTitle']", "[class*='paperTitle']",
                     "[class*='title']", ".detail-title", "h2"]:
            try:
                el = await page.query_selector(sel)
                if el:
                    t = (await el.get_attribute("title") or "").strip()
                    if not t or len(t) < 15:
                        t = (await el.inner_text()).strip()
                    if t and len(t) > 15:
                        title = t
                        logger.info(f"  [诊断] 回退选择器 {sel} → {t[:120]}")
                        break
            except Exception:
                continue

    # 最终回退：#root 第一行长文本
    if not title or len(title) < 15:
        try:
            root_text = await page.evaluate("() => document.getElementById('root')?.innerText || ''")
            lines = [l.strip() for l in root_text.split("\n") if len(l.strip()) > 20]
            if lines:
                title = lines[0]
                logger.info(f"  [诊断] #root首行 → {title[:120]}")
        except Exception:
            pass

    if not title or len(title) < 10:
        logger.info(f"  [诊断] 未能提取到有效标题")
        return

    # ── 标题清洗 ──
    title = re.sub(r'^\d+[、.．]\s*', '', title).strip()
    title = re.sub(r'\s*[:：\-–—]\s*[A-Z][.\s]*[A-Za-z]+(?:\s+et\s+al\.?)?\s*$', '', title)
    title = re.sub(r'\s*[-–—]\s*\d{4}\s*$', '', title)
    title = re.sub(r'\s*[:：]\s*[A-Z]\.\s*\w+(?:\s+(?:and|&)\s+[A-Z]\.\s*\w+)?\s*$', '', title)

    # ── 双语标题清洗：保留中文，剥离英文翻译 ──
    # SPIS 对中文论文常显示 "中文标题 / English Title"
    # 检测 "/ " 两侧语言，保留中文侧
    if ' / ' in title:
        parts = title.split(' / ')
        if len(parts) == 2:
            left, right = parts
            left_has_cn = bool(re.search(r'[一-鿿]', left))
            right_has_cn = bool(re.search(r'[一-鿿]', right))
            if left_has_cn and not right_has_cn:
                title = left  # "中文 / English" → 保留中文
                logger.info(f"  [标题清洗] 剥离英文翻译: {title[:80]}")
            elif right_has_cn and not left_has_cn:
                title = right  # "English / 中文" → 保留中文
                logger.info(f"  [标题清洗] 剥离英文前缀: {title[:80]}")

    if title.startswith("…") or title.startswith("..."):
        core = title.lstrip("…。.，, \t")
        if len(core) > 10:
            for line in full_text.split("\n"):
                line_clean = line.strip()
                if core[:30] in line_clean and len(line_clean) > len(title):
                    title = line_clean
                    logger.info(f"  [诊断] 省略号标题补全: {title[:120]}")
                    break

    # ── 安全校验 ──
    title_lower = title.lower()
    old_lower = old_title.lower().rstrip("…。.，, \t")

    is_truncated = (old_lower.startswith("…") or old_lower.endswith("…") or len(old_lower) < 25)
    is_partial = old_lower.startswith("…") and len(old_lower) > 10

    old_words = set(old_lower.split())
    new_words = set(title_lower.split())
    word_overlap = len(old_words & new_words) / max(len(old_words), 1) if old_words else 0

    old_words_list = old_lower.split()
    if len(old_words_list) >= 3:
        old_without_last_2 = " ".join(old_words_list[:-2])
        substr_match = old_without_last_2 in title_lower if len(old_without_last_2) >= 15 else False
    else:
        substr_match = old_lower[:20] in title_lower if len(old_lower) >= 15 else False

    is_related = substr_match or word_overlap > 0.6
    title_actually_changed = title != old_title

    if len(title) >= len(old_title) and title_actually_changed and \
       (is_truncated or is_partial or is_related or not old_title):
        result["title"] = title
        logger.info(f"  ✓ 标题补全: {title[:120]}")
    elif len(title) > len(old_title) and word_overlap > 0.4:
        result["title"] = title
        logger.info(f"  ✓ 标题补全(放宽): {title[:120]}")
    elif not title_actually_changed:
        logger.info(f"  [诊断] 标题已完整 (len={len(title)}), 无需补全")
    else:
        logger.info(f"  [诊断] 标题未替换: 新({len(title)})>旧({len(old_title)})?"
                    f"={len(title)>len(old_title)} 残缺={is_truncated} 词重叠={word_overlap*100:.0f}%")

    # ── 2. 提取期刊名（V13：label来源优先存储，白名单仅用于规范化大小写）──
    old_journal = result.get("journal", "")
    journal_found = False

    # 策略A：通过 "来源：" label 定位（SPIS 详情页 DOM 结构）
    source_journal_raw = ""  # label 提取的原始刊名
    try:
        source_link = await page.evaluate("""() => {
            const labels = document.querySelectorAll('.summary-label');
            for (const label of labels) {
                if (label.textContent.trim() === '来源：') {
                    const parent = label.closest('.item');
                    if (parent) {
                        const link = parent.querySelector('a.jump-link, a.default-text');
                        if (link) {
                            const title = link.getAttribute('title') || '';
                            if (title) return title.trim();
                            return link.textContent.trim();
                        }
                        const span = parent.querySelector('span');
                        if (span) return span.textContent.trim();
                    }
                }
            }
            return '';
        }""")
        if source_link and len(source_link) > 3:
            source_journal_raw = source_link.strip()
            # 优先在白名单中查找（获取规范大小写），找不到也保留原始文本
            j_from_source = _find_whitelist_journal_in_text(source_journal_raw)
            if j_from_source:
                result["journal"] = j_from_source
                journal_found = True
                logger.info(f"  ✓ 期刊补全(label来源): {j_from_source}")
            else:
                # 非白名单期刊也存储，label 来源是权威的
                # V14修复: 只要 label 提取到了有效刊名，就存储它。
                # 不再用 len() 比较——旧刊名可能更长但错误（被 _resolve_truncated_journal 猜错）。
                # label 提取的是 SPIS 详情页的真实刊名，优先级最高。
                result["journal"] = source_journal_raw
                journal_found = True
                logger.info(f"  ✓ 期刊补全(label来源,非白名单): {source_journal_raw}")
    except Exception as e:
        logger.debug(f"  期刊label提取异常: {e}")

    # 策略B：全文搜索白名单期刊名（仅在 label 未找到时）
    if not journal_found:
        found = _find_whitelist_journal_in_text(full_text)
        if not found:
            top_text = "\n".join(full_text.split("\n")[:30])
            found = _find_whitelist_journal_in_text(top_text)
        if found:
            if not old_journal or len(found) > len(old_journal) or not journal_in_whitelist(old_journal):
                result["journal"] = found
                journal_found = True
                logger.info(f"  ✓ 期刊补全(全文搜索): {found}")

    # 策略C：截断期刊名补全（仅当 label 和全文搜索均失败，且旧刊名明显截断）
    if not journal_found and old_journal and len(old_journal) < 30:
        partial = old_journal.strip().lower()
        if len(partial) >= 5:
            # 如果 label 提取到了原始文本但不在白名单，优先用它（而不是猜测）
            if source_journal_raw and len(source_journal_raw) > len(old_journal):
                result["journal"] = source_journal_raw
                journal_found = True
                logger.info(f"  ✓ 期刊补全(label原始): {source_journal_raw}")
            else:
                resolved = _resolve_truncated_journal(partial)
                if resolved:
                    result["journal"] = resolved
                    journal_found = True
                    logger.info(f"  ✓ 期刊补全(截断修复): {resolved}")

    # ── 3. 提取 DOI（V12：通过 "DOI：" label 定位）──
    if not result.get("doi"):
        doi_extracted = False
        try:
            doi_val = await page.evaluate("""() => {
                const labels = document.querySelectorAll('.summary-label');
                for (const label of labels) {
                    if (label.textContent.trim().startsWith('DOI')) {
                        const parent = label.closest('.item');
                        if (parent) {
                            const link = parent.querySelector('a.jump-link');
                            if (link) {
                                const title = link.getAttribute('title') || '';
                                if (title && title.includes('doi.org')) return title;
                                const href = link.getAttribute('href') || '';
                                if (href && href.includes('doi.org')) return href;
                                return link.textContent.trim();
                            }
                        }
                    }
                }
                return '';
            }""")
            if doi_val and "10." in doi_val:
                # 提取标准 DOI 格式
                m = re.search(r"(10\.\d{4,}/[^\s\"'<>；;，,]+)", doi_val)
                if m:
                    result["doi"] = m.group(1).rstrip(".,;)")
                    doi_extracted = True
                    logger.info(f"  ✓ DOI补全(label): {result['doi']}")
        except Exception as e:
            logger.debug(f"  DOI label提取异常: {e}")

        # 回退：全文正则匹配
        if not doi_extracted:
            m = re.search(r"(10\.\d{4,}/[^\s\"'<>；;，,]+)", full_text)
            if m:
                result["doi"] = m.group(1).rstrip(".,;)")
                doi_extracted = True
                logger.info(f"  ✓ DOI补全(正则): {result['doi']}")

    # ── 4. 提取年份（V12：来源行内 span）──
    if not result.get("year"):
        try:
            yr_val = await page.evaluate("""() => {
                const labels = document.querySelectorAll('.summary-label');
                for (const label of labels) {
                    if (label.textContent.trim() === '来源：') {
                        const parent = label.closest('.item');
                        if (parent) {
                            const spans = parent.querySelectorAll('span');
                            for (const span of spans) {
                                const t = span.textContent.trim();
                                if (/^20[2-9]\\d$/.test(t)) return t;
                            }
                            // 回退：匹配文本中的年份
                            const text = parent.textContent;
                            const m = text.match(/20[2-9]\\d/);
                            if (m) return m[0];
                        }
                    }
                }
                return '';
            }""")
            if yr_val and yr_val.isdigit() and 2020 <= int(yr_val) <= 2030:
                result["year"] = yr_val
                logger.info(f"  ✓ 年份补全: {yr_val}")
        except Exception:
            pass

    # ── 4. 提取年份 ──
    if not result.get("year"):
        for m in re.finditer(r"\b(20[2-9]\d)\b", full_text[:500]):
            yr = int(m.group(1))
            if 2020 <= yr <= 2030:
                result["year"] = str(yr)
                break

    # ── 5. 提取作者 ──
    if not result.get("authors") and not result.get("author"):
        try:
            authors = await page.evaluate("""() => {
                // 策略A：通过 "作者：" label 定位（与期刊/DOI提取一致）
                const labels = document.querySelectorAll('.summary-label');
                for (const label of labels) {
                    if (label.textContent.trim() === '作者：') {
                        const parent = label.closest('.item');
                        if (parent) {
                            // 作者可能在 a.jump-link 或 span 中
                            const links = parent.querySelectorAll('a.jump-link, a.default-text');
                            if (links.length > 0) {
                                return Array.from(links).map(l => l.textContent.trim()).join(', ');
                            }
                            const span = parent.querySelector('span');
                            if (span) return span.textContent.trim();
                            // 直接取父元素文本（去掉"作者："前缀）
                            const text = parent.textContent.trim();
                            return text.replace(/^作者：\\s*/, '');
                        }
                    }
                }
                // 策略B：查找含 "作者" 文本的任意元素
                const allEls = document.querySelectorAll('.item, .summary-item, [class*=\"author\"]');
                for (const el of allEls) {
                    const text = el.textContent.trim();
                    if (text.startsWith('作者：') || text.startsWith('作者:')) {
                        return text.replace(/^作者[：:]\\s*/, '');
                    }
                }
                return '';
            }""")
            if authors and len(authors) > 2:
                result["authors"] = authors.strip()
                result["author"] = authors.strip()  # 兼容两种 key
                logger.info(f"  ✓ 作者补全: {authors[:100]}")
        except Exception:
            pass


# ============================================================================
# 详情页下载链接检测 & 文献求助自动提交
# ============================================================================
async def _check_detail_download_url(tab) -> Optional[str]:
    """检测详情页是否有可用的下载链接。返回下载 URL 或 None。"""
    try:
        download_url = await tab.evaluate("""() => {
            // 方式1: 下载按钮中的链接
            const downloadBtn = document.querySelector('.action-button.download a');
            if (downloadBtn && downloadBtn.href) return downloadBtn.href;

            // 方式2: 下载地址卡片中的链接
            const downloadCard = document.querySelector('[id*="download-card"]');
            if (downloadCard) {
                const links = downloadCard.querySelectorAll('a.jump-link');
                for (const link of links) {
                    if (link.href && link.href.length > 20) return link.href;
                }
            }

            // 方式3: ResearchGate 等外部下载链接
            const cardContent = document.querySelector('.card-content');
            if (cardContent) {
                const links = cardContent.querySelectorAll('a');
                for (const link of links) {
                    if (link.href && (link.href.includes('researchgate') ||
                                      link.href.includes('.pdf') ||
                                      link.href.includes('download'))) {
                        return link.href;
                    }
                }
            }
            return '';
        }""")
        if download_url and len(download_url) > 10:
            return download_url.strip()
    except Exception:
        pass
    return None


async def _auto_submit_literature_help(tab, article: Dict[str, str], help_email: str) -> bool:
    """自动提交 SPIS 文献求助表单（邮箱已由系统预设，无需填写）。

    模拟真人操作节奏：慢速滚动 → 勾选条款 → 确认 → 等待。
    返回 True 表示提交成功，False 表示失败。
    """
    logger.info(f"  🔍 [文献求助] 开始自动提交: {article['title'][:60]}")

    try:
        # ── Step 1: 切换到「文献求助」tab ──
        help_tab_clicked = await tab.evaluate("""() => {
            const tabs = document.querySelectorAll('.ant-tabs-tab');
            for (const t of tabs) {
                if (t.textContent.includes('文献求助') || t.textContent.includes('help')) {
                    t.click();
                    return true;
                }
            }
            const cardTabs = document.querySelectorAll('[data-node-key="help-card"]');
            if (cardTabs.length > 0) { cardTabs[0].click(); return true; }
            return false;
        }""")
        if help_tab_clicked:
            logger.info("  ✓ 已点击「文献求助」tab")
        else:
            logger.warning("  ⚠ 未找到文献求助 tab，尝试继续...")
        await asyncio.sleep(random.uniform(1.5, 2.5))

        # ── Step 2: 模拟真人滚动 ──
        await tab.evaluate("""() => {
            const d = 100 + Math.random() * 200;
            window.scrollBy({top: d, behavior: 'smooth'});
        }""")
        await asyncio.sleep(random.uniform(0.8, 1.5))

        # ── Step 3: 检查并填写邮箱（Playwright 逐字输入，模拟真人）──
        email_filled = False
        # 先诊断：打印页面上所有可见的 input 元素，方便定位
        try:
            diag_inputs = await tab.evaluate("""() => {
                const inputs = document.querySelectorAll('input:not([type="hidden"])');
                const result = [];
                inputs.forEach((inp, i) => {
                    if (inp.offsetParent !== null) {
                        result.push({
                            idx: i,
                            type: inp.type || 'text',
                            placeholder: inp.placeholder || '',
                            value: inp.value || '',
                            className: inp.className || '',
                            id: inp.id || '',
                            name: inp.name || '',
                            ariaLabel: inp.getAttribute('aria-label') || '',
                        });
                    }
                });
                return result;
            }""")
            if diag_inputs:
                logger.info(f"  [诊断] 页面上可见 input 共 {len(diag_inputs)} 个:")
                for inp in diag_inputs:
                    logger.info(f"    [{inp['idx']}] type={inp['type']} placeholder='{inp['placeholder']}' "
                                f"value='{inp['value'][:30]}' class='{inp['className'][:60]}' "
                                f"id='{inp['id']}' name='{inp['name']}'")
            else:
                logger.warning("  [诊断] 页面上没有可见的 input 元素")
        except Exception as e:
            logger.warning(f"  [诊断] input 诊断失败: {e}")

        # 宽泛查找邮箱输入框：遍历所有可见 input，找 type=email 或含"邮箱"/"email"属性的
        try:
            all_inputs = await tab.query_selector_all('input:not([type="hidden"])')
            for inp in all_inputs:
                try:
                    if not await inp.is_visible():
                        continue
                    inp_type = (await inp.get_attribute("type") or "").lower()
                    placeholder = (await inp.get_attribute("placeholder") or "").lower()
                    aria_label = (await inp.get_attribute("aria-label") or "").lower()
                    inp_name = (await inp.get_attribute("name") or "").lower()
                    inp_id = (await inp.get_attribute("id") or "").lower()
                    combined = f"{inp_type} {placeholder} {aria_label} {inp_name} {inp_id}"
                    current_val = (await inp.input_value()).strip()

                    # 已经预设了邮箱 → 跳过填写
                    if current_val and "@" in current_val:
                        logger.info(f"  ✓ 邮箱已预设: {current_val}")
                        email_filled = True
                        break

                    # 判断这行是否像邮箱输入框
                    is_email_field = (
                        inp_type == "email"
                        or "邮箱" in combined or "email" in combined or "mail" in combined
                        or "e-mail" in combined
                    )
                    if not is_email_field:
                        continue

                    # 已确认是邮箱字段 → 填写
                    if current_val:
                        logger.info(f"  [诊断] 邮箱字段当前值='{current_val}'（非邮箱格式），覆盖填写")

                    await inp.click()
                    await asyncio.sleep(random.uniform(0.2, 0.4))
                    await tab.keyboard.press("Control+a")
                    await asyncio.sleep(random.uniform(0.1, 0.2))
                    await tab.keyboard.type(help_email, delay=random.randint(60, 100))
                    logger.info(f"  ✓ 邮箱已逐字输入: {help_email}")
                    email_filled = True
                    break
                except Exception:
                    continue
        except Exception as e:
            logger.warning(f"  [诊断] 输入框遍历异常: {e}")

        # 回退：直接定位包含"邮箱"文本的元素旁边的 input
        if not email_filled:
            try:
                fallback_filled = await tab.evaluate(f"""(email) => {{
                    // 策略1: 找包含"邮箱"文字的 label/div，再找相邻 input
                    const allEls = document.querySelectorAll('*');
                    for (const el of allEls) {{
                        if (el.children.length === 0 && el.textContent.trim() === '邮箱：') {{
                            // 向上找父容器，再找其中的 input
                            let parent = el.parentElement;
                            for (let i = 0; i < 5 && parent; i++) {{
                                const inp = parent.querySelector('input:not([type="hidden"])');
                                if (inp && inp.offsetParent !== null) {{
                                    inp.focus();
                                    inp.value = '';
                                    inp.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                    // 模拟逐字输入
                                    const chars = email.split('');
                                    chars.forEach((ch, idx) => {{
                                        setTimeout(() => {{
                                            inp.value += ch;
                                            inp.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                        }}, idx * 80);
                                    }});
                                    return true;
                                }}
                                parent = parent.parentElement;
                            }}
                        }}
                    }}
                    // 策略2: Ant Design Form.Item 里找
                    const formItems = document.querySelectorAll('.ant-form-item, .ant-row');
                    for (const item of formItems) {{
                        const label = item.querySelector('.ant-form-item-label label');
                        if (label && label.textContent.includes('邮箱')) {{
                            const inp = item.querySelector('input');
                            if (inp && inp.offsetParent !== null) {{
                                inp.focus();
                                inp.value = '';
                                const chars = email.split('');
                                chars.forEach((ch, idx) => {{
                                    setTimeout(() => {{
                                        inp.value += ch;
                                        inp.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                    }}, idx * 80);
                                }});
                                return true;
                            }}
                        }}
                    }}
                    return false;
                }}""", help_email)
                if fallback_filled:
                    # setTimeout 方式不可靠，但回退策略至少尝试了
                    await asyncio.sleep(len(help_email) * 0.08 + 1.0)
                    logger.info(f"  ✓ 邮箱已填写(回退策略): {help_email}")
                    email_filled = True
            except Exception:
                pass

        if not email_filled:
            logger.warning("  ⚠ 未找到邮箱输入框，尝试继续（可能无需填写）...")
        await asyncio.sleep(random.uniform(0.5, 1.0))

        # ── Step 4: 勾选《文献求助服务条款》──
        # 优先用 Playwright 原生点击（更真实），失败回退到 JS evaluate
        checkbox_checked = False
        try:
            cb_el = await tab.query_selector('.custom-checkbox-img:not(.checked)')
            if not cb_el:
                cb_el = await tab.query_selector('input[type="checkbox"]:not(:checked)')
            if cb_el and await cb_el.is_visible():
                await cb_el.click(delay=random.randint(100, 250))
                checkbox_checked = True
                logger.info("  ✓ 已勾选《文献求助服务条款》(原生点击)")
        except Exception:
            pass

        if not checkbox_checked:
            checkbox_checked = await tab.evaluate("""() => {
                const cbs = document.querySelectorAll('.custom-checkbox-img, input[type="checkbox"]');
                for (const cb of cbs) {
                    const p = cb.closest('.reason-checkbox') || cb.closest('label') || cb.parentElement;
                    const t = (p?.textContent || '').toLowerCase();
                    if (t.includes('条款') || t.includes('服务') || t.includes('agree') || t.includes('同意')) {
                        if (cb.classList?.contains('custom-checkbox-img') && !cb.classList.contains('checked')) { cb.click(); return true; }
                        if (cb.tagName === 'INPUT' && !cb.checked) { cb.click(); return true; }
                    }
                }
                for (const cb of cbs) {
                    if (cb.classList?.contains('custom-checkbox-img') && !cb.classList.contains('checked')) { cb.click(); return true; }
                    if (cb.tagName === 'INPUT' && !cb.checked) { cb.click(); return true; }
                }
                return false;
            }""")
        if checkbox_checked:
            logger.info("  ✓ 已勾选《文献求助服务条款》")
        else:
            logger.warning("  ⚠ 未找到服务条款复选框")
        await asyncio.sleep(random.uniform(0.5, 1.0))

        # ── Step 5: 点击「确认」按钮（Playwright 原生优先）──
        confirm_clicked = False
        try:
            for btn_sel in ['button', '.ant-btn', '.modal-btn']:
                buttons = await tab.query_selector_all(btn_sel)
                for btn in buttons:
                    try:
                        text = (await btn.inner_text()).strip()
                        if text in ('确认', '提交', '确定') and await btn.is_visible():
                            # 先慢速移动到按钮（模拟鼠标轨迹）
                            box = await btn.bounding_box()
                            if box:
                                await tab.mouse.move(
                                    box['x'] + random.uniform(10, box['width'] - 10),
                                    box['y'] + random.uniform(5, box['height'] - 5),
                                    steps=random.randint(3, 8),
                                )
                                await asyncio.sleep(random.uniform(0.2, 0.5))
                            await btn.click(delay=random.randint(50, 150))
                            confirm_clicked = True
                            logger.info(f"  ✓ 已点击「{text}」按钮 (原生)")
                            break
                    except Exception:
                        continue
                if confirm_clicked:
                    break
        except Exception:
            pass

        if not confirm_clicked:
            confirm_clicked = await tab.evaluate("""() => {
                const btns = document.querySelectorAll('button, .ant-btn, .modal-btn');
                for (const b of btns) {
                    const t = b.textContent.trim();
                    if ((t === '确认' || t === '提交' || t === '确定') && b.offsetParent !== null) {
                        b.click(); return true;
                    }
                }
                return false;
            }""")
        if confirm_clicked:
            logger.info("  ✓ 已点击「确认」按钮")
        else:
            logger.warning("  ⚠ 未找到确认按钮")

        # ── Step 6: 等待提交完成 ──
        await asyncio.sleep(random.uniform(2.0, 3.5))

        # 检测成功提示
        success = await tab.evaluate("""() => {
            const b = document.body.innerText;
            return b.includes('成功') || b.includes('已提交') || b.includes('提交成功');
        }""")
        if success:
            logger.info(f"  ✓ [文献求助] 提交成功: {article['title'][:60]}")
        else:
            logger.info(f"  ✓ [文献求助] 已提交（等待确认）: {article['title'][:60]}")

        return True

    except Exception as e:
        logger.error(f"  ✗ [文献求助] 自动提交异常: {e}")
        return False


async def _enrich_articles_from_detail(page, articles: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """逐篇点击文章标题打开详情页，采集完整元数据。

    V14 改进：
    - 有下载链接的详情页不关闭，留待后续手动操作。
    - 无下载链接的文献自动通过「文献求助」提交至预设邮箱。
    - 模拟真人操作节奏，避免触发反爬。
    """
    if not articles:
        return articles

    # 优先处理残缺标题（is_garbled）的文章 — 原地重排
    garbled = [a for a in articles if a.get("is_garbled")]
    normal = [a for a in articles if not a.get("is_garbled")]
    articles[:] = garbled + normal
    if garbled:
        logger.info(f"开始详情页补全 ({len(articles)} 篇, 其中 {len(garbled)} 篇标题残缺优先)...")
    else:
        logger.info(f"开始详情页补全 ({len(articles)} 篇)...")
    logger.info("  键盘: P=暂停 G=继续 E=终止")
    logger.info("  功能: 自动检测下载链接，无下载则自动文献求助 → 18922596828@163.com")
    enriched_list = []
    list_url = page.url
    help_email = config.get("help_email", "18922596828@163.com")
    download_pages = []  # 保存有下载链接的详情页，不关闭
    help_submitted = 0

    for i, art in enumerate(articles):
        # ── 键盘控制 ──
        _poll_keyboard()
        await _kb_wait_if_paused()
        if _kb_is_terminated():
            logger.info("用户终止详情补全，保留已处理结果")
            break

        logger.info(f"详情 [{i+1}/{len(articles)}]: {art['title'][:60]}...")
        enriched = dict(art)

        # 每次循环重新查询 DOM
        try:
            await page.wait_for_selector("article.article", timeout=10000)
        except Exception:
            logger.warning("  列表页 DOM 未恢复，跳过")
            enriched_list.append(enriched)
            continue

        article_els = await page.query_selector_all("article.article")

        # 标题文本指纹匹配
        art_el = None
        target_title = art["title"]
        for el in article_els:
            try:
                el_text = await el.inner_text()
                if target_title[:40] in el_text.strip()[:80].replace("\n", " "):
                    art_el = el
                    break
            except Exception:
                continue
        if not art_el:
            if i < len(article_els):
                art_el = article_els[i]
            else:
                logger.warning(f"  找不到文章元素 #{i+1}")
                enriched_list.append(enriched)
                continue

        current_url = page.url

        # 找到可点击元素
        click_el = None
        for sel in ["div.d-t.jump", "div.allow-ai", "div.jump", "div.d-t"]:
            try:
                el = await art_el.query_selector(sel)
                if el:
                    click_el = el
                    break
            except Exception:
                continue
        if not click_el:
            click_el = art_el

        # ── 点击 + 轮询等导航 ──
        new_page = None
        url_changed = False

        try:
            async with page.context.expect_page(timeout=5000) as new_page_info:
                await click_el.click(force=True)
                # 同时轮询 URL 变化（SPA 路由跳转）
                for _ in range(15):
                    await asyncio.sleep(0.5)
                    try:
                        if page.url != current_url:
                            url_changed = True
                            break
                    except Exception:
                        pass
                    _poll_keyboard()
                    await _kb_wait_if_paused()
                if not url_changed:
                    try:
                        new_page = await new_page_info.value
                    except Exception:
                        pass
        except Exception:
            pass

        # ── 如果还没变化，继续等 ──
        if not url_changed and not new_page:
            for _ in range(20):
                await asyncio.sleep(0.5)
                try:
                    if page.url != current_url:
                        url_changed = True
                        break
                except Exception:
                    pass
                _poll_keyboard()

        logger.info(f"  URL变化={url_changed} 新标签={new_page is not None} "
                    f"({current_url[:60]} → {page.url[:60] if url_changed else '(未变)'})")

        # ── 确定当前操作的 tab（新标签页 or 当前页）──
        detail_tab = new_page if new_page else page

        # ── 提取数据 ──
        if new_page:
            try:
                await new_page.wait_for_selector(
                    "h1, h2, .title, [class*='detail'], strong",
                    timeout=15000
                )
                await asyncio.sleep(1.0)
                await _extract_from_detail(new_page, enriched)
            except Exception as e:
                logger.warning(f"  新标签页提取异常: {e}")
        elif url_changed:
            try:
                await page.wait_for_selector(
                    "h1, h2, .title, [class*='detail'], strong",
                    timeout=15000
                )
            except Exception:
                pass
            await asyncio.sleep(1.0)
            await _extract_from_detail(page, enriched)
        else:
            # 检查弹窗
            modal_sel = ".modal, .dialog, .drawer, [role='dialog']"
            try:
                modal_el = await page.query_selector(modal_sel)
            except Exception:
                modal_el = None
            if modal_el:
                await asyncio.sleep(0.5)
                await _extract_from_detail(page, enriched)
                try:
                    await page.keyboard.press("Escape")
                except Exception:
                    pass
            else:
                logger.info(f"  所有策略均未触发导航，保留列表页数据")

        # ── V14: 检测下载链接 → 决定关闭 or 文献求助 ──
        download_url = await _check_detail_download_url(detail_tab)
        has_download = download_url is not None

        if has_download:
            logger.info(f"  📥 下载链接: {download_url[:80]}...")
            # 有新标签页的就留着不关
            if new_page:
                download_pages.append(new_page)
                logger.info(f"  📌 保留详情页（有下载链接）")
            elif url_changed:
                # SPA 跳转的情况，先回到列表页
                logger.info(f"  📌 有下载链接，返回列表页继续")
                try:
                    await page.go_back(timeout=10000)
                except Exception:
                    await page.goto(list_url, timeout=15000)
                await asyncio.sleep(random.uniform(1.0, 2.0))
        else:
            # 无下载链接 → 自动文献求助
            logger.info(f"  📭 无下载链接 → 自动文献求助")
            if new_page:
                # 在新标签页中操作文献求助
                help_ok = await _auto_submit_literature_help(new_page, enriched, help_email)
                if help_ok:
                    help_submitted += 1
                # 文献求助提交后关闭标签页
                try:
                    await asyncio.sleep(random.uniform(1.0, 2.0))
                    await new_page.close()
                except Exception:
                    pass
            elif url_changed:
                # 当前页操作文献求助
                help_ok = await _auto_submit_literature_help(page, enriched, help_email)
                if help_ok:
                    help_submitted += 1
                # 返回列表页
                await asyncio.sleep(random.uniform(1.0, 2.0))
                try:
                    await page.go_back(timeout=10000)
                except Exception:
                    await page.goto(list_url, timeout=15000)
                await asyncio.sleep(random.uniform(1.0, 2.0))
            else:
                try:
                    if new_page:
                        await new_page.close()
                except Exception:
                    pass

        enriched_list.append(enriched)

    # ── 清理：关闭所有保留的下载详情页 ──
    if download_pages:
        logger.info(f"关闭 {len(download_pages)} 个保留的下载详情页...")
        for dp in download_pages:
            try:
                await dp.close()
            except Exception:
                pass

    # 统计（分别统计标题补全、DOI补全、期刊补全）
    improved_titles = sum(1 for a, e in zip(articles, enriched_list)
                          if e.get("title", "") != a.get("title", ""))
    improved_dois = sum(1 for a, e in zip(articles, enriched_list)
                        if e.get("doi") and not a.get("doi"))
    improved_journals = sum(1 for a, e in zip(articles, enriched_list)
                            if e.get("journal") != a.get("journal"))
    # 统计严重截断标题的修复率
    garbled_count = sum(1 for a in articles if a.get("is_garbled"))
    garbled_fixed = sum(1 for a, e in zip(articles, enriched_list)
                        if a.get("is_garbled") and len(e.get("title", "")) > len(a.get("title", "")))
    if garbled_count:
        logger.info(f"详情补全完成: 标题补全 {improved_titles}, DOI补全 {improved_dois}, "
                    f"期刊补全 {improved_journals}, 残缺修复 {garbled_fixed}/{garbled_count}, "
                    f"文献求助 {help_submitted}")
    else:
        logger.info(f"详情补全完成: 标题补全 {improved_titles}, DOI补全 {improved_dois}, "
                    f"期刊补全 {improved_journals}, 文献求助 {help_submitted}")
    return enriched_list


async def scrape_keyword_until_full(
    page,
    keyword: str,
    global_doi_set: Set[str],
    global_title_set: Set[str],
    target_count: int,
    max_pages: int = MAX_PAGES_PER_KEYWORD,
) -> List[Dict[str, str]]:
    new_collected = []
    consecutive_no_new = 0
    for pg in range(max_pages):
        _poll_keyboard()
        await _kb_wait_if_paused()
        if _kb_is_terminated():
            break
        await asyncio.sleep(random.uniform(MIN_SLEEP, MAX_SLEEP))
        page_articles = await scrape_page(page)
        if ENRICH_FROM_DETAIL and page_articles:
            page_articles = await _enrich_articles_from_detail(page, page_articles)
            # V15: 详情页补全后严格过滤 — 期刊必须在白名单内，确保高价值文献
            filtered_before = len(page_articles)
            filtered_out = []
            kept = []
            for a in page_articles:
                j = a.get("journal", "")
                if journal_in_whitelist(j):
                    kept.append(a)
                else:
                    filtered_out.append(a)
                    logger.info(f"✗ [V15过滤] {a['title'][:60]} | 期刊不在白名单: {j}")
            if filtered_out:
                logger.info(f"V15严格过滤: {len(filtered_out)}/{filtered_before} 篇非白名单期刊已移除，"
                           f"保留 {len(kept)} 篇高价值文献")
            page_articles = kept
        page_added = 0
        for art in page_articles:
            doi = art["doi"].strip().lower()
            norm_title = normalize_title(art["title"])
            if doi and doi in global_doi_set:
                logger.info(f"⊘ [去重DOI] {art['title'][:50]} | DOI 已存在，跳过")
                continue
            title_dup = False
            dup_title = ""
            dup_char_sim = 0.0
            dup_word_jac = 0.0
            for gt in global_title_set:
                is_dup, char_sim, word_jac = is_title_duplicate(norm_title, gt)
                if is_dup:
                    title_dup = True
                    dup_title = gt
                    dup_char_sim = char_sim
                    dup_word_jac = word_jac
                    break
            if title_dup:
                logger.info(f"⊘ [去重标题] {art['title'][:60]} | "
                           f"字符相似{dup_char_sim*100:.0f}% 词Jaccard{dup_word_jac*100:.0f}% → 与: {dup_title[:60]}")
                continue
            new_collected.append(art)
            if doi:
                global_doi_set.add(doi)
            global_title_set.add(norm_title)
            page_added += 1
            logger.info(f"✓ 新文献 (累计{len(new_collected)}): {art['title'][:120]} | {art.get('journal','')} | {art.get('year','')}")

        if page_added == 0:
            consecutive_no_new += 1
            if consecutive_no_new >= 3:
                logger.info(f"关键词 [{keyword}] 连续 {consecutive_no_new} 页无新文献，停止翻页。")
                break
        else:
            consecutive_no_new = 0

        # ── 达到目标后提示用户：继续翻同一关键词，还是停止 ──
        if len(new_collected) >= target_count:
            global _keyboard_state
            print(f"\n  本关键词已收集 {len(new_collected)} 篇（目标 {target_count} 篇），"
                  f"当前第 {pg+1}/{max_pages} 页。")
            print(f"  浏览器保持打开，可手动下载PDF。")
            print(f"  [G] 继续翻页检索（同一关键词）  [E] 停止本关键词")
            _keyboard_state = "paused"
            while _keyboard_state == "paused":
                await asyncio.sleep(0.3)
                _poll_keyboard()
            if _keyboard_state == "terminated":
                break
            # G 被按下：继续当前关键词后续页面，不再每页提示
            target_count = 9999

        # ── 翻页（V9：精确选择器，基于 SPIS 真实 HTML） ──
        # SPIS 分页结构:
        #   <button class="pagination-arrow first disabled"><</button>
        #   <button class="pagination-number active">N</button>
        #   <button class="pagination-number">N+1</button> ...
        #   <button class="pagination-arrow">></button>

        # 取当前页码
        current_num = pg + 1  # 回退默认值
        try:
            active_el = await page.query_selector("button.pagination-number.active")
            if active_el:
                txt = (await active_el.inner_text()).strip()
                if txt.isdigit():
                    current_num = int(txt)
        except Exception as e:
            logger.debug(f"取当前页码失败: {e}")

        next_num = current_num + 1
        clicked = False

        # ── 方式1：找文本为 '>' 的 pagination-arrow 按钮 ──
        #     不用 :not(.first) 避免 CSS 兼容问题，直接按文本匹配
        all_arrows = await page.query_selector_all("button.pagination-arrow")
        next_arrow = None
        for arrow in all_arrows:
            try:
                txt = (await arrow.inner_text()).strip()
                if txt == ">":
                    cls = (await arrow.get_attribute("class") or "").lower()
                    if "first" not in cls and "disabled" not in cls:
                        next_arrow = arrow
                        break
            except Exception:
                continue

        if next_arrow:
            try:
                await next_arrow.scroll_into_view_if_needed()
                await asyncio.sleep(0.3)
                await next_arrow.click()
                logger.debug(f"已点击 > 箭头，等待翻页到第{next_num}页...")
                # 轮询等 active 页码变成 next_num（最多 10 秒）
                for _ in range(20):
                    await asyncio.sleep(0.5)
                    new_active = await page.query_selector("button.pagination-number.active")
                    if new_active:
                        new_txt = (await new_active.inner_text()).strip()
                        if new_txt.isdigit() and int(new_txt) == next_num:
                            clicked = True
                            logger.info(f"翻页成功 → 箭头 (第{current_num}页→第{new_txt}页)")
                            break
                if not clicked:
                    logger.warning(f"箭头点击后未检测到页码变化 (期望第{next_num}页)")
            except Exception as e:
                logger.warning(f"箭头翻页异常: {e}")

        # ── 方式2：点击具体页码按钮 ──
        if not clicked:
            target_btn = None
            all_nums = await page.query_selector_all("button.pagination-number")
            for btn in all_nums:
                try:
                    txt = (await btn.inner_text()).strip()
                    if txt == str(next_num) and "active" not in ((await btn.get_attribute("class") or "").lower()):
                        target_btn = btn
                        break
                except Exception:
                    continue

            if target_btn:
                try:
                    await target_btn.scroll_into_view_if_needed()
                    await asyncio.sleep(0.3)
                    await target_btn.click()
                    logger.debug(f"已点击页码按钮 {next_num}，等待激活...")
                    for _ in range(20):
                        await asyncio.sleep(0.5)
                        new_active = await page.query_selector("button.pagination-number.active")
                        if new_active:
                            new_txt = (await new_active.inner_text()).strip()
                            if new_txt.isdigit() and int(new_txt) == next_num:
                                clicked = True
                                logger.info(f"翻页成功 → 页码按钮 (第{current_num}页→第{new_txt}页)")
                                break
                    if not clicked:
                        logger.warning(f"页码按钮点击后未检测到激活 (期望第{next_num}页)")
                except Exception as e:
                    logger.warning(f"页码翻页异常: {e}")
            else:
                logger.info(f"第{next_num}页按钮不在当前分页窗口内（仅显示前几页），且箭头不可用")

        if not clicked:
            logger.info(f"翻页失败 (第{current_num}页→第{next_num}页)，停止该关键词。")
            break

        # 等文章加载
        try:
            await page.wait_for_selector("article.article", timeout=15000)
        except PlaywrightTimeout:
            pass
        await asyncio.sleep(1)
    return new_collected


# ============================================================================
# PDF 下载
# ============================================================================
async def download_via_unpaywall(doi: str, email: str) -> Optional[str]:
    url = f"https://api.unpaywall.org/v2/{doi}?email={email}"
    resp = await async_request("GET", url)
    if resp.status_code != 200:
        return None
    data = resp.json()
    best = data.get("best_oa_location") or {}
    return best.get("url_for_pdf")


async def download_via_semantic_scholar(doi: str) -> Optional[str]:
    encoded = urllib.parse.quote(doi, safe="")
    url = f"https://api.semanticscholar.org/graph/v1/paper/DOI:{encoded}?fields=openAccessPdf"
    resp = await async_request("GET", url)
    if resp.status_code == 429:
        await asyncio.sleep(2)
        resp = await async_request("GET", url)
    if resp.status_code != 200:
        return None
    data = resp.json()
    oa = data.get("openAccessPdf") or {}
    return oa.get("url")


async def download_pdf(pdf_url: str, title: str) -> Optional[Path]:
    try:
        resp = await async_request("GET", pdf_url, stream=True)
        if resp.status_code != 200:
            return None
        PDF_DIR.mkdir(parents=True, exist_ok=True)
        filename = clean_filename(title) + ".pdf"
        filepath = PDF_DIR / filename
        with open(filepath, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                f.write(chunk)
        logger.info(f"PDF 下载成功: {filename}")
        return filepath
    except Exception as e:
        logger.error(f"下载 PDF 失败: {e}")
        return None


# ============================================================================
# Zotero 入库
# ============================================================================
def _parse_authors(author_str: str) -> List[Dict]:
    """将 AI 返回的作者字符串解析为 Zotero creator 数组。
    自动去除「（第一作者）」「（通讯作者）」等标注。"""
    authors = []
    if not author_str:
        return authors
    # 去除中文括号标注：如 "Name（第一作者）" → "Name"
    author_str = re.sub(r'[（(][^)）]*[）)]', '', author_str)
    # 处理 "Name1, Name2, Name3 and Name4" 格式
    names = re.split(r',\s*(?:and\s+)?|\s+and\s+', author_str)
    for name in names:
        name = name.strip().rstrip(".").strip()
        if not name or len(name) < 2:
            continue
        parts = name.rsplit(" ", 1)
        if len(parts) == 2 and len(parts[0]) > 0 and len(parts[1]) > 0:
            authors.append({"creatorType": "author", "firstName": parts[0], "lastName": parts[1]})
        else:
            authors.append({"creatorType": "author", "name": name})
    return authors


async def zotero_add_item(article: Dict, pdf_path: Path, notes: Optional[Dict] = None) -> bool:
    """
    Zotero 入库：通过 API 创建条目 + 将 PDF 写入本地 Zotero 存储目录。

    适配 WebDAV / 坚果云 用户：
      - 通过 API 创建父条目 + 附件条目（仅元数据）
      - 直接将 PDF 复制到 C:\\Users\\...\\Zotero\\storage\\{att_key}\\
      - Zotero 客户端自动检测新文件并同步到坚果云

    不再使用 Zotero S3 上传（files.zotero.net），因为 WebDAV 用户不从那里下载文件。

    返回 True 仅当条目创建 + 本地文件复制均成功。
    """
    zotero = config.get("zotero", {})
    user_id = zotero.get("user_id")
    api_key = zotero.get("api_key")
    if not user_id or not api_key:
        return False
    coll_key = await get_zotero_collection_key()
    if not coll_key:
        return False

    # Zotero 本地数据目录（默认路径，可通过 config 覆盖）
    zotero_data_dir = Path(config.get("zotero_data_dir", os.path.expandvars(r"%USERPROFILE%\Zotero")))
    storage_dir = zotero_data_dir / "storage"

    base = f"https://api.zotero.org/users/{user_id}"

    # ── 构建元数据（极简：只用爬取获得的标题+DOI，其余由 Zotero 自动检索补全）──
    # 不使用 AI 生成的标题/作者/摘要 — AI 可能附加中文翻译破坏 Green Frog/Jasminum 解析
    title = article.get("title", "")
    # 安全清洗：万一标题含 " / 翻译" 则剥离非主导语言侧
    if ' / ' in title:
        parts = title.split(' / ')
        if len(parts) == 2:
            left, right = parts
            left_cn = bool(re.search(r'[一-鿿]', left))
            right_cn = bool(re.search(r'[一-鿿]', right))
            if left_cn and not right_cn:
                title = left
            elif right_cn and not left_cn:
                title = right
    doi = article.get("doi", "").strip()
    # 期刊和年份从爬取数据取（可靠），AI 不覆盖
    pub_title = article.get("journal", "")
    date = article.get("year", "")

    json_headers = {"Zotero-API-Key": api_key, "Content-Type": "application/json"}

    # ── Step 1: 创建父条目 (journalArticle) — 极简元数据，让 Zotero 根据 DOI 自动补全 ──
    item_data = [{
        "itemType": "journalArticle",
        "title": title,
        "DOI": doi,
        "publicationTitle": pub_title,
        "date": date,
        "collections": [coll_key],
    }]
    resp = await async_request("POST", f"{base}/items", headers=json_headers, json=item_data)
    if resp.status_code != 200:
        logger.error(f"Zotero 创建条目失败: {resp.status_code} {resp.text[:200]}")
        return False
    item_key = resp.json()["success"]["0"]
    logger.info(f"Zotero 条目已创建: {title[:50]}")

    # ── Step 2: 创建附件条目 (imported_file) ──
    #    附带文件元数据（md5/mtime/size），Zotero 客户端根据这些信息识别文件
    pdf_size = pdf_path.stat().st_size
    pdf_md5 = hashlib.md5(pdf_path.read_bytes()).hexdigest()
    mtime_ms = int(pdf_path.stat().st_mtime * 1000)

    attach_data = [{
        "itemType": "attachment",
        "linkMode": "imported_file",
        "title": pdf_path.name,
        "contentType": "application/pdf",
        "filename": pdf_path.name,
        "parentItem": item_key,
    }]
    att_resp = await async_request("POST", f"{base}/items", headers=json_headers, json=attach_data)
    if att_resp.status_code != 200:
        logger.error(f"创建附件条目失败: {att_resp.status_code} {att_resp.text[:200]}")
        logger.error(f"   父条目 {item_key} 已创建但无附件，请手动添加 PDF 至 Zotero")
        return False
    att_result = att_resp.json()
    if "successful" not in att_result:
        logger.error(f"附件返回异常: {att_result}")
        return False
    att_entry = att_result["successful"]["0"]
    att_key = att_entry["key"] if isinstance(att_entry, dict) else att_entry

    # ── Step 3: 将 PDF 复制到 Zotero 本地存储目录 ──
    #    路径格式: {zotero_data}/storage/{att_key}/{filename}
    #    这样 Zotero 客户端能检测到文件，并将其同步到 WebDAV/坚果云
    try:
        target_dir = storage_dir / att_key
        target_dir.mkdir(parents=True, exist_ok=True)
        target_path = target_dir / pdf_path.name

        # 复制文件（保留原始修改时间）
        shutil.copy2(pdf_path, target_path)

        # 验证复制后的 MD5
        target_md5 = hashlib.md5(target_path.read_bytes()).hexdigest()
        if target_md5 != pdf_md5:
            logger.error(f"PDF 复制校验失败: 期望 {pdf_md5[:12]} 实际 {target_md5[:12]}")
            return False

        logger.info(f"PDF 已写入 Zotero 本地存储: {target_path}")
        logger.info(f"   下次 Zotero 同步时，该文件将自动上传至坚果云/WebDAV")
    except Exception as e:
        logger.error(f"PDF 写入 Zotero 存储目录失败: {e}")
        return False

    return True


# ============================================================================
# Zotero 同步确认（人工检查环节）
# ============================================================================
async def confirm_zotero_sync(items: List[Dict]) -> bool:
    """
    在通过 API 将文献提交至 Zotero 后，等待用户确认同步完成。
    只有当用户明确确认 Zotero 已完成入库后，才返回 True。

    非交互模式 (--yes):
      显示 30 秒倒计时，给用户时间手动同步 Zotero 客户端，然后自动继续。
      用户可以按 Ctrl+C 中断以跳过确认。

    返回 True → 用户确认完成，继续后续处理
    返回 False → 用户选择跳过，不要标记为已处理
    """
    if not items:
        return True

    # 检查配置是否启用确认步骤
    if not config.get("confirm_zotero_sync", True):
        logger.info("配置 confirm_zotero_sync=false，跳过 Zotero 同步确认。")
        return True

    print("\n" + "=" * 60)
    print("  [等待] Zotero 入库确认  --  请验证附件已正确上传  ")
    print("=" * 60)
    print(f"以下 {len(items)} 篇文献已通过 API 提交至 Zotero：")
    print("-" * 60)
    for i, item in enumerate(items, 1):
        print(f"  [{i}] {item.get('title', 'Unknown')[:70]}")
        print(f"      DOI: {item.get('doi', 'N/A')}")
        print(f"      期刊: {item.get('journal', 'N/A')} ({item.get('year', 'N/A')})")
    print("=" * 60)
    print()
    print("请执行以下操作以完成 Zotero 入库：")
    print("  1. 打开 Zotero 桌面客户端")
    print("  2. 点击同步按钮 (Sync with zotero.org)")
    print("  3. 在「agent抓取」集合中确认上述文献已出现")
    print("  4. 确认每篇文献的 PDF 附件图标正常（非虚线圈）")
    print()
    print("Zotero 会自动完成：")
    print("  [OK] PDF 文件云端备份")
    print("  [OK] 自动抓取 DOI 元数据（如作者、摘要等补充信息）")
    print()

    if _non_interactive:
        countdown = 30
        print(f"[--yes 模式] {countdown} 秒倒计时，请立即同步 Zotero...")
        print("(按 E 可终止)")
        for remaining in range(countdown, 0, -1):
            print(f"\r  [WAIT] 剩余 {remaining:2d} 秒...", end="", flush=True)
            _poll_keyboard()
            if _kb_is_terminated():
                print("\n⏹ 用户终止")
                return False
            await asyncio.sleep(1)
        print("\r  [OK] 倒计时结束，自动继续。")
        logger.info("非交互模式：倒计时结束，自动确认 Zotero 同步。")
        return True
    else:
        print("确认 Zotero 同步完成且附件正常后，按回车继续...")
        print("(输入 'skip' 跳过确认，本次将 不 标记为已处理)")
        _poll_keyboard()
        if _kb_is_terminated():
            print("\n⏹ 用户终止")
            return False
        try:
            response = input(">>> ").strip().lower()
        except EOFError:
            logger.warning("无法读取输入 (EOFError)，跳过确认。")
            return False
        if response == 'skip':
            logger.warning("用户跳过 Zotero 同步确认 — 本次文献将不会被标记为已处理。")
            return False
        logger.info("用户确认 Zotero 同步完成，继续后续处理。")
        return True


# ============================================================================
# 影响因子查询
# ============================================================================
def get_impact_factor(journal: str) -> str:
    """
    根据期刊名查询影响因子。
    优先级：本地 journal_if.json > OpenAlex API > 返回空字符串
    """
    if not journal or not journal.strip():
        return ""

    # 1) 本地精确匹配
    j_key = journal.lower().strip().rstrip(".").rstrip(",")
    if j_key in _journal_if_map:
        return str(_journal_if_map[j_key])

    # 2) 本地模糊匹配（子串包含）
    for jn, if_val in _journal_if_map.items():
        if j_key in jn or jn in j_key:
            logger.info(f"影响因子模糊匹配: '{journal}' ≈ '{jn}' → IF={if_val}")
            return str(if_val)

    # 3) OpenAlex API 回退（异步版本在需要时由调用方使用）
    logger.info(f"期刊 '{journal}' 未在本地 IF 库中，将留空（可手动补充至 journal_if.json）")
    return ""


async def get_impact_factor_async(journal: str) -> str:
    """
    异步版本：先查本地，再查 OpenAlex API。
    OpenAlex 提供 `2yr_mean_citedness` 作为 JCR IF 的近似值。
    """
    local = get_impact_factor(journal)
    if local:
        return local

    # 尝试 OpenAlex API
    if not journal or not journal.strip():
        return ""
    try:
        encoded = urllib.parse.quote(journal.strip())
        url = f"https://api.openalex.org/sources?search={encoded}&per_page=3"
        async with aiohttp.ClientSession() as session:
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=15)) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    results = data.get("results", [])
                    if results:
                        # 提取 best match 的 2yr_mean_citedness
                        for r in results:
                            display_name = r.get("display_name", "")
                            # 名称相似度检查
                            if journal.lower().strip()[:20] in display_name.lower() \
                               or display_name.lower()[:20] in journal.lower().strip():
                                metrics = r.get("summary_stats", {})
                                # OpenAlex 2yr_mean_citedness 约等于 JCR IF
                                if_val = metrics.get("2yr_mean_citedness")
                                if if_val and if_val > 0:
                                    logger.info(f"OpenAlex IF: '{journal}' → {if_val:.2f}")
                                    return f"{if_val:.2f}"
    except Exception as e:
        logger.warning(f"OpenAlex IF 查询失败: {e}")
    return ""


# ============================================================================
# AI 笔记生成
# ============================================================================
def extract_text_from_pdf(pdf_path: Path) -> str:
    doc = None
    try:
        doc = fitz.open(pdf_path)
        text = ""
        for page in doc:
            text += page.get_text()
        return text
    except Exception as e:
        logger.error(f"提取 PDF 文本失败: {e}")
        return ""
    finally:
        if doc is not None:
            try:
                doc.close()
            except Exception:
                pass


async def generate_notes(text: str) -> Optional[Dict[str, str]]:
    """
    使用 DeepSeek V4 Pro API 对论文全文进行学术级深度提取。
    无回退模式——DeepSeek 不可用时直接报错，确保输出质量。

    防幻觉流水线:
    1. DeepSeek 生成 17 字段笔记
    2. 第一轮快速自检 → PASS 则直接入库
    3. 第一轮 FAIL → 第二轮逐条验证 (PASS/WRONG/UNCLEAR)
    4. 变量交叉校验 (规则引擎，不消耗 API)
    5. 置信度标注 + 复核队列

    返回包含 17 字段的 dict（含 _自检标记, _置信度, _验证详情），失败时返回 None。
    """
    deepseek_key = config.get("deepseek_api_key", "")
    deepseek_model = config.get("deepseek_model", "deepseek-chat")

    if not deepseek_key:
        logger.error("DeepSeek API Key 未配置！请在 config.json 中设置 deepseek_api_key。")
        return None

    logger.info(f"调用 DeepSeek API ({deepseek_model}) 生成笔记...")
    notes = await _generate_via_deepseek(text, deepseek_key, deepseek_model)

    if not notes:
        return None

    all_issues = []

    # ── 1. 自检 + 二次验证 ──
    check_result = await _self_check_notes(text, notes, deepseek_key, deepseek_model)
    confidence = check_result.get("confidence", "high")
    notes["_置信度"] = confidence
    notes["_自检标记"] = ""

    if check_result.get("verified"):
        wrong_items = [v for v in check_result["verified"] if v["verdict"] == "WRONG"]
        unclear_items = [v for v in check_result["verified"] if v["verdict"] == "UNCLEAR"]
        notes["_验证详情"] = check_result["verified"]

        if wrong_items:
            wrong_msg = "; ".join(f"WRONG: {w['detail'][:100]}" for w in wrong_items)
            all_issues.append(wrong_msg)
            logger.warning(f"[防幻觉] {len(wrong_items)} 条 WRONG，建议人工复核后决定是否入库")
        if unclear_items:
            unclear_msg = "; ".join(f"UNCLEAR: {u['detail'][:100]}" for u in unclear_items)
            all_issues.append(unclear_msg)
    else:
        notes["_验证详情"] = []

    if check_result.get("issues") and not check_result.get("verified"):
        # 二次验证未执行（API 失败），保留第一轮问题
        all_issues.extend(check_result["issues"])

    # ── 2. 变量交叉校验 ──
    var_warnings = await _cross_validate_variables(text, notes)
    if var_warnings:
        all_issues.extend(var_warnings)
        # 变量问题 → 降级置信度（自检可能 PASS 但变量校验 FAIL）
        if confidence == "high":
            confidence = "medium"
            notes["_置信度"] = confidence
            logger.warning(f"[防幻觉] 变量校验发现问题，置信度降级: HIGH → MEDIUM")

    # ── 3. 汇总标记 ──
    if all_issues:
        notes["_自检标记"] = " | ".join(all_issues)
        logger.warning(f"[防幻觉] 总共 {len(all_issues)} 个问题/警告，置信度: {confidence.upper()}")
    else:
        logger.info("[防幻觉] 全部通过 ✓ 置信度: HIGH")

    # ── 4. 写入复核队列（有问题就进队列，不管置信度高低）──
    if all_issues or confidence in ("low", "medium"):
        _add_to_review_queue(notes)

    return notes


async def _generate_via_deepseek(text: str, api_key: str, model: str) -> Optional[Dict[str, str]]:
    """通过 DeepSeek V4 Pro API（OpenAI 兼容）进行文献提取。
    使用中文 prompt，直接输出中文 key 的 JSON，无需映射。"""
    # 清洗 PDF 文本中的 surrogate 字符和 Unicode 非字符，防止 API HTTP 400
    text = re.sub(r'[\ud800-\udfff￾-￿]', '', text)
    system_prompt = """你是一位专注于人工智能与市场营销交叉领域的资深教授。你正在深度阅读一篇学术论文，为其制作精读笔记，纳入你的个人学术知识体系。

这份笔记的价值在于：未来你需要回顾这篇论文时，不必重读原文，仅凭笔记就能准确还原论文的核心内容——包括它的理论逻辑、研究设计、关键数据和你的批判性思考。

核心纪律：
1. 只写原文中明确出现的信息。不推断、不补充、不美化。
2. 数字——样本量、统计量、p值、均值、alpha系数——必须与原文严格一致。n=207就是207，不是"约200"。
3. 原文没有提到的内容，不要提及。不要写"原文未报告XXX"——如果原文没报告，你的笔记里自然就没有它。
4. 深度来自对原文的透彻理解与清晰转述，而非添加原文中没有的细节。原文详实则笔记详实，原文简略则笔记简略。
5. 笔记使用中文撰写。专业术语（理论名、变量名、量表名、统计方法名）保留英文并括号标注中文。
6. 仅输出合法的 JSON 对象。"""


    user_prompt = f"""你是一位专注于人工智能与市场营销交叉领域的资深教授，正在为这篇论文制作深度精读笔记。这份笔记是你的个人学术知识储备——未来回顾这篇论文时，你不必重读原文，仅凭笔记就能准确还原它的核心内容。

## 核心纪律
你写下的每一个事实性陈述，都必须能在原文中找到具体出处。数字与原文严格一致。原文没有提到的内容，你的笔记中就不会出现——不需要标注"原文未报告"，它不存在于原文，自然也不存在于你的笔记。不要为了显得"详细"而脑补。原文信息翔实你就写翔实，原文写得简略你就如实简略。

## 笔记语言
中文撰写。专业术语（理论名、变量名、量表名、统计方法名）保留英文并括号标注中文。论文标题保持英文原文。

## 字段要求
返回 JSON，18 个字段：

━━━ 基础信息 ━━━
1. 标题：论文完整英文标题。
2. 作者：按原文顺序列出。标注第一作者和通讯作者："Name（第一作者）, Name（通讯作者）"。原文未标注通讯作者则不加。
3. 第一作者
4. 通讯作者
5. 年份（4位数字）
6. 期刊全称（英文）
7. 影响因子：留空
8. 分区：留空
9. 关键词：用原文关键词的语言。术语统一使用全称，禁止附加缩写（用"large language models"而非"LLMs"或"large language models (LLMs)"）。分号分隔。

━━━ 核心内容 ━━━
10. 研究背景与动机：
    交代清楚：①本文针对的现实痛点或理论缺口；②为什么这个问题重要；③已有研究的核心进展和尚存的矛盾/空白；④本文的切入点。像写文献综述一样有逻辑脉络。

11. 研究问题：
    ①列出核心研究问题（RQ）或假设（H）。②每个假设/问题的理论推导逻辑。③研究的理论框架：核心构念（construct）有哪些、它们之间的关系，以及这些构念在框架中的理论角色（前因/结果/中介/调节）。

12. 变量汇总：
    若论文有明确界定的变量/构念（定量研究通常有自变量、因变量、中介变量、调节变量、控制变量等），逐一列出。每个变量独立成段，包含：
    - 变量名称：英文原文（中文翻译）
    - 变量类型：自变量 / 因变量 / 中介变量 / 调节变量 / 控制变量
    - 概念定义：论文中对该变量的理论定义。若原文有明确定义则直接引用；若原文未单独给出定义，则从测量题项中归纳其操作化含义（如"该构念通过xxx等题项测量，反映的是xxx"）。严禁写"论文中未提供明确的理论定义"——这是敷衍。每个变量都必须有实质性的定义内容。
    - 测量方式：量表名称、题项数、来源文献、Cronbach's alpha 值（原文提供了才写，未提供则不写）

    若论文为质性研究、概念性论文或未设置明确变量的文献，此节留空。不要为了填充而编造变量。

13. 研究方法：
    像教授向同事复述一篇论文的方法——清晰、准确、不遗漏关键细节，但不套模板。按论文自身逻辑组织，自然分段。覆盖：研究设计及理由、数据来源与样本（数字须准确）、实验程序/数据收集流程（如适用）、分析方法及选用理由。（注意：各构念的测量细节已在「变量汇总」中列出，此处无需重复。）

14. 方法论详解：
    这是本篇笔记中技术含量最高的章节。目标：未来你读到这篇笔记时，即使忘了这篇论文用了什么方法，看完这一节就能完全理解其方法论逻辑，甚至可以向学生讲解这些方法。

    **按论文类型，覆盖以下内容（不是你全部都要写——选论文实际用到的）：**

    ━━━ 定量实证（含实验）━━━
    ① **建模方法**：论文使用了什么模型？（如：OLS回归、Logit/Probit、结构方程模型SEM、多层线性模型HLM、固定效应面板、双重差分DID、断点回归RDD、工具变量IV、Heckman选择模型、倾向得分匹配PSM、合成控制法、机器学习（随机森林/XGBoost/神经网络）等。）这个模型的核心思想是什么，为什么适合回答本文的研究问题？
    ② **识别策略**（因果推断论文必须有）：作者如何识别因果关系？用什么 variation？（如：外生冲击、政策变化、自然实验、field experiment、实验室实验。）识别假设是什么？
    ③ **内生性处理**（如有）：论文是否讨论了内生性问题？（遗漏变量、反向因果、测量误差、自选择。）作者如何解决？（工具变量——IV是什么、为什么有效、是否通过弱工具变量检验；Heckman两步法；固定效应；匹配方法；等等。）每种方法的核心逻辑用一两句话讲清楚。
    ④ **模型设定**：因变量和自变量的操作化方式。非线性变换（如对数变换）的理由。交互项的含义。关键控制变量。函数形式的选择理由。
    ⑤ **估计细节**：标准误类型（稳健标准误、聚类标准误——聚类层级及理由）。多重共线性诊断（如有）。模型选择标准（AIC/BIC/交叉验证等，如有）。
    ⑥ **稳健性检验**：作者做了哪些稳健性检验？每种检验的核心逻辑和结论。替换因变量/自变量度量方式、替换样本、替换模型、安慰剂检验、平行趋势检验（DID）、排除替代解释等。不要只列名字——每个检验说明它的目的和通过意味着什么。

    ━━━ 质性研究 ━━━
    ① **方法论取向**：扎根理论/现象学/案例研究/叙事分析/民族志/内容分析等。为什么选择这个取向？
    ② **抽样策略**：目的性抽样/理论抽样/滚雪球/最大变异抽样等。为什么这样抽样？
    ③ **数据收集**：访谈（半结构化/深度/焦点小组）、观察、档案、多源数据。如何确保数据丰富性？
    ④ **编码与分析**：编码策略（开放编码/主轴编码/选择性编码）、主题分析步骤。如何从原始数据提炼主题？理论饱和度如何判断？
    ⑤ **可信度保障**：三角验证、成员检查、同行汇报、审计追踪、反身性（研究者自我反思）等。如何确保研究质量？

    ━━━ 概念性/综述论文 ━━━
    ① **理论构建逻辑**：论文如何从已有文献推导出新框架？核心论证结构是什么？
    ② **文献筛选方法**（系统综述）：数据库、检索式、纳入/排除标准、PRISMA流程图描述。偏倚风险评估。
    ③ **分析方法**（文献计量/元分析）：使用的工具（VOSviewer/Bibliometrix/CiteSpace等）、分析维度（共被引/共词/耦合等）。元分析的效应量计算方式、异质性检验（Q统计量/I²）、出版偏倚检验。

    ━━━ 共通的 ━━━
    ⑦ **方法局限性**：本文方法层面的内在局限（不是泛泛的"样本量小"，而是该方法本身有什么不足？如DID的平行趋势假设、IV的排他性约束、实验的外部效度等）。
    ⑧ **方法贡献**：本文在方法上有什么独到之处？新奇的测量方式？巧妙的研究设计？独特的数据集？

    **关键要求**：这不是简单罗列方法名称——是像教授给学生讲解研究方法课一样，把每个方法的**核心逻辑**、**为什么选它**、**它的关键假设**、**本文如何满足（或不满足）这些假设**讲透。好的方法论讲解让读者"原来如此"，不好的只列名词和引用。

15. 研究结果：
    定量研究（有明确假设）：描述性统计→逐条假设检验（每条独立成段："H1（简述）：支持/不支持。统计量值, p = 值, 效应量。"）→中介/调节效应→附加分析。
    质性研究：按主题组织，每个主题独立成段。
    最后用"核心发现"总结3-5条。

15. 讨论与结论：
    作者如何解释核心结果？解释的逻辑是否合理？与已有文献的关系（一致/不一致？作者如何调和？）。理论含义和实践含义。有值得商榷或质疑之处吗？

16. 创新点：
    本文的边际贡献在哪里？理论创新/方法创新/实践创新？逐条列出，每条另起一行。

17. 局限与展望：
    逐条列出，每条局限和每个未来研究方向各占一行。格式如下：
    - 局限性：
    （1）第一条局限...
    （2）第二条局限...
    - 未来研究方向：
    （1）第一个方向...
    （2）第二个方向...
    作者明确提出的都写，你认为值得探索的补充在最后并标注。（注：不是简单翻译原文列表——每条需包含简要解释，说明为什么这是局限或为什么这个方向值得探索。）

**输出格式**
仅输出一个合法的 JSON 对象，键名与上述字段名完全一致，不要添加任何额外的解释文字或 Markdown 标记。示例结构：
{{
  "标题": "",
  "作者": "",
  "第一作者": "",
  "通讯作者": "",
  "年份": "",
  "期刊": "",
  "影响因子": "",
  "分区": "",
  "关键词": "",
  "研究背景与动机": "",
  "研究问题": "",
  "变量汇总": "",
  "研究方法": "",
  "方法论详解": "",
  "研究结果": "",
  "讨论与结论": "",
  "创新点": "",
  "局限与展望": ""
}}

[论文全文]
{text}"""

    try:
        async with aiohttp.ClientSession() as session:
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            }
            payload = {
                "model": model,
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                "temperature": 0.1,
                "max_tokens": 8192,
            }
            async with session.post(
                "https://api.deepseek.com/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=aiohttp.ClientTimeout(total=120),
            ) as resp:
                if resp.status != 200:
                    logger.error(f"DeepSeek API 错误: HTTP {resp.status}")
                    return None
                result = await resp.json()
                content = result["choices"][0]["message"]["content"]
                json_match = re.search(r"\{.*\}", content, re.DOTALL)
                if not json_match:
                    logger.error(f"DeepSeek 返回中未找到合法 JSON")
                    return None
                raw = json.loads(json_match.group())
                # 直接返回中文 key 的 dict（新 prompt 已直接输出中文 key）
                # 补全缺失字段的默认值
                return {
                    "标题": raw.get("标题", ""),
                    "作者": raw.get("作者", ""),
                    "第一作者": raw.get("第一作者", ""),
                    "通讯作者": raw.get("通讯作者", ""),
                    "年份": raw.get("年份", ""),
                    "期刊": raw.get("期刊", ""),
                    "影响因子": "",  # 由外部 get_impact_factor() 填充
                    "分区": raw.get("分区", ""),
                    "关键词": raw.get("关键词", ""),
                    "研究背景与动机": raw.get("研究背景与动机", ""),
                    "研究问题": raw.get("研究问题", ""),
                    "变量汇总": raw.get("变量汇总", ""),
                    "研究方法": raw.get("研究方法", ""),
                    "方法论详解": raw.get("方法论详解", ""),
                    "研究结果": raw.get("研究结果", ""),
                    "讨论与结论": raw.get("讨论与结论", ""),
                    "创新点": raw.get("创新点", ""),
                    "局限与展望": raw.get("局限与展望", ""),
                }
    except Exception as e:
        logger.error(f"DeepSeek API 调用异常: {e}")
        return None


async def _self_check_notes(text: str, notes: Dict[str, str], api_key: str, model: str) -> Dict[str, any]:
    """深度阅读后自检 + 二次验证。

    第一轮：快速检查（轻量 API 调用）
    第二轮：若 FAIL，逐条二次验证（原文片段 vs 笔记表述 → PASS/WRONG/UNCLEAR）

    返回 {"issues": [...], "verified": [...], "confidence": "high"|"medium"|"low"}
    失败不阻塞主流程。"""
    result = {"issues": [], "verified": [], "confidence": "high"}

    # ═══════════════════════════════════════════════
    # 第一轮：快速自检
    # ═══════════════════════════════════════════════
    text_sample = text[:8000]
    notes_check = {
        "变量汇总": notes.get("变量汇总", ""),
        "研究方法": notes.get("研究方法", ""),
        "方法论详解": notes.get("方法论详解", ""),
        "研究结果": notes.get("研究结果", ""),
    }

    check_prompt = f"""你是学术审稿人。请用一两句话快速检查这份精读笔记是否存在以下问题：

1. **变量遗漏**：原文是定量实证研究，笔记"变量汇总"却为空或明显不完整？
2. **数字错误**：笔记中的样本量、统计量（β/SE/t/p/α/R²）、百分比是否与原文一致？
3. **编造内容**：笔记中是否有原文不存在的事实性陈述？

原文片段（前8000字符）：
{text_sample}

笔记关键字段：
{json.dumps(notes_check, ensure_ascii=False, indent=2)}

请回复：
- 如果笔记质量合格，回复 "PASS"
- 如果发现问题，回复 "FAIL: <具体问题>" (每条一行)"""

    round1_issues = []
    try:
        async with aiohttp.ClientSession() as session:
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            }
            payload = {
                "model": model,
                "messages": [
                    {"role": "system", "content": "你是严谨的学术审稿人。只回复判定结果，不添加客套话。"},
                    {"role": "user", "content": check_prompt},
                ],
                "temperature": 0.0,
                "max_tokens": 512,
            }
            async with session.post(
                "https://api.deepseek.com/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=aiohttp.ClientTimeout(total=60),
            ) as resp:
                if resp.status != 200:
                    logger.warning(f"自检 API 错误: HTTP {resp.status}")
                else:
                    resp_data = await resp.json()
                    content = resp_data["choices"][0]["message"]["content"].strip()
                    if content.upper().startswith("PASS"):
                        logger.info("[自检] 第一轮通过 ✓")
                        return result  # 全过，直接返回
                    else:
                        round1_issues = [l.strip("- ") for l in content.split("\n") if l.strip() and any(
                            kw in l for kw in ["FAIL", "错误", "遗漏", "编造", "不一致", "缺失", "问题"])]
                        # 过滤假阳性：模型说"此项可接受"、"无问题"、"与原文一致"的实际上是通过
                        round1_issues = [
                            iss for iss in round1_issues
                            if not any(skip in iss for skip in ["可接受", "无问题", "无编造", "与原文一致", "一致，无", "合理，无", "正确，无"])
                        ]
                        for issue in round1_issues:
                            logger.warning(f"[自检·第一轮] {issue}")
    except Exception as e:
        logger.warning(f"自检第一轮异常（不阻塞）: {e}")
        return result

    if not round1_issues:
        return result

    # ═══════════════════════════════════════════════
    # 第二轮：逐条二次验证
    # ═══════════════════════════════════════════════
    logger.info("[自检] 第一轮发现 %d 个问题，启动二次验证...", len(round1_issues))
    result["confidence"] = "medium"
    round2_verified = []

    # 提取原文中与问题相关的片段
    verify_prompt = f"""你是学术审稿人，正在进行严格的笔记质量审核。

以下笔记可能存在事实性问题。请逐条验证每条问题是否属实。

**原文**（前12000字符，含方法+结果段落）：
{text[:12000]}

**笔记片段**：
研究结果: {notes.get("研究结果", "")[:1500]}
研究方法: {notes.get("研究方法", "")[:1000]}
变量汇总: {notes.get("变量汇总", "")[:1000]}

**待验证问题**：
{chr(10).join(f"{i+1}. {issue}" for i, issue in enumerate(round1_issues))}

请逐条验证。每条回复格式：
[N] 判定: PASS | WRONG | UNCLEAR
    原文原句: "..."
    笔记表述: "..."
    理由: 一句话"""

    try:
        async with aiohttp.ClientSession() as session:
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            }
            payload = {
                "model": model,
                "messages": [
                    {"role": "system", "content": "你是严格的学术审稿人。只回复判定结果，逐条验证，不跳过任何一条。"},
                    {"role": "user", "content": verify_prompt},
                ],
                "temperature": 0.0,
                "max_tokens": 1024,
            }
            async with session.post(
                "https://api.deepseek.com/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=aiohttp.ClientTimeout(total=90),
            ) as resp:
                if resp.status != 200:
                    logger.warning(f"二次验证 API 错误: HTTP {resp.status}")
                    result["issues"] = round1_issues  # 回退：保留第一轮问题
                    return result
                resp_data = await resp.json()
                content = resp_data["choices"][0]["message"]["content"].strip()

                # 解析二次验证结果
                for line in content.split("\n"):
                    line_stripped = line.strip()
                    if "WRONG" in line_stripped:
                        # 提取问题描述
                        verified_issue = line_stripped.split("WRONG", 1)[-1].strip().lstrip(":").strip()
                        if verified_issue:
                            round2_verified.append({"verdict": "WRONG", "detail": verified_issue[:200]})
                            logger.warning(f"[自检·WRONG] {verified_issue[:120]}")
                    elif "UNCLEAR" in line_stripped:
                        verified_issue = line_stripped.split("UNCLEAR", 1)[-1].strip().lstrip(":").strip()
                        if verified_issue:
                            round2_verified.append({"verdict": "UNCLEAR", "detail": verified_issue[:200]})
                            logger.info(f"[自检·UNCLEAR] {verified_issue[:120]}")
                    elif "PASS" in line_stripped:
                        logger.info(f"[自检·PASS] {line_stripped[:120]}")

                # WRONG 数量决定置信度
                wrong_count = sum(1 for v in round2_verified if v["verdict"] == "WRONG")
                if wrong_count > 0:
                    result["confidence"] = "low"
                    logger.warning(f"[自检] 二次验证发现 {wrong_count} 条 WRONG，置信度: LOW")
                elif any(v["verdict"] == "UNCLEAR" for v in round2_verified):
                    result["confidence"] = "medium"
                    logger.info("[自检] 二次验证存在 UNCLEAR，置信度: MEDIUM")
                else:
                    result["confidence"] = "high"
                    logger.info("[自检] 二次验证全部 PASS/澄清，置信度: HIGH")

                result["issues"] = round1_issues
                result["verified"] = round2_verified
                return result
    except Exception as e:
        logger.warning(f"二次验证异常（不阻塞）: {e}")
        result["issues"] = round1_issues
        return result


async def _cross_validate_variables(text: str, notes: Dict[str, str]) -> List[str]:
    """变量交叉校验：检查 变量汇总 中的变量是否与 研究方法 段落一致。

    策略（纯规则，不调用 API）：
    1. 从 变量汇总 提取变量名列表
    2. 从 研究方法 段落提取变量名
    3. 方法中提到的变量但未在汇总中出现 → [变量遗漏]
    4. 汇总中的变量但全文未再出现 → [变量孤立]
    """
    warnings = []
    var_summary = notes.get("变量汇总", "")
    研究方法 = notes.get("研究方法", "")

    if not var_summary or not 研究方法:
        return warnings

    # 提取变量汇总中的变量名（以 - 或数字开头的行、或含"变量名称"的行）
    var_names_in_summary = set()
    for line in var_summary.split("\n"):
        line = line.strip()
        # 匹配 "变量名称：xxx" 或 "- xxx（IV）" 等模式
        name_match = re.search(r'(?:变量名称[：:]\s*|-\s*)([\w\s\-]+?)(?:[（(]|$)', line)
        if name_match:
            var_name = name_match.group(1).strip()
            if len(var_name) >= 3:  # 过滤太短的
                var_names_in_summary.add(var_name.lower())

    # 提取研究方法段落中提到的变量/构念名
    # 找引号中的术语、常见变量模式
    method_var_candidates = set()
    # 引号中的词
    quoted = re.findall(r'["""]([^"""]+?)["»"]', 研究方法)
    for q in quoted:
        if 3 <= len(q) <= 40:
            method_var_candidates.add(q.lower().strip())
    # 常见模式：XXX scale / XXX construct
    scale_matches = re.findall(r'(\w[\w\s]{2,35}?)\s(?:scale|construct|measure|variable)', 研究方法, re.IGNORECASE)
    for sm in scale_matches:
        method_var_candidates.add(sm.lower().strip())

    # 交叉比对
    if var_names_in_summary and method_var_candidates:
        # 方法中提到的变量是否在汇总中
        for mv in method_var_candidates:
            # 模糊匹配
            found = any(
                mv in vs or vs in mv or
                sum(1 for a, b in zip(mv.split(), vs.split()) if a == b) >= 1
                for vs in var_names_in_summary
            )
            if not found and len(mv) > 5:
                warnings.append(f"[变量遗漏] 研究方法提到「{mv}」但变量汇总未收录")

    if warnings:
        logger.warning(f"[变量交叉校验] 发现 {len(warnings)} 个问题")
    else:
        logger.info("[变量交叉校验] 通过 ✓")

    return warnings


# ============================================================================
# Gemini Vision — 分层识图（结构化输出）
# ============================================================================

async def _recognize_figures_structured(pdf_path: Path, gemini_api_key: str) -> Dict[str, any]:
    """用 Gemini Vision 对 PDF 进行分层识图，输出结构化数据。

    三层 Prompt 策略：
    - 理论框架图 → Mermaid flowchart（可在 Obsidian 渲染）
    - 数据表格 → Markdown table（保留全部数字）
    - 统计结果图 → 结构化列表（路径 → β → SE → p）

    返回 {"mermaid_diagrams": [...], "markdown_tables": [...], "statistical_results": [...], "descriptions": [...]}
    失败返回空 dict。
    """
    result = {
        "mermaid_diagrams": [],
        "markdown_tables": [],
        "statistical_results": [],
        "descriptions": [],
    }

    try:
        from google import genai
        import fitz as fitz_module
    except ImportError:
        logger.warning("Gemini Vision 需要: pip install google-genai PyMuPDF")
        return result

    if not gemini_api_key:
        logger.warning("Gemini API Key 未配置，跳过识图")
        return result

    doc = None
    try:
        client = genai.Client(api_key=gemini_api_key)
        doc = fitz_module.open(pdf_path)
    except Exception as e:
        logger.warning(f"Gemini 初始化失败: {e}")
        return result

    try:
        # ── 智能预扫描：文本检测 Figure/Table 标题 ──
        # 比 PyMuPDF 图片检测更准：矢量图、表格都能覆盖，不依赖嵌入图片
        MAX_PAGES_TO_SCAN = 3

        # 图表标题正则（中英文）
        FIGURE_CAPTION_RE = re.compile(
            r'(?:Figure|Fig\.?|Table|FIGURE|FIG\.?|TABLE)\s*[A-Z]?\d+'  # Fig. 1, Figure S1, Table A2
            r'|图\s*\d+|表\s*\d+'                                          # 中文图表
            r'|(?:Figure|Fig\.?|Table)\s*[IVX]+',                          # 罗马数字 Table IV
            re.IGNORECASE,
        )

        candidate_pages = []
        for page_num in range(len(doc)):
            if len(candidate_pages) >= MAX_PAGES_TO_SCAN:
                break
            try:
                text = doc[page_num].get_text()
                if FIGURE_CAPTION_RE.search(text):
                    candidate_pages.append(page_num)
            except Exception:
                continue

        # 退而：如果没检测到标题但 PDF 有图片，扫图片最多的页
        if not candidate_pages:
            image_counts = []
            for page_num in range(len(doc)):
                try:
                    imgs = doc[page_num].get_images(full=True)
                    if imgs:
                        image_counts.append((page_num, len(imgs)))
                except Exception:
                    continue
            image_counts.sort(key=lambda x: -x[1])
            candidate_pages = [p for p, _ in image_counts[:MAX_PAGES_TO_SCAN]]

        # 再退：扫第 2-3 页
        if not candidate_pages:
            candidate_pages = [p for p in range(1, min(4, len(doc)))]

        logger.info(f"  Gemini 图表检测: {len(candidate_pages)}页有图表标题 → {[p+1 for p in candidate_pages]}")

        # 精简 Prompt（减少 token 加快响应）
        system_prompt = """Extract figures/tables from this academic paper page.
TYPE: framework_diagram | statistical_table | descriptive_table | data_chart | other
CAPTION: [caption]
CONTENT:
[framework: Mermaid flowchart with --> for path, -.-> for moderation, ==> for mediation]
[statistical table: PATH: IV→DV, β=x, SE=x, p=x]
[descriptive table: Markdown table]
[chart: Chinese description]
If none: reply NONE."""

        pages_processed = 0
        for page_num in candidate_pages:
            logger.info(f"  Gemini 第{page_num+1}页...")
            try:
                page = doc[page_num]
                pix = page.get_pixmap(dpi=120)
                img_data = pix.tobytes("png")

                resp = client.models.generate_content(
                    model="gemini-flash-latest",
                    contents=[{
                        "parts": [
                            {"text": system_prompt},
                            {"inline_data": {"mime_type": "image/png", "data": img_data}},
                        ]
                    }],
                )

                text = resp.text.strip() if resp.text else ""
                if not text or "NONE" in text.upper()[:10]:
                    logger.info(f"  第{page_num+1}页: 无图表")
                    continue

                pages_processed += 1
                _parse_gemini_structured_output(text, page_num + 1, result)
                logger.info(f"  第{page_num+1}页: ✓ 有图表")

                if page_num != candidate_pages[-1]:
                    await asyncio.sleep(1)

            except Exception as e:
                logger.warning(f"Gemini 第{page_num+1}页异常: {e}")
                if "429" in str(e) or "RESOURCE_EXHAUSTED" in str(e):
                    await asyncio.sleep(15)
                continue

        logger.info(f"  ✓ Gemini 完成: {pages_processed}页有图表, {sum(len(v) for v in result.values())} 个图表/表格")
    finally:
        if doc is not None:
            try:
                doc.close()
            except Exception:
                pass

    return result


def _parse_gemini_structured_output(text: str, page_num: int, result: Dict[str, any]) -> None:
    """解析 Gemini 返回的结构化文本，按类型归类到 result dict 中。"""
    # 按 ## TYPE: 分割段落
    sections = re.split(r'##\s*TYPE:', text)
    for section in sections:
        if not section.strip():
            continue

        section = section.strip()
        type_match = re.match(r'\s*(framework_diagram|statistical_table|descriptive_table|data_chart|other)', section)
        fig_type = type_match.group(1) if type_match else "other"
        content = section[type_match.end():].strip() if type_match else section

        # 提取 CAPTION
        caption_match = re.search(r'##\s*CAPTION:\s*(.+?)(?:\n|$)', content)
        caption = caption_match.group(1).strip()[:200] if caption_match else ""

        # 提取 CONTENT
        content_match = re.search(r'##\s*CONTENT:\s*\n?(.*)', content, re.DOTALL)
        body = content_match.group(1).strip()[:3000] if content_match else content[:2000]

        entry = {
            "page": page_num,
            "caption": caption,
            "body": body,
        }

        # 按类型归类
        if fig_type == "framework_diagram":
            # 提取 mermaid 代码块（如果有）
            mermaid_match = re.search(r'```mermaid\s*\n(.*?)```', body, re.DOTALL)
            if mermaid_match:
                entry["mermaid_code"] = mermaid_match.group(1).strip()
            result["mermaid_diagrams"].append(entry)
        elif fig_type == "statistical_table":
            # 解析路径系数行
            paths = re.findall(r'PATH:\s*(.+?)(?:\n|$)', body)
            if paths:
                entry["paths"] = [p.strip() for p in paths]
            result["statistical_results"].append(entry)
        elif fig_type == "descriptive_table":
            # 检查是否有 Markdown 表格
            if "|" in body:
                entry["has_markdown_table"] = True
            result["markdown_tables"].append(entry)
        elif fig_type == "data_chart":
            result["descriptions"].append(entry)
        else:
            result["descriptions"].append(entry)


def _format_figure_data(figure_data: Dict[str, any]) -> str:
    """将 Gemini 结构化图表数据格式化为文本，供 DeepSeek 分析。"""
    lines = []

    if figure_data.get("mermaid_diagrams"):
        lines.append(f"## 理论框架图 ({len(figure_data['mermaid_diagrams'])} 个)")
        for i, d in enumerate(figure_data["mermaid_diagrams"], 1):
            lines.append(f"\n### 框架图 {i}: {d.get('caption', '')}")
            if d.get("mermaid_code"):
                lines.append(f"```mermaid\n{d['mermaid_code']}\n```")
            else:
                lines.append(d.get("body", ""))

    if figure_data.get("statistical_results"):
        lines.append(f"\n## 统计结果表 ({len(figure_data['statistical_results'])} 个)")
        for i, d in enumerate(figure_data["statistical_results"], 1):
            lines.append(f"\n### 统计表 {i}: {d.get('caption', '')}")
            if d.get("paths"):
                for p in d["paths"]:
                    lines.append(f"- {p}")
            else:
                lines.append(d.get("body", ""))

    if figure_data.get("markdown_tables"):
        lines.append(f"\n## 描述性统计表 ({len(figure_data['markdown_tables'])} 个)")
        for i, d in enumerate(figure_data["markdown_tables"], 1):
            lines.append(f"\n### 表格 {i}: {d.get('caption', '')}")
            lines.append(d.get("body", ""))

    if figure_data.get("descriptions"):
        lines.append(f"\n## 数据图表描述 ({len(figure_data['descriptions'])} 个)")
        for i, d in enumerate(figure_data["descriptions"], 1):
            lines.append(f"\n### 图表 {i}: {d.get('caption', '')}")
            lines.append(d.get("body", ""))

    return "\n".join(lines)


async def _analyze_figures_with_deepseek(
    paper_excerpt: str,
    figure_summary: str,
    notes: Dict[str, str],
) -> Optional[str]:
    """用 DeepSeek 对 Gemini 提取的图表做结合文献的深度分析。"""
    if not figure_summary or len(figure_summary) < 50:
        return None

    deepseek_key = config.get("deepseek_api_key", "")
    deepseek_model = config.get("deepseek_model", "deepseek-chat")
    if not deepseek_key:
        return None

    title = notes.get("标题", "未知论文")
    research_q = notes.get("研究问题", notes.get("research_questions", ""))
    findings = notes.get("主要发现", notes.get("key_findings", ""))

    prompt = f"""你是一位市场营销×AI营销领域的资深研究者。请对以下论文的图表进行深度分析。

## 论文信息
- 标题: {title}
- 研究问题: {research_q[:300]}
- 主要发现: {findings[:300]}

## 论文文本摘要（前3000字）
{paper_excerpt[:3000]}

## Gemini Vision 提取的图表数据
{figure_summary[:4000]}

请做以下分析（中文，专业术语保留英文）：
1. **理论框架解读**：如果有理论模型图，解读各变量之间的关系，分析其理论逻辑
2. **关键数据发现**：从统计结果表中提取最重要的路径系数、效应量、显著性水平
3. **数据一致性检查**：图表中的数据是否与正文中的发现一致？如有矛盾请指出
4. **图表质量评价**：图表的呈现是否清晰、完整？是否有缺失的信息？
5. **研究启示**：这些图表揭示了什么重要洞见？对后续研究有何启发？

控制在 800 字以内，结构清晰。"""

    try:
        from openai import OpenAI
        client = OpenAI(
            api_key=deepseek_key,
            base_url="https://api.deepseek.com/v1",
        )
        resp = client.chat.completions.create(
            model=deepseek_model,
            messages=[
                {"role": "system", "content": "你是一位学术研究者，擅长分析论文中的图表和数据。"},
                {"role": "user", "content": prompt},
            ],
            temperature=0.3,
            max_tokens=2000,
        )
        return resp.choices[0].message.content
    except Exception as e:
        logger.warning(f"图表分析 DeepSeek 调用失败: {e}")
        return None


def extract_images_from_pdf(pdf_path: Path, output_dir: Path) -> List[str]:
    """
    从 PDF 页面中智能提取学术图表（含图题和图注）。

    核心策略：以图注为锚点向上推算图表边界（而非围绕嵌入图片对象）。
    1. 找到所有 Fig(ure)?/Table/图/表 + 数字 开头的图注文本块
    2. 向上找最近的正文段落（>100字符，gap>12pt），作为图表上界
    3. 合并图注上方所有嵌入位图，按需扩展边界
    4. 矢量图（无嵌入位图但有图注）同样以上方正文为界
    5. 200 DPI 页面区域渲染，完整包含图表+图注

    返回相对路径列表（相对于 Obsidian 笔记所在目录，用于 ![[...]] 嵌入）。
    """
    CAPTION_RE = re.compile(
        r"^(Fig(?:ure)?\.?\s*\d+|Table\.?\s*\d+|图\s*\d+|表\s*\d+)",
        re.IGNORECASE
    )
    # 图注可能的后缀（Note:, Source:, * 等）
    NOTE_LIKE = re.compile(r"^(Note|Source|Data|Notes?)[\s:.*]", re.IGNORECASE)

    image_rel_paths: List[str] = []
    if not pdf_path or not pdf_path.exists():
        return image_rel_paths

    try:
        doc = fitz.open(pdf_path)
        output_dir.mkdir(parents=True, exist_ok=True)
        prefix = clean_filename(pdf_path.stem, 30)
        page_w = doc[0].rect.width
        extracted_count = 0

        for page_idx in range(len(doc)):
            page = doc[page_idx]
            page_h = page.rect.height
            text_blocks = page.get_text("blocks")
            image_infos = page.get_image_info()

            # ── 收集本页所有合格的嵌入位图 ──
            raster_rects: List[fitz.Rect] = []
            for img in (image_infos or []):
                bbox_raw = img["bbox"]
                bbox = fitz.Rect(bbox_raw) if isinstance(bbox_raw, tuple) else bbox_raw
                w_pt, h_pt = bbox.width, bbox.height
                size_kb = img.get("size", 0) / 1024
                if w_pt < 100 or h_pt < 70:
                    continue
                if bbox.y0 < 50 and h_pt < 60:
                    continue
                if size_kb < 12 and w_pt < 180:
                    continue
                raster_rects.append(bbox)

            # ── 找到所有图注文本块（作为锚点） ──
            caption_blocks = []  # [tb_tuple]
            for tb in text_blocks:
                if len(tb) < 5:
                    continue
                text = tb[4].strip()
                if CAPTION_RE.match(text):
                    caption_blocks.append(tb)

            if not caption_blocks:
                # 无图注 → 仅对孤立的特大嵌入图做后备提取
                for bbox in raster_rects:
                    if bbox.width > 250 and bbox.height > 150:
                        crop = fitz.Rect(
                            max(0, bbox.x0 - 8), max(0, bbox.y0 - 8),
                            min(page_w, bbox.x1 + 8), min(page_h, bbox.y1 + 8)
                        )
                        self._render_and_save(page, crop, prefix, output_dir, extracted_count + 1)
                        extracted_count += 1
                        image_rel_paths.append(...)  # this is getting complex, let me restructure
                continue

            # ── 按 y 排序图注，分区处理 ──
            # 每个图注的图表区域 = (zone_top, caption_bottom)
            # zone_top = max(上方正文下界, 上一个图注区域下界 + 间距)
            caption_blocks.sort(key=lambda tb: tb[1])  # 按 y0 排序

            rendered_on_page: List[fitz.Rect] = []
            prev_zone_bottom = 0  # 上一个图注区域的下界

            for cap_tb in caption_blocks:
                cap_rect = fitz.Rect(cap_tb[0], cap_tb[1], cap_tb[2], cap_tb[3])
                cap_text = cap_tb[4].strip()

                if cap_rect.y0 < 80:
                    continue

                # ── 确定图表区域上界 (zone_top) ──
                # 默认：从上一个图注区域下界开始 + 间距
                zone_top = max(60, prev_zone_bottom + 15)

                # 如果上一个图注很远（>200pt），说明中间有正文，重新找上界
                if cap_rect.y0 - prev_zone_bottom > 200:
                    zone_top = max(60, cap_rect.y0 - 350)

                # 向上精确定位：找最近的正文段落下界
                for other_tb in sorted(text_blocks, key=lambda b: b[3], reverse=True):
                    if len(other_tb) < 5:
                        continue
                    ot_rect = fitz.Rect(other_tb[0], other_tb[1], other_tb[2], other_tb[3])
                    ot_text = other_tb[4].strip()

                    # 必须在图注上方且在 zone_top 之上
                    if ot_rect.y1 >= cap_rect.y0 - 8 or ot_rect.y1 <= zone_top:
                        continue

                    # 跳过图注和 note 行
                    if CAPTION_RE.match(ot_text) or NOTE_LIKE.match(ot_text):
                        continue

                    # 正文段落（>100字符）→ 更新上界
                    if len(ot_text) > 100:
                        zone_top = ot_rect.y1 + 10
                        break

                    # 短文本可能是小节标题
                    if len(ot_text) > 15 and ot_rect.y1 > zone_top:
                        zone_top = ot_rect.y1 + 6

                # ── 找到属于这个图注的所有嵌入图（在 zone_top 和图注之间）──
                figure_images = [
                    r for r in raster_rects
                    if zone_top - 5 < r.y0 and r.y1 < cap_rect.y0 + 15
                ]

                # ── 计算裁剪区域 ──
                if figure_images:
                    # 有嵌入图：以嵌入图的实际边界为准
                    fig_left = min(r.x0 for r in figure_images)
                    fig_right = max(r.x1 for r in figure_images)
                    fig_top = min(r.y0 for r in figure_images)
                else:
                    # 纯矢量图：以图注的宽度为准，从 zone_top 开始
                    fig_left = cap_rect.x0
                    fig_right = cap_rect.x1
                    fig_top = zone_top
                    # 确保矢量图区域有最小高度
                    if cap_rect.y0 - fig_top < 60:
                        fig_top = cap_rect.y0 - 250

                # ── 向下扩展图注区域 ──
                caption_bottom = cap_rect.y1
                # 包含 Note/Source 等紧跟在图注后的续行
                for tb2 in text_blocks:
                    if len(tb2) < 5:
                        continue
                    tb2_rect = fitz.Rect(tb2[0], tb2[1], tb2[2], tb2[3])
                    tb2_text = tb2[4].strip()
                    # 紧接在图注下方（gap < 10pt），且是 note/source 行或图注延续
                    gap = tb2_rect.y0 - caption_bottom
                    if 1 < gap < 10:
                        if NOTE_LIKE.match(tb2_text) or (
                            not CAPTION_RE.match(tb2_text)
                            and not re.match(r"^\d+\.?\s", tb2_text)
                            and len(tb2_text) < 300
                        ):
                            h_ov = min(tb2_rect.x1, fig_right + 20) - max(tb2_rect.x0, fig_left - 20)
                            if h_ov > 20:
                                caption_bottom = max(caption_bottom, tb2_rect.y1)

                # ── 构建最终裁剪矩形 ──
                crop_rect = fitz.Rect(
                    max(0, fig_left - 12),
                    max(0, fig_top - 8),
                    min(page_w, fig_right + 12),
                    min(page_h, caption_bottom + 10),
                )

                # ── 空间去重 ──
                dup = False
                for prev in rendered_on_page:
                    ox = max(0, min(crop_rect.x1, prev.x1) - max(crop_rect.x0, prev.x0))
                    oy = max(0, min(crop_rect.y1, prev.y1) - max(crop_rect.y0, prev.y0))
                    if ox > 80 and oy > 60:
                        dup = True
                        break
                if dup:
                    continue

                # 最小尺寸检查
                if crop_rect.width < 120 or crop_rect.height < 80:
                    continue

                rendered_on_page.append(crop_rect)
                prev_zone_bottom = caption_bottom  # 更新分区下界，下一个图注从此开始

                # ── 渲染 ──
                try:
                    mat = fitz.Matrix(2.78, 2.78)  # ~200 DPI
                    pix = page.get_pixmap(matrix=mat, clip=crop_rect)
                    extracted_count += 1
                    filename = f"{prefix}_fig{extracted_count}.png"
                    filepath = output_dir / filename
                    while filepath.exists():
                        extracted_count += 1
                        filename = f"{prefix}_fig{extracted_count}.png"
                        filepath = output_dir / filename
                    pix.save(filepath)
                    image_rel_paths.append(f"attachments/{filename}")
                    logger.info(
                        f"  图表 #{extracted_count}: Pg{page_idx+1} "
                        f"({crop_rect.width:.0f}x{crop_rect.height:.0f}pt)"
                        + (f" | {cap_text[:90]}" if cap_text else "")
                    )
                except Exception as e:
                    logger.warning(f"  图表渲染失败 Pg{page_idx+1}: {e}")
                    continue

        doc.close()

        if image_rel_paths:
            logger.info(f"从 PDF 提取 {len(image_rel_paths)} 张图表: {pdf_path.name}")
        else:
            logger.debug(f"PDF 中未发现有效图表: {pdf_path.name}")

    except Exception as e:
        logger.warning(f"PDF 图片提取失败 ({pdf_path.name}): {e}")

    return image_rel_paths


def _normalize_concept(term: str) -> str:
    """归一化概念术语，确保 Obsidian 图谱中同一概念的 wikilink 写法统一。
    - 去除括号内的缩写
    - 英文 → 全小写，但 AI 领域通用缩写保持大写
    - 常见全称术语 → 学界通用缩写（全称/缩写统一为一个节点）
    - 中文不受影响"""
    # 去除括号内的缩写说明
    term = re.sub(r'\s*[（(][^）)]*[）)]', '', term)
    # 合并多余空白
    term = re.sub(r'\s+', ' ', term).strip()
    # 全小写（中文不变）
    term = term.lower()

    # 常见全称 → 学界通用缩写（全称/缩写统一为缩写节点）
    _FULL_TO_STANDARD = {
        "artificial intelligence": "AI",
        "artificial general intelligence": "AGI",
        "generative artificial intelligence": "generative AI",
        "large language models": "LLMs",
        "large language model": "LLM",
        "retrieval-augmented generation": "RAG",
    }
    if term in _FULL_TO_STANDARD:
        return _FULL_TO_STANDARD[term]

    # AI 领域通用缩写保持大写
    _KNOWN_ACRONYMS = {"ai", "agi", "llm", "llms", "rag", "nlp", "rlhf", "gan"}
    words = term.split()
    words = [w.upper() if w in _KNOWN_ACRONYMS else w for w in words]
    return " ".join(words)


def _find_obsidian_note_by_doi(doi: str) -> Optional[Path]:
    """在 Obsidian 笔记库中查找匹配 DOI 的已有笔记，返回 Path 或 None。
    用于防止中英关键词检索产生重复笔记。
    """
    if not doi or not OBSIDIAN_DIR.exists():
        return None
    doi_lower = doi.strip().lower()
    # 缓存搜索结果（同一次运行内复用）
    if not hasattr(_find_obsidian_note_by_doi, "_cache"):
        _find_obsidian_note_by_doi._cache = {}
    if doi_lower in _find_obsidian_note_by_doi._cache:
        return _find_obsidian_note_by_doi._cache[doi_lower]
    for md_file in OBSIDIAN_DIR.rglob("*.md"):
        if md_file.name == "_index.md":
            continue
        try:
            content = md_file.read_text(encoding="utf-8")
            for line in content.split("\n", 30):
                if line.startswith("doi:") and doi_lower in line.lower():
                    _find_obsidian_note_by_doi._cache[doi_lower] = md_file
                    return md_file
        except Exception:
            continue
    _find_obsidian_note_by_doi._cache[doi_lower] = None
    return None


def write_obsidian_note(notes: Dict[str, str], pdf_path: Path, force: bool = False) -> bool:
    try:
        title = notes.get("标题", "Untitled")
        # 安全网：剥离双语标题中非主导语言的翻译后缀
        # （上游 _extract_from_detail 已做清洗，此处兜底 DeepSeek 可能附加的翻译）
        if ' / ' in title:
            parts = title.split(' / ')
            if len(parts) == 2:
                left, right = parts
                left_cn = bool(re.search(r'[一-鿿]', left))
                right_cn = bool(re.search(r'[一-鿿]', right))
                if left_cn and not right_cn:
                    title = left
                elif right_cn and not left_cn:
                    title = right
        year = notes.get("年份", datetime.date.today().year)
        journal = notes.get("期刊", "unknown")
        doi = notes.get("doi", "")
        keywords = notes.get("关键词", "")
        safe_journal = clean_filename(journal, 30)
        safe_title = clean_filename(title, 120)
        year_str = str(year) if year and year != "原文未提及" else "Unknown"
        year_dir = OBSIDIAN_DIR / year_str / safe_journal
        year_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{year_str}_{safe_title}.md"
        filepath = year_dir / filename

        # ── DOI 去重：同一 DOI 已有笔记则跳过（防止中英关键词产生重复笔记）──
        if doi and not force:
            existing = _find_obsidian_note_by_doi(doi)
            if existing and existing != filepath:
                logger.info(f"  ⊘ DOI {doi} 已有笔记 {existing.name}，跳过重复创建")
                # 如果新文件指向不同路径（如 新论文待处理/ vs Zotero storage），
                # 不创建新文件，但旧文件可能 PDF 路径不准确 → 不覆盖
                return False

        if filepath.exists() and not force:
            logger.info(f"Obsidian 笔记已存在，跳过: {filepath.name}")
            return False
        if filepath.exists() and force:
            logger.info(f"Obsidian 笔记已存在，force=True 覆写: {filepath.name}")

        # 构建 tags
        tags = ["AI", "marketing", safe_journal]
        kw_list = re.split(r"[,;，；、]", keywords) if keywords else []
        for kw in kw_list[:5]:
            kw_clean = _normalize_concept(kw).replace(" ", "-")
            if kw_clean and kw_clean not in tags:
                tags.append(kw_clean)

        impact_factor = notes.get("影响因子", "")
        first_author = notes.get("第一作者", "")
        corresponding_author = notes.get("通讯作者", "")
        zone = notes.get("分区", "")
        frontmatter = f"""---
title: "{title}"
authors: "{notes.get('作者', '')}"
first_author: "{first_author}"
corresponding_author: "{corresponding_author}"
year: {year}
journal: "{journal}"
impact_factor: "{impact_factor}"
zone: "{zone}"
doi: "{doi}"
keywords: "{keywords}"
pdf: "{pdf_path.as_posix() if pdf_path else ''}"
reading_mode: ""
reading_date: ""
tags: [{', '.join(tags)}]
---
"""
        body = f"""## 研究背景与动机
{notes.get('研究背景与动机', '')}

## 研究问题
{notes.get('研究问题', '')}

## 变量汇总
{notes.get('变量汇总', '')}

## 研究方法
{notes.get('研究方法', '')}

## 方法论详解
{notes.get('方法论详解', '')}

## 研究结果
{notes.get('研究结果', '')}

## 讨论与结论
{notes.get('讨论与结论', '')}

## 创新点
{notes.get('创新点', '')}

## 局限与展望
{notes.get('局限与展望', '')}

## 图表分析
{notes.get('图表分析', '（暂无图表分析 — 可能 Gemini API Key 未配置或 PDF 中无图表）')}

---
## 相关概念
"""
        # 只从关键词字段提取概念（关键词是干净的列表，理论框架/变量是段落）
        kw_val = notes.get("关键词", "")
        concepts = []
        if kw_val:
            for item in re.split(r"[;,，；、]", kw_val):
                item = _normalize_concept(item)
                if 2 <= len(item) <= 50:
                    concepts.append(item)
        seen = set()
        unique_concepts = [c for c in concepts if not (c.lower() in seen or seen.add(c.lower()))]
        if unique_concepts:
            for c in unique_concepts:
                body += f"- [[{c}]]\n"
        else:
            body += "（暂无）\n"

        # ── 提取 PDF 图表并生成图表索引 ──
        attachments_dir = year_dir / "attachments"
        image_rel_paths = extract_images_from_pdf(pdf_path, attachments_dir)
        if image_rel_paths:
            body += "\n---\n## 图表索引\n"
            for img_path in image_rel_paths:
                body += f"\n![[{img_path}]]\n"
            logger.info(f"  嵌入 {len(image_rel_paths)} 张图表")

        content = frontmatter + "\n" + body
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(content)
        logger.info(f"Obsidian 笔记写入: {filepath}")
        index_path = OBSIDIAN_DIR / "_index.md"
        with open(index_path, "a", encoding="utf-8") as idx:
            idx.write(f"- [[{year_str}/{safe_journal}/{filename.replace('.md', '')}|{title}]]\n")
        return True
    except Exception as e:
        logger.error(f"写入 Obsidian 失败: {e}")
        return False


# ============================================================================
# Excel
# ============================================================================
def _load_excel_dois() -> set[str]:
    """读取 Excel 中已有的所有 DOI，用于去重过滤。"""
    dois: set[str] = set()
    DOI_COL = 9  # NOTE_FIELDS 中 "doi" 的 0-based 索引
    if not EXCEL_PATH.exists():
        return dois
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        for row_idx in range(2, ws.max_row + 1):
            doi_val = str(ws.cell(row_idx, DOI_COL + 1).value or "").strip().lower()
            if doi_val:
                dois.add(doi_val)
        wb.close()
    except Exception as e:
        logger.warning(f"读取 Excel DOI 失败: {e}")
    return dois


def _load_excel_titles() -> set[str]:
    """读取 Excel 中已有的所有标题（normalize 后），用于去重过滤。"""
    titles: set[str] = set()
    TITLE_COL = 1  # NOTE_FIELDS 中 "标题" 的 0-based 索引
    if not EXCEL_PATH.exists():
        return titles
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        for row_idx in range(2, ws.max_row + 1):
            title_val = str(ws.cell(row_idx, TITLE_COL + 1).value or "").strip()
            if title_val:
                titles.add(normalize_title(title_val))
        wb.close()
    except Exception as e:
        logger.warning(f"读取 Excel 标题失败: {e}")
    return titles


def _sanitize_excel_text(value: str) -> str:
    """清洗 Excel 不兼容的字符。
    XML 1.0 不允许 \\x00-\\x08, \\x0B, \\x0C, \\x0E-\\x1F, \\x7F-\\x9F,
    以及 surrogate 字符 (\\uD800-\\uDFFF) 和 Unicode 非字符 (\\uFFFE-\\uFFFF)。"""
    if not value:
        return ""
    # 移除非法 XML 控制字符 + surrogate + Unicode 非字符，保留 \\t \\n \\r
    cleaned = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F\ud800-\udfff￾-￿]", "", str(value))
    return cleaned


def append_to_excel(notes_list: List[Dict[str, str]]):
    headers = NOTE_FIELDS + ["PDF路径", "入库时间"]
    DOI_COL = 9  # NOTE_FIELDS 中 "doi" 的 0-based 索引
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

    if not EXCEL_PATH.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        logger.info("Excel 表头已创建")
        next_seq = 1
    else:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        # 检测表头：第一行第一列匹配当前表头即认为已有表头，不重复添加
        first_cell = str(ws.cell(1, 1).value or "").strip()
        if first_cell != "序号":
            # 表头不对 — 可能是旧版"编号"表头或空表，插入新表头
            ws.insert_rows(1)
            for col_idx, h in enumerate(headers, 1):
                ws.cell(1, col_idx, h)
            logger.info("Excel 表头已更新为「序号」")
        # 计算下一序号 = 最后一行序号 + 1
        last_row = ws.max_row
        if last_row > 1:
            last_seq_val = ws.cell(last_row, 1).value
            try:
                next_seq = int(last_seq_val) + 1 if last_seq_val is not None else 1
            except (ValueError, TypeError):
                next_seq = last_row  # 回退：用行号
        else:
            next_seq = 1

    # 读取已有 DOI 集合，用于去重
    existing_dois: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        doi_val = str(ws.cell(row_idx, DOI_COL + 1).value or "").strip().lower()
        if doi_val:
            existing_dois.add(doi_val)

    skipped = 0
    appended = 0
    for notes in notes_list:
        doi = (notes.get("doi") or "").strip().lower()
        if doi and doi in existing_dois:
            skipped += 1
            logger.info(f"⊘ [Excel去重] {notes.get('标题', '?')[:50]} | DOI {doi} 已存在，跳过")
            continue
        row = [_sanitize_excel_text(notes.get(f, "")) for f in NOTE_FIELDS]
        # 自动填写序号
        row[0] = str(next_seq)
        next_seq += 1
        row.append(_sanitize_excel_text(notes.get("pdf_path", "")))
        row.append(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws.append(row)
        if doi:
            existing_dois.add(doi)
        appended += 1
    wb.save(EXCEL_PATH)
    if skipped:
        logger.info(f"Excel 去重跳过 {skipped} 条，追加 {appended} 条记录")
    else:
        logger.info(f"Excel 追加 {appended} 条记录")


def update_excel_by_doi(notes_list: List[Dict[str, str]]):
    """按 DOI 匹配更新 Excel 已有行，无匹配则追加新行。用于重新深度阅读时更新知识库。"""
    headers = NOTE_FIELDS + ["PDF路径", "入库时间"]
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

    if not EXCEL_PATH.exists():
        # 表都不存在，直接创建
        append_to_excel(notes_list)
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # 确保表头存在
    first_cell = str(ws.cell(1, 1).value or "").strip()
    if first_cell != "序号":
        ws.insert_rows(1)
        for col_idx, h in enumerate(headers, 1):
            ws.cell(1, col_idx, h)
        logger.info("Excel 表头已更新为「序号」")

    # 计算下一序号
    last_row = ws.max_row
    if last_row > 1:
        last_seq_val = ws.cell(last_row, 1).value
        try:
            next_seq = int(last_seq_val) + 1 if last_seq_val is not None else 1
        except (ValueError, TypeError):
            next_seq = last_row
    else:
        next_seq = 1

    # 构建 DOI → 行号 映射（DOI 在第 10 列，0-based index=9）
    doi_col = 9  # NOTE_FIELDS 中 "doi" 的索引
    doi_to_row: dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        existing_doi = str(ws.cell(row_idx, doi_col + 1).value or "").strip().lower()
        if existing_doi:
            doi_to_row[existing_doi] = row_idx

    updated = 0
    appended = 0
    for notes in notes_list:
        doi = (notes.get("doi") or "").strip().lower()
        row_data = [_sanitize_excel_text(notes.get(f, "")) for f in NOTE_FIELDS]
        # 保留原标题的序号（不覆盖），仅在追加时自动填序号
        row_data.append(_sanitize_excel_text(notes.get("pdf_path", "")))
        row_data.append(datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        if doi and doi in doi_to_row:
            # 更新已有行 — 不覆盖序号（列1）
            row_idx = doi_to_row[doi]
            for col_idx, val in enumerate(row_data, 1):
                if col_idx == 1:
                    continue  # 保留原标题序号
                ws.cell(row_idx, col_idx, val)
            updated += 1
            logger.info(f"Excel 更新: {doi}")
        else:
            # 追加新行 — 自动填序号
            row_data[0] = str(next_seq)
            next_seq += 1
            ws.append(row_data)
            appended += 1
            logger.info(f"Excel 追加: {doi or '无DOI'}")

    wb.save(EXCEL_PATH)
    logger.info(f"Excel 更新完成：{updated} 条更新, {appended} 条追加")


# ============================================================================
# PDF 清理功能（新增）
# ============================================================================
async def cleanup_processed_pdfs():
    """
    扫描 PDF_DIR，查找已存在于 processed_log 中的文献所对应的 PDF，
    并在用户确认后删除它们，避免堆积。
    """
    if not PDF_DIR.exists():
        print("新论文待处理文件夹不存在，无需清理。")
        return

    pdf_files = list(PDF_DIR.glob("*.pdf"))
    if not pdf_files:
        print("文件夹中没有 PDF 文件。")
        return

    # 加载处理日志中的 DOI 集合
    processed_dois, _ = load_processed_log()
    if not processed_dois:
        print("processed_log.json 中没有记录，无法判断哪些 PDF 已完成。")
        return

    to_delete = []
    for pdf in pdf_files:
        doi = extract_doi_from_pdf(pdf)
        if doi and doi.strip().lower() in processed_dois:
            to_delete.append((pdf, doi))

    if not to_delete:
        print("未发现可清理的 PDF（所有 PDF 均未在处理日志中）。")
        return

    print("\n===== 可清理的 PDF（已在 Zotero 入库并阅读） =====")
    for pdf, doi in to_delete:
        print(f"  {pdf.name}  (DOI: {doi})")
    print("=" * 50)
    if not _non_interactive:
        confirm = input("确认删除以上文件？(y/n): ").strip().lower()
        if confirm != "y":
            print("已取消清理。")
            return
    else:
        print("--yes 模式：跳过清理（保留 PDF 文件）。")
        return

    deleted_count = 0
    for pdf, _ in to_delete:
        try:
            pdf.unlink()
            logger.info(f"已删除: {pdf.name}")
            deleted_count += 1
        except Exception as e:
            logger.error(f"删除失败 {pdf.name}: {e}")
    print(f"清理完成，共删除 {deleted_count} 个文件。")


# ============================================================================
# 共享检索逻辑：关键词搜索 → SPIS 翻页 → 收集新文献
# ============================================================================
async def _scrape_new_articles(dry_run: bool = False) -> List[Dict]:
    """所有模式共享的检索内核：构建去重集合 → 遍历关键词 → 翻页抓取。
    返回新文献列表；dry_run=True 时仅打印不返回。"""
    all_keywords = print_keywords()
    if not all_keywords:
        return []

    try:
        zotero_dois, zotero_titles = await fetch_zotero_existing_dois()
    except Exception as e:
        logger.warning(f"Zotero 连接失败（{e}），仅使用本地日志去重。")
        zotero_dois, zotero_titles = set(), set()
    log_dois, log_titles = load_processed_log()
    excel_dois = _load_excel_dois()
    excel_titles = _load_excel_titles()
    global_doi_set = zotero_dois | log_dois | excel_dois
    global_title_set = zotero_titles | log_titles | excel_titles
    logger.info(f"去重集合：DOI={len(global_doi_set)} 标题={len(global_title_set)} (含Excel {len(excel_dois)} DOI + {len(excel_titles)} 标题)")

    collected_new = []
    for idx, (keyword, scope) in enumerate(all_keywords, 1):
        _poll_keyboard()
        await _kb_wait_if_paused()
        if _kb_is_terminated():
            print("检索已终止。")
            break
        # 计算本关键词还需收集多少篇
        remaining = max(TARGET_NEW_COUNT - len(collected_new), 5)  # 至少收集5篇/关键词
        p, browser, context, page = await open_spis_and_wait(keyword, idx, len(all_keywords))
        try:
            new_kw = await scrape_keyword_until_full(
                page, keyword,
                global_doi_set, global_title_set,
                remaining
            )
            collected_new.extend(new_kw)
            logger.info(f"关键词 [{keyword}] 贡献 {len(new_kw)} 篇，累计 {len(collected_new)}")
            # ── 打印该关键词贡献的完整文献列表 ──
            if new_kw:
                print(f"\n{'='*70}")
                print(f"  [{idx}/{len(all_keywords)}] 关键词 \"{keyword}\" — 本组文献 ({len(new_kw)}篇):")
                print(f"{'='*70}")
                for j, art in enumerate(new_kw, 1):
                    doi_str = f"  DOI: {art['doi']}" if art.get('doi') else ""
                    print(f"  [{j}] {art['title']}")
                    print(f"      期刊: {art.get('journal', '?')} | 年份: {art.get('year', '?')}{doi_str}")
                print(f"{'='*70}\n")

            # ── 达到目标后不自动关闭浏览器，让用户选择是否继续 ──
            if len(collected_new) >= TARGET_NEW_COUNT:
                global _keyboard_state
                # 如果关键词内已按 E 终止，直接跳出，不再重复提示
                if _keyboard_state != "terminated":
                    print(f"\n已收集 {len(collected_new)} 篇新文献。浏览器保持打开，可手动下载PDF。")
                    print(f"[G] 继续检索（下一个关键词）  [E] 终止检索并关闭浏览器")
                    _keyboard_state = "paused"
                    while _keyboard_state == "paused":
                        await asyncio.sleep(0.3)
                        _poll_keyboard()
                if _keyboard_state == "terminated":
                    print("用户选择终止检索。")
                    break
                # G 被按下：_keyboard_state 变为 "running"，继续下一个关键词
                print("继续检索...")
        finally:
            await context.close()
            await browser.close()
            await p.stop()

    if dry_run:
        print("\n[Dry-run] 检索到的新文献如下（未下载）：")
        for art in collected_new:
            print(f"  {art['title'][:120]} | {art.get('journal','')} | {art.get('year','')}")
    return collected_new


# ============================================================================
# 【Agent 模式】自主检索 — 穷尽关键词 × 年限过滤 × 下载分流
# ============================================================================

async def _scrape_new_articles_autonomous(
    year_start: int = 2025,
    year_end: int = 2026,
    headless: bool = False,
    keyword_cursor: dict = None,
    global_paper_limit: int = 5,
    max_pages_per_kw: int = 10,
    journal_filter: list = None,
    keyword_override: str = None,
    run_logger = None,
) -> dict:
    """【Agent 自主模式】从游标位置开始，逐关键词检索，直到收满 global_paper_limit 篇。

    keyword_cursor: 从 config.json 读的游标，指定从哪个关键词开始
    keyword_override: 若提供，忽略游标，仅用此一个词检索（游标不更新）
    journal_filter: 用户指定的期刊白名单（覆盖默认白名单）
    global_paper_limit: 本次运行总共收多少篇
    """
    # ── 关键词列表：override 优先，否则从游标开始 ──
    if keyword_override:
        all_keywords = [(keyword_override, "自定义")]
        cursor_ci, cursor_ki = 0, 0
        is_override = True
    else:
        all_keywords = print_keywords()
        if not all_keywords:
            return {"collected": [], "by_keyword": {}, "with_links": [], "without_links": [], "help_submitted": 0}
        if keyword_cursor is None:
            keyword_cursor = _read_keyword_cursor()
        cursor_ci = keyword_cursor.get("category_index", 0)
        cursor_ki = keyword_cursor.get("keyword_index", 0)
        is_override = False

    # ── 构建去重集合 ──
    try:
        zotero_dois, zotero_titles = await fetch_zotero_existing_dois()
    except Exception as e:
        logger.warning(f"Zotero 连接失败（{e}），仅使用本地日志去重。")
        zotero_dois, zotero_titles = set(), set()
    log_dois, log_titles = load_processed_log()
    excel_dois = _load_excel_dois()
    excel_titles = _load_excel_titles()
    global_doi_set = zotero_dois | log_dois | excel_dois
    global_title_set = zotero_titles | log_titles | excel_titles
    logger.info(f"[Agent] 去重集合: DOI={len(global_doi_set)} 标题={len(global_title_set)}")
    logger.info(f"[Agent] 全局上限: {global_paper_limit} 篇 | 每词最大翻页: {max_pages_per_kw}")

    all_collected = []
    by_keyword = {}
    papers_with_links = []
    papers_without_links = []
    total_help_submitted = 0
    global_count = [0]  # 可变计数器，跨关键词传递
    keywords_exhausted = False

    # ── 从游标位置开始迭代 ──
    skipped_before_cursor = 0
    for idx, (keyword, scope) in enumerate(all_keywords, 1):
        # 非 override 模式下，跳过游标之前的关键词
        if not is_override:
            kw_cat, kw_idx_in_cat = _get_category_index(scope, idx - 1)
            if kw_cat < cursor_ci or (kw_cat == cursor_ci and kw_idx_in_cat < cursor_ki):
                skipped_before_cursor += 1
                continue

        # ── 全局上限检查 ──
        if global_count[0] >= global_paper_limit:
            logger.info(f"[Agent] 全局已收满 {global_paper_limit} 篇，停止。")
            break

        logger.info(f"\n{'='*60}")
        logger.info(f"[Agent] 关键词 [{idx}/{len(all_keywords)}]: {keyword}")
        logger.info(f"   游标: cat={cursor_ci} ki={cursor_ki} | 已收: {global_count[0]}/{global_paper_limit}")
        logger.info(f"{'='*60}")

        if run_logger:
            run_logger.log_keyword_progress(keyword, cursor_ci, cursor_ki, global_count[0])

        p, context, page = await open_spis_autonomous(
            keyword, year_start=year_start, year_end=year_end, headless=headless
        )
        kw_collected = []
        try:
            kw_collected = await _scrape_keyword_exhaustive(
                page, keyword, global_doi_set, global_title_set,
                year_start=year_start, year_end=year_end,
                max_pages=max_pages_per_kw,
                global_paper_limit=global_paper_limit,
                global_collected_count=global_count,
                journal_filter=journal_filter,
                run_logger=run_logger,
            )
            all_collected.extend(kw_collected)
            by_keyword[keyword] = kw_collected
            logger.info(f"[Agent] 关键词 [{keyword}] 贡献 {len(kw_collected)} 篇，累计 {len(all_collected)}/{global_paper_limit}")

            if kw_collected:
                print(f"\n{'='*70}")
                print(f"  [{idx}/{len(all_keywords)}] 「{keyword}」— {len(kw_collected)}篇:")
                print(f"{'='*70}")
                for j, art in enumerate(kw_collected, 1):
                    doi_str = f"  DOI: {art['doi']}" if art.get('doi') else ""
                    link_info = "🔗有下载" if art.get('download_url') else "📭已求助"
                    print(f"  [{j}] {link_info} {art['title'][:100]}")
                    print(f"      期刊: {art.get('journal', '?')} | {art.get('year', '?')}{doi_str}")
                print(f"{'='*70}\n")

        finally:
            await context.close()
            await p.stop()

        # ── 分流 ──
        for art in kw_collected:
            if art.get("download_url"):
                papers_with_links.append(art)
            else:
                papers_without_links.append(art)
        total_help_submitted += sum(1 for a in kw_collected if a.get("help_submitted"))

        # ── 更新游标（override 模式不更新）──
        if not is_override:
            kw_cat, kw_idx_in_cat = _get_category_index(scope, idx - 1)
            _write_keyword_cursor(kw_cat, kw_idx_in_cat + 1)
            cursor_ci = kw_cat
            cursor_ki = kw_idx_in_cat + 1  # 下一个

        # ── 信号检查 ──
        if run_logger:
            await _check_signal_files(run_logger)

    # ── 检查是否全部穷尽 ──
    if not is_override:
        next_kw, _, _, _ = _get_next_keyword(_read_keyword_cursor())
        if next_kw is None:
            keywords_exhausted = True
            logger.info("[Agent] 🎉 所有关键词已遍历完毕！")
            if run_logger:
                run_logger.log_keywords_exhausted()

    # ── 保存待处理清单（有下载链接的论文）──
    if papers_with_links:
        pending_file = PDF_DIR / "pending_manual.json"
        # 合并已有清单
        existing_pending = []
        if pending_file.exists():
            try:
                with open(pending_file, "r", encoding="utf-8") as f:
                    existing_pending = json.load(f)
            except Exception:
                pass
        # 去重后保存
        existing_titles = {a.get("title", "")[:80] for a in existing_pending}
        new_to_add = [a for a in papers_with_links if a.get("title", "")[:80] not in existing_titles]
        combined = existing_pending + new_to_add
        with open(pending_file, "w", encoding="utf-8") as f:
            json.dump(combined, f, ensure_ascii=False, indent=2)
        logger.info(f"[Agent] pending_manual.json: 原有 {len(existing_pending)} + 新增 {len(new_to_add)} = {len(combined)} 篇")

    result = {
        "collected": all_collected,
        "by_keyword": by_keyword,
        "with_links": papers_with_links,
        "without_links": papers_without_links,
        "help_submitted": total_help_submitted,
        "keywords_exhausted": keywords_exhausted,
    }
    return result


async def _scrape_keyword_exhaustive(
    page,
    keyword: str,
    global_doi_set: Set[str],
    global_title_set: Set[str],
    year_start: int = 2025,
    year_end: int = 2026,
    target_per_kw: int = 5,
    max_pages: int = MAX_PAGES_PER_KEYWORD,
    global_paper_limit: int = None,
    global_collected_count: list = None,
    journal_filter: list = None,
    run_logger = None,
) -> List[Dict[str, str]]:
    """【Agent 模式】翻页收集论文，达到全局上限 or 连续3页无新 or 翻完 max_pages 即停止。

    global_paper_limit: 本次运行总论文上限（跨关键词），None=不限制
    global_collected_count: 可变列表 [count] 用于跨关键词累加计数
    journal_filter: 可选，仅保留这些期刊的论文（覆盖白名单）
    """
    new_collected = []
    consecutive_no_new = 0

    for pg in range(max_pages):
        # ── 信号检查 ──
        if run_logger:
            await _check_signal_files(run_logger)

        await asyncio.sleep(random.uniform(MIN_SLEEP, MAX_SLEEP))
        page_articles = await scrape_page(page)

        # ── 后置年份过滤（兜底：SPIS 年份筛选未生效时）──
        if page_articles:
            before_filter = len(page_articles)
            page_articles = [
                a for a in page_articles
                if _parse_article_year(a.get("year", ""), year_start, year_end)
            ]
            if len(page_articles) < before_filter:
                logger.info(f"  年份过滤 {year_start}-{year_end}: {before_filter} → {len(page_articles)} 篇")

        # ── 详情页补全（含下载链接检测 + 文献求助）──
        if ENRICH_FROM_DETAIL and page_articles:
            page_articles = await _enrich_articles_autonomous(page, page_articles)

            # ── 期刊过滤 ──
            kept = []
            for a in page_articles:
                j = a.get("journal", "")
                # 如果用户指定了 journal_filter，用它；否则用白名单
                if journal_filter:
                    matched = any(
                        jf.lower() in j.lower() or j.lower() in jf.lower()
                        for jf in journal_filter
                    )
                    if matched:
                        kept.append(a)
                    else:
                        logger.info(f"✗ [期刊过滤] {a['title'][:60]} | 不在指定期刊: {j}")
                elif journal_in_whitelist(j):
                    kept.append(a)
                else:
                    logger.info(f"✗ [V15过滤] {a['title'][:60]} | 期刊不在白名单: {j}")
            if len(kept) < len(page_articles):
                filter_name = "期刊过滤" if journal_filter else "V15过滤"
                logger.info(f"  {filter_name}: {len(page_articles) - len(kept)} 篇移除，保留 {len(kept)} 篇")
            page_articles = kept

        # ── 去重 ──
        page_added = 0
        for art in page_articles:
            # ── 全局上限检查（去重循环内检查，避免超收）──
            if global_paper_limit is not None and global_collected_count is not None:
                if global_collected_count[0] >= global_paper_limit:
                    logger.info(f"[Agent] 已收满全局 {global_paper_limit} 篇，停止。")
                    return new_collected

            doi = art["doi"].strip().lower()
            norm_title = normalize_title(art["title"])
            if doi and doi in global_doi_set:
                logger.info(f"⊘ [去重DOI] {art['title'][:50]}")
                continue
            title_dup = False
            for gt in global_title_set:
                is_dup, _, _ = is_title_duplicate(norm_title, gt)
                if is_dup:
                    title_dup = True
                    break
            if title_dup:
                logger.info(f"⊘ [去重标题] {art['title'][:60]}")
                continue
            new_collected.append(art)
            if doi:
                global_doi_set.add(doi)
            global_title_set.add(norm_title)
            page_added += 1
            if global_collected_count is not None:
                global_collected_count[0] += 1
            logger.info(f"✓ [{len(new_collected)}/{global_collected_count[0] if global_collected_count else '?'}] {art['title'][:120]} | {art.get('journal','')} | {art.get('year','')}")

            # ── 全局达标检查 ──
            if global_paper_limit is not None and global_collected_count is not None:
                if global_collected_count[0] >= global_paper_limit:
                    logger.info(f"[Agent] 已收满全局 {global_paper_limit} 篇，完成。")
                    return new_collected

        # ── 翻页终止判断（连续3页无新即停止）──
        if page_added == 0:
            consecutive_no_new += 1
            if consecutive_no_new >= 3:
                logger.info(f"[Agent] 关键词 [{keyword}] 连续 {consecutive_no_new} 页无新文献，穷尽完成。")
                break
        else:
            consecutive_no_new = 0

        # ── 翻页 ──
        if not await _click_next_page(page, pg):
            logger.info(f"[Agent] 关键词 [{keyword}] 翻页结束（第{pg+1}页是最后一页）")
            break

    logger.info(f"[Agent] 关键词 [{keyword}] 完成: {len(new_collected)} 篇新文献")
    return new_collected


def _parse_article_year(year_str: str, year_start: int, year_end: int) -> bool:
    """检查文章年份是否在目标范围内。"""
    if not year_str:
        return True  # 无年份信息 → 保留（后续人工判断）
    try:
        yr = int(str(year_str).strip()[:4])
        return year_start <= yr <= year_end
    except (ValueError, TypeError):
        return True  # 无法解析年份 → 保留


async def _click_next_page(page, current_pg: int) -> bool:
    """翻到下一页。成功返回 True，失败（已到最后一页）返回 False。"""
    # 取当前页码
    current_num = current_pg + 1
    try:
        active_el = await page.query_selector("button.pagination-number.active")
        if active_el:
            txt = (await active_el.inner_text()).strip()
            if txt.isdigit():
                current_num = int(txt)
    except Exception:
        pass

    next_num = current_num + 1

    # ── 方式1: '>' 箭头 ──
    all_arrows = await page.query_selector_all("button.pagination-arrow")
    for arrow in all_arrows:
        try:
            txt = (await arrow.inner_text()).strip()
            if txt == ">":
                cls = (await arrow.get_attribute("class") or "").lower()
                if "first" not in cls and "disabled" not in cls:
                    await arrow.scroll_into_view_if_needed()
                    await asyncio.sleep(0.3)
                    await arrow.click()
                    # 轮询等激活
                    for _ in range(20):
                        await asyncio.sleep(0.5)
                        new_active = await page.query_selector("button.pagination-number.active")
                        if new_active:
                            new_txt = (await new_active.inner_text()).strip()
                            if new_txt.isdigit() and int(new_txt) == next_num:
                                logger.info(f"  → 翻页: {current_num} → {new_txt}")
                                await asyncio.sleep(1)
                                return True
                    logger.warning(f"  箭头点击后未检测到页码变化")
        except Exception:
            continue

    # ── 方式2: 页码按钮 ──
    all_nums = await page.query_selector_all("button.pagination-number")
    for btn in all_nums:
        try:
            txt = (await btn.inner_text()).strip()
            cls = (await btn.get_attribute("class") or "").lower()
            if txt == str(next_num) and "active" not in cls:
                await btn.scroll_into_view_if_needed()
                await asyncio.sleep(0.3)
                await btn.click()
                for _ in range(20):
                    await asyncio.sleep(0.5)
                    new_active = await page.query_selector("button.pagination-number.active")
                    if new_active:
                        new_txt = (await new_active.inner_text()).strip()
                        if new_txt.isdigit() and int(new_txt) == next_num:
                            logger.info(f"  → 翻页: {current_num} → {new_txt}")
                            await asyncio.sleep(1)
                            return True
        except Exception:
            continue

    logger.info(f"  翻页失败（第{current_num}页可能是最后一页）")
    return False


async def _enrich_articles_autonomous(page, articles: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """【Agent 模式】详情页补全 + 下载链接检测 + 文献求助，无需键盘交互。"""
    if not articles:
        return articles

    # 残缺标题优先
    garbled = [a for a in articles if a.get("is_garbled")]
    normal = [a for a in articles if not a.get("is_garbled")]
    articles[:] = garbled + normal
    if garbled:
        logger.info(f"  详情页补全 ({len(articles)} 篇, {len(garbled)} 篇残缺优先)...")
    else:
        logger.info(f"  详情页补全 ({len(articles)} 篇)...")

    enriched_list = []
    list_url = page.url
    help_email = config.get("help_email", "18922596828@163.com")

    for i, art in enumerate(articles):
        logger.info(f"  详情 [{i+1}/{len(articles)}]: {art['title'][:60]}...")
        enriched = dict(art)

        # 恢复列表页 DOM
        try:
            await page.wait_for_selector("article.article", timeout=10000)
        except Exception:
            enriched_list.append(enriched)
            continue

        article_els = await page.query_selector_all("article.article")

        # 标题指纹匹配
        art_el = None
        target_title = art["title"]
        for el in article_els:
            try:
                el_text = await el.inner_text()
                if target_title[:40] in el_text.strip()[:80].replace("\n", " "):
                    art_el = el
                    break
            except Exception:
                continue
        if not art_el:
            if i < len(article_els):
                art_el = article_els[i]
            else:
                enriched_list.append(enriched)
                continue

        current_url = page.url

        # 找可点击元素
        click_el = None
        for sel in ["div.d-t.jump", "div.allow-ai", "div.jump", "div.d-t"]:
            try:
                el = await art_el.query_selector(sel)
                if el:
                    click_el = el
                    break
            except Exception:
                continue
        if not click_el:
            click_el = art_el

        # ── 点击 + 轮询 ──
        new_page = None
        url_changed = False
        try:
            async with page.context.expect_page(timeout=5000) as new_page_info:
                await click_el.click(force=True)
                for _ in range(15):
                    await asyncio.sleep(0.5)
                    try:
                        if page.url != current_url:
                            url_changed = True
                            break
                    except Exception:
                        pass
                if not url_changed:
                    try:
                        new_page = await new_page_info.value
                    except Exception:
                        pass
        except Exception:
            pass

        if not url_changed and not new_page:
            for _ in range(20):
                await asyncio.sleep(0.5)
                try:
                    if page.url != current_url:
                        url_changed = True
                        break
                except Exception:
                    pass

        # ── 提取数据 ──
        detail_tab = new_page if new_page else page
        if new_page:
            try:
                await new_page.wait_for_selector("h1, h2, .title, [class*='detail'], strong", timeout=15000)
                await asyncio.sleep(1.0)
                await _extract_from_detail(new_page, enriched)
            except Exception as e:
                logger.warning(f"  新标签提取异常: {e}")
        elif url_changed:
            try:
                await page.wait_for_selector("h1, h2, .title, [class*='detail'], strong", timeout=15000)
            except Exception:
                pass
            await asyncio.sleep(1.0)
            await _extract_from_detail(page, enriched)
        else:
            modal_sel = ".modal, .dialog, .drawer, [role='dialog']"
            try:
                modal_el = await page.query_selector(modal_sel)
            except Exception:
                modal_el = None
            if modal_el:
                await asyncio.sleep(0.5)
                await _extract_from_detail(page, enriched)
                try:
                    await page.keyboard.press("Escape")
                except Exception:
                    pass

        # ── 期刊名白名单补全（移植自助手模式 _enrich_articles_from_detail）──
        journal = enriched.get("journal", "")
        if not journal or not journal_in_whitelist(journal):
            # 从 detail_tab 全文搜索白名单期刊
            try:
                full_text = await detail_tab.evaluate("() => document.body.innerText")
                found = _find_whitelist_journal_in_text(full_text[:500])
                if found:
                    logger.info(f"  期刊补全: '{journal}' → '{found}'")
                    enriched["journal"] = found
                    journal = found
            except Exception:
                pass
        if journal and not journal_in_whitelist(journal):
            resolved = _resolve_truncated_journal(journal)
            if resolved:
                logger.info(f"  截断修复: '{journal}' → '{resolved}'")
                enriched["journal"] = resolved

        # ── 下载链接检测 → 分流 ──
        download_url = await _check_detail_download_url(detail_tab)
        has_download = download_url is not None

        if has_download:
            enriched["download_url"] = download_url
            logger.info(f"  📥 下载链接: {download_url[:80]}...")
            if new_page:
                try:
                    await new_page.close()
                except Exception:
                    pass
            elif url_changed:
                try:
                    await page.go_back(timeout=10000)
                except Exception:
                    await page.goto(list_url, timeout=15000)
                await asyncio.sleep(random.uniform(1.0, 2.0))
        else:
            # 无下载链接 → 自动文献求助
            logger.info(f"  📭 无下载链接 → 自动文献求助")
            help_ok = await _auto_submit_literature_help(detail_tab, enriched, help_email)
            enriched["help_submitted"] = help_ok
            if new_page:
                try:
                    await asyncio.sleep(random.uniform(1.0, 2.0))
                    await new_page.close()
                except Exception:
                    pass
            elif url_changed:
                try:
                    await page.go_back(timeout=10000)
                except Exception:
                    await page.goto(list_url, timeout=15000)
                await asyncio.sleep(random.uniform(1.0, 2.0))

        enriched_list.append(enriched)

    # 统计
    improved_titles = sum(1 for a, e in zip(articles, enriched_list)
                          if e.get("title", "") != a.get("title", ""))
    improved_dois = sum(1 for a, e in zip(articles, enriched_list)
                        if e.get("doi") and not a.get("doi"))
    with_links = sum(1 for a in enriched_list if a.get("download_url"))
    help_count = sum(1 for a in enriched_list if a.get("help_submitted"))
    logger.info(f"  详情补全完成: 标题+{improved_titles} DOI+{improved_dois} "
                f"下载链接{with_links} 文献求助{help_count}")

    return enriched_list


# ============================================================================
# 【Agent 模式】每日简报生成 & 自主检索入口
# ============================================================================

DAILY_DIR = OBSIDIAN_DIR / "Daily"
DAILY_DIR.mkdir(parents=True, exist_ok=True)


def generate_daily_briefing(result: dict, session_start: str = "") -> str:
    """生成自主检索每日简报，保存为 Obsidian 兼容的 .md 文件。

    Args:
        result: _scrape_new_articles_autonomous 的返回值
        session_start: 检索开始时间（ISO格式字符串）

    Returns:
        简报文件路径
    """
    from datetime import datetime

    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M")
    weekday = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][now.weekday()]

    collected = result.get("collected", [])
    by_keyword = result.get("by_keyword", {})
    with_links = result.get("with_links", [])
    without_links = result.get("without_links", [])
    help_submitted = result.get("help_submitted", 0)

    # ── 构建简报内容 ──
    lines = [
        "---",
        f"date: {date_str}",
        f"time: {time_str}",
        f"weekday: {weekday}",
        f"type: 文献检索简报",
        f"total_collected: {len(collected)}",
        f"with_download_links: {len(with_links)}",
        f"literature_help_submitted: {help_submitted}",
        "keywords_used:",
    ]
    for kw in by_keyword:
        lines.append(f"  - \"{kw}\"")

    lines.extend([
        "---",
        "",
        f"# 📖 LitCall 自主检索简报 — {date_str} {weekday} {time_str}",
        "",
        "## 📊 概览",
        "",
        f"| 指标 | 数值 |",
        f"|------|------|",
        f"| 检索关键词 | {len(by_keyword)} 个 (4 broad + 10 narrow) |",
        f"| 收集新文献 | **{len(collected)}** 篇 |",
        f"| 有下载链接 🔗 | {len(with_links)} 篇 → 待手动下载 |",
        f"| 已提交文献求助 📭 | {help_submitted} 篇 → 等待邮件送达 |",
        f"| 年份筛选 | 2025-2026 |",
        "",
    ])

    # ── 按关键词分组 ──
    if by_keyword:
        lines.append("## 📋 按关键词分布")
        lines.append("")
        total_from_kw = 0
        for kw, papers in by_keyword.items():
            count = len(papers)
            total_from_kw += count
            has_link = sum(1 for p in papers if p.get("download_url"))
            no_link = count - has_link
            lines.append(f"### {kw}")
            lines.append(f"- 共 {count} 篇（🔗有下载: {has_link} | 📭已求助: {no_link}）")
            lines.append("")
            for j, art in enumerate(papers, 1):
                link_icon = "🔗" if art.get("download_url") else "📭"
                title = art.get("title", "?")[:120]
                journal = art.get("journal", "?")
                year = art.get("year", "?")
                doi = art.get("doi", "")
                doi_link = f" [DOI](https://doi.org/{doi})" if doi else ""
                lines.append(f"{j}. {link_icon} **{title}**")
                lines.append(f"   - *{journal}* ({year}){doi_link}")
            lines.append("")

    # ── 待手动下载 (有链接) ──
    if with_links:
        lines.append("## 🔗 待手动下载")
        lines.append("")
        lines.append("以下论文在 SPIS 上有下载链接，请在浏览器中手动下载 PDF：")
        lines.append("")
        lines.append("> **操作提示**：下一次检索会自动复用浏览器登录态（cookies 已保存），")
        lines.append("> 打开 SPIS 后逐个搜索论文标题即可找到下载入口。")
        lines.append("")
        for j, art in enumerate(with_links, 1):
            title = art.get("title", "?")[:150]
            journal = art.get("journal", "?")
            year = art.get("year", "?")
            dl = art.get("download_url", "")[:80]
            lines.append(f"{j}. **{title}**")
            lines.append(f"   - *{journal}* ({year})")
            if dl:
                lines.append(f"   - 下载页: [{dl[:60]}...]({dl})")
            lines.append("")
        # 汇总到 pending_manual.json 的提示
        lines.append(f"> 📁 以上 {len(with_links)} 篇已保存至 `新论文待处理/pending_manual.json`")
        lines.append("")

    # ── 已提交文献求助 (无链接) ──
    if without_links:
        lines.append("## 📭 已自动提交文献求助")
        lines.append("")
        lines.append("以下论文在 SPIS 上无下载链接，已自动通过「文献求助」提交至 **18922596828@163.com**：")
        lines.append("")
        lines.append("> ⏳ 文献求助通常需要 1-3 个工作日处理，请留意邮箱。")
        lines.append("")
        for j, art in enumerate(without_links, 1):
            title = art.get("title", "?")[:150]
            journal = art.get("journal", "?")
            year = art.get("year", "?")
            status = "✅ 已提交" if art.get("help_submitted") else "⚠ 提交失败"
            lines.append(f"{j}. {status} — **{title}**")
            lines.append(f"   - *{journal}* ({year})")
            lines.append("")

    # ── 下一步操作 ──
    lines.extend([
        "## 📌 下一步操作",
        "",
        "### 对于有下载链接的论文：",
        f"1. 打开 SPIS (https://spis.hnlat.com)，搜索论文标题",
        "2. 进入详情页，点击下载按钮下载 PDF",
        f"3. 将下载的 PDF 放入 `新论文待处理/` 文件夹",
        f"4. 在 LitCall 中选择「深度阅读」进行自动处理",
        "",
        "### 对于已提交文献求助的论文：",
        "1. 等待 1-3 个工作日，检查 18922596828@163.com 邮箱",
        "2. 收到 PDF 后放入 `新论文待处理/` 文件夹",
        "3. 在 LitCall 中处理",
        "",
        "---",
        f"*本简报由 LitCall Agent 自动生成 · {date_str} {time_str}*",
    ])

    content = "\n".join(lines)

    # ── 保存文件 ──
    filename = f"{date_str}-检索简报.md"
    filepath = DAILY_DIR / filename
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)

    logger.info(f"[Agent] 每日简报已保存: {filepath}")
    return str(filepath)


async def autonomous_retrieval_session(
    year_start: int = 2025,
    year_end: int = 2026,
    headless: bool = False,
    keyword_cursor: dict = None,
    global_paper_limit: int = 5,
    max_pages_per_kw: int = 10,
    journal_filter: list = None,
    keyword_override: str = None,
    run_logger = None,
) -> dict:
    """【Agent 自主检索 — Phase 1】从游标位置开始，逐关键词检索。

    流程：
    1. 从 keyword_cursor 指向的关键词开始
    2. 收满 global_paper_limit 篇即停（跨关键词总计）
    3. 每个关键词完成后推进游标
    4. 年份过滤 + 期刊白名单 + 下载链接检测
    5. 生成每日简报 → agent抓取/Daily/
    """
    from datetime import datetime
    session_start = datetime.now().isoformat()

    logger.info("=" * 60)
    logger.info("📖 LitCall Phase 1: 自主文献检索")
    logger.info(f"   时间: {session_start}")
    logger.info(f"   年份: {year_start}-{year_end}")
    logger.info(f"   模式: 每运行收 {global_paper_limit} 篇（跨关键词总计）")
    if keyword_override:
        logger.info(f"   自定义关键词: {keyword_override}")
    logger.info("=" * 60)

    result = await _scrape_new_articles_autonomous(
        year_start=year_start,
        year_end=year_end,
        headless=headless,
        keyword_cursor=keyword_cursor,
        global_paper_limit=global_paper_limit,
        max_pages_per_kw=max_pages_per_kw,
        journal_filter=journal_filter,
        keyword_override=keyword_override,
        run_logger=run_logger,
    )

    briefing_path = generate_daily_briefing(result, session_start)

    collected = result.get("collected", [])
    with_links = result.get("with_links", [])
    without_links = result.get("without_links", [])
    help_submitted = result.get("help_submitted", 0)

    print("\n" + "=" * 60)
    print("  🦉 Phase 1 完成 — 检索报告")
    print("=" * 60)
    print(f"  收集新文献:        {len(collected)} 篇")
    print(f"  ├ 有下载链接 🔗:   {len(with_links)} 篇 → 待手动下载")
    print(f"  └ 已提交求助 📭:   {help_submitted} 篇 → 等待邮件")
    print(f"  每日简报:          {briefing_path}")
    print(f"  待处理清单:        {PDF_DIR / 'pending_manual.json'}")
    print("=" * 60)

    if with_links:
        print(f"\n📥 有下载链接 ({len(with_links)} 篇) — 请在 SPIS 手动下载 PDF 放入 新论文待处理/")
        for j, art in enumerate(with_links, 1):
            print(f"  [{j}] {art['title'][:100]}")
            print(f"      {art.get('journal','?')} ({art.get('year','?')})")

    if without_links:
        print(f"\n📭 已文献求助 ({help_submitted} 篇)")

    result["briefing_path"] = briefing_path
    return result


# ═══════════════════════════════════════════════════════════════
# 【Agent 模式】Phase 2: PDF 自动深度阅读 + 入库
# ═══════════════════════════════════════════════════════════════

async def process_pending_pdfs(progress_callback=None) -> dict:
    """【Agent 自主深度阅读】扫描 新论文待处理/*.pdf，全自动处理。

    对每篇 PDF：
    1. 提取 DOI + 文本
    2. DeepSeek 结构化深度阅读（17字段）
    3. Zotero 入库
    4. Obsidian 笔记
    5. Excel 汇总
    6. 标记 processed_log
    7. 删除 PDF ✅

    Returns:
        dict with keys: processed (list), failed (list), zotero_ok (int), deleted (int)
    """
    pdf_files = list(PDF_DIR.glob("*.pdf"))
    if not pdf_files:
        logger.info("[Phase 2] 新论文待处理/ 中没有 PDF 文件，跳过深度阅读。")
        return {"processed": [], "failed": [], "zotero_ok": 0, "deleted": 0, "empty": True}

    # ── 预处理去重 ──
    log_dois, _ = load_processed_log()
    excel_dois = _load_excel_dois()
    already_dois = log_dois | excel_dois
    new_pdfs = []
    skipped = []
    for pdf in pdf_files:
        doi = extract_doi_from_pdf(pdf)
        if doi and doi.strip().lower() in already_dois:
            skipped.append(pdf)
            logger.info(f"⊘ [去重] {pdf.name} 已处理过，跳过")
        else:
            new_pdfs.append(pdf)
    pdf_files = new_pdfs

    if not pdf_files:
        logger.info("[Phase 2] 所有 PDF 均已处理过。")
        # 清理跳过的（已在库中的）PDF
        for pdf in skipped:
            try:
                pdf.unlink()
                logger.info(f"🗑 已删除重复PDF: {pdf.name}")
            except Exception:
                pass
        return {"processed": [], "failed": [], "zotero_ok": 0, "deleted": len(skipped), "empty": True}

    NOTES_DIR.mkdir(parents=True, exist_ok=True)

    results_processed = []
    results_failed = []
    zotero_ok = 0
    deleted_count = 0

    for i, pdf in enumerate(pdf_files, 1):
        logger.info(f"\n[Phase 2] [{i}/{len(pdf_files)}] {pdf.name}")
        if progress_callback:
            progress_callback(i, len(pdf_files), pdf.name, "processing")

        try:
            # 1. 提取 DOI
            doi = extract_doi_from_pdf(pdf)
            if not doi:
                logger.warning(f"  ⊘ DOI 提取失败")
                results_failed.append({"file": pdf.name, "reason": "DOI提取失败"})
                if progress_callback:
                    progress_callback(i, len(pdf_files), pdf.name, "error: DOI")
                continue

            # 2. 提取 PDF 文本
            text = extract_text_from_pdf(pdf)
            if not text:
                logger.warning(f"  ⊘ 文本提取失败")
                results_failed.append({"file": pdf.name, "reason": "文本提取失败"})
                if progress_callback:
                    progress_callback(i, len(pdf_files), pdf.name, "error: text")
                continue

            # 3. DeepSeek 深度阅读
            notes = await generate_notes(text)
            if not notes:
                logger.error(f"  ⊘ DeepSeek 阅读失败")
                results_failed.append({"file": pdf.name, "reason": "DeepSeek阅读失败"})
                if progress_callback:
                    progress_callback(i, len(pdf_files), pdf.name, "error: AI")
                continue

            # 3b. Gemini Vision 图表识别 + DeepSeek 图表分析
            gemini_key = config.get("gemini_api_key", "")
            if gemini_key:
                try:
                    logger.info(f"  🔍 Gemini Vision 图表识别...")
                    figure_data = await _recognize_figures_structured(pdf, gemini_key)
                    if figure_data:
                        total_figs = sum(len(v) for v in figure_data.values())
                        if total_figs > 0:
                            notes["_figures"] = figure_data
                            logger.info(f"  ✓ Gemini 识别 {total_figs} 个图表/表格")

                            # 用 DeepSeek 对图表做结合文献的分析
                            fig_summary = _format_figure_data(figure_data)
                            analysis = await _analyze_figures_with_deepseek(
                                text[:3000], fig_summary, notes
                            )
                            if analysis:
                                notes["图表分析"] = analysis
                                logger.info(f"  ✓ 图表分析完成")
                except Exception as e:
                    logger.warning(f"  ⚠ Gemini 图表识别异常: {e}")

            # 4. 填充影响因子
            journal_name = notes.get("期刊", "")
            if journal_name and journal_name != "原文未提及":
                if_val = get_impact_factor(journal_name)
                if if_val:
                    notes["影响因子"] = if_val

            # 5. Obsidian 笔记
            try:
                write_obsidian_note(notes, pdf)
                logger.info(f"  ✓ Obsidian 笔记")
            except Exception as e:
                logger.warning(f"  ⚠ Obsidian 写入异常: {e}")

            # 6. Excel 汇总（按DOI去重更新，避免重复行）
            try:
                update_excel_by_doi([notes])
                logger.info(f"  ✓ Excel 汇总")
            except Exception as e:
                logger.warning(f"  ⚠ Excel 写入异常: {e}")

            # 7. Zotero 入库
            zotero_ok_this = False
            try:
                article = {
                    "title": notes.get("标题", ""),
                    "doi": doi,
                    "year": notes.get("年份", ""),
                    "journal": journal_name,
                }
                zotero_result = await zotero_add_item(article, pdf, notes)
                if zotero_result:
                    zotero_ok += 1
                    zotero_ok_this = True
                    logger.info(f"  ✓ Zotero 入库")
                else:
                    logger.error(f"  ✗ Zotero 入库失败！PDF 保留待重试: {pdf.name}")
                    logger.error(f"     请检查 Zotero API Key 权限 / 网络 / 存储空间")
            except Exception as e:
                logger.error(f"  ✗ Zotero 入库异常: {e}，PDF 保留待重试")

            # 8. 删除 PDF（仅当 Zotero 入库成功）
            if zotero_ok_this:
                # 标记已处理（仅 Zotero 成功后标记，失败则下次重试）
                save_processed_log([{
                    "doi": doi,
                    "title": notes.get("标题", ""),
                    "file": pdf.name,
                    "year": notes.get("年份", ""),
                    "journal": journal_name,
                }])

                try:
                    pdf.unlink()
                    deleted_count += 1
                    logger.info(f"  🗑 PDF 已删除: {pdf.name}")
                except Exception as e:
                    logger.warning(f"  ⚠ PDF 删除失败: {e}")

                results_processed.append({
                    "file": pdf.name,
                    "doi": doi,
                    "title": notes.get("标题", "")[:120],
                    "journal": journal_name,
                    "year": notes.get("年份", ""),
                    "core_summary": _extract_core_summary(notes),
                    "authors": notes.get("作者", notes.get("authors", "")),
                })
            else:
                # Zotero 失败 → 不标记已处理、不删 PDF、记录为失败
                results_failed.append({
                    "file": pdf.name,
                    "reason": "Zotero入库失败（PDF已保留，下次运行会重试）",
                    "doi": doi,
                })

            if progress_callback:
                progress_callback(i, len(pdf_files), pdf.name, "done")

        except Exception as e:
            logger.error(f"  ✗ 处理异常: {e}")
            results_failed.append({"file": pdf.name, "reason": str(e)})
            if progress_callback:
                progress_callback(i, len(pdf_files), pdf.name, f"error: {e}")

    # ── 清理残留的重复 PDF ──
    for pdf in skipped:
        try:
            pdf.unlink()
            deleted_count += 1
            logger.info(f"🗑 清理重复PDF: {pdf.name}")
        except Exception:
            pass

    # ── Zotero 验证清单 ──
    if results_processed:
        logger.info("=" * 50)
        logger.info("📋 Zotero 入库验证清单（请在 Zotero 桌面端确认以下文献已正确入库）：")
        logger.info("=" * 50)
        for j, item in enumerate(results_processed, 1):
            logger.info(f"  [{j}] {item.get('title', '?')[:80]}")
            logger.info(f"      {item.get('journal', '?')} ({item.get('year', '?')}) | DOI: {item.get('doi', '?')}")
        logger.info("=" * 50)
        logger.info("⚠ 如发现入库异常，请手动从 processed_log.json 中移除对应 DOI 后重新处理。")

    result = {
        "processed": results_processed,
        "failed": results_failed,
        "zotero_ok": zotero_ok,
        "deleted": deleted_count,
        "zotero_checklist": results_processed,  # 供 UI 显示验证清单
        "empty": False,
    }
    return result


# ============================================================================
# 【Agent 模式】结构化运行日志 — AgentRunLogger
# ============================================================================

def _extract_core_summary(notes: dict) -> str:
    """从 DeepSeek 生成的 notes 中提取 2-3 句核心摘要。
    优先级：研究结论 > 研究贡献 > 研究摘要/研究目的 > 摘要首段
    限制 200 字以内。
    """
    candidates = [
        notes.get("研究结论", ""),
        notes.get("研究贡献", ""),
        notes.get("研究摘要", ""),
        notes.get("研究目的", ""),
    ]
    for c in candidates:
        if c and len(c.strip()) > 20:
            # 取前 2-3 句（以中英文句号、分号分割）
            import re
            sentences = re.split(r'[。；;.\n]', c.strip())
            summary = ""
            count = 0
            for s in sentences:
                s = s.strip()
                if len(s) > 8:
                    summary += s + "。"
                    count += 1
                    if count >= 3:
                        break
            if len(summary) > 20:
                return summary[:200]
    # 兜底：取摘要首段
    abstract = notes.get("摘要", "")
    if abstract and len(abstract) > 30:
        return abstract[:200]
    return ""


# ═══════════════════════════════════════════════════════════════
# VPN 连接检测（SPIS 基于学校 IP 白名单认证）
# ═══════════════════════════════════════════════════════════════

async def _check_vpn_connected(page) -> bool:
    """检测 SPIS 是否已通过 VPN 连通（学校 IP 自动识别）。
    返回 True = VPN 连通，False = SPIS 显示登录页（VPN 未连）。
    """
    try:
        is_login_page = await page.evaluate("""() => {
            const body = document.body.innerText || '';
            // VPN未连接时 SPIS 显示登录页面，特征文字：
            //   "账号登录" "手机号登录" "微信扫码" "第三方" "当前IP"
            const loginMarkers = ['账号登录', '手机号登录', '微信扫码', '当前IP'];
            const found = loginMarkers.filter(m => body.includes(m));
            if (found.length >= 2) return true;  // 至少匹配2个 → 确认是登录页
            // 兜底：URL 中包含 login / auth
            if (window.location.href.includes('login') || window.location.href.includes('auth')) return true;
            return false;
        }""")
        return not is_login_page
    except Exception:
        return False


async def _wait_for_vpn(
    headless: bool = False,
    timeout: int = 300,
    progress_callback=None,
) -> tuple:
    """等待 VPN 连接。每 15 秒检测一次 SPIS 是否可访问。

    Args:
        headless: 是否无头模式
        timeout: 最长等待秒数（默认 300 = 5 分钟）
        progress_callback: Streamlit 进度回调

    Returns:
        (connected: bool, browser_page, browser_context, playwright_instance)
        - connected=True  → 复用返回的 page/context/playwright（VPN 已通）
        - connected=False → 全部为 None（超时，调用方应关闭资源）

    调用方负责关闭返回的 context/playwright。
    """
    from datetime import datetime
    from playwright.async_api import async_playwright

    logger.info("🔍 检测 VPN 连接状态（SPIS 学校 IP 认证）...")
    if progress_callback:
        progress_callback("vpn_check", "🔍 检测 VPN 连接状态...", {"status": "checking"})

    playwright = await async_playwright().start()
    user_data_dir = BASE_DIR / "browser_data"

    try:
        if headless:
            context = await playwright.chromium.launch_persistent_context(
                str(user_data_dir),
                headless=True,
                args=["--disable-blink-features=AutomationControlled"],
            )
        else:
            context = await playwright.chromium.launch_persistent_context(
                str(user_data_dir),
                headless=False,
                args=["--disable-blink-features=AutomationControlled"],
            )
    except Exception as e:
        logger.warning(f"Persistent context 失败，使用普通模式: {e}")
        browser = await playwright.chromium.launch(headless=headless)
        context = await browser.new_context()

    page = await context.new_page()

    # 首次打开 SPIS
    try:
        await page.goto(
            "https://spis.hnlat.com",
            timeout=60000,
            wait_until="domcontentloaded",
        )
    except Exception as e:
        logger.warning(f"⚠ SPIS 无法访问: {e}")

    start = datetime.now()
    deadline = start.timestamp() + timeout

    while datetime.now().timestamp() < deadline:
        connected = await _check_vpn_connected(page)
        elapsed = (datetime.now() - start).total_seconds()

        if connected:
            logger.info(f"✅ VPN 已连接，SPIS 可访问（耗时 {elapsed:.0f} 秒）")
            if progress_callback:
                progress_callback("vpn_ok", f"✅ VPN 已连接（{elapsed:.0f} 秒）", {"status": "connected"})
            return True, page, context, playwright

        remaining = timeout - elapsed
        logger.info(f"⏳ VPN 未连接 — SPIS 显示登录页，{remaining:.0f} 秒后超时跳过检索")
        if progress_callback:
            progress_callback(
                "vpn_waiting",
                f"⏳ 等待 VPN 连接... {remaining:.0f} 秒后超时",
                {"status": "waiting", "remaining": int(remaining)},
            )

        # 每 15 秒刷新页面重试
        await asyncio.sleep(15)
        try:
            await page.reload(timeout=30000, wait_until="domcontentloaded")
        except Exception:
            try:
                await page.goto(
                    "https://spis.hnlat.com",
                    timeout=30000,
                    wait_until="domcontentloaded",
                )
            except Exception:
                pass  # 网络不通，继续等

    # 超时
    logger.warning("⏰ VPN 连接超时（5分钟），跳过 Phase 1 检索，继续 Phase 2+3")
    if progress_callback:
        progress_callback("vpn_timeout", "⏰ VPN 超时 — 跳过检索", {"status": "timeout"})

    try:
        await page.close()
        await context.close()
    except Exception:
        pass

    return False, None, None, None


class AgentRunLogger:
    """结构化运行日志：每次 Agent 会话生成一个 JSON 文件到 运行日志/runs/。

    用法:
        run_logger = AgentRunLogger(2025, 2026, keywords_dict)
        run_logger.log_phase1_result(papers_data, with_links, help_submitted)
        run_logger.log_phase2_result(processed, failed)
        run_logger.log_phase3_done(index_size)
        run_logger.finalize("completed")
    """

    def __init__(self, year_start: int, year_end: int, keywords: dict):
        from datetime import datetime
        self.run_id = datetime.now().strftime("%Y-%m-%d-%H%M%S")
        self._started = datetime.now().isoformat()

        broad = keywords.get("broad", []) if isinstance(keywords, dict) else []
        narrow = keywords.get("narrow", []) if isinstance(keywords, dict) else []
        all_kw = broad + narrow

        self._data = {
            "run_id": self.run_id,
            "started_at": self._started,
            "ended_at": None,
            "status": "running",
            "config": {
                "year_start": year_start,
                "year_end": year_end,
                "keywords_count": len(all_kw),
                "keywords": all_kw,
            },
            "phase1": {"status": "pending"},
            "phase2": {"status": "pending"},
            "phase3": {"status": "pending"},
            "briefing_path": "",
        }
        self._path = RUNS_DIR / f"{self.run_id}.json"
        self._path.parent.mkdir(parents=True, exist_ok=True)
        self._save()

    def _save(self):
        """增量写回 JSON 文件（实时可读）。"""
        try:
            self._path.write_text(
                json.dumps(self._data, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        except Exception:
            pass  # 写入失败不阻塞主流程

    # ── VPN ──

    def log_vpn_check(self, connected: bool, waited_seconds: float = 0):
        """记录 VPN 检测结果。"""
        self._data["vpn"] = {
            "connected": connected,
            "waited_seconds": round(waited_seconds, 1),
            "skipped": not connected,
        }
        self._save()

    # ── Phase 1 ──

    def log_phase1_start(self):
        self._data["phase1"]["status"] = "running"
        self._save()

    def log_phase1_skipped(self, reason: str = ""):
        """Phase 1 被跳过（如 VPN 未连接）。"""
        self._data["phase1"] = {
            "status": "skipped",
            "reason": reason or "VPN 未连接，SPIS 不可访问",
            "papers_total": 0,
            "papers_with_links": 0,
            "papers_without_links": 0,
            "help_submitted": 0,
            "papers": [],
        }
        self._save()

    def log_phase1_result(self, by_keyword: dict, with_links: list,
                          without_links: list, help_submitted: int):
        """从 _scrape_new_articles_autonomous 的结果填充 Phase 1 数据。"""
        all_papers = []
        for kw, papers in by_keyword.items():
            for art in papers:
                all_papers.append({
                    "title": art.get("title", "")[:200],
                    "authors": art.get("authors", art.get("author", "")),
                    "year": art.get("year", ""),
                    "journal": art.get("journal", ""),
                    "doi": art.get("doi", ""),
                    "has_download_link": bool(art.get("download_url")),
                    "download_url": art.get("download_url", ""),
                    "help_submitted": bool(art.get("help_submitted")),
                    "keyword": kw,
                })
        self._data["phase1"] = {
            "status": "completed",
            "papers_total": len(all_papers),
            "papers_with_links": len(with_links),
            "papers_without_links": len(without_links),
            "help_submitted": help_submitted,
            "papers": all_papers,
        }
        self._save()

    # ── Phase 2 ──

    def log_phase2_start(self):
        self._data["phase2"]["status"] = "running"
        self._data["phase2"]["papers"] = []
        self._save()

    def log_phase2_paper_processing(self, title: str, doi: str):
        """实时记录当前正在处理的论文（Streamlit 可读）。"""
        if "papers" not in self._data["phase2"]:
            self._data["phase2"]["papers"] = []
        self._data["phase2"]["_currently_processing"] = {
            "title": title[:200],
            "doi": doi,
        }
        self._save()

    def log_phase2_result(self, processed: list, failed: list):
        """processed: [{file, doi, title, journal, year, core_summary}, ...]
        failed: [{file, reason}, ...]
        """
        papers = []
        for item in processed:
            papers.append({
                "title": item.get("title", "")[:200],
                "doi": item.get("doi", ""),
                "authors": item.get("authors", ""),
                "year": item.get("year", ""),
                "journal": item.get("journal", ""),
                "status": "success",
                "core_summary": item.get("core_summary", ""),
                "failure_reason": None,
            })
        for item in failed:
            papers.append({
                "title": item.get("file", "")[:200],
                "doi": "",
                "authors": "",
                "year": "",
                "journal": "",
                "status": "failed",
                "core_summary": "",
                "failure_reason": item.get("reason", ""),
            })
        self._data["phase2"] = {
            "status": "completed",
            "papers_processed": len(processed),
            "papers_failed": len(failed),
            "papers_skipped": 0,
            "papers": papers,
        }
        self._data["phase2"].pop("_currently_processing", None)
        self._save()

    # ── Phase 3 ──

    def log_phase3_done(self, index_size: int):
        self._data["phase3"] = {
            "status": "completed",
            "index_size": index_size,
        }
        self._save()

    # ── 关键词游标 ──

    def log_keyword_progress(self, keyword: str, category_index: int,
                             keyword_index: int, papers_collected: int):
        """记录当前游标进度（实时可读）。"""
        self._data["keyword_cursor"] = {
            "current_keyword": keyword,
            "category_index": category_index,
            "keyword_index": keyword_index,
            "papers_collected_this_keyword": papers_collected,
        }
        self._save()

    def log_keywords_exhausted(self):
        """标记所有关键词已穷尽。"""
        self._data["keywords_exhausted"] = True
        self._save()

    # ── 简报 & 收尾 ──

    def set_briefing(self, path: str):
        self._data["briefing_path"] = path
        self._save()

    def finalize(self, status: str = "completed"):
        """标记运行结束。status: 'completed' | 'failed'"""
        from datetime import datetime
        self._data["status"] = status
        self._data["ended_at"] = datetime.now().isoformat()
        # 清理临时字段
        self._data.get("phase2", {}).pop("_currently_processing", None)
        self._save()


# ═══════════════════════════════════════════════════════════════
# 【Agent 模式】Phase 1+2+3 一体化：检索 → 深度阅读 → 同步
# ═══════════════════════════════════════════════════════════════

async def full_autonomous_session(
    year_start: int = 2025,
    year_end: int = 2026,
    headless: bool = False,
    progress_callback=None,
    keyword_cursor: dict = None,
    global_paper_limit: int = 5,
    max_pages_per_kw: int = 10,
    journal_filter: list = None,
    keyword_override: str = None,
) -> dict:
    """【Agent 完整会话】检索 + 深度阅读 + 同步刷新，一条龙。

    Phase 0: VPN 连通性检测（最多等5分钟）
    Phase 1: SPIS 检索 → 从游标位置开始 → 收满 global_paper_limit 篇即停 → 下载分流 → 简报
    Phase 2: 扫描 新论文待处理/ → 深度阅读 → 入库 → 删PDF
    Phase 3: 重建索引 + 刷新 UI

    progress_callback(phase, msg, details) — 用于 Streamlit 实时进度
    keyword_override: 若提供，忽略游标，仅用此关键词检索一次
    """
    from datetime import datetime
    session_start = datetime.now().isoformat()

    # ── 初始化结构化运行日志 ──
    run_logger = AgentRunLogger(
        year_start=year_start,
        year_end=year_end,
        keywords=config.get("keywords", {}),
    )

    logger.info("=" * 60)
    logger.info("📖 LitCall 完整自主会话启动")
    logger.info("=" * 60)

    # ═══ Phase 0: VPN 连通性检测 ═══
    vpn_start = datetime.now()
    vpn_ok, vpn_page, vpn_context, vpn_playwright = await _wait_for_vpn(
        headless=headless,
        timeout=300,
        progress_callback=progress_callback,
    )
    vpn_waited = (datetime.now() - vpn_start).total_seconds()
    run_logger.log_vpn_check(vpn_ok, vpn_waited)

    # VPN 检测浏览器用完即关（Phase 1 会自己启动新的）
    if vpn_context is not None:
        try:
            await vpn_context.close()
        except Exception:
            pass

    # ═══ Phase 1: 检索 ═══
    if vpn_ok:
        run_logger.log_phase1_start()
        if progress_callback:
            progress_callback("phase1", "Phase 1: SPIS 文献检索启动...", {})

        retrieval_result = await autonomous_retrieval_session(
            year_start=year_start,
            year_end=year_end,
            headless=headless,
            keyword_cursor=keyword_cursor,
            global_paper_limit=global_paper_limit,
            max_pages_per_kw=max_pages_per_kw,
            journal_filter=journal_filter,
            keyword_override=keyword_override,
            run_logger=run_logger,
        )

        # Phase 1 → Phase 2 信号检查
        try:
            await _check_signal_files(run_logger)
        except AgentSignalError:
            run_logger.finalize("terminated")
            raise

        # 写入运行日志 Phase 1
        run_logger.log_phase1_result(
            by_keyword=retrieval_result.get("by_keyword", {}),
            with_links=retrieval_result.get("with_links", []),
            without_links=retrieval_result.get("without_links", []),
            help_submitted=retrieval_result.get("help_submitted", 0),
        )
        run_logger.set_briefing(retrieval_result.get("briefing_path", ""))

        if progress_callback:
            progress_callback("phase1_done", "Phase 1 完成", {
                "collected": len(retrieval_result.get("collected", [])),
                "with_links": len(retrieval_result.get("with_links", [])),
                "help_submitted": retrieval_result.get("help_submitted", 0),
            })
    else:
        # VPN 未连接 → 跳过检索，继续 Phase 2+3
        run_logger.log_phase1_skipped("VPN 未连接（5分钟超时），SPIS 显示登录页")
        if progress_callback:
            progress_callback("phase1_skip", "⏭ Phase 1 跳过（VPN 未连接）", {})
        retrieval_result = {
            "by_keyword": {}, "collected": [], "with_links": [],
            "without_links": [], "help_submitted": 0, "briefing_path": "",
        }

    # ═══ Phase 2: 深度阅读 ═══
    run_logger.log_phase2_start()
    if progress_callback:
        progress_callback("phase2", "Phase 2: 扫描待处理 PDF...", {})

    pdf_result = await process_pending_pdfs(
        progress_callback=(
            lambda i, total, name, status: progress_callback("phase2_progress", f"[{i}/{total}] {name}", {"status": status})
            if progress_callback else None
        )
    )

    # 写入运行日志 Phase 2
    run_logger.log_phase2_result(
        processed=pdf_result.get("processed", []),
        failed=pdf_result.get("failed", []),
    )

    # ═══ Phase 3: 刷新 ═══
    if progress_callback:
        progress_callback("phase3", "Phase 3: 刷新知识库索引 & 更新网页数据...", {})

    # 重建论文索引
    global _qa_paper_index
    _build_paper_index(force_rebuild=True)
    logger.info("[Phase 3] 论文索引已刷新")

    index_size = len(_qa_paper_index) if _qa_paper_index else 0
    run_logger.log_phase3_done(index_size)
    run_logger.finalize("completed")

    if progress_callback:
        progress_callback("phase3_done", "Phase 3 完成", {
            "index_size": index_size,
        })

    # ═══ 最终报告 ═══
    collected = retrieval_result.get("collected", [])
    with_links = retrieval_result.get("with_links", [])
    processed = pdf_result.get("processed", [])
    failed = pdf_result.get("failed", [])

    print("\n" + "=" * 60)
    print("  📖 LitCall 完整会话 — 最终报告")
    print("=" * 60)
    print(f"  Phase 1 检索:")
    print(f"    新文献:   {len(collected)} 篇")
    print(f"    🔗有链接: {len(with_links)} 篇 → 请在 SPIS 手动下载 PDF")
    print(f"    📭已求助: {retrieval_result.get('help_submitted', 0)} 篇")
    print(f"  Phase 2 深度阅读:")
    print(f"    已处理:   {len(processed)} 篇 → 已入库 + PDF 已删除")
    print(f"    失败:     {len(failed)} 篇")
    print(f"  Phase 3 同步:")
    print(f"    知识库索引已刷新")
    print(f"  每日简报: {retrieval_result.get('briefing_path', '?')}")
    print("=" * 60)

    if with_links:
        print(f"\n⚠️  重要提醒：有 {len(with_links)} 篇论文需要你手动下载 PDF！")
        print("   请打开 SPIS 搜索论文标题，下载 PDF 后放入 新论文待处理/ 文件夹。")
        print("   下次运行 Agent 检索时会自动处理。")

    return {
        "retrieval": retrieval_result,
        "pdf_processing": pdf_result,
        "combined_summary": {
            "new_papers_found": len(collected),
            "papers_to_download": len(with_links),
            "papers_processed": len(processed),
            "papers_failed": len(failed),
            "help_submitted": retrieval_result.get("help_submitted", 0),
            "briefing_path": retrieval_result.get("briefing_path", ""),
        },
        "run_id": run_logger.run_id,
    }


# ============================================================================
# 选项1：仅检索新文献
# ============================================================================
async def scrape_only_flow():
    """仅 SPIS 检索 → 保存待处理清单，不做下载/阅读/入库。"""
    collected_new = await _scrape_new_articles()
    if not collected_new:
        print("\n未检索到新文献。")
        return

    # 保存待人工处理清单
    pending_file = PDF_DIR / "pending_manual.json"
    with open(pending_file, "w", encoding="utf-8") as f:
        json.dump(collected_new, f, ensure_ascii=False, indent=2)

    print("\n" + "=" * 60)
    print("  检索完成报告")
    print("=" * 60)
    print(f"  检索新文献:       {len(collected_new)} 篇")
    print(f"  待处理清单:        {pending_file}")
    print("=" * 60)
    print("\n待处理文献：")
    for art in collected_new:
        print(f"  · {art['title'][:120]} | {art.get('journal','')} ({art.get('year','')})")
    print(f"\n[下一步] 请从 SPIS 下载 PDF → 放入 新论文待处理/ 文件夹 → 选择菜单 [2] 深度阅读")
    logger.info(f"检索完成：{len(collected_new)} 篇新文献已写入 {pending_file}")


# ============================================================================
# 选项2：深度阅读 + 知识库入库（Obsidian + Excel）
# ============================================================================
async def deep_read_only_flow():
    """PDF 提取文本 → DeepSeek V4 Pro 深度阅读 → Obsidian 笔记 + Excel 知识库 + 笔记 JSON 暂存。
    Zotero 是独立的文献管理模块，此步骤不做 Zotero 入库。"""
    pdf_files = list(PDF_DIR.glob("*.pdf"))
    if not pdf_files:
        print("新论文待处理/ 文件夹中没有 PDF 文件。")
        print("请先从 SPIS 下载文献 PDF 放入该文件夹。")
        return

    # ── 预处理去重：跳过已处理的 PDF ──
    log_dois, _ = load_processed_log()
    excel_dois = _load_excel_dois()
    already_dois = log_dois | excel_dois
    new_pdfs = []
    skipped_pdfs = []
    for pdf in pdf_files:
        doi = extract_doi_from_pdf(pdf)
        if doi and doi.strip().lower() in already_dois:
            skipped_pdfs.append(pdf.name)
            logger.info(f"⊘ [预处理去重] {pdf.name} | 已在 processed_log 或 Excel 中，跳过")
        else:
            new_pdfs.append(pdf)
    if skipped_pdfs:
        print(f"\n跳过 {len(skipped_pdfs)} 篇已处理文献：")
        for name in skipped_pdfs:
            print(f"  ⊘ {name}")
    pdf_files = new_pdfs
    if not pdf_files:
        print("所有 PDF 均已处理过，无需重复。")
        return

    NOTES_DIR.mkdir(parents=True, exist_ok=True)

    # 键盘控制
    global _keyboard_state
    _keyboard_state = "running"
    print("键盘控制: P=暂停 G=继续 E=终止")

    results = []
    failed = []
    for i, pdf in enumerate(pdf_files, 1):
        _poll_keyboard()
        await _kb_wait_if_paused()
        if _kb_is_terminated():
            print("\n⏹ 用户终止深度阅读，保留已处理结果")
            break

        print(f"\n[{i}/{len(pdf_files)}] {pdf.name}")
        logger.info(f"深度阅读: {pdf.name}")

        # 1. 提取 DOI
        doi = extract_doi_from_pdf(pdf)
        if not doi:
            logger.warning(f"  ⊘ 无法从 PDF 提取 DOI: {pdf.name}")
            failed.append((pdf.name, "DOI 提取失败"))
            continue

        # 2. 提取 PDF 文本
        text = extract_text_from_pdf(pdf)
        if not text:
            logger.warning(f"  ⊘ PDF 文本提取失败: {pdf.name}")
            failed.append((pdf.name, "PDF 文本提取失败"))
            continue

        # 3. DeepSeek V4 Pro 深度阅读
        notes = await generate_notes(text)
        if not notes:
            logger.error(f"  ⊘ DeepSeek AI 阅读失败: {pdf.name}")
            failed.append((pdf.name, "DeepSeek AI 阅读失败"))
            continue

        # 4. 填充影响因子
        journal_name = notes.get("期刊", "")
        if journal_name and journal_name != "原文未提及":
            if_val = get_impact_factor(journal_name)
            if if_val:
                notes["影响因子"] = if_val
                logger.info(f"  影响因子: {journal_name} → IF={if_val}")

        # 5. 写入 Obsidian 笔记（知识库）
        notes["doi"] = doi
        notes["pdf_path"] = str(pdf)
        write_obsidian_note(notes, pdf)
        logger.info(f"  ✓ Obsidian 笔记已写入")

        # 6. 保存笔记 JSON（供后续 Zotero 入库使用）
        notes["_read_date"] = datetime.datetime.now().isoformat()
        safe_doi = doi.replace("/", "_").replace("\\", "_")
        note_file = NOTES_DIR / f"{safe_doi}.json"
        with open(note_file, "w", encoding="utf-8") as f:
            json.dump(notes, f, ensure_ascii=False, indent=2)

        results.append(notes)
        logger.info(f"  ✓ 笔记 JSON 已保存: {note_file.name}")

    # 批量写入 Excel 知识库（自动检测空表→加表头；已有内容→按表头追加）
    if results:
        append_to_excel(results)
        logger.info(f"✓ Excel 知识库追加 {len(results)} 条")

    print("\n" + "=" * 60)
    print("  深度阅读 + 知识库入库 完成")
    print("=" * 60)
    print(f"  成功: {len(results)}/{len(pdf_files)} 篇")
    print(f"  失败: {len(failed)}/{len(pdf_files)} 篇")
    print(f"  Obsidian 笔记: {len(results)} 篇 → agent抓取/")
    print(f"  Excel 记录:    {len(results)} 条 → agent文献汇总.xlsx")
    print(f"  笔记 JSON:     {NOTES_DIR}")
    print("=" * 60)
    if results:
        print("\n已处理文献：")
        for r in results:
            print(f"  · {r.get('标题', '?')[:70]}")
    if failed:
        print("\n失败文献：")
        for name, reason in failed:
            print(f"  · {name} | {reason}")
    print(f"\n[下一步] 选择菜单 [3] Zotero 文献管理库入库")


# ============================================================================
# 选项3：Zotero 文献管理库入库
# ============================================================================
async def 入库_only_flow():
    """读取 NOTES_DIR 中的笔记 JSON → Zotero API 入库（文献管理库）。
    Obsidian/Excel 知识库已在步骤 [2] 完成，此处只做 Zotero 文献同步。"""
    if not NOTES_DIR.exists():
        print("没有找到笔记目录。请先运行 [2] 深度阅读。")
        return

    note_files = sorted(NOTES_DIR.glob("*.json"))
    if not note_files:
        print("没有找到笔记 JSON 文件。请先运行 [2] 深度阅读。")
        return

    # 加载所有笔记
    pending = []
    for nf in note_files:
        try:
            with open(nf, "r", encoding="utf-8") as f:
                notes = json.load(f)
            pending.append((nf, notes))
        except Exception as e:
            logger.warning(f"读取笔记 JSON 失败: {nf.name} ({e})")

    if not pending:
        print("没有可用的笔记 JSON。")
        return

    print(f"\n找到 {len(pending)} 篇笔记待入库 Zotero（文献管理库）：")
    for nf, notes in pending:
        print(f"  · {notes.get('标题', nf.stem)[:70]} | DOI: {notes.get('doi', 'N/A')}")

    # 查询 Zotero 已存在
    zotero_dois, _ = await fetch_zotero_existing_dois()

    # ── Phase 1: Zotero 入库 ──
    logger.info("=" * 40)
    logger.info("Phase 1/2: Zotero 入库 (笔记JSON → Zotero API)")
    logger.info("=" * 40)
    # 键盘控制
    global _keyboard_state
    _keyboard_state = "running"
    print("键盘控制: P=暂停 G=继续 E=终止")
    zotero_items = []
    zotero_failed = []
    skipped_duplicates = []

    for nf, notes in pending:
        _poll_keyboard()
        await _kb_wait_if_paused()
        if _kb_is_terminated():
            print("\n⏹ 用户终止 Zotero 入库")
            break

        doi = notes.get("doi", "")
        title = notes.get("标题", nf.stem)

        if doi and doi in zotero_dois:
            logger.info(f"DOI {doi} 已在 Zotero 中，跳过。")
            skipped_duplicates.append((nf, notes))
            continue

        pdf_path_str = notes.get("pdf_path", "")
        pdf_path = Path(pdf_path_str) if pdf_path_str else None
        if not pdf_path or not pdf_path.exists():
            logger.warning(f"PDF 文件不存在: {pdf_path_str}")
            zotero_failed.append((notes, "PDF 文件不存在"))
            continue

        article = {
            "title": title,
            "doi": doi,
            "journal": notes.get("期刊", ""),
            "year": notes.get("年份", ""),
        }

        ok = await zotero_add_item(article, pdf_path, notes)
        if not ok:
            logger.error(f"Zotero 入库失败: {title[:50]}")
            zotero_failed.append((notes, "Zotero API 入库失败"))
            continue
        zotero_items.append((nf, notes))
        logger.info(f"✓ Zotero 入库成功 ({len(zotero_items)}/{len(pending)}): {title[:50]}")

    # ── Phase 2: 等待用户确认 Zotero 同步完成 ──
    logger.info("=" * 40)
    logger.info("Phase 2/2: Zotero 同步确认 (人工检查)")
    logger.info("=" * 40)
    confirmed = False
    if zotero_items:
        zotero_articles = [{
            "title": it[1].get("标题", ""),
            "doi": it[1].get("doi", ""),
            "journal": it[1].get("期刊", ""),
            "year": it[1].get("年份", ""),
        } for it in zotero_items]
        confirmed = await confirm_zotero_sync(zotero_articles)
    else:
        logger.warning("没有文献成功入库 Zotero，跳过确认步骤。")

    # 记录 processed_log（已确认 Zotero 入库的文献）
    if confirmed and zotero_items:
        save_processed_log([{"doi": it[1].get("doi", ""), "title": it[1].get("标题", "")} for it in zotero_items])
        logger.info(f"processed_log 已更新: {len(zotero_items)} 条")
    elif not confirmed and zotero_items:
        print("\n[WARN] 用户跳过确认，本次入库的文献不会标记为已处理。")

    # ── 报告 ──
    print(f"\nZotero 文献管理库入库完成：")
    print(f"  Zotero 入库成功: {len(zotero_items)} 篇")
    print(f"  Zotero 入库失败: {len(zotero_failed)} 篇")
    print(f"  Zotero 已存在:   {len(skipped_duplicates)} 篇")
    print(f"  用户确认同步:    {'是' if confirmed else '否/跳过'}")
    if zotero_failed:
        print(f"\n[WARN] 入库失败文献：")
        for notes, reason in zotero_failed:
            print(f"  · {notes.get('标题', '?')[:70]} | 原因: {reason}")

    # 清理询问
    await cleanup_processed_pdfs()


# ============================================================================
# 主流程（选项5：检索 → 下载 → 阅读 → 入库 全流程）
# ============================================================================
async def main_flow(dry_run: bool = False):
    collected_new = await _scrape_new_articles(dry_run=dry_run)
    if dry_run or not collected_new:
        return

    email = config.get("unpaywall_email", "18922596828@163.com")
    success_pairs = []
    manual_list = []

    # 键盘控制
    global _keyboard_state
    _keyboard_state = "running"
    print("键盘控制: P=暂停 G=继续 E=终止")

    for art in collected_new:
        _poll_keyboard()
        await _kb_wait_if_paused()
        if _kb_is_terminated():
            print("\n⏹ 用户终止下载环节，剩余文献转入待处理清单")
            manual_list.extend([a for a in collected_new if a not in [pair[0] for pair in success_pairs] and a not in manual_list])
            # 把当前art也加入（如果还没处理）
            if art not in [pair[0] for pair in success_pairs] and art not in manual_list:
                manual_list.append(art)
            break

        doi = art["doi"]
        if not doi:
            manual_list.append(art)   # 无 DOI，需人工下载
            continue
        pdf_url = await download_via_unpaywall(doi, email)
        if not pdf_url:
            pdf_url = await download_via_semantic_scholar(doi)
        if pdf_url:
            path = await download_pdf(pdf_url, art["title"])
            if path:
                success_pairs.append((art, path))
                continue
        manual_list.append(art)

    # ────────────────────────────────────────────────────────────────
    # Phase 1: Zotero 入库
    #   对每篇已下载的 PDF：提取文本 → DeepSeek 阅读 → Zotero API 上传
    #   zotero_add_item() 仅在条目创建 + PDF 上传均成功时返回 True
    # ────────────────────────────────────────────────────────────────
    logger.info("=" * 40)
    logger.info("Phase 1/3: Zotero 入库 (PDF → AI笔记 → Zotero API)")
    logger.info("=" * 40)
    zotero_items = []   # 成功入库的条目列表
    zotero_failed = []  # 入库失败的条目列表
    for art, pdf_path in success_pairs:
        _poll_keyboard()
        await _kb_wait_if_paused()
        if _kb_is_terminated():
            print("\n⏹ 用户终止 Zotero 入库环节")
            # 未处理的文献转入失败列表
            remaining = [(a, p) for a, p in success_pairs if (a, p) not in [(item[0], item[1]) for item in zotero_items]]
            zotero_failed.extend([(a, p, "用户终止") for a, p in remaining])
            break
        # 1. 提取 PDF 文本
        text = extract_text_from_pdf(pdf_path)
        if not text:
            logger.warning(f"PDF 文本提取失败: {pdf_path}")
            zotero_failed.append((art, pdf_path, "PDF 文本提取失败"))
            continue

        # 2. AI 阅读生成笔记（DeepSeek V4 Pro）
        notes = await generate_notes(text)
        if not notes:
            logger.error(f"DeepSeek AI 阅读失败: {pdf_path.name}，跳过该文献（不写入 Zotero/Obsidian/Excel）")
            zotero_failed.append((art, pdf_path, "DeepSeek AI 阅读失败"))
            continue

        # 2.5 填充影响因子（从本地 IF 库或 OpenAlex 获取）
        journal_name = notes.get("期刊", "") or art.get("journal", "")
        if journal_name and journal_name != "原文未提及":
            if_val = get_impact_factor(journal_name)
            if if_val:
                notes["影响因子"] = if_val
                logger.info(f"  影响因子: {journal_name} → IF={if_val}")

        # 3. Zotero API 入库（创建条目 + 上传 PDF 附件）
        ok = await zotero_add_item(art, pdf_path, notes)
        if not ok:
            logger.error(f"Zotero 入库失败: {art['title'][:50]}")
            zotero_failed.append((art, pdf_path, "Zotero API 入库失败"))
            continue
        zotero_items.append((art, pdf_path, notes))
        logger.info(f"✓ Zotero 入库成功 ({len(zotero_items)}/{len(success_pairs)}): {art['title'][:50]}")

    # ────────────────────────────────────────────────────────────────
    # Phase 2: 等待用户确认 Zotero 同步完成
    #   Zotero 客户端同步后才会在本地显示 PDF 附件
    #   用户确认前，不会写入 processed_log（避免标记不完整条目）
    # ────────────────────────────────────────────────────────────────
    logger.info("=" * 40)
    logger.info("Phase 2/3: Zotero 同步确认 (人工检查)")
    logger.info("=" * 40)
    confirmed = False
    if zotero_items:
        confirmed = await confirm_zotero_sync([item[0] for item in zotero_items])
    else:
        logger.warning("没有文献成功入库 Zotero，跳过确认步骤。")

    # ────────────────────────────────────────────────────────────────
    # Phase 3: 后续处理 (Obsidian + Excel + 处理日志)
    #   仅在用户确认 Zotero 入库完成后才执行
    #   processed_log 记录 = 附件已验证存在的文献
    # ────────────────────────────────────────────────────────────────
    logger.info("=" * 40)
    logger.info("Phase 3/3: 笔记生成 (Obsidian + Excel + 日志)")
    logger.info("=" * 40)
    notes_for_excel = []
    processed_records = []
    if confirmed and zotero_items:
        for art, pdf_path, notes in zotero_items:
            notes["doi"] = art["doi"]
            notes["pdf_path"] = str(pdf_path)
            write_obsidian_note(notes, pdf_path)
            notes_for_excel.append(notes)
            processed_records.append({"doi": art["doi"], "title": art["title"]})
        if notes_for_excel:
            append_to_excel(notes_for_excel)
        if processed_records:
            save_processed_log(processed_records)
    elif not confirmed and zotero_items:
        print("\n[WARN] 用户跳过确认，本次 Zotero 入库的文献不会写入 Obsidian 笔记和处理日志。")
        print("  请手动打开 Zotero 确认附件正常后，使用 --process-existing 重新处理。")

    # ── 保存待人工处理清单 ──
    pending_file = PDF_DIR / "pending_manual.json"
    with open(pending_file, "w", encoding="utf-8") as f:
        json.dump(manual_list, f, ensure_ascii=False, indent=2)

    # ── 本次运行报告 ──
    print("\n" + "=" * 60)
    print("  本次运行报告")
    print("=" * 60)
    print(f"  检索新文献:       {len(collected_new)} 篇")
    print(f"  PDF 自动下载:     {len(success_pairs)} 篇")
    print(f"  Zotero API 入库:  {len(zotero_items)} 篇 (成功) / {len(zotero_failed)} 篇 (失败)")
    print(f"  用户确认同步:     {'是' if confirmed else '否/跳过'}")
    print(f"  Obsidian 笔记:    {len(notes_for_excel)} 篇")
    print(f"  Excel 记录:       {len(notes_for_excel)} 条")
    print(f"  已写入处理日志:   {len(processed_records)} 条")
    print(f"  待人工处理:       {len(manual_list)} 篇")
    print("=" * 60)
    if manual_list:
        print("\n待人工处理文献（请从 SPIS 下载 PDF 后放入 新论文待处理/ 文件夹）：")
        for art in manual_list:
            print(f"  · {art['title'][:70]} | {art.get('journal','')} ({art.get('year','')})")
    if zotero_failed:
        print("\n[WARN] Zotero 入库失败文献（请检查网络/Zotero API 后重试）：")
        for art, pdf_path, reason in zotero_failed:
            print(f"  · {art['title'][:70]} | 原因: {reason}")
            print(f"    PDF: {pdf_path}")
    print(f"\n待办清单已保存至 {pending_file}")

    # 清理询问
    await cleanup_processed_pdfs()


# ============================================================================
# 选项4：深度阅读 + Zotero 入库（一键完成，兼容旧 --process-existing）
# ============================================================================
async def process_existing():
    """--process-existing 模式：全自动处理 PDF 文件夹中的所有文献。
    无任何交互式确认，适合批量自动化运行。"""
    global _non_interactive
    _non_interactive = True  # 强制非交互模式，全自动运行
    logger.info("--process-existing 模式启动（全自动，无交互确认）")
    pdf_files = list(PDF_DIR.glob("*.pdf"))
    if not pdf_files:
        print("没有 PDF 文件。")
        return

    # ── 预处理去重：跳过已在 processed_log 或 Excel 中的 PDF ──
    log_dois, _ = load_processed_log()
    excel_dois = _load_excel_dois()
    already_dois = log_dois | excel_dois
    fresh_pdfs = []
    skipped_count = 0
    for pdf in pdf_files:
        doi = extract_doi_from_pdf(pdf)
        if doi and doi.strip().lower() in already_dois:
            skipped_count += 1
            logger.info(f"⊘ [预处理去重] {pdf.name} | DOI {doi} 已处理，跳过")
        else:
            fresh_pdfs.append(pdf)
    if skipped_count:
        print(f"跳过 {skipped_count} 篇已处理文献（processed_log / Excel 中已存在）")
    pdf_files = fresh_pdfs
    if not pdf_files:
        print("所有 PDF 均已处理过，无需重复。")
        return

    pending = []
    for pdf in pdf_files:
        doi = extract_doi_from_pdf(pdf)
        if not doi:
            logger.warning(f"无法从 PDF 提取 DOI: {pdf.name}")
            continue
        pending.append({"doi": doi, "pdf_path": pdf, "title": pdf.stem})

    if not pending:
        print("没有可识别的 PDF（无法提取 DOI）。")
        return

    print("\n===== 可处理 PDF 列表 =====")
    for item in pending:
        print(f"  DOI: {item['doi']} | 文件: {item['pdf_path'].name}")
    print("全自动模式：自动处理以上所有 PDF...")
    logger.info(f"共 {len(pending)} 个 PDF 待处理")

    zotero_dois, _ = await fetch_zotero_existing_dois()

    # ────────────────────────────────────────────────────────────────
    # Phase 1: Zotero 入库
    # ────────────────────────────────────────────────────────────────
    logger.info("=" * 40)
    logger.info("Phase 1/3: Zotero 入库 (PDF → AI笔记 → Zotero API)")
    logger.info("=" * 40)
    # 键盘控制
    global _keyboard_state
    _keyboard_state = "running"
    print("键盘控制: P=暂停 G=继续 E=终止")
    zotero_items = []
    zotero_failed = []
    skipped_duplicates = []
    for item in pending:
        _poll_keyboard()
        await _kb_wait_if_paused()
        if _kb_is_terminated():
            print("\n⏹ 用户终止处理")
            break

        if item["doi"] in zotero_dois:
            logger.info(f"DOI {item['doi']} 已在 Zotero 中，跳过。")
            skipped_duplicates.append(item)
            continue

        # 1. 提取 PDF 文本
        text = extract_text_from_pdf(item["pdf_path"])
        if not text:
            logger.warning(f"PDF 无文本: {item['pdf_path'].name}")
            zotero_failed.append((item, "PDF 文本提取失败"))
            continue

        # 2. AI 阅读生成笔记（DeepSeek V4 Pro）
        notes = await generate_notes(text)
        if not notes:
            logger.error(f"DeepSeek AI 阅读失败: {item['pdf_path'].name}，跳过该文献（不写入 Zotero/Obsidian/Excel）")
            zotero_failed.append((item, "DeepSeek AI 阅读失败"))
            continue

        # 2.5 填充影响因子（从本地 IF 库或 OpenAlex 获取）
        journal_name = notes.get("期刊", "")
        if journal_name and journal_name != "原文未提及":
            if_val = get_impact_factor(journal_name)
            if if_val:
                notes["影响因子"] = if_val
                logger.info(f"  影响因子: {journal_name} → IF={if_val}")

        # 3. Zotero API 入库
        article = {
            "title": notes.get("标题") or item["title"],
            "doi": item["doi"],
            "journal": notes.get("期刊", ""),
            "year": notes.get("年份", ""),
        }
        ok = await zotero_add_item(article, item["pdf_path"], notes)
        if not ok:
            logger.error(f"Zotero 入库失败: {item['pdf_path'].name}")
            zotero_failed.append((item, "Zotero API 入库失败"))
            continue
        zotero_items.append((item, notes))
        logger.info(f"✓ Zotero 入库成功 ({len(zotero_items)}): {item['pdf_path'].name}")

    # ────────────────────────────────────────────────────────────────
    # Phase 2: 等待用户确认 Zotero 同步完成
    # ────────────────────────────────────────────────────────────────
    logger.info("=" * 40)
    logger.info("Phase 2/3: Zotero 同步确认 (人工检查)")
    logger.info("=" * 40)
    confirmed = False
    if zotero_items:
        zotero_articles = [{
            "title": it[0]["title"],
            "doi": it[0]["doi"],
            "journal": it[1].get("期刊", ""),
            "year": it[1].get("年份", ""),
        } for it in zotero_items]
        confirmed = await confirm_zotero_sync(zotero_articles)
    else:
        logger.warning("没有文献成功入库 Zotero，跳过确认步骤。")

    # ────────────────────────────────────────────────────────────────
    # Phase 3: 后续处理 (Obsidian + Excel + 日志)
    # ────────────────────────────────────────────────────────────────
    logger.info("=" * 40)
    logger.info("Phase 3/3: 笔记生成 (Obsidian + Excel + 日志)")
    logger.info("=" * 40)
    notes_list = []
    if confirmed and zotero_items:
        for item, notes in zotero_items:
            notes["doi"] = item["doi"]
            notes["pdf_path"] = str(item["pdf_path"])
            write_obsidian_note(notes, item["pdf_path"])
            notes_list.append(notes)
        if notes_list:
            append_to_excel(notes_list)
            save_processed_log([{"doi": n["doi"], "title": n["标题"]} for n in notes_list])
    elif not confirmed and zotero_items:
        print("\n[WARN] 用户跳过确认，本次入库的文献不会写入 Obsidian 笔记和处理日志。")

    print(f"\n处理完成：")
    print(f"  Zotero 入库成功: {len(zotero_items)} 篇")
    print(f"  Zotero 入库失败: {len(zotero_failed)} 篇")
    print(f"  Zotero 已存在:   {len(skipped_duplicates)} 篇")
    print(f"  用户确认同步:    {'是' if confirmed else '否/跳过'}")
    print(f"  Obsidian 笔记:   {len(notes_list)} 篇")
    if zotero_failed:
        print(f"\n[WARN] 入库失败文献：")
        for item, reason in zotero_failed:
            print(f"  · {item['pdf_path'].name} | 原因: {reason}")

    # 清理询问
    await cleanup_processed_pdfs()


# ============================================================================
# 选项7：从 Zotero 集合重新深度阅读
# ============================================================================
async def re_read_from_zotero_flow():
    """从 Zotero「agent抓取」集合中读取已入库文献 → 重新 DeepSeek 深度阅读
    → 覆写 Obsidian 笔记 + 按 DOI 更新 Excel。不需要手动搬 PDF。"""
    zotero = config.get("zotero", {})
    user_id = zotero.get("user_id")
    api_key = zotero.get("api_key")
    zotero_data_dir = Path(config.get("zotero_data_dir", os.path.expandvars(r"%USERPROFILE%\Zotero")))
    storage_dir = zotero_data_dir / "storage"

    if not user_id or not api_key:
        print("Zotero 未配置，无法读取已入库文献。")
        return

    coll_key = await get_zotero_collection_key()
    if not coll_key:
        print("无法获取 Zotero「agent抓取」集合。")
        return

    base = f"https://api.zotero.org/users/{user_id}"
    headers = {"Zotero-API-Key": api_key}

    # ── 1. 分页获取集合中所有条目 ──
    print("正在从 Zotero 获取已入库文献列表...")
    all_items = []
    start = 0
    limit = 100
    while True:
        url = f"{base}/collections/{coll_key}/items?limit={limit}&start={start}"
        resp = await async_request("GET", url, headers=headers)
        if resp.status_code != 200:
            break
        batch = resp.json()
        if not batch:
            break
        all_items.extend(batch)
        if len(batch) < limit:
            break
        start += limit

    logger.info(f"Zotero 集合中共有 {len(all_items)} 个条目")

    # ── 2. 找到每个条目的 PDF 附件 ──
    pending = []
    for item in all_items:
        item_data = item.get("data", {})
        item_key = item.get("key")
        doi = item_data.get("DOI", "")
        title = item_data.get("title", "")

        children_url = f"{base}/items/{item_key}/children"
        child_resp = await async_request("GET", children_url, headers=headers)
        if child_resp.status_code != 200:
            continue

        for child in child_resp.json():
            child_data = child.get("data", {})
            if child_data.get("itemType") == "attachment" and child_data.get("linkMode") == "imported_file":
                att_key = child.get("key")
                filename = child_data.get("filename", "")
                pdf_path = storage_dir / att_key / filename
                if pdf_path.exists():
                    pending.append({
                        "item_key": item_key,
                        "doi": doi,
                        "title": title,
                        "pdf_path": pdf_path,
                    })
                    break

    if not pending:
        print("未找到可重读的 PDF 文件（Zotero 附件可能尚未同步到本地）。")
        return

    print(f"\n找到 {len(pending)} 篇可重读文献：")
    # 辅助：确保字符串可在 GBK 控制台安全打印
    def _console_safe(s: str) -> str:
        # 先剥离 Unicode non-characters 和 surrogate
        s = re.sub(r'[￾-￿\ud800-\udfff]', '', s)
        try:
            return s.encode('gbk', errors='replace').decode('gbk')
        except Exception:
            return s.encode('ascii', errors='replace').decode('ascii')
    for i, p in enumerate(pending, 1):
        print(f"  [{i}] {_console_safe(p['title'][:70])}")
        print(f"      DOI: {_console_safe(p['doi'] or 'N/A')} | PDF: {_console_safe(p['pdf_path'].name)}")

    if not _non_interactive:
        confirm = input(f"\n确认对以上 {len(pending)} 篇重新深度阅读？(y/n): ").strip().lower()
        if confirm != "y":
            print("已取消。")
            return

    # ── 3. 逐篇重新阅读 ──
    NOTES_DIR.mkdir(parents=True, exist_ok=True)
    results = []
    failed = []

    for i, item in enumerate(pending, 1):
        print(f"\n[{i}/{len(pending)}] {item['title'][:60]}")
        logger.info(f"重新深度阅读: {item['title'][:50]}")

        pdf_path = item["pdf_path"]

        # DOI
        doi = extract_doi_from_pdf(pdf_path) or item["doi"]
        if not doi:
            logger.warning(f"  ⊘ 无法提取 DOI: {pdf_path.name}")
            failed.append((item["title"], "DOI 提取失败"))
            continue

        # 提取文本
        text = extract_text_from_pdf(pdf_path)
        if not text:
            logger.warning(f"  ⊘ PDF 文本提取失败: {pdf_path.name}")
            failed.append((item["title"], "PDF 文本提取失败"))
            continue

        # DeepSeek V4 Pro 深度阅读
        notes = await generate_notes(text)
        if not notes:
            logger.error(f"  ⊘ DeepSeek AI 阅读失败: {pdf_path.name}")
            failed.append((item["title"], "DeepSeek AI 阅读失败"))
            continue

        # Gemini Vision 图表识别 + DeepSeek 图表分析
        gemini_key = config.get("gemini_api_key", "")
        if gemini_key:
            try:
                logger.info(f"  Gemini Vision 图表识别...")
                figure_data = await _recognize_figures_structured(pdf_path, gemini_key)
                if figure_data:
                    total_figs = sum(len(v) for v in figure_data.values())
                    if total_figs > 0:
                        notes["_figures"] = figure_data
                        logger.info(f"  ✓ Gemini 识别 {total_figs} 个图表/表格")

                        # 用 DeepSeek 对图表做结合文献的分析
                        fig_summary = _format_figure_data(figure_data)
                        analysis = await _analyze_figures_with_deepseek(
                            text[:3000], fig_summary, notes
                        )
                        if analysis:
                            notes["图表分析"] = analysis
                            logger.info(f"  ✓ 图表分析完成")
            except Exception as e:
                logger.warning(f"  Gemini 图表识别异常: {e}")

        # 填充影响因子
        journal_name = notes.get("期刊", "")
        if journal_name and journal_name != "原文未提及":
            if_val = get_impact_factor(journal_name)
            if if_val:
                notes["影响因子"] = if_val
                logger.info(f"  影响因子: {journal_name} → IF={if_val}")

        # 覆写 Obsidian 笔记
        notes["doi"] = doi
        notes["pdf_path"] = str(pdf_path)
        write_obsidian_note(notes, pdf_path, force=True)
        logger.info(f"  ✓ Obsidian 笔记已覆写")

        # 保存笔记 JSON
        notes["_read_date"] = datetime.datetime.now().isoformat()
        safe_doi = doi.replace("/", "_").replace("\\", "_")
        note_file = NOTES_DIR / f"{safe_doi}.json"
        with open(note_file, "w", encoding="utf-8") as f:
            json.dump(notes, f, ensure_ascii=False, indent=2)

        results.append(notes)
        logger.info(f"  ✓ 重新阅读完成 ({i}/{len(pending)})")

    # 更新 Excel（按 DOI 匹配更新）
    if results:
        update_excel_by_doi(results)
        logger.info(f"✓ Excel 知识库更新 {len(results)} 条")

    # 更新 processed_log（使 [6] 清理功能可以识别这些已处理文献）
    if results:
        save_processed_log([{"doi": r.get("doi", ""), "title": r.get("标题", "")} for r in results])
        logger.info(f"✓ processed_log 已更新 {len(results)} 条")

    # ── 报告 ──
    print("\n" + "=" * 60)
    print("  重新深度阅读完成")
    print("=" * 60)
    print(f"  成功: {len(results)}/{len(pending)} 篇")
    print(f"  失败: {len(failed)}/{len(pending)} 篇")
    print(f"  Obsidian 覆写: {len(results)} 篇")
    print(f"  Excel 更新:    {len(results)} 条")
    print("=" * 60)
    if results:
        print("\n已重新阅读：")
        for r in results:
            safe = r.get('标题', '?')[:70].encode('gbk', errors='replace').decode('gbk', errors='replace')
            print(f"  · {safe}")
    if failed:
        print("\n失败文献：")
        for title, reason in failed:
            safe_t = title[:70].encode('gbk', errors='replace').decode('gbk', errors='replace')
            print(f"  · {safe_t} | {reason}")


# ============================================================================
# 选项8：知识库问答 (Knowledge Base Q&A) — 教授级学术AI导师
# ============================================================================

# --- 模块级状态 ---
_qa_paper_index: list = []          # 内存论文索引，首次进入时构建
_qa_chat_history: list = []         # 对话历史 [{"role": "user"|"assistant", "content": ...}]
_qa_last_papers: list = []          # 上一轮检索到的论文
_qa_use_external: bool = True       # 是否允许使用外部知识
_qa_top_n: int = 8                  # 检索论文数量

# --- 英文停用词 ---
_QA_STOPWORDS: frozenset = frozenset({
    "the", "and", "for", "are", "but", "not", "you", "all", "can", "had",
    "her", "was", "one", "our", "out", "has", "have", "been", "some", "its",
    "with", "that", "this", "from", "they", "will", "would", "there", "their",
    "which", "were", "when", "what", "who", "how", "where", "into", "more",
    "than", "also", "about", "over", "after", "each", "other", "only", "most",
    "may", "such", "these", "them", "then", "should", "could",
})

# Q&A 笔记保存目录
_QA_NOTES_DIR = OBSIDIAN_DIR / "QA笔记"
# Q&A 对话历史持久化文件
_QA_HISTORY_FILE = SCRIPT_DIR / "qa_history.json"

# 防幻觉复核队列
_REVIEW_QUEUE_FILE = SCRIPT_DIR / "review_queue.json"


def _add_to_review_queue(notes: Dict[str, str]) -> None:
    """将 🟡 或 🔴 置信度的笔记加入人工复核队列。"""
    try:
        queue = []
        if _REVIEW_QUEUE_FILE.exists():
            queue = json.loads(_REVIEW_QUEUE_FILE.read_text(encoding="utf-8"))
        entry = {
            "title": notes.get("标题", "")[:120],
            "first_author": notes.get("第一作者", ""),
            "year": notes.get("年份", ""),
            "journal": notes.get("期刊", ""),
            "doi": notes.get("doi", ""),
            "confidence": notes.get("_置信度", "medium"),
            "issues": notes.get("_自检标记", ""),
            "added": datetime.datetime.now().isoformat(),
        }
        # 去重：同 DOI 只保留最新
        existing_doi = entry.get("doi", "").strip().lower()
        queue = [q for q in queue if q.get("doi", "").strip().lower() != existing_doi or not existing_doi]
        queue.append(entry)
        _REVIEW_QUEUE_FILE.write_text(json.dumps(queue, ensure_ascii=False, indent=2), encoding="utf-8")
        logger.info(f"[复核队列] 已加入: {entry['title'][:50]}")
    except Exception as e:
        logger.warning(f"复核队列写入失败: {e}")


def _get_review_queue() -> List[Dict]:
    """读取人工复核队列。"""
    try:
        if _REVIEW_QUEUE_FILE.exists():
            return json.loads(_REVIEW_QUEUE_FILE.read_text(encoding="utf-8"))
    except Exception:
        pass
    return []


def _remove_from_review_queue(doi: str) -> bool:
    """从复核队列中移除一条记录（人工复核完成）。"""
    try:
        if not _REVIEW_QUEUE_FILE.exists():
            return False
        queue = json.loads(_REVIEW_QUEUE_FILE.read_text(encoding="utf-8"))
        doi_lower = doi.strip().lower()
        new_queue = [q for q in queue if q.get("doi", "").strip().lower() != doi_lower]
        _REVIEW_QUEUE_FILE.write_text(json.dumps(new_queue, ensure_ascii=False, indent=2), encoding="utf-8")
        return len(new_queue) < len(queue)
    except Exception:
        return False


async def _retry_fix_notes(doi: str, issues: str) -> bool:
    """【复核队列修复】用现有笔记 + 问题描述让 DeepSeek 针对性修复。

    Args:
        doi: 论文 DOI
        issues: _自检标记 字段的问题描述（如 "FAIL: 变量汇总字段为空"）

    Returns:
        True 表示修复成功（无新问题），False 表示修复失败（仍有问题或无法修复）
    """
    import aiohttp

    if not doi or not issues:
        return False

    doi_lower = doi.strip().lower()

    # 1. 找到对应的 JSON 笔记
    notes_file = None
    existing_notes = None
    if NOTES_DIR.exists():
        for f in NOTES_DIR.glob("*.json"):
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                if data.get("doi", "").strip().lower() == doi_lower:
                    notes_file = f
                    existing_notes = data
                    break
            except Exception:
                continue

    if not existing_notes:
        logger.warning(f"[修复] 未找到 DOI={doi} 的 JSON 笔记，无法修复")
        return False

    # 2. 构建修复 prompt
    fix_prompt = f"""你是学术审稿人。以下是一篇论文的结构化笔记，自检发现了以下问题：

**问题**：
{issues}

**现有笔记**：
- 标题: {existing_notes.get("标题", "")}
- 作者: {existing_notes.get("作者", "")}
- 摘要: {existing_notes.get("摘要", "")[:1500]}
- 研究背景与动机: {existing_notes.get("研究背景与动机", "")[:1000]}
- 研究问题: {existing_notes.get("研究问题", "")[:1000]}
- 研究方法: {existing_notes.get("研究方法", "")[:1500]}
- 研究结果: {existing_notes.get("研究结果", "")[:1500]}
- 研究结论: {existing_notes.get("研究结论", "")[:1000]}
- 变量汇总: {existing_notes.get("变量汇总", "")}

请针对上述问题，只修复有问题的字段。只输出修复后的字段（JSON 格式），不要输出其他内容。
格式：{{"字段名": "修复后的内容", ...}}

注意：
1. 如果问题是"变量汇总为空"，请从研究方法和研究结果段落中提取变量信息（变量名称、类型、定义、测量方式）
2. 只修复问题中提到的字段，其他字段不要改动
3. 如果原文确实没有相关信息（如质性研究无定量变量），请输出空值"""

    # 3. 调用 DeepSeek
    deepseek_key = config.get("deepseek_api_key", "")
    if not deepseek_key:
        logger.error("[修复] DeepSeek API Key 未配置")
        return False

    try:
        async with aiohttp.ClientSession() as session:
            headers = {
                "Authorization": f"Bearer {deepseek_key}",
                "Content-Type": "application/json",
            }
            payload = {
                "model": config.get("deepseek_model", "deepseek-chat"),
                "messages": [
                    {"role": "system", "content": "你是学术审稿人，只输出修复后的 JSON，不输出任何解释。"},
                    {"role": "user", "content": fix_prompt},
                ],
                "temperature": 0.1,
                "max_tokens": 2048,
            }
            async with session.post(
                "https://api.deepseek.com/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=aiohttp.ClientTimeout(total=120),
            ) as resp:
                if resp.status != 200:
                    logger.error(f"[修复] API 错误: HTTP {resp.status}")
                    return False
                resp_data = await resp.json()
                content = resp_data["choices"][0]["message"]["content"].strip()

                # 解析修复后的 JSON
                import re as _re
                json_match = _re.search(r'\{.*\}', content, _re.DOTALL)
                if not json_match:
                    logger.error(f"[修复] 无法解析修复结果: {content[:200]}")
                    return False
                try:
                    fixes = json.loads(json_match.group())
                except json.JSONDecodeError:
                    logger.error(f"[修复] JSON 解析失败: {content[:200]}")
                    return False

                if not fixes:
                    logger.info("[修复] DeepSeek 判断无需修复（原文确实无相关信息）")
                    _remove_from_review_queue(doi)
                    return True

                # 4. 更新现有笔记
                updated_fields = []
                for key, val in fixes.items():
                    if val and isinstance(val, str) and len(val.strip()) > 5:
                        existing_notes[key] = val
                        updated_fields.append(key)
                        logger.info(f"[修复] ✓ {key} 已更新 ({len(val)} 字)")

                if not updated_fields:
                    logger.info("[修复] 无有效修复内容")
                    _remove_from_review_queue(doi)
                    return True

                # 5. 保存 JSON 笔记
                if notes_file:
                    notes_file.write_text(
                        json.dumps(existing_notes, ensure_ascii=False, indent=2),
                        encoding="utf-8",
                    )

                # 6. 更新 Obsidian 笔记（用 force=True 覆写）
                try:
                    write_obsidian_note(existing_notes, None, force=True)
                    logger.info("[修复] Obsidian 笔记已更新")
                except Exception as e:
                    logger.warning(f"[修复] Obsidian 更新失败: {e}")

                # 7. 更新 Excel
                try:
                    update_excel_by_doi([existing_notes])
                    logger.info("[修复] Excel 已更新")
                except Exception as e:
                    logger.warning(f"[修复] Excel 更新失败: {e}")

                # 8. 清除 _自检标记（修复后重新评估留给下次 generate_notes）
                if "_自检标记" in existing_notes:
                    del existing_notes["_自检标记"]
                if "_置信度" in existing_notes:
                    existing_notes["_置信度"] = "high"

                # 9. 移出复核队列
                _remove_from_review_queue(doi)
                logger.info(f"[修复] ✅ {existing_notes.get('标题','')[:60]} 修复完成，已移出复核队列")
                return True

    except Exception as e:
        logger.error(f"[修复] 异常: {e}")
        return False


def _save_qa_history() -> None:
    """静默保存对话历史到 JSON 文件，用于跨会话恢复。"""
    try:
        with open(_QA_HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(_qa_chat_history, f, ensure_ascii=False, indent=2)
    except Exception:
        pass  # 自动保存失败不阻塞用户操作


# ═══════════════════════════════════════════════════════
# 多会话管理
# ═══════════════════════════════════════════════════════
_QA_SESSIONS_DIR = SCRIPT_DIR / "qa_sessions"
_QA_SESSIONS_DIR.mkdir(parents=True, exist_ok=True)


def _list_qa_sessions() -> List[Dict]:
    """列出所有已保存的 QA 会话（按修改时间倒序）。"""
    sessions = []
    try:
        for f in sorted(_QA_SESSIONS_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
            data = json.loads(f.read_text(encoding="utf-8"))
            sessions.append({
                "id": f.stem,
                "name": data.get("name", f.stem[:30]),
                "created": data.get("created", ""),
                "updated": datetime.datetime.fromtimestamp(f.stat().st_mtime).isoformat(),
                "message_count": len(data.get("messages", [])),
                "first_question": data.get("first_question", ""),
            })
    except Exception:
        pass
    return sessions


def _save_qa_session(session_id: str, name: str, messages: List[Dict]) -> None:
    """保存一个 QA 会话到独立文件。"""
    try:
        now = datetime.datetime.now().isoformat()
        first_q = ""
        for m in messages:
            if m.get("role") == "user":
                first_q = m.get("content", "")[:80]
                break
        data = {
            "name": name,
            "created": now if not (_QA_SESSIONS_DIR / f"{session_id}.json").exists()
                      else json.loads((_QA_SESSIONS_DIR / f"{session_id}.json").read_text(encoding="utf-8")).get("created", now),
            "updated": now,
            "first_question": first_q,
            "messages": messages,
        }
        (_QA_SESSIONS_DIR / f"{session_id}.json").write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        logger.warning(f"保存会话失败: {e}")


def _load_qa_session(session_id: str) -> Optional[List[Dict]]:
    """加载指定会话的消息列表。"""
    try:
        path = _QA_SESSIONS_DIR / f"{session_id}.json"
        if path.exists():
            data = json.loads(path.read_text(encoding="utf-8"))
            return data.get("messages", [])
    except Exception:
        pass
    return None


def _delete_qa_session(session_id: str) -> bool:
    """删除指定会话。"""
    try:
        path = _QA_SESSIONS_DIR / f"{session_id}.json"
        if path.exists():
            path.unlink()
            return True
    except Exception:
        pass
    return False


def _generate_session_name(messages: List[Dict]) -> str:
    """根据对话内容自动生成会话名称。"""
    for m in messages:
        if m.get("role") == "user":
            q = m["content"].strip()
            # 截取第一个有意义的问题作为名称
            q_clean = re.sub(r'\d{4}-\d{4}年\s*', '', q)  # 去年份前缀
            return q_clean[:60] + ("..." if len(q_clean) > 60 else "")
    return f"会话 {datetime.datetime.now().strftime('%m-%d %H:%M')}"


def _tokenize(text: str) -> list[str]:
    """中英混合分词：英文单词 + 中文字符二元组。"""
    tokens: list[str] = []
    if not text:
        return tokens
    text_lower = text.lower()
    # 英文单词 (≥2字母)
    for m in re.finditer(r"[a-z]{2,}", text_lower):
        word = m.group()
        if word not in _QA_STOPWORDS:
            tokens.append(word)
    # 中文字符二元组 (bigram 比 unigram 更精确)
    chinese = re.findall(r"[一-鿿]", text)
    for i in range(len(chinese) - 1):
        tokens.append(chinese[i] + chinese[i + 1])
    return tokens


# ============================================================================
# 语义检索：字符 n-gram TF-IDF，零依赖（仅需 numpy）
# 在 token 匹配之上叠加一层语义相似度，捕捉同义词/近义词/跨语言模式
# ============================================================================
_semantic_vocab: dict = {}          # n-gram → column index
_semantic_vectors: "np.ndarray | None" = None  # (n_papers, n_vocab) TF-IDF 矩阵
_semantic_idf: "np.ndarray | None" = None      # (n_vocab,) IDF 权重


def _char_ngrams(text: str, n: int = 3) -> list[str]:
    """字符 n-gram 提取。中英文统一处理，捕捉子词和跨词模式。"""
    text = text.lower()
    text = re.sub(r'\s+', ' ', text)
    if len(text) < n:
        return [text] if text.strip() else []
    return [text[i:i + n] for i in range(len(text) - n + 1)]


def _build_semantic_index(papers: list[dict]):
    """构建字符 3-gram + 4-gram TF-IDF-L2 矩阵，用于语义相似度计算。"""
    global _semantic_vocab, _semantic_vectors, _semantic_idf
    from collections import Counter

    # 收集所有论文的 n-gram 集合
    all_ngrams: list[set[str]] = []
    for p in papers:
        text = p.get("search_text", "")
        ngrams = set(_char_ngrams(text, n=3) + _char_ngrams(text, n=4))
        all_ngrams.append(ngrams)

    # 文档频率 → 词汇表（只保留出现在 ≥2 篇文献中的 n-gram，控制噪音）
    df = Counter()
    for ngrams in all_ngrams:
        for ng in ngrams:
            df[ng] += 1
    vocab_items = [(ng, cnt) for ng, cnt in df.items() if cnt >= 2]
    vocab_items.sort(key=lambda x: -x[1])
    vocab_items = vocab_items[:8000]  # 上限 8000 特征

    _semantic_vocab = {ng: i for i, (ng, _) in enumerate(vocab_items)}
    n_papers = len(papers)
    n_vocab = len(_semantic_vocab)

    if n_vocab == 0:
        logger.warning("[QA] 语义索引词汇为空，回退到纯 token 检索。")
        _semantic_vectors = None
        _semantic_idf = None
        return

    # TF 矩阵
    tf_matrix = np.zeros((n_papers, n_vocab), dtype=np.float32)
    for i, ngrams in enumerate(all_ngrams):
        for ng in ngrams:
            if ng in _semantic_vocab:
                tf_matrix[i, _semantic_vocab[ng]] += 1

    # IDF
    df_counts = np.sum(tf_matrix > 0, axis=0).astype(np.float32)
    _semantic_idf = np.log((n_papers + 1) / (df_counts + 1)) + 1.0

    # TF-IDF + L2 归一化
    _semantic_vectors = tf_matrix * _semantic_idf
    norms = np.linalg.norm(_semantic_vectors, axis=1, keepdims=True)
    norms[norms == 0] = 1.0
    _semantic_vectors = _semantic_vectors / norms

    logger.info(f"[QA] 语义索引：{n_vocab} 个 n-gram 特征, {n_papers} 篇文献")


def _semantic_scores(query: str, paper_indices: list[int]) -> np.ndarray:
    """计算查询与指定论文的语义余弦相似度。返回 (len(paper_indices),) 的分数数组。"""
    if _semantic_vectors is None or not _semantic_vocab:
        return np.zeros(len(paper_indices))

    query_ngrams = set(_char_ngrams(query, n=3) + _char_ngrams(query, n=4))
    q_vec = np.zeros(len(_semantic_vocab), dtype=np.float32)
    for ng in query_ngrams:
        if ng in _semantic_vocab:
            q_vec[_semantic_vocab[ng]] += 1

    q_vec = q_vec * _semantic_idf
    q_norm = np.linalg.norm(q_vec)
    if q_norm > 0:
        q_vec = q_vec / q_norm

    return np.dot(_semantic_vectors[paper_indices], q_vec)


def _verify_answer_claims(answer: str, retrieved_papers: list[dict],
                          all_papers: list[dict]) -> list[str]:
    """验证 DeepSeek 回答中对本地文献的引用是否准确。

    扫描回答中的 (作者, 年份) 模式，与本地文献索引交叉比对。
    返回验证警告列表，用于提醒用户存疑的引用。
    """
    warnings: list[str] = []

    # 提取 (Author, Year) 引用模式
    # 英文: "De Bock (2026)", "Gelbrich et al. (2026)", "Zhang & Li (2025)", "Smith and Jones (2023)"
    cite_pattern = re.compile(
        r'([A-Z][a-zà-ü]+(?:\s[A-Z][a-zà-ü]+)?'     # 单/双词姓氏 (De Bock, Van Dijk)
        r'(?:\s(?:and|&)\s[A-Z][a-zà-ü]+(?:\s[A-Z][a-zà-ü]+)?)?'  # 可选第二作者
        r'(?:\set\s+al\.?)?)'                         # 可选 et al.（之后不跟名字）
        r'\s*[\(（](\d{4})[\)）]',
    )
    # 中文: "张三（2025）", "李四等（2024）"
    cn_cite_pattern = re.compile(
        r'([一-鿿]{2,4}(?:等)?)\s*[\(（](\d{4})[\)）]',
    )

    # 构建本地索引查找表: {lastname_firstname: {year: paper}}
    local_index: dict[str, dict[str, dict]] = {}
    for p in all_papers:
        authors = p.get("authors", "")
        year = str(p.get("year", ""))
        if not authors or not year:
            continue
        # 取第一作者姓氏
        first_author = authors.split(",")[0].split(";")[0].strip()
        lastname = first_author.split()[-1] if first_author else ""
        key = lastname.lower()
        if key not in local_index:
            local_index[key] = {}
        if year not in local_index[key]:
            local_index[key][year] = p

    retrieved_keys: set[tuple[str, str]] = set()
    for p in retrieved_papers:
        authors = p.get("authors", "")
        year = str(p.get("year", ""))
        first_author = authors.split(",")[0].split(";")[0].strip()
        lastname = first_author.split()[-1] if first_author else ""
        retrieved_keys.add((lastname.lower(), year))

    # 检查英文引用
    for m in cite_pattern.finditer(answer):
        name = m.group(1)
        year = m.group(2)
        lastname = name.split()[-1].split()[-1].lower()  # handles "et al."
        full_ref = f"{name} ({year})"

        if (lastname, year) in retrieved_keys:
            continue  # 已检索到，可直接对照原文核实
        if lastname in local_index and year in local_index[lastname]:
            # 存在于本地库但未被检索到
            warnings.append(
                f"[本地未检索] {full_ref} — 该文献在本地库中但未被本次检索匹配到，"
                f"建议核实引用内容: {local_index[lastname][year]['title'][:60]}"
            )
        elif lastname in local_index:
            # 作者存在但年份不同
            avail_years = sorted(local_index[lastname].keys())
            warnings.append(
                f"[年份不匹配] {full_ref} — 本地库中 {name} 的文献年份为 {', '.join(avail_years)}，"
                f"未找到 {year} 年的记录。可能来自外部知识。"
            )

    return warnings


def _parse_obsidian_note(filepath: Path) -> Optional[dict]:
    """解析单篇 Obsidian 笔记：提取 YAML frontmatter + 7个章节正文。"""
    try:
        content = filepath.read_text(encoding="utf-8")
    except Exception:
        return None

    # --- 提取 YAML frontmatter（第一个 --- ... --- 块）---
    fm_match = re.match(r"^---\s*\n(.*?)\n---\s*\n", content, re.DOTALL)
    if not fm_match:
        return None  # 无 frontmatter 的不是论文笔记
    fm_text = fm_match.group(1)
    body_text = content[fm_match.end():]

    # 简易 YAML 解析（flat key: value，无嵌套）
    paper: dict = {}
    for line in fm_text.split("\n"):
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        m = re.match(r"^(\w+):\s*(.*)", line)
        if not m:
            continue
        key = m.group(1).strip()
        val = m.group(2).strip()
        # 去掉首尾引号
        if len(val) >= 2 and val[0] == val[-1] and val[0] in ('"', "'"):
            val = val[1:-1]
        # tags 字段特殊处理：列表 → 分号分隔字符串
        if key == "tags" and val.startswith("["):
            items = re.findall(r"[\w-]+", val)
            val = "; ".join(items)
        paper[key] = val

    title = paper.get("title", "")
    if not title or title == "Untitled":
        return None  # 不是有效论文笔记

    # --- 提取章节正文 ---
    SECTION_NAMES = [
        "研究背景与动机", "研究问题", "变量汇总", "研究方法",
        "方法论详解", "研究结果", "讨论与结论", "创新点", "局限与展望",
    ]
    sections: dict = {}       # 截断版，用于检索打分
    sections_full: dict = {}  # 完整版，发送给 DeepSeek
    for i, sec_name in enumerate(SECTION_NAMES):
        # 匹配 "## 研究背景与动机" 之类的 header
        pattern = rf"^##\s+{re.escape(sec_name)}\s*\n(.*?)(?=^##\s|\Z)"
        m = re.search(pattern, body_text, re.DOTALL | re.MULTILINE)
        if m:
            text = m.group(1).strip()
            sections_full[sec_name] = text  # 保留完整原文
            # 检索用截断版：控制内存占用，800 字符对 token 匹配已足够
            if len(text) > 800:
                text = text[:800] + "…"
            sections[sec_name] = text
        else:
            sections[sec_name] = ""
            sections_full[sec_name] = ""

    # --- 构建 search_text（用于检索打分）---
    search_parts = [title]
    kw = paper.get("keywords", "")
    if kw:
        search_parts.append(kw.replace(";", " ").replace("，", " "))
    jn = paper.get("journal", "")
    if jn:
        search_parts.append(jn)
    # 章节标题（帮助匹配中文查询）
    search_parts.append(" ".join(SECTION_NAMES))
    for sn in SECTION_NAMES:
        if sections.get(sn):
            search_parts.append(sections[sn])  # 截断版用于检索（本地匹配，不占API token）
    search_text = " ".join(search_parts)

    # --- 构建 full_body（用于发送给 DeepSeek 的上下文）---
    # 使用完整章节正文，确保教授级问答能获取论文全部细节
    body_parts = []
    for sn in SECTION_NAMES:
        if sections_full.get(sn):
            body_parts.append(f"## {sn}\n{sections_full[sn]}")
    full_body = "\n\n".join(body_parts)

    relpath = filepath.relative_to(OBSIDIAN_DIR)
    return {
        "path": str(filepath),
        "relpath": str(relpath).replace("\\", "/"),
        "title": title,
        "authors": paper.get("authors", ""),
        "first_author": paper.get("first_author", ""),
        "year": paper.get("year", ""),
        "journal": paper.get("journal", ""),
        "impact_factor": paper.get("impact_factor", ""),
        "keywords": paper.get("keywords", ""),
        "tags": paper.get("tags", ""),
        "doi": paper.get("doi", ""),
        "sections": sections,
        "full_body": full_body,
        "search_text": search_text,
    }


def _build_paper_index(force_rebuild: bool = False) -> list[dict]:
    """遍历 Obsidian vault，构建内存论文索引。首次构建后缓存。"""
    global _qa_paper_index
    if _qa_paper_index and not force_rebuild:
        return _qa_paper_index

    papers: list[dict] = []
    md_files = list(OBSIDIAN_DIR.glob("**/*.md"))
    for fp in md_files:
        # 跳过非论文文件
        if fp.name == "_index.md":
            continue
        if "concepts" in fp.parts:
            continue
        if ".obsidian" in fp.parts:
            continue
        paper = _parse_obsidian_note(fp)
        if paper:
            papers.append(paper)

    # 按年份降序 + 影响因子降序排列
    def _sort_key(p: dict) -> tuple:
        try:
            yr = int(p.get("year", "0"))
        except (ValueError, TypeError):
            yr = 0
        try:
            if_val = float(p.get("impact_factor", "0"))
        except (ValueError, TypeError):
            if_val = 0.0
        return (-yr, -if_val)

    papers.sort(key=_sort_key)
    _qa_paper_index = papers
    logger.info(f"[QA] 论文索引构建完成：{len(papers)} 篇")

    # 同步构建语义索引（字符 n-gram TF-IDF）
    _build_semantic_index(papers)

    return papers


# 中英文学术概念映射：用户用中文提问，但论文元数据多是英文
_CONCEPT_MAP: dict = {
    # ── AI 技术与智能体 ──
    "智能体": "agent chatbot conversational agent virtual assistant embodied agent autonomous agent multi-agent",
    "聊天机器人": "chatbot conversational agent virtual assistant dialogue system",
    "对话式ai": "conversational ai conversational agent chatbot dialogue system nlp",
    "生成式ai": "generative ai genai chatgpt large language model llm gpt foundation model text generation image generation",
    "大语言模型": "large language model llm gpt foundation model transformer bert pretrained model",
    "自然语言处理": "natural language processing nlp text analysis sentiment analysis computational linguistics",
    "机器学习": "machine learning deep learning neural network supervised unsupervised reinforcement learning",
    "推荐系统": "recommender system recommendation algorithm personalization collaborative filtering",
    "计算机视觉": "computer vision image recognition visual analysis facial recognition",
    "人机协同": "human ai collaboration human in the loop human machine collaboration hybrid intelligence augmentation",
    "人机交互": "human computer interaction hci human ai interaction user interface ux",
    "具身认知": "embodied cognition sensory perception grounding situated cognition",
    "可解释性": "explainability explainable ai xai transparency interpretability black box",
    # ── 消费者心理与行为 ──
    "消费者信任": "consumer trust trust in ai perceived trustworthiness credibility reliability confidence",
    "消费者行为": "consumer behavior purchase intention decision making adoption acceptance willingness to pay",
    "消费者态度": "consumer attitude perception evaluation satisfaction loyalty engagement",
    "消费者心理": "consumer psychology cognition affect emotion motivation attitude persuasion",
    "个性化": "personalization personalized recommendation customization targeting segmentation tailoring",
    "隐私": "privacy data privacy information privacy surveillance data protection gdpr",
    "算法偏见": "algorithmic bias fairness discrimination ethics justice equity",
    "拟人化": "anthropomorphism humanization human-like warmth social presence",
    "用户体验": "user experience ux usability satisfaction engagement flow",
    "情感": "emotion affect sentiment mood feeling emotional response arousal valence",
    # ── 品牌与广告 ──
    "品牌管理": "brand management brand equity brand attitude brand trust brand loyalty brand personality",
    "品牌": "brand branding brand identity brand image brand positioning brand awareness",
    "广告": "advertising ad advertisement persuasion ad effectiveness ad creativity",
    "社交媒体": "social media influencer content creator platform engagement virality",
    "内容营销": "content marketing content creation storytelling user generated content",
    "口碑": "word of mouth ewom electronic word of mouth review rating online review",
    # ── 营销战略 ──
    "数字营销": "digital marketing online marketing social media marketing mobile marketing omnichannel",
    "服务营销": "service marketing customer experience service quality satisfaction loyalty",
    "顾客体验": "customer experience consumer experience service experience journey touchpoint",
    "价值共创": "value co-creation co-creation customer participation engagement",
    "客户关系": "customer relationship crm loyalty retention churn engagement",
    "定价": "pricing price promotion discount willingness to pay revenue",
    "全渠道": "omnichannel multichannel channel integration retail online offline",
    # ── 研究方法 ──
    "研究空白": "research gap future research research agenda future direction limitation opportunity",
    "研究缺口": "research gap future research research agenda underexplored overlooked",
    "文献综述": "literature review systematic review bibliometric analysis meta analysis review",
    "理论模型": "theoretical model conceptual framework theory model framework mechanism",
    "实验设计": "experiment experimental design randomized controlled trial field experiment lab experiment",
    "问卷调查": "survey questionnaire scale measurement construct psychometric",
    "质性研究": "qualitative research interview focus group thematic analysis grounded theory ethnography",
    "定量研究": "quantitative research statistical analysis regression sem structural equation modeling",
    "元分析": "meta-analysis meta analysis systematic review effect size",
    "眼动": "eye tracking eye movement visual attention gaze fixation",
    "神经科学": "neuroscience fmri eeg neuromarketing biometric physiological",
    "文本分析": "text analysis text mining content analysis natural language processing topic modeling",
    "大数据": "big data data mining predictive analytics machine learning data driven",
    # ── 理论与概念 ──
    "技术接受模型": "technology acceptance model tam utaut technology adoption is acceptance",
    "信任理论": "trust theory trustworthiness credibility benevolence integrity ability",
    "社会存在": "social presence parasocial relationship social interaction social cues",
    "自我决定": "self-determination autonomy competence relatedness intrinsic motivation",
    "认知负荷": "cognitive load information processing attention mental effort",
    "心流": "flow flow theory optimal experience immersion engagement",
    "期望确认": "expectation confirmation satisfaction continuance is continuance",
    "创新扩散": "innovation diffusion technology adoption doi diffusion of innovation",
    "信号理论": "signaling theory signal credibility quality signal information asymmetry",
    "归因理论": "attribution theory causal attribution blame responsibility",
    "调节聚焦": "regulatory focus promotion prevention regulatory fit",
    "解释水平": "construal level psychological distance abstraction concreteness",
    "抗拒理论": "reactance psychological reactance freedom threat persuasion resistance",
    # ── 伦理与社会 ──
    "伦理": "ethics ethical ai ethics responsible ai fairness accountability transparency",
    "可持续": "sustainability sustainable esg environmental social green",
    "数字鸿沟": "digital divide digital literacy digital inequality access",
    "虚假信息": "misinformation disinformation fake news deepfake deception",
    # ── 期刊/发表 ──
    "顶刊": "top journal journal of marketing journal of consumer research marketing science",
    "utd": "utd24 ft50 top tier journal management science",
}
_CONCEPT_RE = re.compile(
    "|".join(re.escape(k) for k in sorted(_CONCEPT_MAP.keys(), key=len, reverse=True)),
    re.IGNORECASE,
)


def _expand_query(query: str) -> str:
    """将中文查询中的概念映射为对应的英文学术词汇，提升跨语言检索效果。"""
    expanded = _CONCEPT_RE.sub(lambda m: m.group() + " " + _CONCEPT_MAP.get(m.group().lower(), ""), query)
    return expanded


def _parse_year_filter(query: str) -> tuple[int | None, int | None]:
    """从查询中提取年份过滤条件。支持：2025-2026年、2025年到2026年、近两年等。"""
    # "2025年到2026年" / "2025-2026年" / "2025年至2026年"
    m = re.search(r"(\d{4})\s*[年\-至到]\s*(?:20)?(\d{2,4})\s*年?", query)
    if m:
        y1 = int(m.group(1))
        y2 = int(m.group(2))
        if y2 < 100:
            y2 = 2000 + y2
        return (min(y1, y2), max(y1, y2))
    # "2025年" (单年)
    m = re.search(r"(?<!\d)(\d{4})\s*年", query)
    if m:
        y = int(m.group(1))
        return (y, y)
    # "近N年" / "最近N年"
    m = re.search(r"[近最]近?\s*(\d+)\s*年", query)
    if m:
        n = int(m.group(1))
        import datetime
        current_year = datetime.datetime.now().year
        return (current_year - n + 1, current_year)
    return (None, None)


def _score_paper(query: str, paper: dict) -> float:
    """加权 token overlap 打分。查询会先做概念扩展。"""
    expanded_query = _expand_query(query)
    query_tokens = _tokenize(expanded_query)
    if not query_tokens:
        return 0.0

    def _overlap(target_text: str) -> float:
        target_tokens = _tokenize(target_text)
        if not target_tokens:
            return 0.0
        q_set = set(query_tokens)
        t_set = set(target_tokens)
        inter = q_set & t_set
        return len(inter) / len(q_set)  # 查询 token 的覆盖率

    score = (
        _overlap(paper.get("title", "")) * 4.0 +
        _overlap(paper.get("keywords", "")) * 3.0 +
        _overlap(paper.get("search_text", "")) * 1.5 +
        _overlap(paper.get("journal", "")) * 1.0
    )

    # 高影响因子期刊微幅加成
    try:
        if_val = float(paper.get("impact_factor", "0"))
        if if_val >= 8.0:
            score *= 1.05
    except (ValueError, TypeError):
        pass

    return score


def _retrieve_papers(query: str) -> list[dict]:
    """检索最相关的 N 篇本地文献。Token 匹配 + 语义相似度混合打分，支持年份过滤和智能回退。"""
    global _qa_paper_index, _qa_top_n
    if not _qa_paper_index:
        _build_paper_index()

    def _blended_topn(cands: list[dict], n: int) -> list[dict]:
        """Token + 语义混合打分，取 top-N。"""
        if not cands:
            return []
        # Token 分数
        token_scored = _score_and_rank(query, cands)
        # 语义分数 — 查找候选在完整索引中的位置
        cand_indices = []
        for p, _ in token_scored:
            try:
                cand_indices.append(_qa_paper_index.index(p))
            except ValueError:
                cand_indices.append(0)
        sem_scores = _semantic_scores(query, cand_indices)
        # 归一化 + 混合
        t_max = max(s for _, s in token_scored) if token_scored else 1.0
        s_max = float(np.max(sem_scores)) if len(sem_scores) > 0 and np.max(sem_scores) > 0 else 1.0
        blended = []
        for i, (paper, ts) in enumerate(token_scored):
            tn = ts / t_max if t_max > 0 else 0
            sn = sem_scores[i] / s_max if s_max > 0 else 0
            blended.append((paper, 0.55 * tn + 0.45 * sn))
        blended.sort(key=lambda x: x[1], reverse=True)
        return [p for p, s in blended[:n] if s > 0.03]

    # 年份过滤
    y_min, y_max = _parse_year_filter(query)
    candidates = _qa_paper_index
    year_note = ""
    if y_min is not None:
        filtered = [p for p in candidates
                    if _try_parse_year(p.get("year", "")) and
                    y_min <= _try_parse_year(p.get("year", "")) <= y_max]
        if filtered:
            candidates = filtered
            year_note = f" {y_min}-{y_max}"
        else:
            print(f"  [检索] 年份 {y_min}-{y_max} 无匹配，已放宽年份限制。")

    # 第一轮：混合检索
    results = _blended_topn(candidates, _qa_top_n)

    # 智能回退：如果结果太少（<3条），去掉年份过滤重试
    if len(results) < 3 and y_min is not None:
        results = _blended_topn(_qa_paper_index, _qa_top_n)
        if results:
            print(f"  [检索] 精确匹配较少，已扩大检索范围。")

    # 最终回退：如果仍无结果，返回最新的高影响力论文
    if not results:
        results = _qa_paper_index[:_qa_top_n]
        print(f"  [检索] 未找到直接匹配，返回最新文献供参考。")

    # 打印年份分布
    years = set()
    for p in results:
        y = p.get("year", "")
        if y:
            years.add(str(y))
    print(f"  [检索] 匹配 {len(results)} 篇{year_note}（年份: {', '.join(sorted(years))}）")

    return results


def _try_parse_year(val: str) -> int:
    try:
        return int(str(val).strip()[:4])
    except (ValueError, TypeError):
        return 0


def _score_and_rank(query: str, candidates: list[dict]) -> list[tuple[dict, float]]:
    scored = [(p, _score_paper(query, p)) for p in candidates]
    scored.sort(key=lambda x: x[1], reverse=True)
    return scored


def _build_qa_messages(question: str, context_papers: list[dict],
                       chat_history: list[dict]) -> list[dict]:
    """构建发送给 DeepSeek 的完整 messages 数组。"""

    system_prompt = """你是一位在市场营销与人工智能营销交叉领域执教多年的资深教授。你的学生（一位博士生）会向你请教各种学术问题——从找文献、写综述，到理论讨论、方法选择、研究设计。你认真对待每一个问题，因为你知道这关系到学生的学术成长。

## 你的学术素养
- 理论功底深厚：能追溯概念的学术源头，比较不同学派的核心观点与争议
- 方法严谨：能评价研究设计的优劣，指出因果推断的局限，建议更合适的方法
- 深入浅出：能把复杂的概念、模型、方法拆解开来逐步讲解，让学生真正理解
- 视野开阔：既能深入一个具体问题，也能跳出该问题看到更广阔的学术图景和研究机会
- 诚实谦逊：不确定的地方明确标注"这一点我并非完全确定"，不假装全知

## 你的知识来源
1. [本地文献] —— 用户已精读并整理的论文笔记，是你回答的首要根据地
2. [外部知识] —— 你自己的学术积累（经典理论、经典文献、方法知识、学科共识）

## 根据问题类型调整回答方式

当学生请你**查找/推荐文献**时：
- 先列出本地文献库中匹配的论文，简要说明每篇与问题的关联
- 再推荐值得检索的外部文献（标注"建议检索"），包括作者、大致方向、推荐理由
- 说明检索策略：可以搜哪些关键词、关注哪些期刊

当学生请你**写文献综述**时：
- 按理论脉络或研究主题组织，而非逐篇罗列
- 指出不同研究的共识、分歧、演化趋势
- 最后给出总结性评述：这个领域整体走到了哪里，还缺什么

当学生与你**讨论理论或概念**时：
- 从经典定义讲到前沿发展，追溯概念的学术演化
- 讲清核心假设、实证支持、适用边界
- 联系学生已有的本地文献：哪些论文支持了这个理论，哪些提出了挑战
- 用具体例子帮助理解抽象概念

当学生请你**比较方法或评价研究设计**时：
- 说清每种方法的逻辑、假设、优劣、适用场景
- 如果本地文献中有使用这些方法的论文，引用它们作为案例
- 给出具体建议：在学生的研究情境下，哪种方法更合适，为什么

当学生请你**找研究空白**时：
- 基于本地文献的逻辑链条，指出哪些问题已被充分研究、哪些还缺
- 区分"没人做过所以是空白"和"有人尝试过但没解决所以是机会"
- 建议可行的研究切入点和理论视角

当学生与你**开放式学术讨论**时：
- 展现批判性思维，可以提出与主流观点不同的看法
- 引用证据支持你的论点，区分"学界共识"、"少数派观点"和"我个人认为"
- 鼓励学生思考：这背后更深层的问题是什么？

## 回答规则
1. 始终标注信息来源：[本地文献] / [外部知识]
2. 不确定的内容标注"待验证"或"我对此并非完全确定"
3. 推荐外部文献时注明"建议检索"——你的学生需要知道哪些是已验证的、哪些是需要自己去核实的
4. 中文回答，专业术语保留英文"""

    messages: list[dict] = [{"role": "system", "content": system_prompt}]

    # 最近 6 轮对话历史（12 条消息）
    if chat_history:
        messages.extend(chat_history[-12:])

    # 构建本地文献上下文
    if context_papers:
        context_blocks: list[str] = []
        for i, p in enumerate(context_papers, 1):
            block = f"""### [{i}] {p['title']}
**作者**: {p.get('authors', 'N/A')}
**期刊**: {p.get('journal', 'N/A')} ({p.get('year', 'N/A')}) | IF={p.get('impact_factor', 'N/A')}
**关键词**: {p.get('keywords', 'N/A')}

{p.get('full_body', '')}"""
            context_blocks.append(block)
        context_text = "\n\n---\n\n".join(context_blocks)

        user_message = f"""以下是你本地文献库中与学生问题相关的论文深度阅读笔记：

{context_text}

---
**学生的问题**: {question}

请根据问题类型，以教授的身份回答。如果本地文献足以回答，以本地文献为主、外部知识为辅。如果本地文献不足以完全回答，请调用你的知识储备进行补充，明确标注每部分信息的来源。"""
    else:
        user_message = f"""**学生的问题**: {question}

本地文献库中未找到与该问题直接相关的论文。请基于你的学术知识储备回答，标注为[外部知识]。如果合适，建议学生检索哪些方向的文献、使用哪些关键词、关注哪些期刊。"""

    messages.append({"role": "user", "content": user_message})
    return messages


async def _chat_via_deepseek(messages: list[dict]) -> Optional[str]:
    """通用 DeepSeek 对话函数。复用现有 API 调用模式。"""
    deepseek_key = config.get("deepseek_api_key", "")
    deepseek_model = config.get("deepseek_model", "deepseek-chat")
    if not deepseek_key:
        logger.error("[QA] DeepSeek API Key 未配置！")
        return None

    # 清洗消息中的 surrogate 字符（部分旧笔记可能有编码残留）
    def _clean_text(s: str) -> str:
        if not s:
            return ""
        # 移除孤立 surrogate 字符 (U+D800-U+DFFF)
        return re.sub(r'[\ud800-\udfff]', '', s)

    clean_messages = []
    for m in messages:
        clean_messages.append({
            "role": m["role"],
            "content": _clean_text(m.get("content", "")),
        })

    try:
        async with aiohttp.ClientSession() as session:
            headers = {
                "Authorization": f"Bearer {deepseek_key}",
                "Content-Type": "application/json",
            }
            payload = {
                "model": deepseek_model,
                "messages": clean_messages,
                "temperature": 0.5,
                "max_tokens": 4096,
            }
            async with session.post(
                "https://api.deepseek.com/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=aiohttp.ClientTimeout(total=180),
            ) as resp:
                if resp.status != 200:
                    error_text = await resp.text()
                    logger.error(f"[QA] DeepSeek API 错误: HTTP {resp.status} - {error_text[:200]}")
                    return None
                result = await resp.json()
                return result["choices"][0]["message"]["content"]
    except Exception as e:
        logger.error(f"[QA] DeepSeek API 调用异常: {e}")
        return None


def _print_qa_help():
    """打印 Q&A 模式帮助。"""
    print("""
╔══════════════════════════════════════════════════════╗
║  知识库问答 — 教授级学术AI导师                        ║
╠══════════════════════════════════════════════════════╣
║  直接输入问题即可获得基于本地文献+外部知识的深度回答    ║
║                                                      ║
║  特殊命令:                                            ║
║    /exit      — 返回主菜单                            ║
║    /save      — 保存当前对话到 Obsidian QA笔记         ║
║    /clear     — 清空对话历史                          ║
║    /papers    — 重新显示上一轮检索到的本地文献          ║
║    /n <数字>  — 设置检索文献数量 (默认8, 范围3-15)     ║
║    /external  — 切换是否允许外部知识 (当前: {state})   ║
║    /help      — 显示此帮助                            ║
║                                                      ║
║  信息来源标注:                                        ║
║    [本地]本地文献 = 来自你的 Obsidian 论文笔记          ║
║    [外部]外部知识 = 来自 DeepSeek 的学术知识储备       ║
╚══════════════════════════════════════════════════════╝
""".format(state="开启" if _qa_use_external else "关闭"))


def _save_qa_chat(chat_history: list[dict]):
    """保存当前 Q&A 对话到 Obsidian QA笔记目录。"""
    if not chat_history:
        print("  [QA] 没有对话内容可保存。")
        return

    _QA_NOTES_DIR.mkdir(parents=True, exist_ok=True)
    now = datetime.datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    filename = f"QA_{timestamp}.md"
    filepath = _QA_NOTES_DIR / filename

    # 提取首个问题作为摘要
    first_q = ""
    for msg in chat_history:
        if msg["role"] == "user":
            first_q = msg["content"][:80]
            break

    # 收集引用文献的 wikilinks
    ref_wikilinks: list[str] = []
    if _qa_last_papers:
        for p in _qa_last_papers:
            relpath = p.get("relpath", "")
            if relpath:
                ref_name = relpath.replace(".md", "").replace("/", "/")
                ref_wikilinks.append(f"  - [[{relpath.replace('.md', '')}|{p['title'][:60]}]]")

    # 构建内容
    lines = [
        "---",
        f"date: {now.strftime('%Y-%m-%d %H:%M')}",
        f"question_summary: \"{first_q}\"",
        "tags: [QA笔记]",
        "---",
        "",
        "# 知识库问答记录",
        "",
    ]

    for msg in chat_history:
        if msg["role"] == "user":
            lines.append(f"## 🤔 问题\n\n{msg['content']}\n")
        elif msg["role"] == "assistant":
            lines.append(f"## 🤖 回答\n\n{msg['content']}\n")

    if ref_wikilinks:
        lines.append("## 引用的本地文献\n")
        lines.extend(ref_wikilinks)
        lines.append("")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"  [QA] 对话已保存到: {filepath}")
    logger.info(f"[QA] 对话已保存: {filepath}")


# ═══════════════════════════════════════════════════════════════
#  Agent 模式 — 规划 + 工具执行
# ═══════════════════════════════════════════════════════════════

# Agent 可用工具注册表
AGENT_TOOLS = {
    "search_papers": {
        "description": "检索本地文献库。参数：query(搜索查询), n(返回篇数,默认8)。返回论文列表（标题、作者、年份、期刊、摘要）。",
        "function": "_tool_search_papers",
    },
    "read_paper": {
        "description": "深度阅读一篇论文的完整笔记。参数：title(论文标题，模糊匹配)。返回完整笔记（18个字段）。",
        "function": "_tool_read_paper",
    },
    "find_theories": {
        "description": "从本地文献库中找出与某个主题相关的理论框架。参数：topic(研究主题)。返回理论列表及来源论文。",
        "function": "_tool_find_theories",
    },
    "find_gaps": {
        "description": "从本地文献的未来研究方向中提取与某个主题相关的研究空白。参数：topic(研究主题)。返回研究方向列表。",
        "function": "_tool_find_gaps",
    },
    "extract_methods": {
        "description": "提取文献中使用的分析方法。参数：papers(论文标题列表，最多5篇)。返回每篇论文的方法摘要。",
        "function": "_tool_extract_methods",
    },
    "search_all_papers": {
        "description": "【全面检索】检索本地文献库中与query相关的ALL论文（不限制返回数量）。适用于理论梳理、文献综述、研究空白分析等需要穷尽所有相关文献的任务。参数：query(搜索查询)。返回所有匹配论文的完整列表。",
        "function": "_tool_search_all_papers",
    },
    "search_external": {
        "description": "调用AI的外部知识回答本地文献未覆盖的问题。参数：question(需要外部知识回答的问题)。返回基于AI知识的回答。",
        "function": "_tool_search_external",
    },
}

AGENT_PLANNER_SYSTEM = """你是一个学术研究任务的规划器。用户会给你一个研究目标，你需要把它拆解为可执行的步骤序列。

你可以使用的工具只有以下这些（不要编造其他工具）：
{tool_descriptions}

可用数据：本地文献库论文，每篇有完整的结构化笔记（标题、作者、理论、变量、方法、结果、未来研究方向等）。使用 search_all_papers 可获取全部相关论文。

请输出一个 JSON 数组，每个元素是一个步骤：
[
  {{"tool": "search_papers", "args": {{"query": "...", "n": 8}}, "reason": "为什么需要这步"}},
  {{"tool": "read_paper", "args": {{"title": "..."}}, "reason": "..."}},
  ...
]

## 工具选择指南（重要！）
- 快速查找、了解概况 → search_papers（返回 top-N，默认8篇）
- 理论梳理、文献综述、研究空白分析 → search_all_papers（返回所有相关论文，不截断）
- 查找某个主题的理论框架 → find_theories（已在全库搜索，无需配合 search_all_papers）
- 查找未来研究方向 → find_gaps（已在全库搜索，无需配合 search_all_papers）
- 深度阅读单篇论文 → read_paper
- 提取研究方法 → extract_methods
- 本地文献无法覆盖时补充 → search_external

## 规则
1. 简单的事实性问题、概念解释、一两句话能回答的问题 → 输出空数组 []（走快速问答模式）
2. 只有需要多篇文献综合、跨步骤推理、需要"找→读→分析→写"链路的问题才需要规划
3. 第一步通常是 search_papers 或 search_all_papers 或 find_theories，最后不要加"write/summarize"步骤（系统会自动综合）
4. 步骤 2-5 个，不要过度拆分
5. search_external 只在本地文献明确无法覆盖时使用，不要作为首选
6. 当用户明确要求"所有""全部""穷尽""全面"检索时，必须使用 search_all_papers 而非 search_papers

只输出 JSON 数组，不要输出任何其他文字。"""


def _agent_get_tool_descriptions() -> str:
    return "\n".join(f"- {name}: {info['description']}" for name, info in AGENT_TOOLS.items())


# ═══ Tool Implementations ═══

def _tool_search_papers(query: str, n: int = 8) -> list[dict]:
    """检索本地文献库。"""
    global _qa_top_n
    old_n = _qa_top_n
    _qa_top_n = n
    try:
        papers = _retrieve_papers(query)
        return [
            {
                "title": p.get("title", "?")[:150],
                "first_author": p.get("first_author", "?"),
                "year": p.get("year", "?"),
                "journal": p.get("journal", "?")[:50],
                "keywords": p.get("keywords", "")[:80],
                "abstract_snippet": p.get("abstract", "")[:200] if p.get("abstract") else "",
            }
            for p in papers
        ]
    finally:
        _qa_top_n = old_n


def _tool_search_all_papers(query: str) -> list[dict]:
    """【全面检索】检索本地文献库中所有相关论文，不限制返回数量。

    适用于理论梳理、文献综述、研究空白分析等需要穷尽所有相关文献的任务。
    与 search_papers 的区别：search_papers 返回 top-N（默认8篇），本工具返回全部匹配论文。
    """
    global _qa_top_n, _qa_paper_index
    if not _qa_paper_index:
        _build_paper_index()

    old_n = _qa_top_n
    _qa_top_n = len(_qa_paper_index)  # 返回全部
    try:
        papers = _retrieve_papers(query)
        return [
            {
                "title": p.get("title", "?")[:200],
                "first_author": p.get("first_author", "?"),
                "year": p.get("year", "?"),
                "journal": p.get("journal", "?")[:80],
                "keywords": p.get("keywords", "")[:120],
                "abstract_snippet": p.get("abstract", "")[:300] if p.get("abstract") else "",
                "theories": p.get("theories", "")[:200],
                "future_directions": p.get("future_directions", "")[:200],
            }
            for p in papers
        ]
    finally:
        _qa_top_n = old_n


def _tool_read_paper(title: str) -> Optional[dict]:
    """深度阅读一篇论文的完整笔记。"""
    global _qa_paper_index
    if not _qa_paper_index:
        _build_paper_index()

    # 模糊匹配标题
    best = None
    best_score = 0
    title_lower = title.lower().strip()
    for p in _qa_paper_index:
        p_title = p.get("title", "").lower()
        # Token overlap
        query_tokens = set(title_lower.split())
        title_tokens = set(p_title.split())
        overlap = len(query_tokens & title_tokens)
        if overlap > best_score:
            best_score = overlap
            best = p
        # 子串匹配加分
        if title_lower[:30] in p_title or p_title[:30] in title_lower:
            best_score = max(best_score, len(title_lower) // 4)

    if best and best_score >= 2:
        return {
            "title": best.get("title", "?")[:200],
            "first_author": best.get("first_author", "?"),
            "year": best.get("year", "?"),
            "journal": best.get("journal", "?"),
            "research_questions": best.get("research_questions", "")[:300],
            "theories": best.get("theories", "")[:200],
            "variables": best.get("variables", "")[:200],
            "methods": best.get("methods", "")[:300],
            "key_findings": best.get("key_findings", "")[:400],
            "limitations": best.get("limitations", "")[:200],
            "future_directions": best.get("future_directions", "")[:200],
        }
    return None


def _tool_find_theories(topic: str) -> list[dict]:
    """从本地文献中找出与某个主题相关的理论框架。遍历全库所有论文，不限制返回数量。"""
    global _qa_paper_index
    if not _qa_paper_index:
        _build_paper_index()

    results = []
    topic_lower = topic.lower()
    for p in _qa_paper_index:
        text = " ".join([
            p.get("title", ""), p.get("abstract", ""),
            p.get("research_questions", ""), p.get("discussion", ""),
        ]).lower()
        if topic_lower in text or any(tok in text for tok in topic_lower.split() if len(tok) > 2):
            theories = p.get("theories", "")
            if theories and len(theories) > 5:
                results.append({
                    "paper": f"{p.get('first_author','?')} ({p.get('year','?')})",
                    "title": p.get("title", "?")[:120],
                    "theories": theories[:300],
                })
    # 按年份降序排列，返回全部结果
    results.sort(key=lambda r: r["paper"], reverse=True)
    return results  # 不再截断，返回全部


def _tool_find_gaps(topic: str) -> list[dict]:
    """从本地文献的未来研究方向中提取与某个主题相关的研究空白。遍历全库，不限制返回数量。"""
    global _qa_paper_index
    if not _qa_paper_index:
        _build_paper_index()

    results = []
    topic_lower = topic.lower()
    for p in _qa_paper_index:
        gaps = p.get("future_directions", "")
        if gaps and len(gaps) > 20:
            if topic_lower in gaps.lower() or any(tok in gaps.lower() for tok in topic_lower.split() if len(tok) > 2):
                results.append({
                    "paper": f"{p.get('first_author','?')} ({p.get('year','?')})",
                    "gaps": gaps[:300],
                })
    results.sort(key=lambda r: r["paper"], reverse=True)
    return results  # 不再截断，返回全部


def _tool_extract_methods(papers: list[str]) -> list[dict]:
    """提取指定论文的研究方法。"""
    global _qa_paper_index
    if not _qa_paper_index:
        _build_paper_index()

    results = []
    for query_title in papers[:5]:
        best = _tool_read_paper(query_title)
        if best:
            results.append({
                "paper": f"{best['first_author']} ({best['year']})",
                "title": best["title"][:120],
                "methods": best.get("methods", "未找到方法信息")[:300],
            })
    return results


async def _tool_search_external(question: str) -> Optional[str]:
    """调用 AI 的外部知识。"""
    msgs = [
        {"role": "system", "content": "你在市场营销×AI营销交叉领域有深厚的学术积累。基于你的知识回答问题。中文，专业术语保留英文。如果不知道就诚实说不知道。"},
        {"role": "user", "content": question},
    ]
    return await _chat_via_deepseek(msgs)


# ═══ Agent Core: Plan → Execute → Synthesize ═══

async def agent_plan(goal: str, chat_history: list[dict]) -> Optional[list[dict]]:
    """用 DeepSeek 将用户目标拆解为可执行的工具调用序列。"""
    tool_descriptions = _agent_get_tool_descriptions()
    system_prompt = AGENT_PLANNER_SYSTEM.format(tool_descriptions=tool_descriptions)

    # 包含最近对话上下文
    context = ""
    if chat_history:
        recent = chat_history[-6:]
        context = "\n最近对话：\n" + "\n".join(
            f"{'用户' if m['role']=='user' else '系统'}: {m['content'][:200]}"
            for m in recent
        )

    user_prompt = f"研究目标：{goal}{context}\n\n请输出执行计划（JSON数组）："

    msgs = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    try:
        raw = await _chat_via_deepseek(msgs)
        if not raw:
            return None
        # 提取 JSON 数组
        json_match = re.search(r'\[.*\]', raw, re.DOTALL)
        if not json_match:
            return None
        plan = json.loads(json_match.group())
        if not isinstance(plan, list):
            return None
        # 验证每个步骤的 tool 是否合法
        valid_plan = []
        for step in plan:
            if isinstance(step, dict) and step.get("tool") in AGENT_TOOLS:
                valid_plan.append(step)
        return valid_plan if valid_plan else None
    except Exception as e:
        logger.warning(f"[Agent] 规划失败: {e}")
        return None


def _execute_tool_sync(tool_name: str, args: dict) -> dict:
    """同步执行单个工具调用，返回结构化结果。"""
    tool_info = AGENT_TOOLS.get(tool_name)
    if not tool_info:
        return {"error": f"未知工具: {tool_name}"}

    try:
        func = globals().get(tool_info["function"])
        if not func:
            return {"error": f"工具未实现: {tool_info['function']}"}

        if tool_name == "search_external":
            # 异步工具需要特殊处理
            return {"error": "search_external 需异步调用，请使用 async 版本"}
        else:
            result = func(**args)
            return {"tool": tool_name, "args": args, "result": result, "ok": True}
    except Exception as e:
        return {"tool": tool_name, "args": args, "error": str(e), "ok": False}


async def agent_execute_plan(plan: list[dict], progress_callback=None) -> list[dict]:
    """执行规划好的工具调用序列。"""
    results = []
    for i, step in enumerate(plan):
        tool_name = step["tool"]
        args = step.get("args", {})
        reason = step.get("reason", "")

        if progress_callback:
            progress_callback(i, len(plan), tool_name, args, reason, "running")

        if tool_name == "search_external":
            # 异步执行
            try:
                answer = await _tool_search_external(args.get("question", ""))
                result = {"tool": tool_name, "args": args, "result": answer, "ok": True}
            except Exception as e:
                result = {"tool": tool_name, "args": args, "error": str(e), "ok": False}
        else:
            # 同步工具在 executor 中运行（避免阻塞）
            import concurrent.futures
            loop = asyncio.get_event_loop()
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
                result = await loop.run_in_executor(
                    pool, _execute_tool_sync, tool_name, args
                )

        if progress_callback:
            status = "done" if result.get("ok") else "error"
            progress_callback(i, len(plan), tool_name, args, reason, status)

        results.append(result)
    return results


async def agent_synthesize(goal: str, plan: list[dict], step_results: list[dict]) -> Optional[str]:
    """将所有步骤的结果综合为最终回答。"""
    # 构建步骤结果摘要
    summary_parts = []
    for i, (step, result) in enumerate(zip(plan, step_results)):
        tool = step["tool"]
        reason = step.get("reason", "")
        if result.get("ok") and result.get("result"):
            r = result["result"]
            if isinstance(r, list):
                summary_parts.append(f"步骤{i+1} [{tool}] ({reason}): 返回了 {len(r)} 条结果\n" +
                    "\n".join(f"  - {json.dumps(item, ensure_ascii=False)[:300]}" for item in r[:5]))
            elif isinstance(r, dict):
                summary_parts.append(f"步骤{i+1} [{tool}] ({reason}):\n  {json.dumps(r, ensure_ascii=False)[:800]}")
            elif isinstance(r, str):
                summary_parts.append(f"步骤{i+1} [{tool}] ({reason}):\n  {r[:800]}")
        elif result.get("error"):
            summary_parts.append(f"步骤{i+1} [{tool}] ({reason}): ❌ {result['error']}")
        else:
            summary_parts.append(f"步骤{i+1} [{tool}] ({reason}): 无结果")

    steps_summary = "\n\n".join(summary_parts)

    synth_prompt = f"""你是一位在市场营销×AI营销交叉领域有深厚学术积累的研究者。请根据以下分步骤检索到的信息，综合回答用户的研究目标。

研究目标：{goal}

执行步骤与结果：
{steps_summary}

请提供一份结构清晰、学术严谨的综合回答：
- 如果是文献综述类：按主题组织，指出共识与分歧，最后给出整体评述
- 如果是理论分析类：讲清概念、核心假设、实证证据、适用边界
- 如果是方法比较类：比较优劣和适用场景
- 如果步骤结果不足以全面回答，诚实说明并提供进一步建议
- 标注信息来源：本地文献用📄，外部知识用🌐
- 中文回答，专业术语保留英文"""

    msgs = [
        {"role": "system", "content": "你是一位学术研究者，综合检索结果给出严谨的回答。中文，专业术语保留英文。"},
        {"role": "user", "content": synth_prompt},
    ]
    return await _chat_via_deepseek(msgs)


async def flow_qa():
    """选项 [8]：知识库问答 — 教授级学术AI导师。"""
    global _qa_paper_index, _qa_chat_history, _qa_last_papers
    global _qa_use_external, _qa_top_n

    # 构建索引
    if not _qa_paper_index:
        print("\n[QA] 正在构建知识库索引...")
        _build_paper_index()
        print(f"[OK] 已索引 {len(_qa_paper_index)} 篇文献笔记\n")

    # 尝试恢复上次对话历史
    _qa_chat_history = []
    if _QA_HISTORY_FILE.exists():
        try:
            with open(_QA_HISTORY_FILE, "r", encoding="utf-8") as f:
                loaded = json.load(f)
                if isinstance(loaded, list):
                    _qa_chat_history = loaded
                    rounds = len(_qa_chat_history) // 2
                    if rounds > 0:
                        print(f"\n[v] 已恢复上次对话 ({rounds} 轮)。输入 /clear 清空历史。")
        except Exception:
            pass

    print("=" * 60)
    print("  知识库问答 — 教授级学术AI导师")
    print(f"  本地文献: {len(_qa_paper_index)} 篇 | 外部知识: {'开启' if _qa_use_external else '关闭'}")
    print("=" * 60)
    print("  直接输入问题进行提问。输入 /help 查看帮助。")
    print("=" * 60)

    while True:
        # 读取用户输入
        try:
            question = input("\n[Q] > ").strip()
        except (EOFError, KeyboardInterrupt):
            _save_qa_history()
            print("\n已退出问答模式。")
            break

        if not question:
            continue

        # ── 特殊命令处理 ──
        if question.startswith("/"):
            cmd = question.split()[0].lower()
            if cmd in ("/exit", "/quit"):
                _save_qa_history()
                print("已退出问答模式。")
                break
            elif cmd == "/help":
                _print_qa_help()
                continue
            elif cmd == "/clear":
                _qa_chat_history = []
                _qa_last_papers = []
                _save_qa_history()
                print("[v] 对话历史已清空。")
                continue
            elif cmd == "/papers":
                if _qa_last_papers:
                    print(f"\n上一轮检索到的本地文献 ({len(_qa_last_papers)} 篇):")
                    for i, p in enumerate(_qa_last_papers, 1):
                        print(f"  [{i}] {p['title'][:70]}")
                        print(f"      {p.get('first_author', '?')} ({p.get('year', '?')}) | "
                              f"{p.get('journal', '?')[:40]} | IF={p.get('impact_factor', '?')}")
                else:
                    print("尚无检索记录。")
                continue
            elif cmd == "/external":
                _qa_use_external = not _qa_use_external
                state = "开启" if _qa_use_external else "关闭"
                print(f"[v] 外部知识已{state}。")
                if not _qa_use_external:
                    print("  (将仅基于本地文献回答，不足时诚实告知)")
                continue
            elif cmd == "/save":
                _save_qa_chat(_qa_chat_history)
                continue
            elif cmd == "/n":
                parts = question.split()
                if len(parts) >= 2:
                    try:
                        n = int(parts[1])
                        if 3 <= n <= 15:
                            _qa_top_n = n
                            print(f"[v] 检索数量已设置为 {n} 篇。")
                        else:
                            print("检索数量范围: 3-15")
                    except ValueError:
                        print("用法: /n <数字>，如 /n 10")
                else:
                    print("用法: /n <数字>，如 /n 10")
                continue
            else:
                print(f"未知命令: {cmd}。输入 /help 查看可用命令。")
                continue

        # ── 检索本地文献 ──
        relevant = _retrieve_papers(question)
        _qa_last_papers = relevant

        if relevant:
            print(f"\n[检索] 匹配到 {len(relevant)} 篇本地文献:")
            for i, p in enumerate(relevant, 1):
                print(f"  [{i}] {p['title'][:70]}")
                print(f"      {p.get('first_author', '?')} ({p.get('year', '?')}) | "
                      f"{p.get('journal', '?')[:40]}")
        else:
            print("\n[检索] 本地文献中未找到直接匹配的论文。")
            if _qa_use_external:
                print("   将以外部知识回答，建议后续补充相关文献。")

        # ── 构建 messages 并调用 DeepSeek ──
        if not _qa_use_external and not relevant:
            print("\n[DeepSeek] 本地文献无匹配，且外部知识已关闭。请尝试换个问题或开启外部知识 (/external)。")
            continue

        print("\n[DeepSeek] 正在调用 DeepSeek V4 Pro 生成回答...")
        messages = _build_qa_messages(question, relevant, _qa_chat_history)
        answer = await _chat_via_deepseek(messages)

        if not answer:
            print("\n[ERROR] DeepSeek API 调用失败，请检查网络和 API Key 配置。可重试。")
            continue

        # ── 显示回答 ──
        # 清洗 GBK 无法编码的字符（上下标、emoji 等）
        def _safe_print(text: str) -> None:
            try:
                print(text)
            except UnicodeEncodeError:
                # 逐字符打印，跳过无法编码的
                for ch in text:
                    try:
                        print(ch, end="")
                    except UnicodeEncodeError:
                        print("?", end="")
                print()

        print("\n" + "─" * 50)
        _safe_print(answer)
        print("─" * 50)

        # ── 引用验证：检查回答中对本地文献的归因是否准确 ──
        if relevant:
            warnings = _verify_answer_claims(answer, relevant, _qa_paper_index)
            if warnings:
                print("\n[验证] 以下引用需要注意：")
                for w in warnings:
                    print(f"  {w}")

        # ── 更新对话历史 ──
        _qa_chat_history.append({"role": "user", "content": question})
        _qa_chat_history.append({"role": "assistant", "content": answer})

        # 限制历史长度（最近 10 轮 = 20 条）
        if len(_qa_chat_history) > 20:
            _qa_chat_history = _qa_chat_history[-20:]
        _save_qa_history()


# ============================================================================
# 入口
# ============================================================================
async def main():
    global _non_interactive
    args = sys.argv[1:]
    if "--yes" in args or "-y" in args:
        _non_interactive = True
        args = [a for a in args if a not in ("--yes", "-y")]
        logger.info("非交互模式已启用 (--yes)：将自动确认所有提示，Phase 2 使用 30 秒倒计时。")

    # ── 命令行参数模式（兼容旧用法 / 脚本调用）──
    if args:
        if args[0] == "--process-existing":
            await process_existing()
        elif args[0] == "--cleanup":
            await cleanup_processed_pdfs()
        elif args[0] == "--dry-run":
            await main_flow(dry_run=True)
        elif args[0] == "--scrape-only":
            await scrape_only_flow()
        elif args[0] == "--read-only":
            await deep_read_only_flow()
        elif args[0] == "--入库":
            await 入库_only_flow()
        elif args[0] == "--re-read":
            await re_read_from_zotero_flow()
        elif args[0] == "--qa":
            await flow_qa()
        else:
            print("未知参数。可用: --dry-run, --scrape-only, --read-only, --入库, --process-existing, --cleanup, --re-read, --qa [--yes]")
        return

    # ── 交互式菜单模式 ──
    MENU = """
╔══════════════════════════════════════════════════╗
║     SPIS 半自动文献处理系统                        ║
╠══════════════════════════════════════════════════╣
║  [1] 检索新文献         SPIS → 待处理清单          ║
║  [2] 深度阅读 + 入库     PDF → DeepSeek → Obsidian ║
║                          + Excel + Zotero (主力)   ║
║  [3] 清理已处理的 PDF                             ║
║  [4] 批量重读           Zotero集合 → 覆写笔记       ║
║                          (Prompt升级时使用)         ║
║  [5] 知识库问答          教授级学术AI导师            ║
║  [0] 退出                                        ║
╠══════════════════════════════════════════════════╣
║  键盘控制 (检索/补全过程中):                       ║
║    P = 暂停    G = 继续    E = 终止               ║
╚══════════════════════════════════════════════════╝"""

    while True:
        print(MENU)
        try:
            choice = input("请选择运行模式 [0-5]: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n再见。")
            break

        if choice == "1":
            print("\n>>> 启动 [1] 检索新文献\n")
            await scrape_only_flow()
        elif choice == "2":
            print("\n>>> 启动 [2] 深度阅读 + 入库 (Obsidian + Excel + Zotero)\n")
            await deep_read_only_flow()
            # 深度阅读完成后，询问是否继续 Zotero 入库
            if not _non_interactive:
                do_zotero = input("\n是否继续 Zotero 入库？(y/n): ").strip().lower()
                if do_zotero == 'y':
                    print()
                    await 入库_only_flow()
                else:
                    print("跳过 Zotero 入库。可稍后用 [3] 手动入库 → 实际已合并到 [2]，请直接重新运行 [2] 或使用 --入库。")
            else:
                await 入库_only_flow()
        elif choice == "3":
            print("\n>>> 启动 [3] 清理已处理的 PDF\n")
            await cleanup_processed_pdfs()
        elif choice == "4":
            print("\n>>> 启动 [4] 批量重读（Zotero 集合 → 覆写笔记）\n")
            await re_read_from_zotero_flow()
        elif choice == "5":
            print("\n>>> 启动 [5] 知识库问答\n")
            await flow_qa()
        elif choice == "0":
            print("再见。")
            break
        else:
            print("无效选项，请输入 0-5。")

if __name__ == "__main__":
    asyncio.run(main())
